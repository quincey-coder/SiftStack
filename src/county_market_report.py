"""County-agnostic 7-sheet market research workbook generator.

Merges an authoritative Sift Market Finder extraction (ZIP + neighborhood
investor data) with a flat per-county public-data bundle (FRED, Census, BLS,
crime, econ-dev, distress, metro housing). Sheets 4 (Economic), 5 (Crime),
6 (Recommendations) are FULLY populated, not placeholders.

Importable: generate_report(...). Also has a CLI for single-county use.

Flat bundle shape expected:
{
  "fred": {"fred_national_dom_days": int},
  "metro_redfin": {median_sale_price, *_yoy_pct, median_ppsf, ppsf_yoy_pct,
                   homes_sold, homes_sold_yoy_pct, median_dom, dom_yoy_change_days,
                   sale_to_list_pct, pct_above_list, pct_price_drops, observation_period},
  "demographics": {...}, "employment": {...}, "crime": {...},
  "econ_dev": {...}, "distress": {...},
  "synthesis": {"market_assessment":[...]}   # optional
}
"""

import argparse
import json
import statistics
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SECTION_FONT = Font(bold=True, size=13, color="2F5496")
SUB_FONT = Font(bold=True, size=11, color="333333")
LABEL_FONT = Font(size=11, color="555555")
THIN = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
              top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PCT1 = '0.0%'


def hrow(ws, r, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN


def border_row(ws, r, n):
    for c in range(1, n + 1):
        ws.cell(row=r, column=c).border = THIN


def autofit(ws, min_w=12, max_w=46):
    for col in ws.columns:
        mx, letter = 0, None
        for cell in col:
            if letter is None:
                try:
                    letter = get_column_letter(cell.column)
                except Exception:
                    pass
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        if letter:
            ws.column_dimensions[letter].width = max(min_w, min(mx + 2, max_w))


def money(v):
    return f"${v:,.0f}" if isinstance(v, (int, float)) and v else (v or "")


def fmt_int(v):
    return f"{v:,}" if isinstance(v, (int, float)) else (v or "N/A")


def pct(v):
    return f"{v}%" if isinstance(v, (int, float)) else (v or "")


def supply_months(hom, sold):
    return round(hom / sold, 1) if sold else None


def spread_pct(sale, value):
    return round((sale - value) / value, 4) if value else None


def wholesaling_score(inv, dom, value, fred):
    diff = (dom - fred) if (dom and fred) else 0
    if inv >= 30 and diff <= -15 and value and value < 350000:
        return "★★★★★"
    if inv >= 20 and diff < 0 and value and value < 400000:
        return "★★★★☆"
    if inv >= 30 and diff < 0:
        return "★★★★☆"
    if inv >= 10 and diff <= 0:
        return "★★★☆☆"
    if inv >= 5:
        return "★★☆☆☆"
    return "★☆☆☆☆"


def prettify(k):
    import re
    k = str(k).strip()
    m = re.match(r'^(\d[\d,]*)mo$', k)
    if m:
        return "$" + format(int(m.group(1).replace(",", "")), ",") + "/mo"
    if re.match(r'^\d{2,4}$', k):
        return f"{int(k) / 100:.1f}%"
    return k


def summ_lookup(summ, label, default):
    for k, v in (summ or {}).items():
        if isinstance(v, str) and v.strip().lower() == label.lower():
            return prettify(k)
        if isinstance(k, str) and k.strip().lower() == label.lower():
            return v
    return default


def homeownership(summ, demo):
    if demo.get("homeownership_rate_pct"):
        return f"{demo['homeownership_rate_pct']}%"
    for k, v in (summ or {}).items():
        if isinstance(v, str) and v.lower().startswith("renters"):
            return f"Sift: {v}"
    return "N/A"


def derive_assessment(inv6, hom, sold, zips, fred, bundle):
    """Three core ratings from authoritative Sift data + qualitative ratings
    derived from the public bundle (overridden by synthesis if present)."""
    monthly = inv6 / 6
    inv_rating = "HIGH" if inv6 >= 500 else ("MODERATE" if inv6 >= 100 else "LOW")
    sup = supply_months(hom, sold) or 0
    sup_rating = "HIGH" if sup < 4 else ("MODERATE" if sup <= 6 else "LOW")
    doms = [x.get("median_days_on_market") for x in zips if x.get("median_days_on_market")]
    med_dom = statistics.median(doms) if doms else 0
    vel = "HIGH" if med_dom < fred else ("MODERATE" if med_dom <= fred * 1.5 else "LOW")
    out = [
        {"category": "Investor Activity", "rating": inv_rating,
         "commentary": f"{round(monthly)} monthly investor transactions ({inv6:,} over 6 mo)"},
        {"category": "Market Velocity", "rating": vel,
         "commentary": f"County median DOM ~{med_dom} vs FRED national {fred}"},
        {"category": "Supply Tightness", "rating": sup_rating,
         "commentary": f"{sup} months of inventory ({hom:,} on market / {sold:,} sold)"},
    ]
    # Qualitative ratings derived from the bundle
    redfin = bundle.get("metro_redfin") or {}
    demo = bundle.get("demographics") or {}
    emp = bundle.get("employment") or {}
    appr = redfin.get("median_sale_price_yoy_pct")
    if isinstance(appr, (int, float)):
        ar = "STRONG" if appr >= 5 else ("STABLE" if appr >= 0 else "DECLINING")
        out.append({"category": "Price Appreciation", "rating": ar,
                    "commentary": f"County median sale price {appr:+}% YoY (Redfin)"})
    g = demo.get("annual_growth_rate_pct")
    if g is None and isinstance(demo.get("population_growth_since_2020_pct"), (int, float)):
        g = round(demo["population_growth_since_2020_pct"] / 5, 2)
    if isinstance(g, (int, float)):
        gr = "HIGH" if g >= 1.0 else ("MODERATE" if g >= 0.2 else "LOW")
        out.append({"category": "Population Growth", "rating": gr,
                    "commentary": f"~{g}%/yr ({demo.get('population_growth_since_2020_pct','?')}% since 2020)"})
    u = emp.get("unemployment_rate_pct")
    if isinstance(u, (int, float)):
        er = "STRONG" if u <= 4.0 else ("STABLE" if u <= 5.0 else "SOFT")
        ny = emp.get("nonfarm_yoy_pct")
        ntxt = f", nonfarm {ny:+}% YoY" if isinstance(ny, (int, float)) else ""
        out.append({"category": "Employment", "rating": er,
                    "commentary": f"{emp.get('msa_name','MSA')} unemployment {u}%{ntxt}"})
    return out


# ── Sheet builders ────────────────────────────────────────────────────

def _zip_or_nbr_sheet(wb, title, rows, key, fred):
    ws = wb.create_sheet(title)
    label = "ZIP Code" if key == "zip_code" else "Neighborhood"
    hrow(ws, 1, [label, "6-Mo Inv Trans", "Homes on Market", "Homes Sold/Mo", "Median DOM",
                 "DOM vs National", "Median Home Value", "Median Sale Price", "Spread %",
                 "Supply Months", "Wholesaling Score"])
    ws.freeze_panes = "A2"
    r = 2
    for x in sorted(rows, key=lambda v: v.get("total_inv_trans_6mo") or 0, reverse=True):
        inv = x.get("total_inv_trans_6mo") or 0
        dom = x.get("median_days_on_market") or 0
        hom = x.get("homes_on_market") or 0
        sold = x.get("homes_sold_last_month") or 0
        value = x.get("median_home_value") or 0
        sale = x.get("median_sale_price") or 0
        diff = (dom - fred) if dom else None
        spread = spread_pct(sale, value)
        sup = supply_months(hom, sold)
        ws.cell(row=r, column=1, value=x.get(key))
        ws.cell(row=r, column=2, value=inv)
        ws.cell(row=r, column=3, value=hom)
        ws.cell(row=r, column=4, value=sold)
        ws.cell(row=r, column=5, value=dom)
        dc = ws.cell(row=r, column=6, value=diff)
        if diff is not None:
            dc.fill = GREEN if diff < 0 else (YELLOW if diff <= fred * 0.15 else RED)
        ws.cell(row=r, column=7, value=x.get("median_home_value_raw"))
        ws.cell(row=r, column=8, value=x.get("median_sale_price_raw"))
        sc = ws.cell(row=r, column=9)
        if spread is not None:
            sc.value = spread
            sc.number_format = PCT1
            if spread < 0:
                sc.fill = GREEN
        spc = ws.cell(row=r, column=10, value=sup)
        if sup is not None:
            spc.fill = GREEN if sup < 3 else (YELLOW if sup <= 6 else RED)
        ws.cell(row=r, column=11, value=wholesaling_score(inv, dom, value, fred))
        if inv >= 60:
            for c in range(1, 12):
                ws.cell(row=r, column=c).font = Font(bold=True)
        border_row(ws, r, 11)
        r += 1
    autofit(ws)


def sheet_exec(wb, sift, bundle, fred, county, st):
    ws = wb.active
    ws.title = "Executive Summary"
    z = sift["zip_data"]
    zact = [x for x in z if (x.get("total_inv_trans_6mo") or 0) > 0]
    n = sift["neighborhood_data"]
    hom = sum(x.get("homes_on_market") or 0 for x in z)
    sold = sum(x.get("homes_sold_last_month") or 0 for x in z)
    inv6 = sum(x.get("total_inv_trans_6mo") or 0 for x in z)
    vals = [x["median_home_value"] for x in zact if x.get("median_home_value")]
    med_val = statistics.median(vals) if vals else 0
    summ = sift.get("summary", {})
    redfin = bundle.get("metro_redfin") or {}
    demo = bundle.get("demographics") or {}
    emp = bundle.get("employment") or {}

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"{county.upper()} COUNTY, {st} - MARKET RESEARCH REPORT"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now():%B %d, %Y}"
    ws["A3"] = "Data Source: REI Sift Market Finder (proprietary investor data) + FRED, U.S. Census, BLS, Redfin, FBI/Sheriff, ATTOM"

    row = 5
    ws.cell(row=row, column=1, value="SECTION A - COUNTY OVERVIEW").font = SECTION_FONT
    row += 1
    hrow(ws, row, ["Metric", "Value", "Notes"])
    row += 1
    overview = [
        ("Population", fmt_int(demo.get("population_latest")),
         f"+{demo.get('population_growth_since_2020_pct')}% since 2020" if demo.get("population_growth_since_2020_pct") is not None else "Census"),
        ("Median Home Value", money(med_val), "Sift (median of ZIP medians)"),
        ("Median Household Income", money(demo.get("median_household_income")), "Census ACS"),
        ("Unemployment Rate", pct(emp.get("unemployment_rate_pct")),
         f"BLS {emp.get('unemployment_obs_month','')} ({emp.get('msa_name','MSA')})"),
        ("Homes on Market", f"{hom:,}", "Sift (sum of ZIPs)"),
        ("Monthly Investor Transactions", f"{round(inv6/6):,}", f"6-mo total {inv6:,} / 6"),
        ("Homes Sold Last Month", f"{sold:,}", "Sift (sum of ZIPs)"),
        ("Market Rent", summ_lookup(summ, "Market Rent", "N/A"), "Sift"),
        ("Gross Rental Yield", summ_lookup(summ, "Gross Rental Yield", "N/A"), "Sift"),
        ("Homeownership Rate", homeownership(summ, demo), "Sift / Census"),
        ("Months of Supply", f"{supply_months(hom, sold)}", "Homes on Market / Homes Sold (balanced=3-6)"),
        ("National Days on Market", f"{fred} days", f"FRED MEDDAYONMARUS {datetime.now():%b %Y}"),
        ("County Median Sale Price (Redfin)", money(redfin.get("median_sale_price")),
         f"{redfin.get('median_sale_price_yoy_pct')}% YoY" if redfin.get("median_sale_price_yoy_pct") is not None else "Redfin"),
    ]
    for metric, value, notes in overview:
        ws.cell(row=row, column=1, value=metric).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value if value not in (None, "") else "N/A").font = Font(bold=True, size=11)
        ws.cell(row=row, column=3, value=notes)
        border_row(ws, row, 3)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="SECTION B - MARKET ASSESSMENT").font = SECTION_FONT
    row += 1
    hrow(ws, row, ["Category", "Rating", "Commentary"])
    row += 1
    syn_ma = (bundle.get("synthesis") or {}).get("market_assessment")
    assessment = syn_ma if syn_ma else derive_assessment(inv6, hom, sold, z, fred, bundle)
    for item in assessment:
        rt = (item.get("rating") or "").upper()
        ws.cell(row=row, column=1, value=item.get("category"))
        rc = ws.cell(row=row, column=2, value=item.get("rating"))
        if rt in ("HIGH", "STRONG", "IMPROVING"):
            rc.fill = GREEN
        elif rt in ("MODERATE", "STABLE", "BALANCED"):
            rc.fill = YELLOW
        elif rt in ("LOW", "DECLINING", "WEAK", "SOFT"):
            rc.fill = RED
        ws.cell(row=row, column=3, value=item.get("commentary"))
        border_row(ws, row, 3)
        row += 1

    for ttl, src, kk in [("SECTION C - TOP 5 ZIP CODES FOR WHOLESALING", z, "zip_code"),
                         ("SECTION D - TOP 5 NEIGHBORHOODS FOR WHOLESALING", n, "neighborhood")]:
        row += 1
        ws.cell(row=row, column=1, value=ttl).font = SECTION_FONT
        row += 1
        hdr = "ZIP Code" if kk == "zip_code" else "Neighborhood"
        hrow(ws, row, ["Rank", hdr, "6-Mo Inv Trans", "Median Home Value", "Days on Market"])
        row += 1
        for i, x in enumerate(sorted(src, key=lambda v: v.get("total_inv_trans_6mo") or 0, reverse=True)[:5], 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=x.get(kk))
            ws.cell(row=row, column=3, value=x.get("total_inv_trans_6mo"))
            ws.cell(row=row, column=4, value=x.get("median_home_value_raw"))
            ws.cell(row=row, column=5, value=x.get("median_days_on_market"))
            border_row(ws, row, 5)
            row += 1
    autofit(ws)


def sheet_econ(wb, bundle, county, st):
    ws = wb.create_sheet("Economic Indicators")
    emp = bundle.get("employment") or {}
    demo = bundle.get("demographics") or {}
    redfin = bundle.get("metro_redfin") or {}
    ws["A1"] = f"ECONOMIC INDICATORS - {county.upper()} COUNTY, {st}"
    ws["A1"].font = SECTION_FONT
    r = 3
    ws.cell(row=r, column=1, value=f"Section A - Employment (BLS, {emp.get('msa_name','MSA')})").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Metric", "Value", "Notes"])
    r += 1
    for m, v, nt in [
        ("MSA", emp.get("msa_name", "N/A"), emp.get("counties_covered", "")),
        ("Civilian Labor Force", fmt_int(emp.get("civilian_labor_force")), ""),
        ("Unemployment Rate", pct(emp.get("unemployment_rate_pct")), emp.get("unemployment_obs_month", "")),
        ("Total Nonfarm Jobs", fmt_int(emp.get("total_nonfarm_jobs")),
         (pct(emp.get("nonfarm_yoy_pct")) + " YoY") if emp.get("nonfarm_yoy_pct") is not None else ""),
    ]:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=nt)
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Employment by Sector").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Sector", "Jobs (000s)", "12-Mo Change %"])
    r += 1
    for s in emp.get("sectors") or []:
        ws.cell(row=r, column=1, value=s.get("sector"))
        ws.cell(row=r, column=2, value=s.get("jobs_000s"))
        cc = ws.cell(row=r, column=3, value=pct(s.get("change_12mo_pct")))
        if isinstance(s.get("change_12mo_pct"), (int, float)):
            cc.fill = GREEN if s["change_12mo_pct"] > 0 else (RED if s["change_12mo_pct"] < 0 else YELLOW)
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Major Employers (region)").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Employer", "Sector", "Approx. Employees"])
    r += 1
    for e in emp.get("major_employers") or []:
        ws.cell(row=r, column=1, value=e.get("employer"))
        ws.cell(row=r, column=2, value=e.get("sector"))
        ws.cell(row=r, column=3, value=str(e.get("approx_employees") or ""))
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section B - Demographics (Census)").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Metric", "Value", "Notes"])
    r += 1
    for m, v, nt in [
        ("Population", fmt_int(demo.get("population_latest")), demo.get("population_year", "")),
        ("Growth Since 2020", pct(demo.get("population_growth_since_2020_pct")), ""),
        ("Annual Growth Rate", pct(demo.get("annual_growth_rate_pct")), ""),
        ("Median Age", demo.get("median_age"), ""),
        ("Median Household Income", money(demo.get("median_household_income")), ""),
        ("Per Capita Income", money(demo.get("per_capita_income")), ""),
        ("Poverty Rate", pct(demo.get("poverty_rate_pct")), ""),
        ("Bachelor's Degree or Higher", pct(demo.get("bachelors_or_higher_pct")), ""),
        ("Owner-Occupied Housing", pct(demo.get("owner_occupied_pct")), ""),
        ("Homeownership Rate", pct(demo.get("homeownership_rate_pct")),
         f"Renters {demo.get('renter_pct')}%" if demo.get("renter_pct") else ""),
        ("Median Home Value (Census)", money(demo.get("median_home_value_census")), "ACS estimate"),
    ]:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=v if v not in (None, "") else "N/A")
        ws.cell(row=r, column=3, value=nt)
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section C - Housing Market (Redfin county)").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Metric", "Value", "YoY Change"])
    r += 1
    for m, v, nt in [
        ("Median Sale Price", money(redfin.get("median_sale_price")), pct(redfin.get("median_sale_price_yoy_pct"))),
        ("Median Price / Sq Ft", money(redfin.get("median_ppsf")), pct(redfin.get("ppsf_yoy_pct"))),
        ("Homes Sold", fmt_int(redfin.get("homes_sold")), pct(redfin.get("homes_sold_yoy_pct"))),
        ("Median Days on Market", redfin.get("median_dom"),
         f"{redfin.get('dom_yoy_change_days')} days" if redfin.get("dom_yoy_change_days") is not None else ""),
        ("Sale-to-List Price", pct(redfin.get("sale_to_list_pct")), ""),
        ("Homes Above List Price", pct(redfin.get("pct_above_list")), ""),
        ("Homes with Price Drops", pct(redfin.get("pct_price_drops")), ""),
    ]:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=v if v not in (None, "") else "N/A")
        ws.cell(row=r, column=3, value=nt)
        border_row(ws, r, 3)
        r += 1
    autofit(ws)


def sheet_crime(wb, bundle, county, st):
    ws = wb.create_sheet("Crime & Safety")
    crime = bundle.get("crime") or {}
    ws["A1"] = f"CRIME & SAFETY - {county.upper()} COUNTY, {st}"
    ws["A1"].font = SECTION_FONT
    if crime.get("overall_trend"):
        ws["A2"] = f"Overall trend: {crime['overall_trend']}"
    r = 4
    ws.cell(row=r, column=1, value="Section A - Crime Statistics").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Crime Type", "Prior Year", "Current Year", "Change %"])
    r += 1
    for c in crime.get("crime_stats") or []:
        ws.cell(row=r, column=1, value=c.get("crime_type"))
        ws.cell(row=r, column=2, value=c.get("prior_year"))
        ws.cell(row=r, column=3, value=c.get("current_year"))
        ch = ws.cell(row=r, column=4, value=pct(c.get("change_pct")))
        if isinstance(c.get("change_pct"), (int, float)):
            ch.fill = GREEN if c["change_pct"] < 0 else (RED if c["change_pct"] > 0 else YELLOW)
        border_row(ws, r, 4)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section B - Violent / Homicide Trend").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Year", "Value", "Metric"])
    r += 1
    for m in crime.get("violent_trend") or []:
        ws.cell(row=r, column=1, value=m.get("year"))
        ws.cell(row=r, column=2, value=m.get("value"))
        ws.cell(row=r, column=3, value=m.get("metric"))
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section C - Safety Assessment by Area").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Area", "Safety Rating", "Notes"])
    r += 1
    for a in crime.get("area_safety") or []:
        ws.cell(row=r, column=1, value=a.get("area"))
        rc = ws.cell(row=r, column=2, value=a.get("safety_rating"))
        rt = (a.get("safety_rating") or "").upper()
        if "HIGH" in rt:
            rc.fill = GREEN
        elif "MOD" in rt:
            rc.fill = YELLOW
        elif "LOW" in rt:
            rc.fill = RED
        ws.cell(row=r, column=3, value=a.get("notes"))
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section D - Key Insights for Investors").font = SUB_FONT
    r += 1
    for ins in crime.get("investor_insights") or []:
        ws.cell(row=r, column=1, value=f"- {ins}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    autofit(ws)


def sheet_reco(wb, sift, bundle, fred, county, st):
    ws = wb.create_sheet("Investment Recommendations")
    z = sift["zip_data"]
    n = sift["neighborhood_data"]
    syn = bundle.get("synthesis") or {}
    ws["A1"] = f"INVESTMENT RECOMMENDATIONS - {county.upper()} COUNTY, {st}"
    ws["A1"].font = SECTION_FONT
    zsorted = sorted(z, key=lambda v: v.get("total_inv_trans_6mo") or 0, reverse=True)

    def rationale(x):
        inv = x.get("total_inv_trans_6mo") or 0
        dom = x.get("median_days_on_market") or 0
        value = x.get("median_home_value") or 0
        dl = "below-avg DOM" if dom and dom < fred else "at/above-avg DOM"
        price = "low price point" if value and value < 200000 else ("moderate prices" if value and value < 350000 else "higher prices")
        return f"{inv} inv trans (6mo), {dl}, {price}"

    r = 3
    ws.cell(row=r, column=1, value="Section A - Tier 1: Highest-Priority ZIP Codes").font = SUB_FONT
    r += 1
    hrow(ws, r, ["ZIP Code", "Inv Trans", "Median Value", "DOM", "Supply Mo", "Rationale"])
    r += 1
    for x in [v for v in zsorted if (v.get("total_inv_trans_6mo") or 0) >= 60][:8]:
        hom, sold = x.get("homes_on_market") or 0, x.get("homes_sold_last_month") or 0
        ws.cell(row=r, column=1, value=x["zip_code"])
        ws.cell(row=r, column=2, value=x.get("total_inv_trans_6mo"))
        ws.cell(row=r, column=3, value=x.get("median_home_value_raw"))
        ws.cell(row=r, column=4, value=x.get("median_days_on_market"))
        ws.cell(row=r, column=5, value=supply_months(hom, sold))
        ws.cell(row=r, column=6, value=rationale(x))
        border_row(ws, r, 6)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section B - Tier 1: Highest-Priority Neighborhoods").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Neighborhood", "Inv Trans", "Median Value", "DOM", "Supply Mo", "Rationale"])
    r += 1
    for x in sorted(n, key=lambda v: v.get("total_inv_trans_6mo") or 0, reverse=True)[:8]:
        hom, sold = x.get("homes_on_market") or 0, x.get("homes_sold_last_month") or 0
        ws.cell(row=r, column=1, value=x["neighborhood"])
        ws.cell(row=r, column=2, value=x.get("total_inv_trans_6mo"))
        ws.cell(row=r, column=3, value=x.get("median_home_value_raw"))
        ws.cell(row=r, column=4, value=x.get("median_days_on_market"))
        ws.cell(row=r, column=5, value=supply_months(hom, sold))
        ws.cell(row=r, column=6, value=rationale(x))
        border_row(ws, r, 6)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section C - Tier 2: Secondary ZIP Opportunities").font = SUB_FONT
    r += 1
    hrow(ws, r, ["ZIP Code", "Inv Trans", "Median Value", "DOM", "Supply Mo", "Rationale"])
    r += 1
    for x in [v for v in zsorted if 30 <= (v.get("total_inv_trans_6mo") or 0) < 60][:8]:
        hom, sold = x.get("homes_on_market") or 0, x.get("homes_sold_last_month") or 0
        ws.cell(row=r, column=1, value=x["zip_code"])
        ws.cell(row=r, column=2, value=x.get("total_inv_trans_6mo"))
        ws.cell(row=r, column=3, value=x.get("median_home_value_raw"))
        ws.cell(row=r, column=4, value=x.get("median_days_on_market"))
        ws.cell(row=r, column=5, value=supply_months(hom, sold))
        ws.cell(row=r, column=6, value=rationale(x))
        border_row(ws, r, 6)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section D - Market Timing Considerations").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Factor", "Current Status", "Implication"])
    r += 1
    timing = syn.get("market_timing") or []
    if not timing:
        hom = sum(x.get("homes_on_market") or 0 for x in z)
        sold = sum(x.get("homes_sold_last_month") or 0 for x in z)
        redfin = bundle.get("metro_redfin") or {}
        timing = [
            {"factor": "National DOM", "status": f"{fred} days (FRED)", "implication": "Velocity benchmark"},
            {"factor": "Supply", "status": f"{supply_months(hom, sold)} months", "implication": "Under 3 = seller's; 3-6 = balanced"},
            {"factor": "Price Trend", "status": f"{redfin.get('median_sale_price_yoy_pct','?')}% YoY", "implication": "Direction of county prices"},
        ]
    for t in timing:
        ws.cell(row=r, column=1, value=t.get("factor"))
        ws.cell(row=r, column=2, value=t.get("status"))
        ws.cell(row=r, column=3, value=t.get("implication"))
        border_row(ws, r, 3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section E - Recommended Strategy").font = SUB_FONT
    r += 1
    steps = syn.get("strategy_steps") or []
    if not steps:
        t1 = ", ".join(x["zip_code"] for x in zsorted[:5])
        steps = [f"Focus marketing on the highest-activity ZIPs: {t1}",
                 "Prioritize properties with above-average DOM for motivated sellers",
                 "Match price-band targeting to the wholesale buyer pool"]
    for i, s in enumerate(steps, 1):
        ws.cell(row=r, column=1, value=f"{i}. {s}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    autofit(ws)


def sheet_sources(wb, sift, bundle, fred, county, st, disclaimers):
    ws = wb.create_sheet("Data Sources")
    ws["A1"] = "DATA SOURCES & METHODOLOGY"
    ws["A1"].font = SECTION_FONT
    today = datetime.now().strftime("%Y-%m-%d")
    nz = len(sift["zip_data"])
    nn = len(sift["neighborhood_data"])
    r = 3
    ws.cell(row=r, column=1, value="Section A - Primary Data Sources").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Source", "Data Type", "Date Retrieved", "URL / Notes"])
    r += 1
    for s in [
        ("REI Sift Market Finder", f"Investor transactions, home values, DOM, inventory ({nz} ZIPs, {nn} neighborhoods)", today, "app.reisift.io/market-finder"),
        ("FRED (St. Louis Fed)", "National + Ohio median Days on Market", today, "fred.stlouisfed.org"),
        ("U.S. Census Bureau", "Demographics, population, income", today, f"census.gov/quickfacts/{county.lower()}countyohio"),
        ("Bureau of Labor Statistics", "Employment, unemployment, sectors", today, "bls.gov/eag"),
        ("Redfin", "County housing trends, prices, DOM", today, "redfin.com"),
        ("FBI UCR / County Sheriff", "Crime statistics", today, "fbi.gov / county sheriff"),
        ("ATTOM / RealtyTrac", "Foreclosure & distress activity", today, "attomdata.com"),
    ]:
        for c, v in enumerate(s, 1):
            ws.cell(row=r, column=c, value=v)
        border_row(ws, r, 4)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section B - Methodology").font = SUB_FONT
    r += 1
    hrow(ws, r, ["Component", "Description"])
    r += 1
    for m, dsc in [
        ("Wholesaling Score", "Composite: investor transactions (primary), DOM vs national (secondary), price accessibility (tertiary)"),
        ("DOM vs National", f"Median DOM minus FRED national median ({fred} days as of {datetime.now():%b %Y})"),
        ("Price Spread", "(Median Sale Price - Median Home Value) / Median Home Value"),
        ("Supply Months", "Homes on Market / Homes Sold Last Month - under 3 = seller's, 3-6 = balanced, over 6 = buyer's"),
        ("Tier Classification", "Tier 1 = 60+ inv trans (6mo, ~10/mo); Tier 2 = 30-59; derived from Sift data"),
    ]:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=dsc)
        border_row(ws, r, 2)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Section C - Disclaimers").font = SUB_FONT
    r += 1
    for d in disclaimers:
        ws.cell(row=r, column=1, value=f"- {d}")
        r += 1
    autofit(ws)


OHIO_DISCLAIMERS = [
    "Data is current as of retrieval date and subject to change",
    "REI Sift data represents proprietary investor transaction tracking (6-month window)",
    "Public figures (Census/BLS/crime) often lag real-time market conditions",
    "Ohio is a judicial-foreclosure / county sheriff-sale state; tax-delinquent property moves via tax-lien certificate and tax-foreclosure auctions - verify county-specific process and timelines",
    "Older Midwest housing stock carries lead-paint, knob-and-tube, and foundation/basement risks - budget inspections accordingly",
    "Investment recommendations are informational; always conduct independent due diligence",
]


def generate_report(sift, bundle, county, st, fred, out, disclaimers=None):
    wb = openpyxl.Workbook()
    sheet_exec(wb, sift, bundle, fred, county, st)
    _zip_or_nbr_sheet(wb, "ZIP Code Analysis", list(sift["zip_data"]), "zip_code", fred)
    _zip_or_nbr_sheet(wb, "Neighborhood Analysis", list(sift["neighborhood_data"]), "neighborhood", fred)
    sheet_econ(wb, bundle, county, st)
    sheet_crime(wb, bundle, county, st)
    sheet_reco(wb, sift, bundle, fred, county, st)
    sheet_sources(wb, sift, bundle, fred, county, st, disclaimers or OHIO_DISCLAIMERS)
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sift", required=True)
    ap.add_argument("--bundle", default="")
    ap.add_argument("--county", required=True)
    ap.add_argument("--state", required=True)
    ap.add_argument("--fred-dom", type=int, default=52)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    sift = json.loads(Path(args.sift).read_text(encoding="utf-8"))
    bundle = json.loads(Path(args.bundle).read_text(encoding="utf-8")) if args.bundle and Path(args.bundle).exists() else {}
    fred = args.fred_dom
    bf = (bundle.get("fred") or {}).get("fred_national_dom_days")
    if isinstance(bf, (int, float)) and bf > 0:
        fred = int(bf)
    out = generate_report(sift, bundle, args.county, args.state, fred, args.out)
    print(f"Saved: {out} (FRED {fred})")


if __name__ == "__main__":
    main()
