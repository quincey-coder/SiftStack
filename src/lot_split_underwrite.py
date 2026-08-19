"""Ground-up development underwrite: demo the existing house, split the lot,
build new spec homes, sell. Built for 306 N Morgan St, Friendsville TN --
the post-walkthrough exit engine only scores wholesale/wholetail/flip/rental,
it has no build-new lane, so this is a standalone model.

Every cost line cites its source in the workbook itself. Where a figure is a
national/regional proxy rather than a verified Blount County number, the
workbook says so -- same discipline as the rest of the deal-analysis stack:
no invented numbers, no silent precision.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY, BLUE, GREEN, GOLD, RED, GREY = "0A1130", "316AFF", "1B9E5A", "B8860B", "C0392B", "666666"


def _money(v) -> str:
    if not isinstance(v, (int, float)):
        return str(v or "")
    return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"


def _psf(v) -> str:
    return f"${v:,.0f}/sf"


def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=NAVY)
    if sub:
        c = ws.cell(row=2, column=1, value=sub)
        c.font = Font(size=10, color=GREY)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _subhead(ws, row, text, col=1, color=BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=11, color=color)
    return row + 1


def _widths(ws, widths, start=1):
    for i, w in enumerate(widths, start):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══ 1. Deal facts (verified, 2026-07-25) ═══════════════════════════════

SUBJECT = {
    "address": "306 N Morgan St, Friendsville, TN 37737",
    "parcel": "043N B 02400 000",
    "lot_sqft": 51_841,          # 1.19 ac, Blount County GIS, live parcel pull
    "existing_sqft": 1_200,      # county card, 1920-built
    "purchase_price": 135_000,   # asking
}

PLAN = {
    "n_lots": 2,                 # zoning: 20,000 sf min in City of Friendsville R-1;
                                  # 51,841 / 20,000 = 2.59 -> 2 lots by right, not 3
    "home_sqft": 1_500,          # target spec: 3bd/2ba builder-to-mid grade,
                                  # sized to the ~25,900 sf conforming lots
}

# ══ 2. Cost assumptions, one line = one source ═════════════════════════
# point = what the model uses; lo/hi = the sensitivity band shown on Sale & Profit.
# Sources noted inline; full citations render on the Sources sheet.

SHARED_COSTS = [
    # item, point, lo, hi, source
    ("Demolition (existing 1,200 sf house)", 9_000, 7_200, 12_000,
     "Knoxville-area demo benchmark $5-10/sf, most jobs $8,400-$17,700 total "
     "(HomeBlue, DemoProsPlus). Partially-collapsed rear wing adds handling risk, "
     "point estimate skews to the upper-middle of the per-sf range."),
    ("Survey + 2-lot minor subdivision plat", 3_500, 1_500, 5_000,
     "Blount County minor subdivision (<5 lots, no new road/utility) survey+plat "
     "range; exact recording fee unconfirmed with the county."),
]

PER_LOT_SITE_DEV = [
    ("Water tap (Town of Friendsville, new 5/8\" service + account fee)", 2_600, 2_600, 2_600,
     "Friendsville Water Works published rate: $2,550 tap + $50 account fee. Verified."),
    ("Septic system, full install (3-bed, tank+drainfield+perc+excavation)", 9_500, 7_000, 12_000,
     "TN/Knoxville regional install range. GATE: county GIS codes this parcel as "
     "septic (no public sewer); a 2026-07-18 site note says the property 'just got "
     "city sewer' -- confirm with Friendsville Water Works. If sewer is real, this "
     "line drops to a tap fee instead."),
    ("Electric service (Fort Loudoun Electric Cooperative)", 750, 50, 1_500,
     "FLEC published new-service fee is nominal ($5 membership + $10 service + "
     "temp pole ~$35); point estimate adds a margin for an underground drop or "
     "distance charge FLEC prices per-site."),
    ("Driveway/culvert + grading + clearing", 8_500, 5_000, 12_000,
     "TN regional: culvert $1,500-5,000, gravel drive $3,000-8,000, grading "
     "$1-3/sf, clearing $1,000-5,000/acre -- combined range for a half-acre in-town lot."),
    ("Building permit + plan review", 1_750, 1_000, 2_500,
     "Blount County Development Services ICC-valuation fee schedule, estimated "
     "against a ~$250-300K build valuation. Verify exact bracket with the county."),
]

HARD_COST_PSF = {"lo": 125, "point": 140, "hi": 155}
# NAHB 2024 national avg $162/sf on a 2,647 sf average home; smaller homes run
# HIGHER $/sf (fixed kitchen/bath/mechanical cost over less area). East South
# Central (incl. TN) prices below the national average. Recommended underwriting
# range for 1,400-1,600 sf builder-to-mid grade in this market: $125-155/sf.

PER_LOT_SOFT_COST_RULES = {
    "plans": 1_750,                 # stock plans, 1,400-1,600 sf: $1,000-2,500
    "gl_insurance_per_month": 300,  # contractor GL, $80-600/mo range midpoint
    "construction_months": 6,
    "staking": 850,                 # survey/staking $500-1,200
    "insurance_pct_of_hard": {"lo": 0.01, "point": 0.03, "hi": 0.05},   # builder's risk
    "contingency_pct_of_hard": {"lo": 0.10, "point": 0.12, "hi": 0.15},  # standard spec-build range
}

FINANCING = {
    "rate": 0.11,        # 10-12% typical for ground-up spec construction
    "points": 0.025,     # 1-4 points quoted, 2.5% used as a blended point estimate
    "term_months": 11,   # 9-14 months permit-to-close midpoint
    "source": "Hard-money/spec construction-to-perm: 8-15% (10-12% typical for "
              "ground-up), 1-4 points, 9-12 month draw term (Crestmont Capital, "
              "Builder Finance). National proxy, no TN-specific spec-loan quote found.",
}

SELLING_COST_PCT = 0.08   # matches the SELLING_COST_PCT convention in post_walkthrough.py

SALE_SCENARIOS = [
    ("Conservative (in-town, no acreage premium)", 200,
     "Anchored off the true in-town resale comps (165 Cedar Crest Ln $184/sf dated "
     "condition, 112 N Farnum $171/sf) plus a modest new-construction premium. "
     "Deliberately does NOT use the area's acreage/lakefront comps -- same bleed "
     "risk flagged on the ARV in the rehab underwrite."),
    ("Moderate", 220,
     "Midpoint between the in-town discount case and the two real Friendsville "
     "new-construction sales below."),
    ("Comps-parity (actual Friendsville new construction)", 245,
     "The two real new-construction sales found in Friendsville: 1235 Marble Hill "
     "Rd $233/sf (1,904 sf, 1.3 ac) and a 2,200 sf sale at $254/sf. Both carry some "
     "acreage premium this in-town lot won't fully command -- treat as a ceiling, "
     "not a base case."),
]

# ── Path B: split the lot, sell the two RAW (unbuilt) lots to a builder ──
# No construction, no water tap/electric/driveway (that's the builder's job after
# they own it) -- just get to two legal, sellable parcels. Much cheaper, much
# faster, and the acquisition price is the lever that actually decides it.

WHOLESALE_PREP_COSTS = [
    ("Demolition (existing house)", 9_000,
     "Same demo line as the build-new path -- a builder buying a lot still wants "
     "it cleared, not a condemned structure on it."),
    ("Survey + 2-lot minor subdivision plat", 3_500,
     "Same as the build-new path."),
    ("Septic/soil suitability eval, both lots (perc test only, not full install)",
     2_700,
     "Blount Co. Environmental Health perc/soil test $700-2,000/lot, NOT a full "
     "septic install -- the buyer's builder installs their own system. "
     "$1,350/lot x 2 point estimate."),
]

WHOLESALE_FINANCING = {
    "rate": 0.12, "points": 0.02, "term_months": 6,
    "source": "Acquisition-side hard money, matching this stack's own "
              "DEFAULT_HARD_MONEY_RATE/POINTS (deal_analyzer.py). Shorter 6-month "
              "term than the build path -- no construction phase, just demo, "
              "survey, and plat/marketing to a builder buyer.",
}

WHOLESALE_SELLING_COST_PCT = 0.06   # land sale commission/closing, lighter than a retail home sale

LOT_VALUE_SCENARIOS = [
    ("Low", 33_000,
     "A 0.55 ac finished/buildable lot in Tallassee, Blount Co., listed at $32,900 "
     "-- the closest real size analog (our 2 lots run ~0.59 ac each) found in "
     "this market."),
    ("Mid", 55_000,
     "Midpoint: credits this parcel's in-town position + already-run public water "
     "over the Tallassee comp, without assuming a real premium exists."),
    ("High", 80_000,
     "Speculative ceiling if a builder specifically pays up for a shovel-ready, "
     "R-1, water-connected in-town lot given confirmed builder demand nearby "
     "(Smithbilt Homes, Greenleaf Properties, Harmony Investments all active in "
     "Blount Co.). Not backed by a real per-lot sale at this size -- get an LOI "
     "before trusting this number."),
]

MIN_WHOLESALE_FEE = 15_000   # matches post_walkthrough.py's own floor for "worth the file"


# ══ 3. Pro forma math ═══════════════════════════════════════════════════

def solve_build_purchase_price(pf: dict, sale_scenario: dict, target_profit: float) -> float:
    """Purchase price P that hits target_profit under the build-new path, holding
    every other assumption fixed. Financing applies to the whole loan basis, so
    each $1 of P costs 1 + rate*(term/12) + points in all-in cost."""
    marginal = 1 + FINANCING["rate"] * (FINANCING["term_months"] / 12) + FINANCING["points"]
    fixed_all_in_at_zero_purchase = pf["all_in_total"] - marginal * SUBJECT["purchase_price"]
    return (sale_scenario["net_proceeds"] - fixed_all_in_at_zero_purchase - target_profit) / marginal


def build_wholesale_pro_forma() -> dict:
    prep_total = sum(item[1] for item in WHOLESALE_PREP_COSTS)
    marginal = 1 + WHOLESALE_FINANCING["rate"] * (WHOLESALE_FINANCING["term_months"] / 12) + WHOLESALE_FINANCING["points"]

    scenarios = []
    for name, lot_value, basis in LOT_VALUE_SCENARIOS:
        revenue = lot_value * PLAN["n_lots"]
        net_revenue = revenue * (1 - WHOLESALE_SELLING_COST_PCT)
        # target_purchase(target_profit) = (net_revenue - prep_total*marginal_on_prep - target_profit) / marginal
        # prep costs are financed too, so they carry the same marginal multiplier as purchase price.
        breakeven_purchase = (net_revenue - prep_total * marginal) / marginal
        min_fee_purchase = (net_revenue - prep_total * marginal - MIN_WHOLESALE_FEE) / marginal
        scenarios.append({
            "name": name, "lot_value": lot_value, "basis": basis,
            "revenue": revenue, "net_revenue": net_revenue,
            "breakeven_purchase": breakeven_purchase,
            "min_fee_purchase": min_fee_purchase,
        })
    return {"prep_total": prep_total, "marginal": marginal, "scenarios": scenarios}


# ══ 3. Pro forma math ═══════════════════════════════════════════════════

def build_pro_forma() -> dict:
    n = PLAN["n_lots"]
    home_sqft = PLAN["home_sqft"]

    shared_total = sum(item[1] for item in SHARED_COSTS)
    site_dev_per_lot = sum(item[1] for item in PER_LOT_SITE_DEV)
    site_dev_total = site_dev_per_lot * n

    hard_per_home = HARD_COST_PSF["point"] * home_sqft
    hard_total = hard_per_home * n

    r = PER_LOT_SOFT_COST_RULES
    insurance_per_home = r["insurance_pct_of_hard"]["point"] * hard_per_home
    contingency_per_home = r["contingency_pct_of_hard"]["point"] * hard_per_home
    gl_per_home = r["gl_insurance_per_month"] * r["construction_months"]
    soft_per_home = r["plans"] + insurance_per_home + gl_per_home + r["staking"] + contingency_per_home
    soft_total = soft_per_home * n

    loan_basis = SUBJECT["purchase_price"] + shared_total + site_dev_total + hard_total
    interest = loan_basis * FINANCING["rate"] * (FINANCING["term_months"] / 12)
    points_cost = loan_basis * FINANCING["points"]
    financing_total = interest + points_cost

    all_in_total = (SUBJECT["purchase_price"] + shared_total + site_dev_total
                     + hard_total + soft_total + financing_total)

    scenarios = []
    for name, psf, basis in SALE_SCENARIOS:
        sale_per_home = psf * home_sqft
        sale_total = sale_per_home * n
        selling_costs = sale_total * SELLING_COST_PCT
        net_proceeds = sale_total - selling_costs
        profit = net_proceeds - all_in_total
        scenarios.append({
            "name": name, "psf": psf, "basis": basis,
            "sale_per_home": sale_per_home, "sale_total": sale_total,
            "selling_costs": selling_costs, "net_proceeds": net_proceeds,
            "profit": profit, "profit_per_home": profit / n,
            "roi": profit / all_in_total,
        })

    combined_sqft = home_sqft * n
    breakeven_psf = all_in_total / (combined_sqft * (1 - SELLING_COST_PCT))
    best_case_profit = scenarios[-1]["profit"]
    # Raising/lowering purchase price P moves loan basis 1:1, so each $1 of P costs
    # 1 + rate*(term/12) + points in all-in cost. Solve profit(P*) = 0 from profit(P0).
    marginal_cost_per_dollar = 1 + FINANCING["rate"] * (FINANCING["term_months"] / 12) + FINANCING["points"]
    breakeven_purchase = SUBJECT["purchase_price"] + best_case_profit / marginal_cost_per_dollar

    return {
        "shared_total": shared_total, "site_dev_per_lot": site_dev_per_lot,
        "site_dev_total": site_dev_total, "hard_per_home": hard_per_home,
        "hard_total": hard_total, "soft_per_home": soft_per_home,
        "soft_total": soft_total, "insurance_per_home": insurance_per_home,
        "contingency_per_home": contingency_per_home, "gl_per_home": gl_per_home,
        "loan_basis": loan_basis, "interest": interest, "points_cost": points_cost,
        "financing_total": financing_total, "all_in_total": all_in_total,
        "scenarios": scenarios, "breakeven_psf": breakeven_psf,
        "breakeven_purchase": breakeven_purchase,
    }


# ══ 4. Workbook ═══════════════════════════════════════════════════════

def build_seller_walkthrough(ws, pf: dict, wf: dict):
    """Plain-language version of the same numbers, in the order you'd actually
    say them out loud to the seller. Every figure here is pulled from the same
    pro forma as the analyst sheets -- nothing new is assumed."""
    low = wf["scenarios"][0]   # the Tallassee comp, the only real comp we have
    mid = wf["scenarios"][1]
    high = wf["scenarios"][2]

    _title(ws, "Land Value Walkthrough",
           "How we got to our number -- for the seller conversation. Every figure "
           "here also appears on the analyst sheets; nothing is added just for this page.")
    r = 4

    r = _subhead(ws, r, "1. WHAT'S ACTUALLY HERE")
    for line in [
        f"{SUBJECT['lot_sqft']:,} sf, 1.19 acres, zoned R-1 inside the City of Friendsville.",
        f"City ordinance requires 20,000 sf minimum per lot. {SUBJECT['lot_sqft']:,} / "
        f"20,000 = {SUBJECT['lot_sqft']/20_000:.2f} -- this splits into 2 legal lots, "
        "not 3. A 3rd lot isn't allowed without a variance from the city.",
        f"The {SUBJECT['existing_sqft']:,} sf house (built 1920) comes down. It isn't "
        "priced into the land value below.",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    r += 1

    r = _subhead(ws, r, "2. WHY THE HOUSE ITSELF DOESN'T ADD VALUE")
    for line in [
        "The back of the house (kitchen wing and the two-story addition) has a "
        "collapsed floor and a broken roof beam -- that part comes down no matter "
        "what happens to the rest of the property.",
        "A full renovation of the whole house doesn't cash-flow at today's price "
        "once it's priced against what similar homes actually sell for in this "
        "specific pocket of town, not the lake-view homes a few streets over.",
        "Building two brand-new houses on the split lots doesn't work either -- "
        "what it costs to build here runs well above what a new house actually "
        "resells for in this specific small market.",
        "That's why the number below is a LAND number, not a house number or a "
        "construction number.",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    r += 1

    r = _subhead(ws, r, "3. WHAT THE LAND IS ACTUALLY WORTH", color=GREEN)
    c = ws.cell(row=r, column=1, value=(
        f"Real comp: {low['basis']}"))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 30
    r += 2

    r = _header(ws, r, ["", "Amount", ""]) or r + 1
    ladder = [
        ("Two lots at the comp price ($33,000/lot)", low["revenue"], ""),
        ("Less: demolition", -WHOLESALE_PREP_COSTS[0][1], WHOLESALE_PREP_COSTS[0][1]),
        ("Less: survey + subdivide into 2 legal lots", -WHOLESALE_PREP_COSTS[1][1], WHOLESALE_PREP_COSTS[1][1]),
        ("Less: soil/septic testing, both lots", -WHOLESALE_PREP_COSTS[2][1], WHOLESALE_PREP_COSTS[2][1]),
    ]
    for label, signed_val, _ in ladder:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=_money(signed_val))
        c.font = Font(color=RED if signed_val < 0 else None)
        r += 1
    net_land_value = low["revenue"] - wf["prep_total"]
    ws.cell(row=r, column=1, value="Net land value, once split").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(round(net_land_value, -2))).font = Font(bold=True, color=NAVY)
    r += 2

    c = ws.cell(row=r, column=1, value=(
        f"That {_money(round(net_land_value, -2))} is what the two lots are worth AFTER "
        "the split is done and paid for -- not what the property is worth today, sitting "
        "as one lot with a condemned structure on it. We have to put up the demo, survey, "
        "and testing money and carry it for months before those two lots exist to sell."))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 45
    r += 2

    r = _subhead(ws, r, "4. OUR NUMBER", color=RED)
    c = ws.cell(row=r, column=1, value=(
        f"Backing out the cost of carrying that work and a minimum profit for taking on "
        f"the risk, a workable offer lands between {_money(round(low['min_fee_purchase'], -2))} "
        f"and {_money(round(low['breakeven_purchase'], -2))}."))
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 45
    r += 2

    r = _subhead(ws, r, "IF THE SELLER PUSHES BACK ON THE COUNTY VALUE OR THE ASK PRICE", color=GOLD)
    for line in [
        "The county/Zillow-style estimate assumes a normal, livable house. It doesn't "
        "know about the collapsed floor and broken roof beam in back, and it isn't what "
        "actually happens when a property like this gets marketed and sold here.",
        f"If a builder is willing to pay a premium for an already-split, water-connected, "
        f"in-town lot (no confirmed comp at this level yet), the number could stretch toward "
        f"{_money(mid['lot_value'])}-{_money(high['lot_value'])}/lot instead of "
        f"{_money(low['lot_value'])}/lot -- worth getting a builder's actual offer before "
        "committing to a number above our range.",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 30
        r += 1
    _widths(ws, [70, 16, 12, 10, 10, 10])


def build_workbook(pf: dict, wf: dict, out: str):
    wb = Workbook()

    # -- Seller Walkthrough (read this one first) --
    ws = wb.active
    ws.title = "Seller Walkthrough"
    build_seller_walkthrough(ws, pf, wf)

    # -- Overview --
    ws = wb.create_sheet("Overview")
    _title(ws, "306 N Morgan St -- Demo & Build-New Underwrite",
           "Split the 1.19 ac lot into 2 conforming R-1 lots, demo the existing house, "
           "build two 1,500 sf spec homes, sell. Built 2026-07-25.")
    r = 4
    r = _subhead(ws, r, "THE PLAN")
    for label, val in [
        ("Acquisition", _money(SUBJECT["purchase_price"]) + " (asking)"),
        ("Parcel", f"{SUBJECT['parcel']}, {SUBJECT['lot_sqft']:,} sf (1.19 ac)"),
        ("Zoning", "City of Friendsville R-1, 20,000 sf min lot, 100 ft min width -- "
                   f"{SUBJECT['lot_sqft']:,} / 20,000 = {SUBJECT['lot_sqft']/20_000:.2f}, "
                   "2 lots by right, a 3rd needs a variance and is not modeled here"),
        ("Existing structure", f"{SUBJECT['existing_sqft']:,} sf, 1920 -- full demo, "
                                "no salvage value assumed"),
        ("New build spec", f"2 homes x {PLAN['home_sqft']:,} sf, 3bd/2ba, builder-to-mid grade"),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    r = _subhead(ws, r, "HEADLINE VERDICT", color=RED)
    verdict = (
        f"All-in cost to deliver both finished homes: {_money(pf['all_in_total'])}. "
        f"Breakeven sale price is {_psf(pf['breakeven_psf'])} combined -- HIGHER than "
        f"the best real new-construction sale found in Friendsville ({_psf(245)}, "
        "1235 Marble Hill Rd / a 2,200 sf comp). Every sale scenario below, including "
        "selling at parity with those actual new-construction sales, is a loss at the "
        f"{_money(SUBJECT['purchase_price'])} asking price. "
        f"For this to break even at the comps-parity price, the acquisition would need "
        f"to land closer to {_money(pf['breakeven_purchase'])}, not {_money(SUBJECT['purchase_price'])}."
    )
    c = ws.cell(row=r, column=1, value=verdict)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(color=RED, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 90
    _widths(ws, [22, 90])

    # -- Land & Site Costs --
    ws = wb.create_sheet("Land & Site Costs")
    _title(ws, "Land & Site Costs")
    r = 3
    r = _subhead(ws, r, "SHARED (whole parcel, not per-lot)")
    r = _header(ws, r, ["Item", "Cost", "Range", "Source / note"]) or r + 1
    for item, point, lo, hi, src in SHARED_COSTS:
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=_money(point))
        ws.cell(row=r, column=3, value=f"{_money(lo)} - {_money(hi)}")
        ws.cell(row=r, column=4, value=src).alignment = Alignment(wrap_text=True)
        r += 1
    ws.cell(row=r, column=1, value="Shared subtotal").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["shared_total"])).font = Font(bold=True, color=NAVY)
    r += 2
    r = _subhead(ws, r, "PER LOT (x2 lots)")
    r = _header(ws, r, ["Item", "Cost", "Range", "Source / note"]) or r + 1
    for item, point, lo, hi, src in PER_LOT_SITE_DEV:
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=_money(point))
        ws.cell(row=r, column=3, value=f"{_money(lo)} - {_money(hi)}")
        ws.cell(row=r, column=4, value=src).alignment = Alignment(wrap_text=True)
        r += 1
    ws.cell(row=r, column=1, value="Per-lot site dev subtotal").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["site_dev_per_lot"])).font = Font(bold=True, color=NAVY)
    r += 1
    ws.cell(row=r, column=1, value="Site dev, both lots").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["site_dev_total"])).font = Font(bold=True, color=NAVY)
    _widths(ws, [42, 14, 20, 60])

    # -- Construction Budget --
    ws = wb.create_sheet("Construction Budget")
    _title(ws, "Construction Budget", f"Per home, {PLAN['home_sqft']:,} sf spec build")
    r = 3
    r = _subhead(ws, r, "HARD COST (materials + labor)")
    ws.cell(row=r, column=1, value="Cost per sf").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"{_psf(HARD_COST_PSF['lo'])} - {_psf(HARD_COST_PSF['hi'])}, "
                                    f"point {_psf(HARD_COST_PSF['point'])}")
    r += 1
    ws.cell(row=r, column=1, value="NAHB 2024 national avg $162/sf on a 2,647 sf avg home "
            "(smaller homes run higher $/sf, fixed kitchen/bath cost over less area); East "
            "South Central region (incl. TN) prices below national average. Regional/national "
            "proxy, no verified Blount-County builder survey.").font = Font(size=9, color=GREY)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    r += 2
    ws.cell(row=r, column=1, value="Hard cost, per home").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["hard_per_home"])).font = Font(bold=True, color=NAVY)
    r += 1
    ws.cell(row=r, column=1, value="Hard cost, both homes").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["hard_total"])).font = Font(bold=True, color=NAVY)
    r += 2
    r = _subhead(ws, r, "SOFT COST, per home")
    rules = PER_LOT_SOFT_COST_RULES
    rows = [
        ("Stock plans (1,400-1,600 sf)", rules["plans"], "$1,000-2,500 range"),
        (f"Builder's risk insurance ({rules['insurance_pct_of_hard']['point']:.0%} of hard cost)",
         pf["insurance_per_home"], "1-5% of hard cost is the typical range"),
        (f"General liability ({rules['gl_insurance_per_month']}/mo x {rules['construction_months']} mo)",
         pf["gl_per_home"], "$80-600/mo contractor GL range"),
        ("Survey / construction staking", rules["staking"], "$500-1,200 range"),
        (f"Contingency ({rules['contingency_pct_of_hard']['point']:.0%} of hard cost)",
         pf["contingency_per_home"], "10-15% of hard cost is the standard spec-build benchmark"),
    ]
    r = _header(ws, r, ["Item", "Cost", "Note"]) or r + 1
    for item, val, note in rows:
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=_money(val))
        ws.cell(row=r, column=3, value=note).font = Font(size=9, color=GREY)
        r += 1
    ws.cell(row=r, column=1, value="Soft cost, per home").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["soft_per_home"])).font = Font(bold=True, color=NAVY)
    r += 1
    ws.cell(row=r, column=1, value="Soft cost, both homes").font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value=_money(pf["soft_total"])).font = Font(bold=True, color=NAVY)
    _widths(ws, [46, 16, 60])

    # -- Financing --
    ws = wb.create_sheet("Financing")
    _title(ws, "Construction Financing")
    r = 3
    for label, val in [
        ("Loan basis (purchase + shared + site dev + hard cost)", _money(pf["loan_basis"])),
        ("Rate", f"{FINANCING['rate']:.0%} annual"),
        ("Points", f"{FINANCING['points']:.1%} origination"),
        ("Term", f"{FINANCING['term_months']} months (permit-to-close midpoint)"),
        ("Interest carry", _money(pf["interest"])),
        ("Points cost", _money(pf["points_cost"])),
        ("Financing total", _money(pf["financing_total"])),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1
    c = ws.cell(row=r, column=1, value=FINANCING["source"])
    c.alignment = Alignment(wrap_text=True)
    c.font = Font(size=9, color=GREY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _widths(ws, [46, 20, 20, 20])

    # -- Sale & Profit --
    ws = wb.create_sheet("Sale & Profit")
    _title(ws, "Sale & Profit -- 3 Scenarios",
           f"All-in cost (fixed across scenarios): {_money(pf['all_in_total'])}")
    r = 3
    r = _header(ws, r, ["Scenario", "$/sf", "Sale, both homes", "Selling costs (8%)",
                        "Net proceeds", "Profit", "Profit/home", "ROI"]) or r + 1
    for s in pf["scenarios"]:
        ws.cell(row=r, column=1, value=s["name"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=2, value=_psf(s["psf"]))
        ws.cell(row=r, column=3, value=_money(s["sale_total"]))
        ws.cell(row=r, column=4, value=_money(s["selling_costs"]))
        ws.cell(row=r, column=5, value=_money(s["net_proceeds"]))
        pc = ws.cell(row=r, column=6, value=_money(s["profit"]))
        pc.font = Font(bold=True, color=GREEN if s["profit"] > 0 else RED)
        ws.cell(row=r, column=7, value=_money(s["profit_per_home"]))
        ws.cell(row=r, column=8, value=f"{s['roi']:.0%}")
        r += 1
    r += 1
    for s in pf["scenarios"]:
        c = ws.cell(row=r, column=1, value=f"{s['name']}: {s['basis']}")
        c.alignment = Alignment(wrap_text=True)
        c.font = Font(size=9, color=GREY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1
    r = _subhead(ws, r, "BREAKEVEN", color=RED)
    ws.cell(row=r, column=1, value="Sale price needed to break even, combined")
    ws.cell(row=r, column=2, value=_psf(round(pf["breakeven_psf"])))
    r += 1
    ws.cell(row=r, column=1, value="Purchase price for $0 profit at the comps-parity sale price")
    ws.cell(row=r, column=2, value=_money(round(pf["breakeven_purchase"], -3)))
    _widths(ws, [26, 12, 18, 16, 16, 16, 14, 10])

    # -- Target Purchase Price --
    ws = wb.create_sheet("Target Purchase Price")
    _title(ws, "What The Contract Price Would Have To Be",
           "Two different plans, same demo + 2-lot split. Purchase price is the "
           "lever that decides whether either one clears the floor below.")
    r = 3
    r = _subhead(ws, r, "PATH A -- build 2 new homes ourselves, then sell", color=RED)
    fixed_cost_ex_purchase = pf["all_in_total"] - (1 + FINANCING["rate"] * (FINANCING["term_months"] / 12)
                                                    + FINANCING["points"]) * SUBJECT["purchase_price"]
    combined_sqft = PLAN["home_sqft"] * PLAN["n_lots"]
    c = ws.cell(row=r, column=1, value=(
        "Purchase price barely moves this one. Fixed costs (site dev + hard "
        f"construction + soft costs + financing) already run {_money(round(fixed_cost_ex_purchase, -3))}, "
        f"against the {combined_sqft:,} combined sf being built. At the Moderate sale "
        "price ($220/sf), the deal is UNDERWATER even at a $0 purchase price. Only "
        "the optimistic comps-parity sale price ($245/sf) clears cost at all, and "
        "even then only a ~11% ROI at zero land cost -- construction cost is the "
        "binding constraint here, not the seller's ask."))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 75
    r += 2
    r = _header(ws, r, ["Sale scenario", "$/sf", "Purchase price for $0 profit",
                        "Purchase price for 15% ROI"]) or r + 1
    for s in pf["scenarios"]:
        be = solve_build_purchase_price(pf, s, 0)
        target = solve_build_purchase_price(pf, s, 0.15 * pf["all_in_total"])
        ws.cell(row=r, column=1, value=s["name"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=2, value=_psf(s["psf"]))
        bc = ws.cell(row=r, column=3, value=_money(round(be, -2)))
        bc.font = Font(color=GREEN if be > 0 else RED)
        tc = ws.cell(row=r, column=4, value=_money(round(target, -2)) if target > 0 else "not reachable")
        tc.font = Font(color=GREEN if target > 0 else RED)
        r += 1
    r += 2

    r = _subhead(ws, r, "PATH B -- split the lot, sell both RAW lots to a builder", color=GREEN)
    c = ws.cell(row=r, column=1, value=(
        "No construction, no water/electric/driveway (the builder buyer does that "
        "after closing). Just demo + survey + minor plat + a soil eval, then sell "
        "two legal parcels. This is the path where the contract price actually "
        "decides the outcome."))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 40
    r += 2
    r = _header(ws, r, ["Lot value scenario", "$/lot", "Revenue, both lots",
                        "Purchase price for $0 profit",
                        f"Purchase price for a {_money(MIN_WHOLESALE_FEE)} minimum fee"]) or r + 1
    for s in wf["scenarios"]:
        ws.cell(row=r, column=1, value=s["name"])
        ws.cell(row=r, column=2, value=_money(s["lot_value"]))
        ws.cell(row=r, column=3, value=_money(s["revenue"]))
        bc = ws.cell(row=r, column=4, value=_money(round(s["breakeven_purchase"], -2)))
        bc.font = Font(bold=True, color=GREEN if s["breakeven_purchase"] > 0 else RED)
        fc = ws.cell(row=r, column=5, value=_money(round(s["min_fee_purchase"], -2)))
        fc.font = Font(bold=True, color=GREEN if s["min_fee_purchase"] > 0 else RED)
        r += 1
    r += 1
    for s in wf["scenarios"]:
        c = ws.cell(row=r, column=1, value=f"{s['name']} (${s['lot_value']:,}/lot): {s['basis']}")
        c.alignment = Alignment(wrap_text=True)
        c.font = Font(size=9, color=GREY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Prep cost (demo + survey/plat + soil eval, both lots)").font = Font(bold=True)
    ws.cell(row=r, column=2, value=_money(wf["prep_total"]))
    r += 1
    c = ws.cell(row=r, column=1, value=(
        f"Financing: {WHOLESALE_FINANCING['rate']:.0%} rate, {WHOLESALE_FINANCING['points']:.0%} "
        f"points, {WHOLESALE_FINANCING['term_months']}-month term (matches this stack's own "
        "acquisition-side hard money default) -- {:.0%} selling cost on the land sale."
    ).format(WHOLESALE_SELLING_COST_PCT))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    r = _subhead(ws, r, "READ", color=BLUE)
    c = ws.cell(row=r, column=1, value=(
        "At the Mid lot-value scenario, the contract would need to come down to "
        f"roughly {_money(round(wf['scenarios'][1]['breakeven_purchase'], -3))} to break even, or "
        f"{_money(round(wf['scenarios'][1]['min_fee_purchase'], -3))} to clear a "
        f"{_money(MIN_WHOLESALE_FEE)} minimum fee -- both well under the $135,000 ask. Only the "
        "High (speculative, no confirmed comp) lot-value scenario gets purchase price "
        "close to the current ask. Get a builder LOI on real per-lot value before "
        "trusting anything past the Low/Mid scenarios."))
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 60
    _widths(ws, [30, 14, 20, 22, 26, 14])

    # -- Sources & Assumptions --
    ws = wb.create_sheet("Sources & Assumptions")
    _title(ws, "Sources & Assumptions")
    r = 3
    r = _subhead(ws, r, "Flagged verified (Blount County / Friendsville specific)")
    for line in [
        "Parcel size, zoning district, dimensional table -- Blount County GIS live parcel pull "
        "+ official Friendsville Zoning Ordinance",
        "Water tap fee -- Town of Friendsville Water Works published rate schedule",
        "Electric hookup base fee -- Fort Loudoun Electric Cooperative (the actual utility "
        "serving Friendsville) published new-service rates",
        "Two real Friendsville new-construction sales (Marble Hill Rd $233/sf, a 2,200 sf "
        "sale at $254/sf) used as the comps-parity ceiling",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1
    r = _subhead(ws, r, "Flagged as national/regional proxy (not Blount-specific)", color=GOLD)
    for line in [
        "Hard construction $/sf -- NAHB national survey + Knoxville-market lead-gen sites, "
        "no verified Blount County builder cost survey exists",
        "Septic full-install cost -- TN/Knoxville regional range",
        "Driveway/culvert/grading/clearing -- TN regional range",
        "Building permit fee -- Blount County fee schedule bracket estimated against a "
        "~$250-300K valuation, not confirmed directly with the county",
        "Construction financing rate/points/term -- national hard-money/spec-construction "
        "benchmarks, no TN-specific spec loan quote obtained",
        "Survey + minor subdivision plat cost -- Blount County process is known (minor "
        "subdivision, <5 lots), exact recording fee not confirmed",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1
    r = _subhead(ws, r, "Open gates before this is bankable", color=RED)
    for line in [
        "Confirm septic vs. sewer directly with Friendsville Water Works -- county GIS says "
        "septic, a 2026-07-18 site note says sewer was just added. This is a real swing item.",
        "Get 1-2 actual contractor bids on hard construction cost for this specific spec -- "
        "the $125-155/sf range is a desk estimate, not a bid.",
        "Get a local builder/realtor read on achievable new-construction $/sf for an in-town, "
        "no-acreage Friendsville lot specifically -- the two real comps used here both carry "
        "some acreage premium this parcel won't fully command.",
        "Confirm the 2-lot split clears Blount County Environmental Health septic/soil "
        "suitability before spending anything on survey/plat.",
    ]:
        c = ws.cell(row=r, column=1, value="- " + line)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    _widths(ws, [100, 10, 10, 10])

    wb.save(out)


def main():
    pf = build_pro_forma()
    wf = build_wholesale_pro_forma()
    out = str(Path.cwd() / "306_N_Morgan_Lot_Split_Underwrite.xlsx")
    build_workbook(pf, wf, out)
    print(f"All-in cost (build path): {_money(pf['all_in_total'])}")
    print(f"Breakeven $/sf: {_psf(round(pf['breakeven_psf']))}")
    for s in pf["scenarios"]:
        be = solve_build_purchase_price(pf, s, 0)
        print(f"Path A {s['name']}: sell {_psf(s['psf'])} -> profit {_money(s['profit'])} "
              f"({s['roi']:.0%} ROI) | $0-profit purchase price: {_money(round(be, -2))}")
    for s in wf["scenarios"]:
        print(f"Path B {s['name']} (${s['lot_value']:,}/lot): breakeven purchase "
              f"{_money(round(s['breakeven_purchase'], -2))}, "
              f"${MIN_WHOLESALE_FEE:,}-fee purchase {_money(round(s['min_fee_purchase'], -2))}")
    print(f"Workbook: {out}")


if __name__ == "__main__":
    main()
