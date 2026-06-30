"""Ingest a city's code-enforcement open-records response into NoticeData.

Cities fulfill Texas PIA requests in three shapes, all handled here:
  - CSV / TSV / Excel  → header-heuristic column mapping (schemas vary wildly)
  - PDF (digital)      → pdfminer text layer → multi-record LLM extraction
  - PDF (scanned)      → pypdfium2 render → Tesseract OCR → multi-record LLM extraction
All paths emit NoticeData(notice_type="code_violation") that flow through the
normal enrichment + DataSift pipeline exactly like every other source.

Used by the Bell/Williamson open-records pipeline once a city sends a file back.
Transport-agnostic: hand it a path, it doesn't care how the file arrived.
"""
import csv
import logging
import re
from datetime import datetime
from pathlib import Path

import config
import llm_client
from notice_parser import NoticeData

logger = logging.getLogger(__name__)

# target field -> keywords matched against normalized headers (first hit wins,
# keywords ordered most-specific first so "property address" beats bare "city")
_COLUMN_KEYWORDS: dict[str, list[str]] = {
    "address": ["propertyaddress", "situsaddress", "situs", "fulladdress",
                "streetaddress", "address", "location", "property"],
    "city": ["propertycity", "sitcity", "city", "municipality"],
    "zip": ["zipcode", "postalcode", "zip", "postal"],
    "parcel_id": ["parcelid", "parcelnumber", "accountnumber", "propertyid",
                  "parcel", "account", "apn", "pid"],
    "date_added": ["dateopened", "opendate", "opened", "datefiled", "filed",
                   "createddate", "created", "casedate", "reporteddate", "date"],
    "violation_description": ["violationtype", "casetype", "complainttype",
                              "violation", "description", "nature", "category",
                              "codetype", "type"],
    "status": ["casestatus", "status", "disposition"],
    "owner_name": ["ownername", "responsibleparty", "owner", "respondent", "name"],
    "compliance_deadline": ["compliancedate", "complyby", "duedate", "deadline",
                            "compliance"],
}

_DATE_FORMATS = [
    "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%m/%d/%y",
    "%m-%d-%Y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y", "%m/%d/%Y %H:%M",
]


def _norm(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (header or "").lower())


def _map_columns(headers: list[str]) -> dict[str, str]:
    """Pick the best source header for each target field."""
    norm_map = {h: _norm(h) for h in headers}
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field, keywords in _COLUMN_KEYWORDS.items():
        for kw in keywords:
            match = next(
                (h for h, n in norm_map.items() if h not in used and kw in n),
                None,
            )
            if match:
                mapping[field] = match
                used.add(match)
                break
    return mapping


def _norm_date(value: str) -> str:
    v = (value or "").strip()
    if not v:
        return ""
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v[:19] if "T" in v else v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # last resort: pull a YYYY-MM-DD or MM/DD/YYYY substring
    m = re.search(r"(\d{4}-\d{2}-\d{2})", v) or re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", v)
    if m:
        return _norm_date(m.group(1))
    return ""


def _read_rows(path: Path) -> tuple[list[str], list[dict]]:
    """Read a CSV or XLSX into (headers, list-of-row-dicts)."""
    suffix = path.suffix.lower()
    if suffix in (".csv", ".txt", ".tsv"):
        delim = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=delim)
            headers = reader.fieldnames or []
            return list(headers), list(reader)
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(rows, [])]
        out = []
        for r in rows:
            out.append({headers[i]: ("" if v is None else str(v))
                        for i, v in enumerate(r) if i < len(headers)})
        return headers, out
    raise ValueError(f"Unsupported file type: {suffix} (expected .csv/.tsv/.xlsx)")


def _case_to_notice(c: dict, city: str, county: str) -> NoticeData | None:
    """Build a code_violation NoticeData from a normalized field dict.

    Shared by the CSV/row path and the PDF/LLM path. Returns None when the case
    has no usable property address.
    """
    address = (c.get("address") or "").strip()
    if not address or address.lower() in ("none", "n/a", "null", "-"):
        return None
    if address.isupper() or address.islower():
        address = address.title()

    desc = (c.get("violation_description") or "").strip()
    status = (c.get("status") or "").strip()
    raw_bits = [b for b in (desc, f"Status: {status}" if status else "") if b]

    return NoticeData(
        date_added=_norm_date(c.get("date_added", "")),
        address=address,
        city=((c.get("city") or "").strip() or city).title(),
        state="TX",
        zip=re.sub(r"\D", "", c.get("zip", "") or "")[:5],
        owner_name=(c.get("owner_name") or "").strip(),
        notice_type="code_violation",
        county=county,
        source_url=f"{city} code enforcement (open-records request)",
        raw_text=" | ".join(raw_bits),
        parcel_id=(c.get("parcel_id") or "").strip(),
        violation_description=desc,
        compliance_deadline=_norm_date(c.get("compliance_deadline", "")),
        case_status=status,
    )


def parse_response_file(path: str | Path, city: str, county: str) -> list[NoticeData]:
    """Parse a city's code-enforcement export (CSV/TSV/Excel/PDF) into NoticeData.

    Output is run through the shared neglect filter (drops closed cases + keeps
    only neglect/distress violation types) — same gate as the Travis scraper.
    """
    path = Path(path)
    if path.suffix.lower() == ".pdf":
        notices = _parse_pdf(path, city, county)
    else:
        notices = _parse_tabular(path, city, county)

    from violation_filter import filter_code_violations
    return filter_code_violations(notices)


def _parse_tabular(path: Path, city: str, county: str) -> list[NoticeData]:
    """CSV / TSV / Excel → NoticeData via header-heuristic column mapping."""
    headers, rows = _read_rows(path)
    if not headers:
        logger.warning("Open-records ingest: %s has no headers", path.name)
        return []

    mapping = _map_columns(headers)
    if "address" not in mapping:
        logger.warning(
            "Open-records ingest: no address-like column found in %s (headers=%s)",
            path.name, headers,
        )
        return []
    logger.info("Open-records ingest (%s/%s): column map = %s", city, county, mapping)

    notices: list[NoticeData] = []
    skipped = 0
    for row in rows:
        c = {field: (row.get(col, "") if col else "")
             for field, col in mapping.items()}
        notice = _case_to_notice(c, city, county)
        if notice is None:
            skipped += 1
            continue
        notices.append(notice)

    logger.info(
        "Open-records ingest (%s/%s): %d records, %d skipped (no address)",
        city, county, len(notices), skipped,
    )
    return notices


# ── PDF path (digital text layer or scanned → OCR → multi-record LLM) ──

_PDF_TEXT_MIN = 120        # chars of text layer below which we treat it as scanned
_PDF_CHUNK = 6000          # char window per LLM extraction call
_PDF_OVERLAP = 400         # window overlap so cases on a boundary aren't lost

_CASES_SYSTEM = (
    "You are a precise data-extraction engine for Texas city/county "
    "code-enforcement records. You output only valid JSON and never invent data."
)

_CASES_PROMPT = """\
From the following {city}, {county} County, Texas code-enforcement / code-compliance \
records, extract EVERY distinct case.

Return ONLY a JSON object of this exact shape:
{{"cases": [{{"address": "", "city": "", "zip": "", "parcel_id": "", "date_added": "", \
"violation_description": "", "status": "", "owner_name": ""}}]}}

Rules:
- "address" = the street address (number + street) of the cited PROPERTY. Required; \
omit any case with no property address.
- "date_added" = date the case was opened/filed (copy whatever format you see).
- "violation_description" = the violation type / nature / description.
- Use "" for any field not present. Return {{"cases": []}} if there are no cases.

Records:
{text}"""


def _extract_pdf_text(path: Path) -> str:
    """Get a PDF's text via pdfminer; fall back to OCR for scanned PDFs."""
    text = ""
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(str(path)) or ""
    except Exception as e:
        logger.warning("Open-records PDF: pdfminer failed on %s: %s", path.name, e)

    if len(text.strip()) >= _PDF_TEXT_MIN:
        return text

    # Sparse/empty text layer → scanned PDF; OCR each page (reuses the photo stack).
    logger.info("Open-records PDF: %s has no usable text layer — OCR fallback", path.name)
    try:
        from pdf_importer import render_pdf_pages
        from image_utils import fix_rotation, ocr_page
        return "\n".join(ocr_page(fix_rotation(img)) for img in render_pdf_pages(path))
    except Exception as e:
        logger.error("Open-records PDF: OCR fallback failed on %s: %s", path.name, e)
        return text


def _extract_cases_llm(text: str, city: str, county: str, api_key: str) -> list[dict]:
    """Pull all code-enforcement cases out of one text chunk via the LLM."""
    prompt = _CASES_PROMPT.format(city=city, county=county, text=text)
    parsed = llm_client.chat_json(prompt, system=_CASES_SYSTEM, max_tokens=4096, api_key=api_key)
    if isinstance(parsed, dict) and isinstance(parsed.get("cases"), list):
        return [c for c in parsed["cases"] if isinstance(c, dict)]
    return []


def _parse_pdf(path: Path, city: str, county: str) -> list[NoticeData]:
    text = _extract_pdf_text(path)
    if not text.strip():
        logger.warning("Open-records PDF: no text extracted from %s", path.name)
        return []

    api_key = config.ANTHROPIC_API_KEY
    if getattr(config, "LLM_BACKEND", "anthropic") == "anthropic" and not api_key:
        logger.warning(
            "Open-records PDF: ANTHROPIC_API_KEY not set — cannot extract %s "
            "(convert to CSV or set the key).", path.name,
        )
        return []

    # Window long documents so each LLM call stays in-budget; dedup across the
    # overlapping boundaries by (address, opened-date).
    notices: list[NoticeData] = []
    seen: set[tuple[str, str]] = set()
    step = max(1, _PDF_CHUNK - _PDF_OVERLAP)
    for start in range(0, len(text), step):
        chunk = text[start:start + _PDF_CHUNK]
        if not chunk.strip():
            continue
        for case in _extract_cases_llm(chunk, city, county, api_key):
            notice = _case_to_notice(case, city, county)
            if notice is None:
                continue
            key = (notice.address.lower(), notice.date_added)
            if key in seen:
                continue
            seen.add(key)
            notices.append(notice)

    logger.info("Open-records PDF (%s/%s): %d cases from %s",
                city, county, len(notices), path.name)
    return notices
