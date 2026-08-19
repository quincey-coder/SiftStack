"""Obituary opportunity ranking for a DataSift/reisift account (built for ty+2).

The premise: obituary records are the LEAST saturated distressed list we hold. A
notice-of-default owner is on every wholesaler's mail drop and dialer. A decedent
home is only reachable after somebody does the research (who died, who inherited,
who signs), so the competition thins out to whoever is willing to do that work.
This ranks the obituary universe so a lean budget goes to the few records where
real equity, real financial pressure, and near-zero competition overlap.

Pipeline:
  1  pull   every record on the account's Obituary list (detail + custom fields)
  2  gate   drop anything that cannot be worked (see GATES below)
  3  score  six weighted components, 0-100 composite, fully auditable
  4  ship   branded Excel workbook, ranked top to bottom

GATES (hard, each one counted and reported on the Excluded sheet):
  - obituary date missing, or fewer than MIN_OBIT_MONTHS months old
  - already sold (status sold, or an MLS sale recorded after the death)
  - upside down (debt swallowing the equity)
  - do-not-mail / opt-out
  - outside the buy box (non single family, or value over MAX_VALUE)

Usage:
  python src/obituary_opportunity.py --pull                    # refresh the cache
  python src/obituary_opportunity.py --out Obituary_Opportunities.xlsx
  python src/obituary_opportunity.py --account datasift-apikey --min-months 3
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── account wiring ────────────────────────────────────────────────────

CLIENTS = Path(
    os.getenv("DEALROOM_API_PATH", "")
)
BASE = "https://apiv2.reisift.io"
OBIT_LIST_TITLE = "Obituary"
DEFAULT_ACCOUNT = "datasift-apikey"  # ty+2 Open API key, no expiry
CACHE = Path("output") / "obituary_raw.json"

NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "CC0000"

# ── model constants ───────────────────────────────────────────────────

MIN_OBIT_MONTHS = 3      # Ty's rule: give probate time to open
MAX_VALUE = 700_000      # buy box ceiling
MIN_VALUE = 1

# Weights are set from the measured distribution of THIS list, not from intuition.
# Ty+2's obituary universe is 63% free and clear and 99% carries no auction-track
# list, so equity and quietness are near-constants: they are what makes the whole
# list worth working, but they barely separate one record from another inside it.
# The variables that actually spread are distress (rare, therefore decisive when
# present), dataflik's investor score, realtor score and house age.
W_DISTRESS, W_FIT, W_EQUITY, W_TIMING, W_SATURATION, W_CONTACT = 28, 22, 20, 12, 10, 8

# Lists that mean "this owner is already being blasted by every other buyer".
AUCTION_TRACK = {
    "Pre-Foreclosures - Notice of Default", "Pre-Foreclosures - Lis Pendens",
    "Pre-Foreclosures - Final Judgement", "Pre-Foreclosures - Notice of Foreclosure",
    "Pre-Foreclosure", "Foreclosure", "Auction", "Bank Owned (REO)",
    "Tax Sale", "FTM Tax Sale", "Public Sale", "Trustee Deed", "Court Order",
}
# Physical-distress lists (real pressure, still quiet).
PHYSICAL_DISTRESS = {"Condemned", "Code Enforcement", "Fire Damage", "Utility Shut-Off"}

SFR_TYPES = {"single family residential", "single family", "sfr", "residential"}

# Statuses that mean the record is finished: we already worked it and the answer
# was no. Spotted live on 205 Shasta Dr, which scored top of the list on its
# fundamentals while sitting at not_interested.
DEAD_STATUSES = {"not_interested", "not interested", "dead", "dead lead", "wrong number",
                 "wrong_number", "do not contact", "do_not_contact", "dnc", "sold"}
# Statuses that mean somebody is already on it. Not a reason to exclude, but a
# reason not to hand it fresh budget as if it were a new lead.
IN_PROGRESS_STATUSES = {"deep prospecting", "prospecting", "warm lead", "hot lead",
                        "offer made", "under contract", "negotiating"}


# ── small helpers ─────────────────────────────────────────────────────

def num(v, default=None):
    """Best-effort float from the API's mixed string/number money fields."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)) or "nan")
    except ValueError:
        return default


def parse_date(v):
    if not v:
        return None
    s = str(v)[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clamp(x, lo=0.0, hi=1.0):
    return max(lo, min(hi, x))


def ramp(x, lo, hi):
    """0 at or below lo, 1 at or above hi, linear between."""
    if x is None or hi <= lo:
        return 0.0
    return clamp((x - lo) / (hi - lo))


# ── API layer (read-only, 429-aware) ──────────────────────────────────

class Reader:
    """Minimal read-only reisift client that respects the server rate limit."""

    def __init__(self, account: str, gap: float = 0.45):
        sys.path.insert(0, str(CLIENTS))
        os.environ["REISIFT_ACCOUNT"] = account
        from reisift_auth import get_headers  # noqa: E402
        self._headers = get_headers
        self.gap = gap

    def _call(self, path, *, method="GET", body=None, override=None, tries=6):
        for _ in range(tries):
            time.sleep(self.gap)
            hdrs = self._headers(
                content_type="application/json" if body is not None else None,
                method_override=override,
            )
            data = json.dumps(body).encode() if body is not None else None
            req = urllib.request.Request(BASE + path, data=data, headers=hdrs, method=method)
            try:
                with urllib.request.urlopen(req, timeout=25) as r:
                    raw = r.read()
                    return json.loads(raw) if raw else {}
            except urllib.error.HTTPError as e:
                txt = e.read().decode("utf-8", "replace")
                if e.code == 429:
                    m = re.search(r"available in (\d+)", txt)
                    wait = min((int(m.group(1)) if m else 60) + 3, 300)
                    print(f"      rate limited, sleeping {wait}s", flush=True)
                    time.sleep(wait)
                    continue
                raise RuntimeError(f"HTTP {e.code} on {path}: {txt[:200]}")
        raise RuntimeError(f"rate-limit retries exhausted on {path}")

    def obituary_list_uuid(self):
        r = self._call("/api/internal/list/?limit=999")
        for l in (r.get("results") or r.get("data") or []):
            if (l.get("title") or "").strip().lower() == OBIT_LIST_TITLE.lower():
                return l["uuid"]
        raise RuntimeError(f"No '{OBIT_LIST_TITLE}' list on this account")

    def list_records(self, list_uuid, ptype):
        out, offset = [], 0
        while True:
            r = self._call(
                "/api/internal/property/", method="POST", override="GET",
                body={"limit": 100, "offset": offset, "ordering": "-list_count",
                      "query": {"must": {"any_lists": [list_uuid], "property_type": ptype}}},
            )
            rows = r.get("results") or r.get("data") or []
            out.extend(rows)
            total = r.get("count") or 0
            offset += 100
            print(f"    {ptype}: {len(out)}/{total}", flush=True)
            if offset >= total or not rows:
                return out

    def detail(self, uuid):
        return self._call(f"/api/internal/property/{uuid}/")

    def custom_fields(self, uuid):
        r = self._call(f"/api/internal/property/{uuid}/custom-field/")
        return r.get("results") or r.get("data") or (r if isinstance(r, list) else [])

    def option_labels(self):
        """option uuid -> label, so select-type custom fields read as words."""
        r = self._call("/api/internal/custom-fields/?limit=999")
        items = r.get("results") or r.get("data") or (r if isinstance(r, list) else [])
        out = {}
        for f in items:
            for o in (f.get("options") or []):
                if o.get("uuid"):
                    out[o["uuid"]] = o.get("label") or o.get("title") or o.get("value")
        return out


def pull(account: str, cache: Path = CACHE):
    """Pull the whole obituary list to a JSON cache. Resumable: reruns only refill gaps."""
    rd = Reader(account)
    cache.parent.mkdir(parents=True, exist_ok=True)
    existing = {}
    if cache.exists():
        for x in json.loads(cache.read_text(encoding="utf-8")).get("records", []):
            existing[x["uuid"]] = x

    print(f"Paginating the {OBIT_LIST_TITLE} list...", flush=True)
    lu = rd.obituary_list_uuid()
    thin = rd.list_records(lu, "clean") + rd.list_records(lu, "incomplete")
    rows, seen = [], set()
    for r in thin:
        if r["uuid"] not in seen:
            seen.add(r["uuid"])
            rows.append(r)
    print(f"  {len(rows)} unique obituary records\n", flush=True)

    print("Hydrating detail + custom fields...", flush=True)
    out, t0 = [], time.time()
    for i, r in enumerate(rows, 1):
        u = r["uuid"]
        prev = existing.get(u, {})
        rec = prev.get("record")
        cf = prev.get("custom_fields")
        err = None
        try:
            if rec is None:
                rec = rd.detail(u)
            if cf is None:
                cf = rd.custom_fields(u)
        except Exception as e:
            err = str(e)[:200]
        out.append({"uuid": u, "thin": r, "record": rec, "custom_fields": cf, "error": err})
        if i % 25 == 0:
            cache.write_text(json.dumps(
                {"pulled": date.today().isoformat(), "account": account, "records": out},
                default=str), encoding="utf-8")
            print(f"    {i}/{len(rows)}  ({time.time()-t0:.0f}s)", flush=True)

    opts = {}
    try:
        opts = rd.option_labels()
    except Exception as e:
        print("  (option labels unavailable:", str(e)[:80], ")")

    cache.write_text(json.dumps(
        {"pulled": date.today().isoformat(), "account": account,
         "option_labels": opts, "records": out}, default=str), encoding="utf-8")
    ok = sum(1 for x in out if x["record"])
    print(f"\nWrote {cache} — {ok}/{len(out)} hydrated, {time.time()-t0:.0f}s")
    return cache


# ── normalize one record into a flat, scoreable dict ──────────────────

def flatten(entry: dict, opt_labels: dict) -> dict:
    """Turn the raw API payload into the flat field set the model scores on."""
    rec = entry.get("record") or {}
    thin = entry.get("thin") or {}
    addr = rec.get("address") or thin.get("address") or {}
    own = rec.get("owner") or thin.get("owner") or {}
    own_addr = own.get("address") or {}

    cf = {}
    for row in (entry.get("custom_fields") or []):
        lbl = ((row.get("custom_field") or {}).get("label") or "").strip()
        val = row.get("value")
        if isinstance(val, str) and val in opt_labels:
            val = opt_labels[val]
        if lbl:
            cf[lbl] = val

    lists = set(rec.get("lists") or [])
    tags = set(rec.get("tags") or [])
    phones = own.get("phones") or []
    phone_tags = {t for p in phones for t in (p.get("tags") or [])}

    est = num(rec.get("estimate_value"))
    eq_pct = num(rec.get("equity_percent"))
    eq_dollars = num(cf.get("Est Equity"))
    if eq_dollars is None and est is not None and eq_pct is not None:
        eq_dollars = est * eq_pct / 100.0

    obit = parse_date(rec.get("last_obituary_date"))
    # Death Date / Date of Death are hand-entered by the deep-prospecting pass and
    # beat the aggregator's obituary-publication date when present.
    dod = parse_date(cf.get("Death Date")) or parse_date(cf.get("Date of Death"))
    anchor = dod or obit

    # "Total Delinquency" is liens PLUS taxes, verified live: 6031 Ridgeview reads
    # 13,766.72 = 12,908.72 lien + 858.00 tax. Reading it as the tax figure double
    # counts the lien and inflates exactly the records this model is meant to find,
    # so the tax amount only ever comes from a real tax field.
    tax_amt = num(cf.get("Tax delinquency amount")) or num(rec.get("tax_delinquent_value"))
    total_debt = num(cf.get("Total Delinquency"))

    tax_yrs = num(cf.get("Number of years tax-linked")) or num(rec.get("year_behind_on_taxes"))
    if tax_yrs is None and rec.get("tax_delinquent_year"):
        y = num(rec.get("tax_delinquent_year"))
        if y and 1990 < y <= date.today().year:
            tax_yrs = date.today().year - y + 1
    if tax_yrs is None:
        # The county pull writes the unpaid years into Notes when no field carries them.
        m_yrs = re.search(r"Unpaid county tax years?:\s*([0-9,\s]+)", str(rec.get("notes") or ""))
        if m_yrs:
            yrs = [y for y in re.findall(r"\d{4}", m_yrs.group(1))]
            if yrs:
                tax_yrs = float(len(set(yrs)))

    lien_amt = num(cf.get("Outstanding Lien Amount"))
    active_liens = None
    m = re.match(r"\s*(\d+)\s*of\s*(\d+)", str(cf.get("Liens Active vs Total") or ""))
    if m:
        active_liens = int(m.group(1))

    attempts = sum(num(rec.get(k), 0) or 0 for k in
                   ("sms_attempts", "predictivecall_attempts", "directmail_attempts", "rvm_attempts"))

    st = (rec.get("structure_type") or "").strip().lower()
    mls = (rec.get("mls") or "").strip()

    return {
        "uuid": entry["uuid"],
        "street": addr.get("street"), "city": addr.get("city"),
        "state": addr.get("state"), "zip": (addr.get("postal_code") or "")[:5],
        "county": addr.get("county"),
        "owner_name": " ".join(x for x in [own.get("first_name"), own.get("last_name")] if x).strip()
                      or own.get("company"),
        "owner_company": own.get("company"),
        "owner_type": own.get("type"),
        "owner_first": own.get("first_name"), "owner_last": own.get("last_name"),
        "owner_deceased": bool(own.get("deceased")),
        "mail_street": own_addr.get("street"), "mail_zip": (own_addr.get("postal_code") or "")[:5],
        "mail_city": own_addr.get("city"), "mail_state": own_addr.get("state"),
        "status": rec.get("status"), "lists": sorted(lists), "tags": sorted(tags),
        "list_count": len(lists),
        "vacant": bool(addr.get("vacant")) or ("Vacant" in lists),
        "estimate_value": est, "equity_pct": eq_pct, "equity_dollars": eq_dollars,
        "last_sold": parse_date(rec.get("last_sold")), "last_sale_price": num(rec.get("last_sale_price")),
        "owned_since": parse_date(rec.get("owned_since")),
        "rental_value": num(rec.get("rental_value")),
        "sqft": num(rec.get("sqft")), "beds": rec.get("bedrooms"), "baths": num(rec.get("bathrooms")),
        "year_built": num(rec.get("year")), "structure_type": rec.get("structure_type"),
        "is_sfr": any(k in st for k in SFR_TYPES) and "multi" not in st,
        "mls": mls,
        "obit_date": obit, "dod": dod, "anchor_date": anchor,
        "probate_open": parse_date(rec.get("probate_open_date")),
        "pr_name": rec.get("personal_representative") or cf.get("Personal Representative"),
        "decedent_name": cf.get("Decedent Name"),
        "decision_maker": cf.get("Decision Maker"),
        "obit_url": cf.get("Obituary URL"),
        "tax_amount": tax_amt, "tax_years": tax_yrs, "total_debt": total_debt,
        "lien_amount": lien_amt, "active_liens": active_liens,
        "liens_active_total": cf.get("Liens Active vs Total"),
        "debt_vs_equity_pct": num(cf.get("Debt vs Equity Pct")),
        "upside_down": str(cf.get("Upside Down") or "").strip().lower() in ("yes", "true"),
        "upside_down_reason": cf.get("Upside Down Reason"),
        "investor_score": num(rec.get("investor_score")),
        "on_market_score": num(rec.get("on_market_score")),
        "realtor_score": num(rec.get("realtor_score")),
        "skiptraced": bool(own.get("skiptraced") or thin.get("skiptraced")),
        "n_phones": len(phones), "phone_tags": sorted(phone_tags),
        "n_emails": len(own.get("emails") or []),
        "dnc": bool(own.get("dnc")), "opt_out": bool(own.get("opt_out")),
        "do_not_mail": bool(rec.get("do_not_mail_ever") or own.get("do_not_mail_ever")),
        "attempts": attempts,
        "heirs_resolved": "Heirs Resolved" in tags,
        "auction_lists": sorted(lists & AUCTION_TRACK),
        "physical_lists": sorted(lists & PHYSICAL_DISTRESS),
        "notes": rec.get("notes"),
        "incomplete": (rec.get("type") or thin.get("property_type")) == "incomplete",
    }


# ── gates ─────────────────────────────────────────────────────────────

def gate(r: dict, today: date, min_months: int) -> str | None:
    """Return the reason this record cannot be worked, or None if it qualifies."""
    if not r["anchor_date"]:
        return "No obituary or death date on the record"
    months = (today - r["anchor_date"]).days / 30.44
    if months < min_months:
        return f"Too recent: {months:.1f} months since death, gate is {min_months}"
    st = (r["status"] or "").strip().lower()
    if st in DEAD_STATUSES or r["mls"].strip().lower() == "sold":
        return f"Already worked and closed out: status {r['status'] or 'sold'}"
    if r["last_sold"] and r["anchor_date"] and r["last_sold"] > r["anchor_date"]:
        return f"Estate already transacted: sold {r['last_sold']} after the death"
    if r["upside_down"]:
        return f"Upside down: {r['upside_down_reason'] or 'debt exceeds equity'}"
    if r["do_not_mail"] or r["opt_out"]:
        return "Do not mail / opted out"
    if r["estimate_value"] is None:
        return "No value on the record (cannot underwrite)"
    if r["estimate_value"] > MAX_VALUE:
        return f"Over buy box: value ${r['estimate_value']:,.0f} > ${MAX_VALUE:,.0f}"
    if r["estimate_value"] < MIN_VALUE:
        return "No value on the record (cannot underwrite)"
    if not r["is_sfr"]:
        return f"Not single family: {r['structure_type'] or 'unknown structure type'}"
    return None


# ── the six scoring components ────────────────────────────────────────

def score_distress(r):
    """Hard financial and physical pressure on the estate. On this list only about
    3% of records carry any of it, which is exactly why it is the heaviest weight:
    a rare signal is a decisive one. Points are additive and capped at the weight."""
    pts, why = 0.0, []

    yr_s = 0.0 if not r["tax_years"] else clamp(r["tax_years"] / 2.0)
    amt_s = ramp(r["tax_amount"], 500, 8_000)
    tax = max(yr_s, amt_s)
    pts += 9 * tax
    if tax > 0:
        bits = []
        if r["tax_years"]:
            bits.append(f"{r['tax_years']:.0f}yr behind on taxes")
        if r["tax_amount"]:
            bits.append(f"${r['tax_amount']:,.0f} tax delinquent")
        why.append(", ".join(bits) or "tax delinquent")

    if r["lien_amount"] and (r["active_liens"] is None or r["active_liens"] > 0):
        pts += 7 * clamp(0.35 + 0.65 * ramp(r["lien_amount"], 500, 12_000))
        why.append(f"${r['lien_amount']:,.0f} in active liens")

    if r["vacant"]:
        pts += 6
        why.append("VACANT")

    if r["physical_lists"]:
        pts += 4 * (1.0 if len(r["physical_lists"]) > 1 else 0.65)
        why.append(", ".join(r["physical_lists"]).lower())

    absentee = ("Absentee Owners" in r["lists"]
                or (r["mail_zip"] and r["zip"] and r["mail_zip"] != r["zip"]))
    if absentee:
        pts += 3
        why.append("absentee: not the occupant's home")

    fragile = [l for l in ("Low Income", "Low Credit Score", "Negative Equity") if l in r["lists"]]
    if fragile:
        pts += 2
        why.append(", ".join(fragile).lower())

    return min(pts, W_DISTRESS), why


def score_fit(r):
    """Would an investor actually buy this, and is it out of retail's reach? A
    free-and-clear $600k house in good shape is a listing, not a wholesale deal.
    dataflik's investor score, the realtor score and the year built are the three
    variables with real spread on this list, so this is where separation comes from."""
    pts, why = 0.0, []

    inv = ramp(r["investor_score"], 20, 90)
    pts += 10 * inv
    if (r["investor_score"] or 0) >= 75:
        why.append(f"investor score {r['investor_score']:.0f}")

    # High realtor score means the heirs will list it and an agent will win it.
    rs = r["realtor_score"]
    if rs is not None:
        pts += 4 * (1.0 - ramp(rs, 30, 90))
        if rs <= 40:
            why.append("agents are not competing for it")

    yb = r["year_built"]
    if yb:
        pts += 4 * (1.0 if yb <= 1960 else 0.6 if yb <= 1980 else 0.25 if yb <= 2000 else 0.0)
        if yb <= 1960:
            why.append(f"built {yb:.0f}, rehab lane")

    if r["rental_value"] and r["estimate_value"]:
        y = r["rental_value"] * 12 / r["estimate_value"] * 100
        pts += 2 * ramp(y, 6, 10)
        if y >= 9:
            why.append(f"{y:.0f}% gross yield, landlord exit")

    # Deal-size band: cheap enough that retail buyers hesitate, rich enough to pay.
    v = r["estimate_value"]
    if v:
        if 120_000 <= v <= 380_000:
            pts += 2
        elif v < 120_000 or v <= 480_000:
            pts += 1

    return min(pts, W_FIT), why


def score_equity(r):
    """Room in the deal. Percent drives feasibility (with no lender, nobody can
    block a discounted sale), dollars drive whether the profit justifies the work.
    Weighted below distress because 63% of this list is already free and clear,
    so equity mostly tells you the list is good, not which record is."""
    pct = ramp(r["equity_pct"], 30, 95)
    dollars = ramp(r["equity_dollars"], 75_000, 350_000)
    why = []
    if (r["equity_pct"] or 0) >= 99:
        why.append("free and clear")
    elif (r["equity_pct"] or 0) >= 70:
        why.append(f"{r['equity_pct']:.0f}% equity")
    if (r["equity_dollars"] or 0) >= 250_000:
        why.append(f"${r['equity_dollars']:,.0f} equity")
    return W_EQUITY * (0.5 * pct + 0.5 * dollars), why


def score_saturation(r):
    """Start at full marks and subtract every reason somebody else is already
    calling this owner. Measured on ty+2: 99% of the obituary list carries no
    auction-track distressor at all, so this component is near-constant and is
    weighted accordingly. It is what makes the LIST valuable, not what ranks
    inside it. Its job here is to push the rare contaminated record down."""
    pts, why = float(W_SATURATION), []

    n_auc = len(r["auction_lists"])
    if n_auc:
        pts -= min(4 + 2 * (n_auc - 1), 7)
        why.append(f"also on {', '.join(r['auction_lists'])}, publicly marketed to everyone")

    if "Tax Delinquent" in r["lists"]:
        pts -= 1
        why.append("county publishes the delinquent roll")

    if (r["on_market_score"] or 0) >= 70 or r["mls"].strip().lower() in (
            "listed", "for sale", "active", "pending"):
        pts -= 2.5
        why.append(f"on market ({r['mls'] or 'high on-market score'})")

    if r["attempts"]:
        pts -= min(r["attempts"] * 0.5, 2)
        why.append(f"we have already touched it {r['attempts']:.0f}x")

    if r["list_count"] >= 6:
        pts -= 1
        why.append(f"on {r['list_count']} lists, heavily flagged in bulk data")

    pts = max(pts, 0.0)
    if pts >= W_SATURATION - 0.01:
        why = ["nobody else is working it: obituary is the only signal on the record"]
    return pts, why


def score_timing(r, today):
    """Where the estate sits in the probate cycle. TN vests real property in the
    heirs at death and the creditor window is 4 months, so months 6-15 is when a
    decedent home is both legally sellable and emotionally ready to be sold."""
    months = (today - r["anchor_date"]).days / 30.44
    if months < 6:
        base = 0.75
    elif months < 9:
        base = 0.95
    elif months < 15:
        base = 1.00
    elif months < 24:
        base = 0.80
    else:
        base = 0.60

    bonus, why = 0.0, [f"{months:.0f} months since death"]
    if r["probate_open"]:
        bonus += 0.15
        why.append(f"probate open {r['probate_open']}")
    if r["pr_name"]:
        bonus += 0.10
        why.append(f"PR known: {r['pr_name']}")
    return W_TIMING * clamp(base + bonus), why


def score_contact(r):
    """How much work stands between this record and a human on the phone. The
    smallest weight on purpose: a cold obituary record costs about $0.24 to resolve
    through the v5 SmartSkip flow, so poor contactability is a cost line, not a
    reason to skip a good house."""
    pts, why = 0.0, []
    if r["heirs_resolved"] or r["decision_maker"]:
        pts += 3
        why.append("heirs already resolved")
    if r["skiptraced"]:
        pts += 1.5
    tags = set(r["phone_tags"])
    if "Dial First" in tags:
        pts += 2
        why.append("Dial First number on file")
    elif "Dial Second" in tags:
        pts += 1.5
    elif r["n_phones"]:
        pts += 0.8
    if r["n_emails"]:
        pts += 0.7
    return min(pts, W_CONTACT), why


# ── judgment flags (things a human must check before working the lead) ──

def flags(r, today):
    out = []
    # The spouse-obituary trap: an obituary attached to the record does NOT mean
    # the owner of record died. Calling a living widow about her dead husband's
    # estate is the worst call this list can produce.
    if not r["decedent_name"]:
        out.append("VERIFY DECEDENT: no decedent name on file, confirm the person who "
                   "died is the owner of record before any heir approach")
    elif r["owner_name"]:
        dl = (r["decedent_name"] or "").strip().lower().split()
        ol = (r["owner_name"] or "").strip().lower().split()
        if dl and ol and dl[-1] != ol[-1]:
            out.append(f"VERIFY DECEDENT: decedent '{r['decedent_name']}' does not share a "
                       f"surname with owner '{r['owner_name']}', likely a spouse or relative")
    if r["owner_company"]:
        out.append("Entity owner: SmartSkip cannot name-trace an LLC or trust, route to "
                   "Enformion BusinessV2 for the principal")
    months = (today - r["anchor_date"]).days / 30.44
    if months >= 6 and not r["probate_open"]:
        out.append("No probate opened past 6 months: either the heirs are dormant "
                   "(your opening) or title is contested, find out which before spending")
    if (r["status"] or "").strip().lower() in IN_PROGRESS_STATUSES:
        out.append(f"Already in your pipeline at status '{r['status']}': pick up the existing "
                   f"thread, do not hand it fresh budget as a new lead")
    if r["incomplete"]:
        out.append("Incomplete record: address or owner data is partial")
    if r["dnc"]:
        out.append("Owner on DNC: mail and door only, no dialer")
    return out


# Thresholds are set from the observed score distribution on ty+2 (p50 51.8,
# p95 59.1, max 68.5), not from a round number. A is roughly the top 4%: the
# short list a lean budget can actually work by hand.
TIER_A, TIER_B, TIER_C = 60.0, 54.0, 46.0


def tier(score):
    if score >= TIER_A:
        return "A"
    if score >= TIER_B:
        return "B"
    if score >= TIER_C:
        return "C"
    return "D"


def assemble(cache: Path, min_months: int, today: date):
    """Score the cache. Returns (qualified sorted desc, excluded, meta)."""
    blob = json.loads(cache.read_text(encoding="utf-8"))
    if isinstance(blob, list):           # scratch-format cache
        entries, opts, pulled = blob, {}, None
    else:
        entries = blob.get("records") or []
        opts = blob.get("option_labels") or {}
        pulled = blob.get("pulled")

    qualified, excluded, unhydrated = [], [], 0
    for e in entries:
        if not e.get("record"):
            unhydrated += 1
            excluded.append({"street": (e.get("thin") or {}).get("address", {}).get("street"),
                             "uuid": e["uuid"], "reason": "Could not load the record from the API"})
            continue
        r = flatten(e, opts)
        why_not = gate(r, today, min_months)
        if why_not:
            r["reason"] = why_not
            excluded.append(r)
            continue

        di, w_di = score_distress(r)
        fi, w_fi = score_fit(r)
        eq, w_eq = score_equity(r)
        ti, w_ti = score_timing(r, today)
        sa, w_sa = score_saturation(r)
        co, w_co = score_contact(r)
        r.update({
            "s_distress": round(di, 1), "s_fit": round(fi, 1), "s_equity": round(eq, 1),
            "s_timing": round(ti, 1), "s_saturation": round(sa, 1), "s_contact": round(co, 1),
            "score": round(di + fi + eq + ti + sa + co, 1),
            "months_since": round((today - r["anchor_date"]).days / 30.44, 1),
            "flags": flags(r, today),
            "hard_distress": bool(r["tax_amount"] or r["tax_years"] or r["lien_amount"]
                                  or r["vacant"] or r["physical_lists"]),
        })
        r["tier"] = tier(r["score"])
        # Rationale: name the strongest drivers in plain words, hard distress first
        # because that is the rare signal a human should react to.
        drivers = [(di + 100, w_di)] if w_di and r["hard_distress"] else []
        drivers += [(x, w) for x, w in ((di, w_di), (fi, w_fi), (eq, w_eq),
                                        (sa, w_sa), (ti, w_ti)) if w and (x, w) not in drivers]
        drivers.sort(key=lambda t: -t[0])
        seen_bits, bits = set(), []
        for _, ws in drivers:
            for b in ws:
                if b not in seen_bits:
                    seen_bits.add(b)
                    bits.append(b)
        r["why"] = "; ".join(bits[:4]) if bits else "no standout driver"
        qualified.append(r)

    qualified.sort(key=lambda x: -x["score"])
    meta = {"pulled": pulled, "total": len(entries), "unhydrated": unhydrated,
            "qualified": len(qualified), "excluded": len(excluded)}
    return qualified, excluded, meta


# ── workbook ──────────────────────────────────────────────────────────

def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=15, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=10, color="666666")


def _header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kv(ws, start, rows):
    r = start
    for pair in rows:
        a, b = (list(pair) + ["", ""])[:2]
        ws.cell(row=r, column=1, value=a)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if a and b == "":
            ws.cell(row=r, column=1).font = Font(bold=True, color=BLUE)
        r += 1
    return r


def _rows(ws, start, rows, money_cols=(), pct_cols=()):
    r = start
    for row in rows:
        for j, v in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if j in money_cols and isinstance(v, (int, float)):
                c.number_format = "$#,##0"
            if j in pct_cols and isinstance(v, (int, float)):
                c.number_format = '0"%"'
        r += 1
    return r


DETAIL_COLS = [
    ("Rank", 6), ("Tier", 5), ("Score", 7), ("Address", 26), ("City", 14), ("Zip", 7),
    ("County", 9), ("Owner of record", 20), ("Value", 11), ("Equity %", 9), ("Equity $", 12),
    ("Beds", 5), ("Baths", 6), ("SqFt", 7), ("Built", 6), ("Rent", 9),
    ("Months since death", 9), ("Vacant", 7),
    ("Tax owed", 10), ("Yrs behind", 8), ("Active liens", 11),
    ("Investor score", 8), ("Realtor score", 8),
    ("Distress pts", 8), ("Fit pts", 7), ("Equity pts", 8), ("Timing pts", 8),
    ("Quiet pts", 8), ("Contact pts", 8),
    ("Competing lists", 24), ("Phones", 7), ("Why it ranks", 46), ("Must verify", 52),
]


def _detail_row(i, r):
    return [
        i, r["tier"], r["score"], r["street"], r["city"], r["zip"], r["county"],
        r["owner_name"], r["estimate_value"], r["equity_pct"], r["equity_dollars"],
        r["beds"], r["baths"], r["sqft"], r["year_built"], r["rental_value"],
        r["months_since"], "yes" if r["vacant"] else "",
        r["tax_amount"], r["tax_years"], r["lien_amount"],
        r["investor_score"], r["realtor_score"],
        r["s_distress"], r["s_fit"], r["s_equity"], r["s_timing"],
        r["s_saturation"], r["s_contact"],
        ", ".join(r["auction_lists"]) or "none", r["n_phones"],
        r["why"], " | ".join(r["flags"]),
    ]


def build_workbook(qualified, excluded, meta, out: Path, min_months: int,
                   today: date, mail_cost: float, touches: int, top_n: int):
    wb = Workbook()

    # 1 Overview -------------------------------------------------------
    ws = wb.active
    ws.title = "1 Overview"
    _title(ws, "Obituary opportunity ranking",
           f"ty+2 account, obituary list, run {today.isoformat()}"
           + (f", data pulled {meta['pulled']}" if meta.get("pulled") else ""))
    _widths(ws, [34, 96])

    a = [x for x in qualified if x["tier"] == "A"]
    b = [x for x in qualified if x["tier"] == "B"]
    quiet = [x for x in qualified if not x["auction_lists"]]
    top = qualified[:top_n]
    all_mail = meta["total"] * mail_cost * touches
    top_mail = len(top) * mail_cost * touches
    dp_cost = len(top) * 0.24

    _kv(ws, 4, [
        ["The premise", ""],
        ["", "A notice-of-default owner is on every wholesaler's mail drop and every dialer, "
             "because the filing is public and instantly machine readable. A decedent home is "
             "only reachable after somebody researches who died, who inherited, and who can "
             "sign. That research is the moat. This ranking spends a lean budget where equity, "
             "real financial pressure, and near-zero competition overlap."],
        ["", ""],
        ["The funnel", ""],
        ["Obituary records on the account", meta["total"]],
        ["Dropped by the gates (see sheet 6)", meta["excluded"]],
        ["Qualified and ranked", meta["qualified"]],
        [f"Tier A (score {TIER_A:.0f}+), the short list to work by hand", len(a)],
        [f"Tier B (score {TIER_B:.0f} to {TIER_A:.0f})", len(b)],
        ["Quiet: obituary carries no auction-track list",
         f"{len(quiet)} of {len(qualified)} ({100*len(quiet)/max(len(qualified),1):.0f}%)"],
        ["Carrying hard distress (tax, lien, vacancy, code)",
         f"{len([x for x in qualified if x['hard_distress']])} of {len(qualified)} "
         f"({100*len([x for x in qualified if x['hard_distress']])/max(len(qualified),1):.0f}%)"],
        ["", ""],
        ["Two findings that shaped the ranking", ""],
        ["Saturation is a list-level advantage, not a ranking variable",
         f"{100*len(quiet)/max(len(qualified),1):.0f}% of the qualified obituary universe "
         "carries no foreclosure, tax sale, auction or trustee-deed flag at all. The premise "
         "is confirmed: almost nobody else has a machine-readable reason to call these owners. "
         "But because it is true of nearly every record, quietness cannot rank them against "
         "each other, which is why saturation is weighted 10 rather than 20. The competitive "
         "edge is in choosing this list over the foreclosure list, and it is already banked."],
        ["The list is clean-title, so the death IS the motivation",
         "63% of qualified records are free and clear and only about 3% carry any tax "
         "delinquency, lien, vacancy or code action. This is not a distressed-debt list, it is "
         "a list of paid-off senior homes whose owner died. Sheet 3 isolates the few records "
         "where hard financial pressure is also present. Everywhere else the seller's reason "
         "to move is the estate itself, so the pitch is speed, certainty and as-is, not rescue."],
        ["", ""],
        ["The budget case", ""],
        ["Assumption", f"${mail_cost:.2f} per mail piece, {touches} touches per record"],
        ["Mail the whole obituary list", f"${all_mail:,.0f}"],
        [f"Mail only the top {len(top)}", f"${top_mail:,.0f}"],
        ["Deep prospect the same top records",
         f"${dp_cost:,.0f} at $0.24 each (SmartSkip, Tracerfy gap-fill, obituary research, "
         f"Trestle tiers)"],
        ["What that buys",
         f"${all_mail - top_mail - dp_cost:,.0f} freed, spent on records where you already "
         f"know who signs instead of postcards to a dead person's mailbox"],
        ["", ""],
        ["How to work it", ""],
        ["1", "Run the deep-prospecting v5 flow on Tier A first. Confirm the decedent IS the "
              "owner of record before anything else. See the Must verify column."],
        ["2", "Every record carries a Must verify note. That column is not optional reading. "
              "The most common failure on this list is calling a living widow about her "
              "husband's estate."],
        ["3", "Call the resolved heirs, do not mail the property address. The decedent's "
              "mailbox is the one place nobody is reading."],
        ["4", "Records with no probate opened past 6 months are the highest upside and the "
              "highest risk. Find out whether the heirs are dormant or the title is contested "
              "before you spend."],
    ])

    # 2 Top opportunities ----------------------------------------------
    ws = wb.create_sheet("2 Top opportunities")
    _title(ws, f"Top {len(top)} obituary opportunities",
           f"Ranked by the composite on sheet 5. Tier A is score {TIER_A:.0f}+, set from this list's own distribution.")
    _widths(ws, [w for _, w in DETAIL_COLS])
    _header(ws, 4, [h for h, _ in DETAIL_COLS])
    _rows(ws, 5, [_detail_row(i, x) for i, x in enumerate(top, 1)],
          money_cols=(9, 11, 16, 19, 21), pct_cols=(10,))
    ws.freeze_panes = "E5"

    # 3 Hard distress ---------------------------------------------------
    hard = [x for x in qualified if x["hard_distress"]]
    ws = wb.create_sheet("3 Hard distress")
    _title(ws, f"{len(hard)} of {len(qualified)} qualified records carry hard distress",
           "Tax delinquency, active liens, vacancy, condemnation or code enforcement. "
           "Everything else on this list is a clean-title free-and-clear home where the "
           "death itself is the only motivation.")
    cols = [("Rank", 6), ("Score", 7), ("Address", 26), ("City", 14), ("Zip", 7),
            ("Value", 11), ("Equity $", 12), ("Vacant", 7), ("Tax owed", 10),
            ("Yrs behind", 8), ("Active liens", 11), ("Liens active/total", 12),
            ("Total debt", 12), ("Debt vs equity %", 10), ("Other lists", 26),
            ("What the county recorded", 70)]
    _widths(ws, [w for _, w in cols])
    _header(ws, 5, [h for h, _ in cols])
    rank = {id(x): i for i, x in enumerate(qualified, 1)}
    _rows(ws, 6, [[
        rank[id(x)], x["score"], x["street"], x["city"], x["zip"],
        x["estimate_value"], x["equity_dollars"], "yes" if x["vacant"] else "",
        x["tax_amount"], x["tax_years"], x["lien_amount"],
        x.get("liens_active_total") or "", x["total_debt"], x["debt_vs_equity_pct"],
        ", ".join(l for l in x["lists"] if l != "Obituary")[:120],
        re.sub(r"\s+", " ", str(x["notes"] or "")).strip()[:600],
    ] for x in hard], money_cols=(6, 7, 9, 11, 13))
    ws.freeze_panes = "C6"

    # 4 All qualified ---------------------------------------------------
    ws = wb.create_sheet("4 All qualified")
    _title(ws, f"All {len(qualified)} qualified obituary records", "Same columns, full ranking.")
    _widths(ws, [w for _, w in DETAIL_COLS])
    _header(ws, 4, [h for h, _ in DETAIL_COLS])
    _rows(ws, 5, [_detail_row(i, x) for i, x in enumerate(qualified, 1)],
          money_cols=(9, 11, 16, 19, 21), pct_cols=(10,))
    ws.freeze_panes = "E5"

    # 4 Scoring model ---------------------------------------------------
    ws = wb.create_sheet("5 Scoring model")
    _title(ws, "How the score is built",
           "Every point is traceable. Change a weight, rerun, compare.")
    _widths(ws, [30, 12, 96])
    _header(ws, 4, ["Component", "Weight", "What it measures and how"])
    _rows(ws, 5, [
        ["How the weights were set", "",
         "From the measured distribution of this list, not from intuition. A first pass "
         "weighted equity 30 and saturation 20 and produced eleven distinct scores across the "
         "top forty records, because on this list equity and quietness are near-constants: "
         "63% of qualified records are free and clear and 99% carry no auction-track list at "
         "all. Those two facts are why the list is worth working, but they cannot rank it. "
         "The weights below put the load on the variables that actually vary."],
        ["", "", ""],
        ["Distress and pressure", W_DISTRESS,
         "Additive, capped at the weight. Tax delinquency 9 (years behind or dollars owed, "
         "whichever is worse, 2+ years is full marks, dollars ramp $500 to $8k). Active liens "
         "7 (released liens do not count, per the release-filtering rule). Vacant 6. Condemned, "
         "code, fire or utility shut-off 4. Absentee, so not the occupant's home, 3. Low income "
         "or low credit or negative equity 2. Heaviest weight precisely BECAUSE it is rare: "
         "only about 3% of this list carries any hard distress, so when it is present it is "
         "the thing that separates an estate that must sell from one that can wait."],
        ["Investor fit and exit", W_FIT,
         "dataflik investor score 10 (ramped 20 to 90). Realtor score inverted 4: a high "
         "realtor score means the heirs will list it and an agent wins it, not you. Year built "
         "4 (pre-1960 full marks, pre-1980 0.6, pre-2000 0.25): older stock means rehab, which "
         "means retail buyers hesitate. Gross rental yield 2 (ramped 6% to 10%) for the "
         "landlord exit. Deal-size band 2, best between $120k and $380k. These three scores "
         "have the widest spread of anything on the record, so this is where separation "
         "actually comes from."],
        ["Equity and room", W_EQUITY,
         "Half equity percent ramped 30% to 95%, half equity dollars ramped $75k to $350k. "
         "Percent drives feasibility: with no lender, nobody can block a discounted sale. "
         "Dollars drive whether the profit justifies the work. Weighted below distress because "
         "most of this list is already free and clear, so equity mostly tells you the list is "
         "good rather than which record is."],
        ["Timing in the probate cycle", W_TIMING,
         "Months since death: under 6 scores 0.75, 6 to 9 scores 0.95, 9 to 15 scores 1.00, "
         "15 to 24 scores 0.80, over 24 scores 0.60. Plus 0.15 if probate is open and 0.10 if "
         "the personal representative is named. Texas vests real property in the heirs at "
         "death (T.C.A. 31-2-103) and the creditor window is 4 months, so months 6 to 15 is "
         "when a decedent home is both legally sellable and emotionally ready to be sold."],
        ["Low saturation", W_SATURATION,
         "Starts at full marks and subtracts competition. On an auction-track list (notice of "
         "default, lis pendens, foreclosure, tax sale, trustee deed, auction, REO) costs 4 for "
         "the first and 2 for each additional, capped at 7. County-published tax delinquent "
         "roll 1. On market or on-market score 70+ costs 2.5. Our own prior touches cost 0.5 "
         "each, capped at 2. On 6+ lists costs 1. Weighted at 10 rather than 20 for an honest "
         "reason: measured on this account, 99% of the obituary list is already quiet, so this "
         "component cannot rank the list, it can only push the rare contaminated record down. "
         "The saturation advantage is real and it is the reason to work this list at all, it "
         "just lives in the choice of list rather than in the order within it."],
        ["Contactability", W_CONTACT,
         "Heirs already resolved 3, skip traced 1.5, Dial First number 2 (Dial Second 1.5, any "
         "phone 0.8), email 0.7. Deliberately the smallest weight: a cold obituary record costs "
         "about $0.24 to resolve through the v5 flow, so poor contactability is a cost line, "
         "not a reason to skip a good house."],
        ["", "", ""],
        ["Gates (hard exclusions)", "",
         f"No obituary or death date. Under {min_months} months since death (Ty's rule: give "
         "probate time to open). Already sold, or an MLS sale recorded after the death. Upside "
         "down. Do not mail or opted out. No value on the record. Value over "
         f"${MAX_VALUE:,.0f}. Not single family."],
        ["Date anchor", "",
         "A hand-entered Death Date or Date of Death beats the aggregator's obituary "
         "publication date when both exist, because the research pass verified it. On this run "
         "no record carried a hand-entered date, so every months-since figure comes from the "
         "aggregator's obituary publication date and is a proxy for the true date of death."],
        ["Known limits, read these", "",
         "Equity percent and value are dataflik estimates, not appraisals, so treat the "
         "ranking as a call order and not as underwriting. The obituary flag only started "
         "flowing 2025-12, so nothing here is older than about 8 months and the 15+ month "
         "bands are empty by construction. Zero records on this account have a probate open "
         "date, a decedent name or a resolved heir, so the research layer is entirely ahead of "
         "you and the timing bonus never fires. Above all: whether the person who died is the "
         "owner of record is NOT verifiable from CRM data, which is why every row carries a "
         "Must verify note."],
    ])
    ws.freeze_panes = "A5"

    # 5 Excluded --------------------------------------------------------
    ws = wb.create_sheet("6 Excluded")
    _title(ws, f"{len(excluded)} records dropped by the gates",
           "Nothing is silently discarded. Every drop has a stated reason.")
    counts = {}
    for x in excluded:
        key = re.sub(r"\$?\d[\d,]*(?:\.\d+)?", "N", x.get("reason") or "unknown")
        counts[key] = counts.get(key, 0) + 1
    _widths(ws, [70, 10, 26, 16, 8, 30])
    _header(ws, 4, ["Reason (grouped)", "Count", "", "", "", ""])
    r = _rows(ws, 5, [[k, v] for k, v in sorted(counts.items(), key=lambda t: -t[1])])
    r += 1
    ws.cell(row=r, column=1, value="Every excluded record").font = Font(bold=True, color=BLUE)
    r += 1
    _header(ws, r, ["Reason", "Months since death", "Address", "City", "Zip", "Owner"])
    _rows(ws, r + 1, [[
        x.get("reason"),
        (round((today - x["anchor_date"]).days / 30.44, 1)
         if x.get("anchor_date") else ""),
        x.get("street"), x.get("city"), x.get("zip"), x.get("owner_name"),
    ] for x in excluded])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--pull", action="store_true", help="refresh the cache from the API first")
    ap.add_argument("--account", default=DEFAULT_ACCOUNT)
    ap.add_argument("--cache", default=str(CACHE))
    ap.add_argument("--out", default="Obituary_Opportunities.xlsx")
    ap.add_argument("--min-months", type=int, default=MIN_OBIT_MONTHS)
    ap.add_argument("--top", type=int, default=50)
    ap.add_argument("--mail-cost", type=float, default=0.75, help="cost per mail piece")
    ap.add_argument("--touches", type=int, default=6, help="mail touches per record")
    ap.add_argument("--today", default=None, help="override the run date, YYYY-MM-DD")
    args = ap.parse_args()

    cache = Path(args.cache)
    today = parse_date(args.today) or date.today()

    if args.pull or not cache.exists():
        pull(args.account, cache)

    qualified, excluded, meta = assemble(cache, args.min_months, today)
    print(f"\n{meta['total']} obituary records | {meta['qualified']} qualified | "
          f"{meta['excluded']} gated out")
    if meta["unhydrated"]:
        print(f"  WARNING: {meta['unhydrated']} records never loaded from the API. "
              f"Rerun with --pull to fill them before trusting the ranking.")

    for i, x in enumerate(qualified[:15], 1):
        print(f"  {i:2}. [{x['tier']}] {x['score']:5.1f}  {str(x['street'])[:28]:28} "
              f"{str(x['city'])[:12]:12} ${x['estimate_value'] or 0:>8,.0f}  {x['why'][:70]}")

    out = build_workbook(qualified, excluded, meta, Path(args.out),
                         args.min_months, today, args.mail_cost, args.touches, args.top)
    print(f"\nWrote {out}")


if __name__ == "__main__":
    main()
