"""Williamson County tax-delinquent scraper.

Downloads "Current Year and Prior Taxes Due (Excel)" from the Wilco tax page
and emits filtered NoticeData. Each XLSX row is one (parcel, tax_year)
combination — a parcel with N years of delinquencies takes N rows. We group
by Quick Ref (parcel ID), sum the per-year balances, and take the earliest
tax year as the first delinquent year.

Source: https://www.wilcotx.gov/761/Property-Tax-Roll-Information-Request
Refresh: ~weekly by Wilco TAC.
Records: ~30K rows → ~13K unique Real-Property parcels after grouping +
  filtering Personal/Mobile/Other prefixes.

Filters applied per-parcel (post-grouping):
  1. Quick Ref prefix == 'R' (Real Property; skip P/M/N).
  2. Years delinquent >= min_years (default 2).
  3. Total fee balance >= min_amount (default $3,000).

Cross-run diff: data/williamson_tax_state/, raw archive: data/williamson_tax_raw/.
Master cross-reference for probate/foreclosure backfill uses the live WCAD
SODA API in cad_lookup.py (already wired) — no master download needed here.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

import wilco_tax_cache
from notice_parser import NoticeData, normalize_court_name
from scrapers import register
from scrapers import tax_delinquent_state as state_mod

logger = logging.getLogger(__name__)


LAST_RUN_DIFF: dict | None = None
LAST_RUN_STATS: dict | None = None
LAST_RUN_REMOVED: dict | None = None
LAST_RUN_REPORT_PATH: Path | None = None
LAST_RUN_RAW_PATH: Path | None = None


DEFAULT_MIN_DELINQUENT_YEARS = 2
DEFAULT_MIN_DELINQUENT_AMOUNT = 3000.0


_HOA_PAT = re.compile(
    r"\b(HOA|HOMEOWNERS|HOMEOWNER'S|OWNERS ASSOC|ASSOCIATION|CONDO ASSN?|"
    r"MANAGEMENT|PROPERTY OWNERS|PROP OWNERS)\b",
    re.IGNORECASE,
)
_BUSINESS_PAT = re.compile(
    r"\b(LLC|L\.L\.C\.|INC|INCORPORATED|CORP|CORPORATION|LTD|LP|"
    r"COMPANY|CO\.?|TRUST|ESTATE|BANK|HOLDINGS|PARTNERS|PARTNERSHIP|"
    r"PROPERTIES|GROUP|FUND|VENTURES|INVESTMENTS)\b",
    re.IGNORECASE,
)


def _is_business(name: str) -> bool:
    return bool(name and _BUSINESS_PAT.search(name))


# Suffixes that Python's .title() mangles — re-uppercase after title-casing.
_UPPERCASE_SUFFIX_PAT = re.compile(
    r"\b(LLC|INC|CORP|LP|LLP|PLLC|PA|PC|CO|MD|DDS|DVM|CPA|ESQ|JR|SR|"
    r"II|III|IV|V|VI|VII|VIII|IX|X|"
    r"USA|TX|TXR|HOA|HUD|VA|FBO|ETUX|ETVIR|ETAL)\b\.?",
    re.IGNORECASE,
)


def _title_case(name: str) -> str:
    """Title-case an ALL-CAPS name, then re-uppercase legal/credential suffixes.

    Python's `.title()` lowercases all-caps suffixes (LLC → Llc, IV → Iv).
    Post-process to restore canonical uppercase form.
    """
    if not name:
        return ""
    out = name.title() if name.isupper() else name
    return _UPPERCASE_SUFFIX_PAT.sub(lambda m: m.group(0).upper(), out)


def _strip_etux_etal(name: str) -> str:
    """Strip trailing ETUX/ETVIR/ETAL clauses + spousal `& <name>` tails.

    Applied to BOTH the individual-name path AND the business-name path —
    Wilco mirrors Bell where the source appends ETAL to either kind of name.
    """
    if not name:
        return ""
    return re.sub(
        r"\s+(?:(?:ETUX|ETVIR|ETAL)\b|&\s).*$",
        "",
        name,
        flags=re.IGNORECASE,
    ).strip()


def _split_situs(situs: str) -> tuple[str, str, str]:
    """Wilco's `Situs Address` packs street + city + zip on one line, e.g.
        '11 KENDALL ST SPC 11 GEORGETOWN 78626'
        '603 WEST DR S5 LEANDER 78641'
        '218 TAYLOR ST THRALL 76578'   ← 'TAYLOR' is also a Wilco city name
    Returns (street, city, zip5). Strategy: peel the trailing zip, then peel
    the longest known TX city name from the right edge — anything left is
    the street. Falls back to raw situs as street with empty city/zip when
    no zip is present.
    """
    if not situs:
        return "", "", ""
    from notice_parser import TX_CITIES

    s = re.sub(r"\s+", " ", situs).strip()
    # Peel trailing zip (5 digits, optionally +4)
    m = re.match(r"^(.*?)\s+(\d{5})(?:-?\d{4})?$", s)
    if not m:
        return s, "", ""
    body = m.group(1).strip()
    zip5 = m.group(2)
    # Peel the longest known city from the right of the body. TX_CITIES is
    # already sorted longest-first to handle multi-word cities like "Round
    # Rock" / "Cedar Park" before single-word substrings.
    body_upper = body.upper()
    for city in TX_CITIES:
        cu = city.upper()
        if body_upper.endswith(" " + cu) or body_upper == cu:
            street = body[: len(body) - len(cu)].strip()
            return street, city, zip5
    # Fallback: assume the last whitespace-token is the city
    parts = body.rsplit(" ", 1)
    if len(parts) == 2:
        return parts[0], parts[1].title(), zip5
    return body, "", zip5


def _wilco_normalize_name(raw: str) -> tuple[str, str]:
    """Normalize Wilco owner name into (display_owner, business_name).

    Wilco names use 2-space LAST  FIRST format (e.g. "RAMIREZ  AGUSTIN")
    or business names. Drops parenthetical annotations like "(TITLED)" and
    strips ETUX/ETAL tails before further processing — applied to BOTH
    business and individual name paths so trailing markers don't survive
    on entity names ("Foo LLC Etal" → "Foo LLC").
    """
    if not raw:
        return "", ""
    cleaned = re.sub(r"\s*\([^)]*\)\s*", " ", raw).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = _strip_etux_etal(cleaned)
    if _is_business(cleaned):
        out = _title_case(cleaned)
        return out, out
    flipped = normalize_court_name(cleaned)
    return _title_case(flipped) if flipped.isupper() else flipped, ""


@register("Williamson", "tax_delinquent")
class WilliamsonTaxDelinquentScraper:
    """Williamson County delinquent property roll → NoticeData."""

    COUNTY = "Williamson"

    def __init__(
        self,
        min_years: int = DEFAULT_MIN_DELINQUENT_YEARS,
        min_amount: float = DEFAULT_MIN_DELINQUENT_AMOUNT,
        fixture_xlsx: str | Path | None = None,
        skip_target_zip: bool = False,
    ):
        self.min_years = min_years
        self.min_amount = min_amount
        self.fixture_xlsx = Path(fixture_xlsx) if fixture_xlsx else None
        self.skip_target_zip = skip_target_zip

    async def scrape(
        self,
        mode: str = "daily",
        since_date: str | None = None,
        max_notices: int | None = None,
    ) -> list[NoticeData]:
        global LAST_RUN_DIFF, LAST_RUN_STATS, LAST_RUN_REMOVED
        global LAST_RUN_REPORT_PATH, LAST_RUN_RAW_PATH
        LAST_RUN_DIFF = None
        LAST_RUN_STATS = None
        LAST_RUN_REMOVED = None
        LAST_RUN_REPORT_PATH = None
        LAST_RUN_RAW_PATH = None

        logger.info(
            "Williamson tax delinquent: filters %d+ years, $%.0f+ owed, target-zip=%s",
            self.min_years, self.min_amount, "OFF" if self.skip_target_zip else "ON",
        )

        # ── Acquire XLSX ──
        if self.fixture_xlsx:
            if not self.fixture_xlsx.exists():
                logger.error("Fixture XLSX not found: %s", self.fixture_xlsx)
                return []
            xlsx_path = self.fixture_xlsx
        else:
            xlsx_path = wilco_tax_cache.download_delinquent_xlsx()
            if not xlsx_path:
                logger.error("Williamson tax delinquent: download failed")
                return []
        raw_bytes = xlsx_path.read_bytes()

        try:
            raw_archive_path, raw_sha = state_mod.archive_raw_bytes(
                self.COUNTY, raw_bytes, suffix=".xlsx",
            )
            LAST_RUN_RAW_PATH = raw_archive_path
            logger.info("Archived raw XLSX: %s", raw_archive_path)
        except Exception as e:
            logger.warning("Raw XLSX archival failed: %s", e)
            raw_archive_path, raw_sha = None, ""

        # ── First pass: bucket rows by Quick Ref ──
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Wilco's data lives on the first sheet (usually "DataExport*"); a
        # second blank "Sheet1" tab sometimes follows. Pick the first sheet
        # that has more than one row.
        ws = None
        for sheet in wb.worksheets:
            if sheet.max_row and sheet.max_row > 1:
                ws = sheet
                break
        if ws is None:
            wb.close()
            logger.error("Williamson XLSX has no data rows")
            return []

        rows = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            wb.close()
            return []
        col = {h: i for i, h in enumerate(header) if h is not None}

        def _g(row, name, default=""):
            i = col.get(name)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return default if v is None else v

        # parcel_id → aggregated dict
        bucket: dict[str, dict] = defaultdict(lambda: {
            "years": [],
            "balance": 0.0,
            "owner": "",
            "situs": "",
            "addr1": "",
            "addr2": "",
            "city": "",
            "state": "",
            "zip": "",
            "legal": "",
        })
        non_real = 0
        no_apn = 0
        for row in rows:
            qr_raw = _g(row, "Quick Ref")
            qr = str(qr_raw).strip()
            if not qr:
                no_apn += 1
                continue
            # Wilco prefixes: R = Real, P = Personal, M = Mobile, N = ???
            if not qr.upper().startswith("R"):
                non_real += 1
                continue
            b = bucket[qr]
            try:
                year = int(_g(row, "Tax Year"))
                b["years"].append(year)
            except (ValueError, TypeError):
                pass
            try:
                fee = float(_g(row, "Fee Balance") or 0)
                b["balance"] += fee
            except (ValueError, TypeError):
                pass
            # Take the longest non-empty seen-so-far for descriptive fields
            for src, dst in [
                ("Owner Name", "owner"),
                ("Situs Address", "situs"),
                ("Addr1", "addr1"),
                ("Addr2", "addr2"),
                ("City", "city"),
                ("State", "state"),
                ("Zip", "zip"),
                ("Legal Desc", "legal"),
            ]:
                v = str(_g(row, src) or "").strip()
                if v and len(v) > len(b[dst]):
                    b[dst] = v
        wb.close()

        # ── Second pass: build NoticeData per parcel ──
        notices: list[NoticeData] = []
        today = datetime.now().strftime("%Y-%m-%d")
        source_apns = set(bucket.keys())
        removed = {
            "no_year": 0, "under_years": 0, "hoa": 0,
            "under_amt": 0, "zip": 0, "no_owner": 0,
            "non_real": non_real, "no_apn": no_apn,
        }

        target_zips: set[str] = set()
        if not self.skip_target_zip:
            try:
                from zip_filter import load_target_zips
                target_zips = set(load_target_zips() or [])
            except Exception:
                pass

        this_year = datetime.now().year
        for parcel_id, b in bucket.items():
            if max_notices and len(notices) >= max_notices:
                break

            years = b["years"]
            if not years:
                removed["no_year"] += 1
                continue
            first_year = min(years)
            # Years delinquent counted from the earliest delinquency to now
            years_count = max(this_year - first_year, len(set(years)))
            if self.min_years > 0 and years_count < self.min_years:
                removed["under_years"] += 1
                continue

            owed = b["balance"]
            if self.min_amount > 0 and owed < self.min_amount:
                removed["under_amt"] += 1
                continue

            owner_raw = b["owner"]
            if not owner_raw:
                removed["no_owner"] += 1
                continue

            display_owner, business_name = _wilco_normalize_name(owner_raw)
            if business_name and _HOA_PAT.search(business_name):
                removed["hoa"] += 1
                continue

            street, city_from_situs, zip_from_situs = _split_situs(b["situs"])
            zip5 = (zip_from_situs or "")[:5]

            if target_zips and zip5 and zip5 not in target_zips:
                removed["zip"] += 1
                continue

            # Mailing address — trust the Addr1/City/State/Zip columns
            mail_street = b["addr1"]
            if b["addr2"]:
                mail_street = f"{mail_street} {b['addr2']}".strip()
            mail_city = b["city"].title() if b["city"] else ""
            mail_state = (b["state"] or "TX").upper()
            mail_zip = b["zip"][:5] if b["zip"] else ""
            mailing_full = ", ".join(
                p for p in [mail_street, mail_city, f"{mail_state} {mail_zip}".strip()]
                if p
            )

            notice = NoticeData(
                notice_type="tax_delinquent",
                county="Williamson",
                state="TX",
                date_added=today,
                address=street.title() if street else "",
                city=city_from_situs,
                zip=zip5,
                owner_name=display_owner or _title_case(owner_raw),
                parcel_id=parcel_id,
                source_url=f"https://search.wcad.org/Property/View/{parcel_id}",
            )
            notice.business_name = business_name
            notice.tax_owner_name = owner_raw
            notice.mailing_address = mailing_full
            if mail_street:
                notice.owner_street = mail_street
                notice.owner_city = mail_city
                notice.owner_state = mail_state
                notice.owner_zip = mail_zip
            notice.tax_delinquent_amount = f"{owed:.2f}"
            notice.tax_delinquent_years = str(years_count)
            notice.raw_text = (
                f"Quick Ref: {parcel_id}\n"
                f"Owner: {owner_raw}\n"
                f"Address: {street}, {city_from_situs}, TX {zip5}\n"
                f"Mailing: {mailing_full}\n"
                f"Delinquent: ${owed:,.2f} ({len(set(years))} year{'s' if len(set(years)) != 1 else ''} since {first_year})\n"
                f"Legal: {b['legal']}"
            )
            notices.append(notice)

        # ── Cross-run diff + state commit ──
        try:
            state = state_mod.load_state(self.COUNTY)
            prev_apns = set(state.get("last_run_apns") or [])
            diff = state_mod.compute_diff(source_apns, prev_apns)
            stats = state_mod.build_stats(notices)

            if raw_archive_path is not None:
                state_mod.commit_run(
                    self.COUNTY, state,
                    current_apns=source_apns,
                    raw_path=raw_archive_path,
                    raw_sha=raw_sha,
                    diff=diff,
                )
            report_path = state_mod.write_report_json(
                self.COUNTY, diff, stats, removed,
                raw_path=raw_archive_path or "",
            )
            LAST_RUN_DIFF = diff.to_dict()
            LAST_RUN_STATS = stats
            LAST_RUN_REMOVED = removed
            LAST_RUN_REPORT_PATH = report_path
            if diff.guardrail_tripped:
                logger.warning(
                    "Williamson tax-delinquent guardrail tripped: %s — prior state preserved",
                    diff.guardrail_reason,
                )
            elif diff.is_first_run:
                logger.info(
                    "Williamson tax-delinquent first run: seeding state with %d APNs",
                    len(source_apns),
                )
            else:
                logger.info(
                    "Williamson tax-delinquent cross-run diff: NEW=%d REPEAT=%d DROPPED=%d "
                    "(dropped = paid/sold — see %s)",
                    diff.new_count, diff.repeat_count, diff.dropped_count, report_path,
                )
        except Exception:
            logger.exception("Williamson tax-delinquent diff/state failed — continuing")

        logger.info(
            "Williamson tax delinquent: %d records kept (filtered: %s)",
            len(notices),
            ", ".join(f"{k}={v}" for k, v in removed.items() if v) or "none",
        )
        return notices
