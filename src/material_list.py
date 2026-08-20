"""Full tier 1/2/3 material list for a rehab, priced live from Home Depot.

Every priced line is a real Home Depot SKU priced to the subject zip via the
SerpApi home_depot engine (delivery_zip = local store pricing), the same
contract as serp_materials_pull.py. Tier grading:

  - graded items: ONE search, wide price band, then tier 1 = 25th percentile
    product, tier 2 = median, tier 3 = 75th percentile. The representative
    product nearest each percentile is named on the sheet, so every number
    traces to a SKU.
  - per-tier items: tiers are different product classes (laminate vs granite
    vs quartz, 3-tab vs architectural), so each tier gets its own search.
  - fixed items: commodity lines (OSB, PEX, mortar); one median price for all
    tiers.
  - allowances: services and lump scopes with no HD SKU (dumpsters, termite
    treatment, fitting kits). Flagged as ALLOWANCE on the sheet.

Quantities default to the 158 Old State walk facts (1,946 sqft finished 3/2,
11 window openings already replaced, 60% of drywall already hung) and are
overridable by CLI.

Usage:
  python src/material_list.py --zip 78704                    # live pull (~94 searches)
  python src/material_list.py --zip 78704 --cached           # reuse saved pull
  python src/material_list.py --master --zip 78704 --lock    # write the locked artifacts
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import statistics
import sys
import time
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from rehab_estimator import DEFAULT_REGION, estimate_rehab  # noqa: E402

logger = logging.getLogger("material_list")

NAVY, BLUE, GREEN, GOLD, GREY = "0A1130", "316AFF", "1B9E5A", "B8860B", "6B7280"

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# Flush the pricing cache every N successful searches. Searches are paid
# and slow; a run killed by a timeout used to lose all of them because the
# cache was written only after the final key. With this, --cached resumes.
CHECKPOINT_EVERY = 5

# Market label stamped into the locked artifacts. The ZIP is what
# sku_pricing validates against; this is the human-readable name.
MARKET_BY_ZIP_PREFIX = {
    "787": "austin",     # Travis / Williamson metro
    "786": "austin",
    "765": "bell",       # Killeen / Temple / Belton
    "379": "knox",       # legacy East Tennessee list
}


def market_for_zip(zip_code: str) -> str:
    return MARKET_BY_ZIP_PREFIX.get(str(zip_code)[:3], str(zip_code))

# ── Search catalog ────────────────────────────────────────────────────
# key: (query, relevance_regex, price_lo, price_hi)
SEARCHES = {
    # roof
    "shingles_3tab": ("3 tab roofing shingles", r"3[- ]?tab", 20, 55),
    "shingles_arch": ("architectural shingles", r"shingle", 25, 60),
    "underlayment": ("synthetic roofing underlayment", r"underlayment", 50, 150),
    "drip_edge": ("aluminum drip edge 10 ft", r"drip", 5, 30),
    "ridge_cap": ("ridge cap shingles", r"ridge|hip", 30, 90),
    "osb": ("7/16 osb sheathing", r"osb|sheathing", 10, 40),
    "roof_nails": ("roofing nails coil", r"roofing nail", 20, 90),
    "pipe_boot": ("roof pipe boot flashing", r"boot|flashing", 5, 40),
    # structure + crawl
    "pt_2x8": ("2 in. x 8 in. x 12 ft. pressure treated lumber",
               r"2 in\. x 8 in\. x 12 ft\.", 10, 80),
    "mortar": ("type s mortar mix 80 lb", r"type s", 5, 25),
    "vapor_barrier": ("20 mil crawl space vapor barrier", r"vapor|barrier|mil", 100, 700),
    "sump_pump": ("1/3 hp sump pump", r"sump", 100, 400),
    "sump_basin": ("18 in sump pump basin", r"basin|sump", 30, 150),
    "seam_tape": ("vapor barrier seam tape", r"tape", 10, 80),
    # windows + exterior
    "window": ("vinyl single hung window 36 x 60", r"single hung|window", 120, 400),
    "fascia": ("primed fascia board 1x6 12 ft", r"fascia|primed|trim|board", 8, 60),
    "paint_ext_5gal": ("5 gallon exterior paint", r"exterior", 100, 350),
    "entry_door": ("36 in fiberglass entry door prehung", r"entry|fiberglass|door", 250, 950),
    # kitchen
    # HD titles read "Base KITCHEN Cabinet", so the regex must not require the
    # words adjacent (the base-kitchen-cabinet gotcha from the 2026-07 pull).
    "base_cab_t1": ("unfinished base cabinet 24 in", r"base.*cabinet", 100, 350),
    "base_cab_t2": ("shaker base cabinet 24 in", r"base.*cabinet", 150, 450),
    "base_cab_t3": ("soft close shaker base cabinet 24 in", r"base.*cabinet", 200, 650),
    "wall_cab_t1": ("unfinished wall cabinet 30 in", r"wall.*cabinet", 80, 300),
    "wall_cab_t2": ("shaker wall cabinet 30 in", r"wall.*cabinet", 100, 350),
    "wall_cab_t3": ("soft close shaker wall cabinet 30 in", r"wall.*cabinet", 150, 550),
    "counter_t1": ("laminate countertop 8 ft", r"laminate|countertop", 80, 450),
    "counter_t2": ("prefab granite countertop", r"granite|quartz|countertop|butcher block", 200, 2000),
    # No quartz search: HD lists only $4 samples online; fabricated quartz is a
    # special-order quote, priced below as a stated allowance.
    "kitchen_sink": ("stainless steel drop in kitchen sink 33", r"sink", 100, 500),
    "kitchen_faucet": ("kitchen faucet pull down", r"faucet", 60, 400),
    "range": ("electric range", r"range|stove", 400, 1600),
    "otr_micro": ("over the range microwave", r"microwave", 150, 700),
    "dishwasher": ("dishwasher stainless", r"dishwasher", 300, 1200),
    "fridge": ("refrigerator stainless", r"refrigerator", 500, 3000),
    "backsplash": ("subway tile 3x6", r"subway|tile", 10, 60),
    # baths
    "bathtub": ("60 in alcove bathtub", r"tub", 250, 900),
    "tub_surround": ("bathtub wall surround", r"surround|wall set", 200, 1200),
    "shower_kit": ("48 in shower kit", r"shower", 400, 1600),
    "vanity36": ("36 in bathroom vanity with top", r"vanity", 250, 1200),
    "vanity24": ("24 in bathroom vanity with top", r"vanity", 150, 800),
    "toilet": ("2 piece elongated toilet", r"\d-piece.*toilet|toilet.*gpf", 100, 450),
    "bath_faucet": ("bathroom sink faucet", r"faucet", 40, 250),
    "shower_valve": ("tub and shower faucet set", r"shower|tub", 60, 400),
    "tile": ("porcelain floor tile 12x24", r"porcelain|tile", 15, 80),
    "bath_fan": ("bathroom exhaust fan", r"exhaust|fan", 30, 200),
    # floors, walls, doors
    "lvp": ("lifeproof luxury vinyl plank flooring", r"vinyl plank|lvp", 30, 120),
    "drywall": ("1/2 in 4x8 drywall", r"drywall|gypsum", 10, 30),
    "joint_compound": ("joint compound 4.5 gal", r"joint compound|mud", 10, 45),
    "drywall_tape": ("drywall joint tape", r"tape", 3, 25),
    "primer_5gal": ("5 gallon primer", r"primer", 60, 250),
    "paint_int_5gal": ("5 gallon interior paint", r"interior|paint", 100, 300),
    "baseboard": ("primed mdf baseboard 12 ft", r"baseboard|mdf|primed", 8, 45),
    "int_door": ("prehung interior door 6 panel", r"prehung|door", 60, 260),
    "door_knob": ("interior door knob", r"knob|lever", 10, 70),
    # mechanicals
    "heat_pump": ("4 ton heat pump split system", r"heat pump|split system", 3000, 9000),
    "water_heater": ("50 gallon electric water heater", r"water heater", 400, 1200),
    "pex_half": ("1/2 in pex pipe 100 ft", r"pex", 30, 130),
    "pex_34": ("3/4 in pex pipe 100 ft", r"pex", 50, 200),
    "panel_200": ("200 amp main breaker load center", r"load center|panel|breaker", 150, 550),
    "romex_12_2": ("12/2 romex 250 ft", r"12/2|romex|wire", 80, 320),
    "romex_14_2": ("14/2 romex 250 ft", r"14/2|romex|wire", 60, 260),
    "flush_mount": ("led flush mount ceiling light", r"flush mount|light", 15, 130),
    "ceiling_fan": ("52 in ceiling fan with light", r"ceiling fan", 60, 260),
    "vanity_light": ("3 light vanity light", r"vanity|light", 30, 220),
    # master-catalog additions (2026-07-31)
    "concrete": ("80 lb concrete mix", r"concrete", 4, 15),
    "pt_4x4": ("4x4x8 pressure treated post", r"4 in\. x 4 in\.", 8, 35),
    "deck_board": ("5/4 pressure treated deck board 12 ft", r"deck|pressure", 8, 35),
    "lockset": ("entry door knob and deadbolt combo", r"deadbolt|entry", 25, 130),
    "patio_slider": ("72 in sliding patio door", r"sliding|patio", 400, 1600),
    "garage_door": ("16 ft x 7 ft garage door", r"garage door", 500, 2500),
    "siding_panel": ("vinyl siding panel dutch lap", r"siding", 5, 45),
    "soffit": ("vinyl soffit panel", r"soffit", 5, 45),
    "gutter": ("aluminum gutter 10 ft", r"gutter", 5, 45),
    "downspout": ("aluminum downspout 10 ft", r"downspout", 5, 45),
    "ext_light": ("outdoor wall light", r"lantern|outdoor|porch|wall light", 10, 150),
    "cab_pulls": ("cabinet pulls 25 pack", r"pull|handle", 12, 130),
    "range_hood": ("30 in range hood", r"range hood|hood", 40, 300),
    "shower_door": ("sliding tub shower door", r"shower door|sliding", 150, 700),
    "vanity48": ("48 in bathroom vanity with top", r"vanity", 350, 1600),
    "medicine_cab": ("medicine cabinet with mirror", r"medicine|mirror", 40, 300),
    "towel_set": ("bath hardware set", r"towel|bath|accessor", 20, 120),
    "r13_batt": ("r-13 insulation batt", r"r-13|insulation", 30, 100),
    "blown_insul": ("blown in insulation", r"blown|insulation", 10, 45),
    "bifold": ("30 in bifold closet door", r"bifold", 40, 220),
    "casing": ("door casing moulding", r"casing", 4, 60),
    "quarter_round": ("quarter round moulding 12 ft", r"quarter round|shoe", 3, 25),
    "closet_kit": ("wire closet organizer kit", r"closet", 20, 250),
    "mini_split": ("12000 btu mini split", r"mini split", 500, 2000),
    "thermostat": ("programmable thermostat", r"thermostat", 25, 220),
    "wh_gas40": ("rheem gas water heater 40 gallon", r"water heater", 400, 1400),
    "gfci": ("20 amp gfci outlet", r"gfci", 15, 70),
    "outlets": ("duplex outlet 10 pack", r"outlet|receptacle", 10, 60),
    "switches": ("single pole switch 10 pack", r"switch", 10, 60),
    "smoke": ("10 year smoke detector", r"smoke", 10, 70),
    "co_det": ("carbon monoxide detector", r"carbon monoxide|combination", 20, 90),
    "mulch": ("brown mulch 2 cu ft", r"mulch", 2, 10),
}
# Per-case items: divide case price by coverage to get per-sqft.
CASE_SQFT = {"lvp": 20.1, "tile": 15.6, "backsplash": 12.5}

# ── Line catalog ──────────────────────────────────────────────────────
# (category, item, qty, unit, pricing, note)
#   qty: scalar or {tier: scalar}; 0/absent tier = line omitted at that tier
#   pricing: ("graded", key) percentile-tiered price off one search
#            ("fixed", key)  one median price, all tiers
#            ("per_tier", {1: key, 2: key, 3: key}) distinct searches
#            ("allow", amount or {tier: amount}) no HD SKU; stated allowance


def build_lines_sanland(q) -> list:
    """3014 Sanland Ave (Burlington): 948 sqft 3/1 ranch on a crawl. Scope is
    driven by the 2026-07-26 video walk: hoarding cleanout, full roof assumed
    until a roofer scopes it, carport framing rot, ceiling failure in 2-3
    rooms, vines on the siding, kitchen and bath dated but intact (kept at
    tier 1), no-window rear bedroom needs an egress window to count."""
    full_rock = round(q["sqft"] * 3.4 / 32)   # whole-house re-rock over paneling
    return [
        ("Demo and cleanout", "Dumpster pulls (30 yd), hoarding cleanout", {1: 4, 2: 4, 3: 5},
         "each", ("allow", 475), "Floor-to-ceiling belongings in every room; junk-out "
         "before any trade can price work"),
        ("Demo and cleanout", "Vine and vegetation clearing", 1, "job", ("allow", 800),
         "Siding condition unverifiable until cleared; re-scope siding after"),
        ("Structural and termite", "Termite/WDO inspection + treatment", 1, "job",
         ("allow", 2500), "Walk flag: heavy vine contact with siding and foundation "
         "is a conducive condition; verify before contract"),
        ("Structural and termite", "Carport framing, PT 2x8x12", 12, "each",
         ("fixed", "pt_2x8"), "Carport roof structure actively failing at fascia and decking"),
        ("Structural and termite", "Carport decking and hardware allowance", 1, "job",
         ("allow", 1500), ""),
        ("Roof", "Shingles", q["shingle_bundles"], "bundle",
         ("per_tier", {1: "shingles_3tab", 2: "shingles_arch", 3: "shingles_arch"}),
         f"{q['roof_squares']} squares. WALK GATE: priced as a FULL roof (multi-room "
         "ceiling failure is the tell); no trusted number until a roofer walks it"),
        ("Roof", "Synthetic underlayment", 2, "roll", ("fixed", "underlayment"), ""),
        ("Roof", "Drip edge 10 ft", 13, "each", ("fixed", "drip_edge"), ""),
        ("Roof", "Ridge cap", 1, "bundle", ("fixed", "ridge_cap"), ""),
        ("Roof", "OSB re-deck allowance", 8, "sheet", ("fixed", "osb"),
         "Long-standing leak; decking rot expected at the failed rooms"),
        ("Roof", "Roofing nails (coil)", 1, "box", ("fixed", "roof_nails"), ""),
        ("Roof", "Pipe boots", 2, "each", ("fixed", "pipe_boot"), ""),
        ("Windows and exterior", "Security bar removal and disposal", 1, "job",
         ("allow", 300), "All windows barred"),
        ("Windows and exterior", "Vinyl single hung windows", {2: 9, 3: 9}, "each",
         ("fixed", "window"), "Tier 1 keeps the existing functioning windows. Count is "
         "the 1-per-100-sqft estimate; verify onsite"),
        ("Windows and exterior", "Egress window, rear bedroom", {2: 1, 3: 1}, "job",
         ("allow", {2: 3000, 3: 3000}), "Legalizes the no-window 3rd bedroom (walk "
         "finding). Investor exit can skip it; retail/wholetail needs the bed count"),
        ("Windows and exterior", "Fascia 1x6x12 primed", 10, "each", ("fixed", "fascia"), ""),
        ("Windows and exterior", "Exterior paint 5 gal", {1: 2, 2: 2, 3: 3}, "bucket",
         ("fixed", "paint_ext_5gal"), ""),
        ("Windows and exterior", "Entry door, 36 in fiberglass prehung", {2: 1, 3: 1},
         "each", ("fixed", "entry_door"), ""),
        ("Windows and exterior", "Siding repair allowance", {2: 1, 3: 1}, "job",
         ("allow", {2: 1500, 3: 3000}), "Unverifiable under the vines; re-scope after clearing"),
        ("Windows and exterior", "Landscaping and curb allowance", 1, "job",
         ("allow", {1: 440, 2: 1320, 3: 2640}), "Engine exterior allowance, Knoxville adjusted"),
        ("Kitchen", "Cabinet paint and hardware refresh", {1: 1}, "job", ("allow", 350),
         "Kitchen is dated but intact (walk); tier 1 keeps the boxes and appliances"),
        ("Kitchen", "Base cabinets 24 in", {2: 5, 3: 5}, "each",
         ("per_tier", {2: "base_cab_t2", 3: "base_cab_t3"}), ""),
        ("Kitchen", "Wall cabinets 30 in", {2: 4, 3: 4}, "each",
         ("per_tier", {2: "wall_cab_t2", 3: "wall_cab_t3"}), ""),
        ("Kitchen", "Countertops", {2: 2}, "section", ("per_tier", {2: "counter_t2"}),
         "~22 sqft run, what HD stocks"),
        ("Kitchen", "Countertops, fabricated quartz", {3: 1}, "job", ("allow", {3: 1600}),
         "22 sqft at ~$70/sf fabricated; HD sells quartz as a special-order quote"),
        ("Kitchen", "Kitchen sink, 33 in stainless drop-in", {2: 1, 3: 1}, "each",
         ("fixed", "kitchen_sink"), ""),
        ("Kitchen", "Kitchen faucet", {2: 1, 3: 1}, "each", ("graded", "kitchen_faucet"), ""),
        ("Kitchen", "Electric range", {2: 1, 3: 1}, "each", ("graded", "range"),
         "Existing older gas range works; verify gas vs electric before buying"),
        ("Kitchen", "Over-the-range microwave", {2: 1, 3: 1}, "each", ("graded", "otr_micro"), ""),
        ("Kitchen", "Dishwasher", {2: 1, 3: 1}, "each", ("graded", "dishwasher"),
         "Verify space and hookup exist in the 1945 kitchen"),
        ("Kitchen", "Refrigerator", {2: 1, 3: 1}, "each", ("graded", "fridge"),
         "Existing white fridge works; tier 1 keeps it"),
        ("Kitchen", "Backsplash, subway tile", {2: 30, 3: 30}, "sqft", ("graded", "backsplash"), ""),
        ("Bathroom", "Tier 1 refresh: reglaze, caulk, hardware", {1: 1}, "job",
         ("allow", 300), "Single bath, dated but intact and undamaged (walk)"),
        ("Bathroom", "Alcove tub 60 in", {2: 1, 3: 1}, "each", ("fixed", "bathtub"), ""),
        ("Bathroom", "Tub wall surround", {2: 1}, "kit", ("graded", "tub_surround"), ""),
        ("Bathroom", "Wall tile in place of surround", {3: 50}, "sqft", ("graded", "tile"),
         "Tier 3 tiles the tub walls"),
        ("Bathroom", "Vanity 24 in with top", {2: 1, 3: 1}, "each", ("graded", "vanity24"), ""),
        ("Bathroom", "Toilet, elongated 2 pc", {2: 1, 3: 1}, "each", ("graded", "toilet"), ""),
        ("Bathroom", "Bath sink faucet", {2: 1, 3: 1}, "each", ("graded", "bath_faucet"), ""),
        ("Bathroom", "Tub/shower valve and trim", {2: 1, 3: 1}, "each",
         ("graded", "shower_valve"), ""),
        ("Bathroom", "Porcelain floor tile 12x24", {2: 45, 3: 45}, "sqft", ("graded", "tile"), ""),
        ("Bathroom", "Exhaust fan", 1, "each", ("graded", "bath_fan"),
         "Add at every tier; none visible in the walk video"),
        ("Flooring", "LVP flooring", q["floor_sqft"], "sqft", ("graded", "lvp"),
         f"{q['sqft']:,} sqft plus 10% waste. Hardwood may be salvageable under the "
         "overlays at tier 3; priced as full LVP until the cleanout exposes the floors"),
        ("Walls and ceilings", "Drywall 1/2 in 4x8", {1: 20, 2: full_rock, 3: full_rock},
         "sheet", ("fixed", "drywall"),
         "Tier 1 repairs the 2-3 failed ceilings only; tiers 2-3 re-rock the house "
         "(replaces the 1960s-70s paneling)"),
        ("Walls and ceilings", "Joint compound 4.5 gal", {1: 4, 2: 10, 3: 10}, "box",
         ("fixed", "joint_compound"), ""),
        ("Walls and ceilings", "Drywall tape", {1: 2, 2: 6, 3: 6}, "roll",
         ("fixed", "drywall_tape"), ""),
        ("Walls and ceilings", "Ceiling insulation replacement", 1, "job", ("allow", 600),
         "Batts over the failed ceilings are assumed wet; replace at the leak rooms"),
        ("Walls and ceilings", "Primer 5 gal", {1: 1, 2: 3, 3: 3}, "bucket",
         ("fixed", "primer_5gal"), "Paneling and new rock both need a full prime"),
        ("Walls and ceilings", "Interior paint 5 gal", {1: 4, 2: 5, 3: 5}, "bucket",
         ("graded", "paint_int_5gal"), ""),
        ("Doors and trim", "Baseboard, primed MDF 12 ft", {2: 32, 3: 32}, "each",
         ("fixed", "baseboard"), "~380 lf; tier 1 keeps and paints existing trim"),
        ("Doors and trim", "Prehung interior doors", {2: 6, 3: 6}, "each",
         ("graded", "int_door"), ""),
        ("Doors and trim", "Interior door hardware", {1: 6, 2: 8, 3: 8}, "each",
         ("graded", "door_knob"), ""),
        ("HVAC", "Service and repair allowance", {1: 1}, "job", ("allow", 500),
         "Tier 1 services what runs; system age unknown from the video"),
        ("HVAC", "Heat pump split system", {2: 1, 3: 1}, "system", ("graded", "heat_pump"),
         "Boxed price excludes heat strip kit, line set, thermostat"),
        ("HVAC", "Heat strip kit, line set, thermostat", {2: 1, 3: 1}, "job", ("allow", 600), ""),
        ("Plumbing", "Repair allowance", {1: 1}, "job", ("allow", 400), ""),
        ("Plumbing", "Water heater, 50 gal electric", {2: 1, 3: 1}, "each",
         ("fixed", "water_heater"), "One WH sighted in the crawl; age unknown"),
        ("Plumbing", "PEX 1/2 in x 100 ft", {2: 3, 3: 3}, "roll", ("fixed", "pex_half"),
         "Unflagged in the walk but a 1945 house; priced as a likely galvanized repipe "
         "at the reno tiers, verify supply material in the crawl"),
        ("Plumbing", "PEX 3/4 in x 100 ft", {2: 1, 3: 1}, "roll", ("fixed", "pex_34"), ""),
        ("Plumbing", "PEX fittings, valves, manifold", {2: 1, 3: 1}, "kit", ("allow", 300), ""),
        ("Electrical", "200A main breaker load center", {2: 1, 3: 1}, "each",
         ("fixed", "panel_200"), ""),
        ("Electrical", "Breakers", {2: 1, 3: 1}, "set", ("allow", 300), ""),
        ("Electrical", "Romex 12/2 x 250 ft", {2: 2, 3: 2}, "roll", ("fixed", "romex_12_2"),
         "Partial rewire while the leak-damaged ceilings are open"),
        ("Electrical", "Romex 14/2 x 250 ft", {2: 2, 3: 2}, "roll", ("fixed", "romex_14_2"), ""),
        ("Electrical", "Devices, plates, GFCIs, smokes", 1, "set", ("allow", 350),
         "Smokes at minimum on every tier"),
        ("Electrical", "LED flush mount lights", {1: 4, 2: 6, 3: 6}, "each",
         ("graded", "flush_mount"), "Rear bedroom is flashlight-lit in the video; it "
         "gets a fixture at every tier"),
        ("Electrical", "Ceiling fans 52 in", {2: 2, 3: 2}, "each", ("graded", "ceiling_fan"), ""),
        ("Electrical", "Vanity light", {2: 1, 3: 1}, "each", ("graded", "vanity_light"), ""),
    ]


def build_lines(q) -> list:
    """158 Old State: q carries the subject-derived quantities."""
    return [
        ("Demo and disposal", "Dumpster pulls (30 yd)", {1: 2, 2: 3, 3: 3}, "each",
         ("allow", 475), "Service, not an HD SKU"),
        ("Structural and termite", "Pressure treated 2x8x12 sill plate", 8, "each",
         ("fixed", "pt_2x8"), "Rotted sill over masonry (walk finding)"),
        ("Structural and termite", "Type S mortar, parging repair", 15, "bag",
         ("fixed", "mortar"), "Failing parging at the CMU wall"),
        ("Structural and termite", "WDO inspection + termite treatment", 1, "job",
         ("allow", 1200), "Service. Walk flag carries a $4,000 package incl framing allowance"),
        ("Structural and termite", "Framing repair allowance at sill", 1, "job",
         ("allow", 1500), "Balance of the walk termite flag; verify in inspection period"),
        ("Crawl space", "20 mil vapor barrier roll", q["vapor_rolls"], "roll",
         ("fixed", "vapor_barrier"),
         "~1,900 sqft crawl at the ~300 sqft rep roll; walk flag: crawl wet at the CMU wall"),
        ("Crawl space", "Sump pump 1/3 hp", 1, "each", ("fixed", "sump_pump"), ""),
        ("Crawl space", "Sump basin", 1, "each", ("fixed", "sump_basin"), ""),
        ("Crawl space", "Seam tape", 2, "roll", ("fixed", "seam_tape"), ""),
        ("Roof", "Shingles", q["shingle_bundles"], "bundle",
         ("per_tier", {1: "shingles_3tab", 2: "shingles_arch", 3: "shingles_arch"}),
         f"{q['roof_squares']} squares at 3 bundles per square"),
        ("Roof", "Synthetic underlayment", 3, "roll", ("fixed", "underlayment"), ""),
        ("Roof", "Drip edge 10 ft", q["drip_pieces"], "each", ("fixed", "drip_edge"), ""),
        ("Roof", "Ridge cap", 2, "bundle", ("fixed", "ridge_cap"), ""),
        ("Roof", "OSB re-deck allowance", 10, "sheet", ("fixed", "osb"), "Rot discovery allowance"),
        ("Roof", "Roofing nails (coil)", 2, "box", ("fixed", "roof_nails"), ""),
        ("Roof", "Pipe boots", 3, "each", ("fixed", "pipe_boot"), ""),
        ("Windows and exterior", "Vinyl single hung window", q["windows_remaining"], "each",
         ("fixed", "window"),
         f"{q['windows_done']} openings already have new vinyl (walk credit); count remaining onsite"),
        ("Windows and exterior", "Fascia 1x6x12 primed", 12, "each", ("fixed", "fascia"), ""),
        ("Windows and exterior", "Exterior paint 5 gal", {1: 2, 2: 3, 3: 4}, "bucket",
         ("fixed", "paint_ext_5gal"), ""),
        ("Windows and exterior", "Entry door, 36 in fiberglass prehung", 1, "each",
         ("fixed", "entry_door"), ""),
        ("Windows and exterior", "Siding repair allowance", {2: 1, 3: 1}, "job",
         ("allow", {2: 2640, 3: 5280}), "Engine exterior allowance, Knoxville adjusted"),
        ("Windows and exterior", "Landscaping and curb allowance", 1, "job",
         ("allow", {1: 440, 2: 1320, 3: 2640}), "Engine exterior allowance, Knoxville adjusted"),
        ("Kitchen", "Base cabinets 24 in", 6, "each",
         ("per_tier", {1: "base_cab_t1", 2: "base_cab_t2", 3: "base_cab_t3"}),
         "T1 unfinished, T2 shaker, T3 soft close"),
        ("Kitchen", "Wall cabinets 30 in", 6, "each",
         ("per_tier", {1: "wall_cab_t1", 2: "wall_cab_t2", 3: "wall_cab_t3"}), ""),
        ("Kitchen", "Countertops", {1: 2, 2: 2}, "section",
         ("per_tier", {1: "counter_t1", 2: "counter_t2"}),
         "~30 sqft run. T1 laminate postform, T2 what HD stocks (butcher block/prefab)"),
        ("Kitchen", "Countertops, fabricated quartz", {3: 1}, "job",
         ("allow", {3: 2100}), "30 sqft at ~$70/sf fabricated. HD sells quartz as a "
         "special-order quote (only $4 samples online), so no SKU price exists"),
        ("Kitchen", "Kitchen sink, 33 in stainless drop-in", 1, "each",
         ("fixed", "kitchen_sink"), ""),
        ("Kitchen", "Kitchen faucet", 1, "each", ("graded", "kitchen_faucet"), ""),
        ("Kitchen", "Electric range", 1, "each", ("graded", "range"), ""),
        ("Kitchen", "Over-the-range microwave", 1, "each", ("graded", "otr_micro"), ""),
        ("Kitchen", "Dishwasher", 1, "each", ("graded", "dishwasher"), ""),
        ("Kitchen", "Refrigerator", 1, "each", ("graded", "fridge"), ""),
        ("Kitchen", "Backsplash, subway tile", {2: 40, 3: 40}, "sqft",
         ("graded", "backsplash"), "Tier 1 skips backsplash (engine rule)"),
        ("Bathrooms", "Alcove tub 60 in (bath 1)", 1, "each", ("fixed", "bathtub"), ""),
        ("Bathrooms", "Tub wall surround (bath 1)", 1, "kit", ("graded", "tub_surround"), ""),
        ("Bathrooms", "Shower kit 48 in (bath 2, NEW)", 1, "kit", ("graded", "shower_kit"),
         "Bath 2 is the reconfig addition (2/1 to 3/2); gated on layout verification"),
        ("Bathrooms", "Vanity 36 in with top (bath 1)", 1, "each", ("graded", "vanity36"), ""),
        ("Bathrooms", "Vanity 24 in with top (bath 2)", 1, "each", ("graded", "vanity24"), ""),
        ("Bathrooms", "Toilets, elongated 2 pc", 2, "each", ("graded", "toilet"), ""),
        ("Bathrooms", "Bath sink faucets", 2, "each", ("graded", "bath_faucet"), ""),
        ("Bathrooms", "Tub/shower valve and trim sets", 2, "each", ("graded", "shower_valve"), ""),
        ("Bathrooms", "Porcelain floor tile 12x24", 90, "sqft", ("graded", "tile"),
         "Both bath floors plus waste"),
        ("Bathrooms", "Wall tile upgrade (bath 1)", {3: 60}, "sqft", ("graded", "tile"),
         "Tier 3 only: tiled surround in place of panel kit accents"),
        ("Bathrooms", "Exhaust fans", 2, "each", ("graded", "bath_fan"), ""),
        ("Flooring", "LVP flooring", q["floor_sqft"], "sqft", ("graded", "lvp"),
         f"{q['sqft']:,} sqft plus 10% waste; case-derived per-sqft price"),
        ("Walls and paint", "Drywall 1/2 in 4x8", q["drywall_sheets"], "sheet",
         ("fixed", "drywall"),
         f"Remaining {100 - q['drywall_done_pct']}% of hang; {q['drywall_done_pct']}% already hung and taped (walk credit)"),
        ("Walls and paint", "Joint compound 4.5 gal", 10, "box", ("fixed", "joint_compound"), ""),
        ("Walls and paint", "Drywall tape", 6, "roll", ("fixed", "drywall_tape"), ""),
        ("Walls and paint", "Primer 5 gal", 3, "bucket", ("fixed", "primer_5gal"),
         "New drywall takes a full prime coat"),
        ("Walls and paint", "Interior paint 5 gal", 8, "bucket", ("graded", "paint_int_5gal"),
         "Walls, ceilings and trim, 2 coats"),
        ("Doors and trim", "Baseboard, primed MDF 12 ft", q["baseboard_sticks"], "each",
         ("fixed", "baseboard"), f"~{q['baseboard_lf']} lf"),
        ("Doors and trim", "Prehung interior doors", 8, "each", ("graded", "int_door"), ""),
        ("Doors and trim", "Interior door hardware", 10, "each", ("graded", "door_knob"), ""),
        ("HVAC", "Heat pump split system", 1, "system", ("graded", "heat_pump"),
         "Boxed price excludes heat strip kit, line set, thermostat"),
        ("HVAC", "Heat strip kit, line set, thermostat", 1, "job", ("allow", 600),
         "Scope flag from the pricing pull"),
        ("Plumbing", "Water heater, 50 gal electric", 1, "each", ("fixed", "water_heater"), ""),
        ("Plumbing", "PEX 1/2 in x 100 ft", 5, "roll", ("fixed", "pex_half"),
         "Galvanized supply repipe (walk finding)"),
        ("Plumbing", "PEX 3/4 in x 100 ft", 2, "roll", ("fixed", "pex_34"), ""),
        ("Plumbing", "PEX fittings, valves, manifold", 1, "kit", ("allow", 400), ""),
        ("Electrical", "200A main breaker load center", 1, "each", ("fixed", "panel_200"), ""),
        ("Electrical", "Breakers", 1, "set", ("allow", 300), ""),
        ("Electrical", "Romex 12/2 x 250 ft", 3, "roll", ("fixed", "romex_12_2"), ""),
        ("Electrical", "Romex 14/2 x 250 ft", 4, "roll", ("fixed", "romex_14_2"), ""),
        ("Electrical", "Devices, plates, GFCIs, smokes", 1, "set", ("allow", 350), ""),
        ("Electrical", "LED flush mount lights", 8, "each", ("graded", "flush_mount"), ""),
        ("Electrical", "Ceiling fans 52 in", 3, "each", ("graded", "ceiling_fan"), ""),
        ("Electrical", "Vanity lights", 2, "each", ("graded", "vanity_light"), ""),
    ]


# ── Master catalog (property-agnostic, the standardization list) ──────
# One row per material we use anywhere. No tiers: each row carries a STANDARD
# pick plus Budget/Upgrade options where a real grade choice exists.
#   pricing: ("graded", key)   Budget/Standard/Upgrade = p25/median/p75 of one search
#            ("fixed", key)    Standard only (commodity; grade is meaningless)
#            ("classes", {"budget": key, "standard": key, "upgrade": key})
#                              different product classes; any slot may be
#                              ("text", "...") for quote-only options
# qty_driver states how the quantity scales on a real project.

MASTER = [
    ("Roofing", "Shingles", "bundle", "roof sqft (living x 1.1) / 100 x 3, +1 sq waste",
     ("classes", {"budget": "shingles_3tab", "standard": "shingles_arch"}),
     "Budget 3-tab only for rentals; retail exits standardize on architectural"),
    ("Roofing", "Synthetic underlayment", "roll", "1 per 10 squares", ("fixed", "underlayment"), ""),
    ("Roofing", "Drip edge 10 ft", "each", "perimeter lf / 10", ("fixed", "drip_edge"), ""),
    ("Roofing", "Ridge cap", "bundle", "ridge lf / 25", ("fixed", "ridge_cap"), ""),
    ("Roofing", "OSB 7/16 sheathing", "sheet", "rot allowance, 8-10 per reroof", ("fixed", "osb"), ""),
    ("Roofing", "Roofing nails (coil)", "box", "1 per 15 squares", ("fixed", "roof_nails"), ""),
    ("Roofing", "Pipe boots", "each", "1 per plumbing stack", ("fixed", "pipe_boot"), ""),
    ("Structure and site", "PT 2x8x12", "each", "sill/deck framing repair", ("fixed", "pt_2x8"), ""),
    ("Structure and site", "PT 4x4x8 post", "each", "deck/porch posts", ("fixed", "pt_4x4"), ""),
    ("Structure and site", "PT 5/4 deck board 12 ft", "each", "deck sqft / 5.5", ("fixed", "deck_board"), ""),
    ("Structure and site", "Type S mortar 80 lb", "bag", "parging/masonry repair", ("fixed", "mortar"), ""),
    ("Structure and site", "Concrete mix 80 lb", "bag", "steps and pad repair", ("fixed", "concrete"), ""),
    ("Structure and site", "20 mil vapor barrier (~300 sqft roll)", "roll", "crawl sqft / 300",
     ("fixed", "vapor_barrier"), ""),
    ("Structure and site", "Sump pump 1/3 hp", "each", "per wet crawl", ("fixed", "sump_pump"), ""),
    ("Structure and site", "Sump basin", "each", "per wet crawl", ("fixed", "sump_basin"), ""),
    ("Structure and site", "Vapor barrier seam tape", "roll", "2 per crawl", ("fixed", "seam_tape"), ""),
    ("Windows and exterior doors", "Vinyl single hung 36x60", "each", "~1 per 100 sqft living",
     ("graded", "window"), ""),
    ("Windows and exterior doors", "Entry door, fiberglass prehung 36 in", "each", "1-2 per house",
     ("graded", "entry_door"), ""),
    ("Windows and exterior doors", "Entry lockset + deadbolt", "each", "per exterior door",
     ("fixed", "lockset"), ""),
    ("Windows and exterior doors", "Sliding patio door 72 in", "each", "as-found", ("fixed", "patio_slider"), ""),
    ("Windows and exterior doors", "Garage door 16x7", "each", "as-found", ("fixed", "garage_door"), ""),
    ("Siding and exterior", "Vinyl siding panel", "each", "~24 panels per square", ("fixed", "siding_panel"), ""),
    ("Siding and exterior", "Vinyl soffit panel", "each", "eave lf / panel coverage", ("fixed", "soffit"), ""),
    ("Siding and exterior", "Fascia 1x6x12 primed", "each", "eave lf / 12", ("fixed", "fascia"), ""),
    ("Siding and exterior", "Aluminum gutter 10 ft", "each", "eave lf / 10", ("fixed", "gutter"),
     "Sectional; seamless sub-out runs ~$8-10/lf installed, see Allowances"),
    ("Siding and exterior", "Aluminum downspout 10 ft", "each", "1 per corner", ("fixed", "downspout"), ""),
    ("Siding and exterior", "Exterior paint 5 gal", "bucket", "1 per ~1,000 sqft wall",
     ("graded", "paint_ext_5gal"), ""),
    ("Siding and exterior", "Exterior wall lantern", "each", "per exterior door", ("graded", "ext_light"), ""),
    ("Kitchen", "Base cabinets 24 in", "each", "5-6 per 10x10 kitchen",
     ("classes", {"budget": "base_cab_t1", "standard": "base_cab_t2", "upgrade": "base_cab_t3"}),
     "Budget unfinished, standard shaker, upgrade soft close"),
    ("Kitchen", "Wall cabinets 30 in", "each", "4-6 per 10x10 kitchen",
     ("classes", {"budget": "wall_cab_t1", "standard": "wall_cab_t2", "upgrade": "wall_cab_t3"}), ""),
    ("Kitchen", "Countertops", "section", "~2-3 sections (25-30 sqft) per kitchen",
     ("classes", {"budget": "counter_t1", "standard": "counter_t2",
                  "upgrade": ("text", "Fabricated quartz, special-order quote ~$70/sf")}),
     "Budget laminate postform, standard what HD stocks (butcher block/prefab)"),
    ("Kitchen", "Cabinet pulls (25 pack)", "pack", "1 per kitchen", ("fixed", "cab_pulls"), ""),
    ("Kitchen", "Kitchen sink 33 in SS drop-in", "each", "1 per kitchen", ("fixed", "kitchen_sink"), ""),
    ("Kitchen", "Kitchen faucet, pull down", "each", "1 per kitchen", ("graded", "kitchen_faucet"), ""),
    ("Kitchen", "Electric range", "each", "1 per kitchen", ("graded", "range"),
     "Verify gas vs electric before ordering"),
    ("Kitchen", "Over-the-range microwave", "each", "1 per kitchen", ("graded", "otr_micro"), ""),
    ("Kitchen", "Range hood 30 in", "each", "where no OTR micro fits", ("graded", "range_hood"), ""),
    ("Kitchen", "Dishwasher", "each", "1 per kitchen (verify hookup)", ("graded", "dishwasher"), ""),
    ("Kitchen", "Refrigerator", "each", "1 per kitchen", ("graded", "fridge"), ""),
    ("Kitchen", "Backsplash, subway tile", "sqft", "~30-40 sqft per kitchen", ("graded", "backsplash"), ""),
    ("Bathrooms", "Alcove tub 60 in", "each", "per tub bath", ("graded", "bathtub"), ""),
    ("Bathrooms", "Tub wall surround", "kit", "per tub bath (standard)", ("graded", "tub_surround"),
     "Upgrade path is tiled walls, priced off the tile line"),
    ("Bathrooms", "Shower kit 48 in", "kit", "per shower bath", ("graded", "shower_kit"), ""),
    ("Bathrooms", "Sliding tub shower door", "each", "retail exits only", ("graded", "shower_door"), ""),
    ("Bathrooms", "Vanity 24 in with top", "each", "small/secondary bath", ("graded", "vanity24"), ""),
    ("Bathrooms", "Vanity 36 in with top", "each", "primary bath", ("graded", "vanity36"), ""),
    ("Bathrooms", "Vanity 48 in with top", "each", "large primary bath", ("graded", "vanity48"), ""),
    ("Bathrooms", "Toilet, elongated 2 pc", "each", "per bath", ("graded", "toilet"), ""),
    ("Bathrooms", "Bath sink faucet", "each", "per bath", ("graded", "bath_faucet"), ""),
    ("Bathrooms", "Tub/shower valve and trim", "each", "per bath", ("graded", "shower_valve"), ""),
    ("Bathrooms", "Porcelain tile 12x24", "sqft", "~45 sqft per bath floor; +50-60 for tiled walls",
     ("graded", "tile"), ""),
    ("Bathrooms", "Medicine cabinet / mirror", "each", "per bath", ("graded", "medicine_cab"), ""),
    ("Bathrooms", "Bath hardware set", "set", "per bath", ("fixed", "towel_set"), ""),
    ("Bathrooms", "Exhaust fan", "each", "per bath", ("graded", "bath_fan"), ""),
    ("Flooring", "LVP flooring", "sqft", "living sqft x 1.1", ("graded", "lvp"),
     "House standard is LVP throughout; carpet only by exception (see Allowances)"),
    ("Walls and paint", "Drywall 1/2 in 4x8", "sheet", "sqft x 3.4 / 32 for a full re-rock",
     ("fixed", "drywall"), ""),
    ("Walls and paint", "Joint compound 4.5 gal", "box", "1 per 10 sheets", ("fixed", "joint_compound"), ""),
    ("Walls and paint", "Drywall tape", "roll", "1 per 17 sheets", ("fixed", "drywall_tape"), ""),
    ("Walls and paint", "R-13 wall batt insulation", "bag", "open walls only", ("fixed", "r13_batt"), ""),
    ("Walls and paint", "Blown-in attic insulation", "bag", "attic sqft / 60 (R-30 top-up)",
     ("fixed", "blown_insul"), ""),
    ("Walls and paint", "Primer 5 gal", "bucket", "1 per ~1,400 sqft new rock", ("fixed", "primer_5gal"), ""),
    ("Walls and paint", "Interior paint 5 gal", "bucket", "1 per ~350 sqft floor, 2 coats",
     ("graded", "paint_int_5gal"), ""),
    ("Doors and trim", "Prehung interior door, 6 panel", "each", "per opening", ("graded", "int_door"), ""),
    ("Doors and trim", "Bifold closet door 30 in", "each", "per closet", ("fixed", "bifold"), ""),
    ("Doors and trim", "Interior knob/lever", "each", "per door +2 spares", ("graded", "door_knob"), ""),
    ("Doors and trim", "Baseboard, primed MDF 12 ft", "each", "perimeter lf / 12", ("fixed", "baseboard"), ""),
    ("Doors and trim", "Door casing kit", "opening", "per door/window opening", ("fixed", "casing"), ""),
    ("Doors and trim", "Quarter round 12 ft", "each", "perimeter lf / 12", ("fixed", "quarter_round"), ""),
    ("Doors and trim", "Closet organizer kit", "closet", "per closet", ("fixed", "closet_kit"), ""),
    ("HVAC", "Heat pump split system", "system", "1 per house (size to sqft)", ("graded", "heat_pump"),
     "Boxed price excludes heat strips, line set, thermostat: add the scope kit allowance"),
    ("HVAC", "Mini split 12k BTU", "each", "additions/bonus rooms", ("fixed", "mini_split"), ""),
    ("HVAC", "Programmable thermostat", "each", "1 per system", ("fixed", "thermostat"), ""),
    ("Plumbing", "Water heater 50 gal electric", "each", "1 per house", ("graded", "water_heater"), ""),
    ("Plumbing", "Water heater 40 gal gas", "each", "where gas is the service", ("fixed", "wh_gas40"), ""),
    ("Plumbing", "PEX 1/2 in x 100 ft", "roll", "3-5 rolls per repipe", ("fixed", "pex_half"), ""),
    ("Plumbing", "PEX 3/4 in x 100 ft", "roll", "1-2 rolls per repipe", ("fixed", "pex_34"), ""),
    ("Electrical", "200A main breaker load center", "each", "per panel swap", ("fixed", "panel_200"), ""),
    ("Electrical", "Romex 12/2 x 250 ft", "roll", "2-3 per rewire", ("fixed", "romex_12_2"), ""),
    ("Electrical", "Romex 14/2 x 250 ft", "roll", "2-4 per rewire", ("fixed", "romex_14_2"), ""),
    ("Electrical", "GFCI outlet 20A", "each", "kitchen/bath/exterior locations", ("fixed", "gfci"), ""),
    ("Electrical", "Duplex outlets (10 pack)", "pack", "1-2 per house", ("fixed", "outlets"), ""),
    ("Electrical", "Switches (10 pack)", "pack", "1 per house", ("fixed", "switches"), ""),
    ("Electrical", "Smoke detector (10 yr)", "each", "per bedroom + 1 per floor", ("fixed", "smoke"), ""),
    ("Electrical", "CO detector", "each", "1 per floor (fuel appliances)", ("fixed", "co_det"), ""),
    ("Electrical", "LED flush mount light", "each", "per room without fan", ("graded", "flush_mount"), ""),
    ("Electrical", "Ceiling fan 52 in", "each", "bedrooms + living", ("graded", "ceiling_fan"), ""),
    ("Electrical", "Vanity light", "each", "per bath", ("graded", "vanity_light"), ""),
    ("Landscape", "Mulch 2 cu ft", "bag", "~15-25 bags per front bed refresh", ("fixed", "mulch"), ""),
]

MASTER_ALLOWANCES = [
    ("Demo and site", "Dumpster pull, 30 yd", 475, "Per pull; 1 per ~1,500 sqft of demo/clutter"),
    ("Structure", "WDO/termite inspection", 150, "Every project inside the inspection period"),
    ("Structure", "Termite treatment", 1200, "When the WDO comes back active"),
    ("Structure", "Framing repair allowance", 1500, "Sill/joist discovery; scoped per walk"),
    ("Kitchen", "Fabricated quartz countertop", 70, "Per sqft, special-order quote; HD lists only samples online"),
    ("Siding and exterior", "Seamless gutter sub-out", 9, "Per lf installed; alternative to sectional"),
    ("HVAC", "Heat pump scope kit", 600, "Heat strips, line set, thermostat per system"),
    ("Plumbing", "PEX fittings, valves, manifold", 400, "Per repipe"),
    ("Electrical", "Breaker set", 300, "Per panel swap"),
    ("Electrical", "Misc devices, plates, low-volt", 150, "Per house"),
    ("Flooring", "Carpet (exception only)", 2.5, "Per sqft installed, quoted; house standard is LVP"),
    ("Landscape", "Landscaping and curb", 1000, "Baseline; scoped per walk"),
]


# ── Pricing pull ──────────────────────────────────────────────────────

# Both spellings are accepted: the underscore form is what the docs and
# .env.example use, the bare form is what the operator .env actually carries.
# NOTE: this is SerpApi (serpapi.com, engine=home_depot), NOT Serper.dev
# (google.serper.dev) which SERPER_API_KEY holds for obituary search. They are
# unrelated vendors and the keys are not interchangeable.
SERPAPI_KEY_NAMES = ("SERPAPI_KEY", "SERPAPIKEY")


def load_key() -> str:
    for name in SERPAPI_KEY_NAMES:
        k = os.getenv(name)
        if k:
            return k.strip()
    env = Path(__file__).parent.parent / ".env"
    if env.exists():
        for line in env.read_text(encoding="utf-8", errors="replace").splitlines():
            line = line.strip()
            for name in SERPAPI_KEY_NAMES:
                if line.startswith(name + "="):
                    val = line.split("=", 1)[1].strip().strip('"').strip("'")
                    if val:
                        return val
    sys.exit("SerpApi key not found. Set one of %s in env or .env "
             "(serpapi.com — NOT the Serper.dev key)." % ", ".join(SERPAPI_KEY_NAMES))


def _write_cache(cache: Path, zip_code: str, used: int, out: dict) -> None:
    cache.parent.mkdir(parents=True, exist_ok=True)
    tmp = cache.with_suffix(cache.suffix + ".tmp")
    tmp.write_text(json.dumps(
        {"zip": zip_code, "date": date.today().isoformat(), "searches_used": used,
         "keys": out}, indent=1), encoding="utf-8")
    tmp.replace(cache)  # atomic: a kill mid-write cannot corrupt the cache


def _search(k: str, term: str, zip_code: str, key: str, sort):
    """One SerpApi home_depot call, 3 attempts. -> (json, sort_label) or (None, _)."""
    params = {"engine": "home_depot", "q": term, "delivery_zip": zip_code,
              "api_key": key}
    if sort:
        params["hd_sort"] = sort
    label = sort or "relevance"
    for attempt in range(3):
        try:
            r = requests.get("https://serpapi.com/search.json", params=params, timeout=90)
        except Exception as exc:
            # SerpApi read-timeouts arrive in waves under burst load
            # (2026-07-31: 26 of 32 in one run); back off and retry.
            logger.warning("[%s/%s] attempt %d: %s", k, label, attempt + 1, exc)
            time.sleep(5 * (attempt + 1))
            continue
        try:
            d = r.json()
        except Exception as exc:
            logger.warning("[%s/%s] bad JSON: %s", k, label, exc)
            return None, label
        if r.status_code != 200 or d.get("error"):
            logger.warning("[%s/%s] HTTP %s %s", k, label, r.status_code,
                           str(d.get("error"))[:80])
            return None, label
        return d, label
    return None, label


def _in_band(d: dict, rx: str, lo: float, hi: float) -> list:
    """Products whose title matches the relevance regex and price is in band."""
    rows = []
    for p in d.get("products", []):
        price, title = p.get("price"), p.get("title") or ""
        if price is None or not re.search(rx, title, re.I):
            continue
        if not (lo <= float(price) <= hi):
            continue
        rows.append({"title": title[:120], "price": float(price),
                     "link": p.get("link") or "", "brand": p.get("brand")})
    return rows


def pull_prices(zip_code: str, cache: Path, cached: bool, needed=None) -> dict:
    """key -> {prices: [...], products: [{title, price, link}...]} local to zip.

    `needed` limits fetching to the keys a build actually uses (a per-property
    profile does not stall on master-only keys that SerpApi is failing on).
    The cache is always merged, never clobbered: keys outside `needed` ride
    along untouched.
    """
    keys = set(needed) if needed else set(SEARCHES)
    prior, prior_used = {}, 0
    if cache.exists():
        data = json.loads(cache.read_text(encoding="utf-8"))
        prior, prior_used = data.get("keys", {}), data.get("searches_used", 0)
    todo = [k for k in SEARCHES if k in keys and (not cached or k not in prior)]
    if cached:
        logger.info("Reusing cached pull %s (%d keys); fetching %d missing",
                    cache, len(prior), len(todo))
    if not todo:
        return prior
    key = load_key()
    out, used, done = dict(prior), prior_used, 0
    logger.info("pulling %d searches for zip %s (checkpoint every %d)",
                len(todo), zip_code, CHECKPOINT_EVERY)
    for k, (term, rx, lo, hi) in SEARCHES.items():
        if k not in todo:
            continue
        d, sort_used = _search(k, term, zip_code, key, "top_sellers")
        if d is None:
            continue
        used += 1
        rows = _in_band(d, rx, lo, hi)
        if not rows:
            # hd_sort=top_sellers sometimes returns the store's GENERIC top
            # sellers rather than results for the query (verified 78704
            # 2026-08-20: "outdoor wall light" -> retaining-wall block and
            # joint compound). Retry once on pure relevance before giving up.
            raw_ts = len(d.get("products", []))
            d2, _ = _search(k, term, zip_code, key, None)
            if d2 is not None:
                used += 1
                rows2 = _in_band(d2, rx, lo, hi)
                if rows2:
                    logger.info("[%s] top_sellers returned %d irrelevant; "
                                "relevance sort recovered %d", k, raw_ts, len(rows2))
                    d, rows, sort_used = d2, rows2, "relevance"
        if rows:
            out[k] = {"query": term, "sort": sort_used,
                      "products": sorted(rows, key=lambda x: x["price"])}
            logger.info("[%s] %d in-band, $%.2f-$%.2f", k, len(rows),
                        rows[0]["price"] if rows else 0, max(x["price"] for x in rows))
        else:
            logger.warning("[%s] no in-band matches (%d raw)", k, len(d.get("products", [])))
        done += 1
        if done % CHECKPOINT_EVERY == 0:
            _write_cache(cache, zip_code, used, out)
            logger.info("checkpoint: %d/%d searched, %d keys priced, %d paid",
                        done, len(todo), len(out), used)
        time.sleep(0.5)
    _write_cache(cache, zip_code, used, out)
    logger.info("searches used: %d | keys priced: %d/%d | cached to %s",
                used, len(out), len(SEARCHES), cache)
    return out


def _pct_product(rows: list, pct: float) -> dict:
    prices = [r["price"] for r in rows]
    target = statistics.quantiles(prices, n=4)[{0.25: 0, 0.5: 1, 0.75: 2}[pct]] \
        if len(prices) >= 4 else statistics.median(prices)
    return min(rows, key=lambda r: abs(r["price"] - target))


def price_line(pricing: tuple, tier: int, prices: dict) -> tuple:
    """-> (unit_price, product_title, link, kind) or (None, reason, '', kind)."""
    kind, ref = pricing
    if kind == "allow":
        amt = ref.get(tier) if isinstance(ref, dict) else ref
        return (amt, "ALLOWANCE (no HD SKU)", "", kind) if amt else (None, "n/a", "", kind)
    key = ref[tier] if kind == "per_tier" else ref
    data = prices.get(key)
    if not data:
        return None, f"not priced (search '{SEARCHES[key][0]}' returned no in-band match)", "", kind
    rows = data["products"]
    if kind == "graded":
        rep = _pct_product(rows, {1: 0.25, 2: 0.5, 3: 0.75}[tier])
    else:
        rep = _pct_product(rows, 0.5)
    unit = rep["price"] / CASE_SQFT[key] if key in CASE_SQFT else rep["price"]
    return round(unit, 2), rep["title"], rep["link"], kind


# ── Workbook ──────────────────────────────────────────────────────────

THIN = Border(bottom=Side(style="thin", color="D5D9E0"))


def _sheet_header(ws, row: int, cols: list):
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, size=10, color=NAVY)
        c.border = Border(bottom=Side(style="medium", color=NAVY))


def _title(ws, text: str, sub: str = ""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=9, color=GREY)


def build_workbook(subject: dict, q: dict, lines: list, prices: dict, meta: dict,
                   out_path: str):
    wb = Workbook()
    wb.remove(wb.active)
    tier_names = {1: "Tier 1 Minimum Viable", 2: "Tier 2 Builder Grade",
                  3: "Tier 3 Mid-Range Upgrade"}
    totals, all_rows = {}, {}

    for tier in (1, 2, 3):
        ws = wb.create_sheet(f"Tier {tier}")
        _title(ws, f"{subject['address']}: material list, {tier_names[tier]}",
               f"Home Depot pricing local to zip {meta['zip']} (SerpApi home_depot engine, "
               f"pulled {meta['date']}). Materials only; labor and GC totals live on the "
               "Repair Numbers sheet of the post-walkthrough workbook.")
        cols = ["Category", "Item", "Qty", "Unit", "Unit $", "Extended $",
                "Priced product (nearest the tier percentile)", "Note"]
        _sheet_header(ws, 4, cols)
        r, cat_start, last_cat, total, unpriced = 5, 5, None, 0.0, []
        rows_out = []

        def close_cat(upto_row):
            nonlocal cat_start
            cat_start = upto_row

        for (cat, item, qty, unit, pricing, note) in lines:
            tier_qty = qty.get(tier, 0) if isinstance(qty, dict) else qty
            if not tier_qty:
                continue
            unit_price, product, link, kind = price_line(pricing, tier, prices)
            if cat != last_cat and last_cat is not None:
                r += 0
            if cat != last_cat:
                ws.cell(row=r, column=1, value=cat).font = Font(bold=True, size=10, color=BLUE)
                last_cat = cat
            ws.cell(row=r, column=2, value=item)
            ws.cell(row=r, column=3, value=tier_qty)
            ws.cell(row=r, column=4, value=unit)
            if unit_price is None:
                ws.cell(row=r, column=7, value=product).font = Font(color=GOLD, size=9)
                unpriced.append(item)
            else:
                ext = round(unit_price * tier_qty)
                total += ext
                uc = ws.cell(row=r, column=5, value=unit_price)
                uc.number_format = "$#,##0.00"
                ec = ws.cell(row=r, column=6, value=ext)
                ec.number_format = "$#,##0"
                pc = ws.cell(row=r, column=7, value=product)
                pc.font = Font(size=9, color=GREY if kind == "allow" else NAVY)
                if link:
                    pc.hyperlink = link
                rows_out.append((cat, item, tier_qty, unit, ext))
            if note:
                ws.cell(row=r, column=8, value=note).font = Font(size=9, color=GREY)
            for j in range(1, 9):
                ws.cell(row=r, column=j).border = THIN
                ws.cell(row=r, column=j).alignment = Alignment(vertical="top", wrap_text=(j in (7, 8)))
            r += 1

        r += 1
        ws.cell(row=r, column=2, value="Materials total").font = Font(bold=True, size=11, color=NAVY)
        tc = ws.cell(row=r, column=6, value=round(total))
        tc.font = Font(bold=True, size=11, color=GREEN)
        tc.number_format = "$#,##0"
        eng = meta["engine_materials"][tier]
        r += 1
        ws.cell(row=r, column=2,
                value=f"Rehab engine materials budget at this tier (full scope, Knoxville): "
                      f"${eng:,.0f}. Knox engine materials are SKU-grounded from the locked "
                      "master list; any residual gap is quantity assumptions, not prices.").font = \
            Font(size=9, color=GREY)
        if unpriced:
            r += 1
            ws.cell(row=r, column=2,
                    value=f"Not priced (no in-band API result): {', '.join(unpriced)}. "
                          "Excluded from the total.").font = Font(size=9, color=GOLD)
        widths = [22, 34, 6, 8, 10, 11, 52, 40]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A5"
        totals[tier] = round(total)
        all_rows[tier] = rows_out

    # Compare sheet
    ws = wb.create_sheet("Tier Compare", 0)
    _title(ws, f"{subject['address']}: tier 1 vs 2 vs 3 materials",
           "Extended materials cost per line. Blank = not in that tier's scope.")
    _sheet_header(ws, 4, ["Category", "Item", "Qty", "Unit", "Tier 1 $", "Tier 2 $", "Tier 3 $"])
    index = {}
    for tier in (1, 2, 3):
        for (cat, item, qty, unit, ext) in all_rows[tier]:
            d = index.setdefault((cat, item), {"unit": unit, "qtys": {}})
            d[tier], d["qtys"][tier] = ext, qty
    r, last_cat = 5, None
    for (cat, item), d in index.items():
        if cat != last_cat:
            ws.cell(row=r, column=1, value=cat).font = Font(bold=True, size=10, color=BLUE)
            last_cat = cat
        ws.cell(row=r, column=2, value=item)
        qtys = d["qtys"]
        # A tier-varying quantity renders as T1/T2/T3 so the compare stays honest.
        qty_disp = qtys[max(qtys)] if len(set(qtys.values())) == 1 else \
            "/".join(str(qtys.get(t, 0)) for t in (1, 2, 3))
        ws.cell(row=r, column=3, value=qty_disp)
        ws.cell(row=r, column=4, value=d["unit"])
        for tier, col in ((1, 5), (2, 6), (3, 7)):
            if tier in d:
                c = ws.cell(row=r, column=col, value=d[tier])
                c.number_format = "$#,##0"
        for j in range(1, 8):
            ws.cell(row=r, column=j).border = THIN
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="Materials total").font = Font(bold=True, size=11, color=NAVY)
    for tier, col in ((1, 5), (2, 6), (3, 7)):
        c = ws.cell(row=r, column=col, value=totals[tier])
        c.font = Font(bold=True, size=11, color=GREEN)
        c.number_format = "$#,##0"
    for j, w in enumerate([22, 34, 6, 8, 11, 11, 11], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    # Read me
    ws = wb.create_sheet("Read Me", 0)
    _title(ws, f"{subject['address']}: full material list, tiers 1-3")
    notes = [
        f"Subject: {subject['address']}, {q['sqft']:,} sqft, priced to the FINISHED "
        f"{subject['beds']}/{subject['baths']:g} configuration.",
        f"Pricing source: Home Depot via the SerpApi home_depot engine, delivery_zip "
        f"{meta['zip']} (local store pricing), pulled {meta['date']}, "
        f"{meta['searches_used']} searches.",
        "Tier grading: one search per item with a wide price band; tier 1 buys the 25th "
        "percentile product, tier 2 the median, tier 3 the 75th. Where tiers are different "
        "product classes (counters, shingles, cabinets) each tier has its own search. "
        "Every priced line names the actual product and links to it.",
        *meta.get("walk_notes", []),
        "Lines marked ALLOWANCE are services or lump scopes with no Home Depot SKU "
        "(dumpsters, termite treatment, fitting kits, engine exterior allowances).",
        "MATERIALS ONLY. Labor, soft costs and the GC vs self-perform totals are on the "
        "Repair Numbers sheet of the post-walkthrough workbook; this list is the "
        "shopping half of those same scenarios.",
        f"Totals: tier 1 ${totals[1]:,} | tier 2 ${totals[2]:,} | tier 3 ${totals[3]:,}.",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=4 + i * 2, column=1, value=n)
        c.font = Font(size=10, color=NAVY if i == len(notes) - 1 else "374151")
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    wb.save(out_path)
    return totals


def _slot(prices: dict, key, pct: float):
    """Resolve one Budget/Standard/Upgrade slot -> (price, title, link) or None."""
    if isinstance(key, tuple) and key[0] == "text":
        return (None, key[1], "")
    data = prices.get(key)
    if not data:
        return None
    rep = _pct_product(data["products"], pct)
    unit = rep["price"] / CASE_SQFT[key] if key in CASE_SQFT else rep["price"]
    return (round(unit, 2), rep["title"], rep["link"])


# ── Locked artifact export (the frozen master list the engine + skill read) ──

DATA_DIR = Path(__file__).parent.parent / "data"


def _row_key(pricing: tuple, item: str) -> str:
    """Stable key for a MASTER row: the search key, or an item slug for classes."""
    kind, ref = pricing
    if isinstance(ref, str):
        return ref
    return re.sub(r"[^a-z0-9]+", "_", item.lower()).strip("_")


def _allow_key(item: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", item.lower()).strip("_")


def _grade_slots(prices: dict, pricing: tuple) -> dict:
    """Resolve all three grade slots for a MASTER row. Unlike the workbook view,
    the lock keeps a collapsed percentile spread (the engine needs a price per
    grade even when the result set was thin)."""
    kind, ref = pricing
    if kind == "graded":
        return {"budget": _slot(prices, ref, 0.25), "standard": _slot(prices, ref, 0.5),
                "upgrade": _slot(prices, ref, 0.75)}
    if kind == "classes":
        slots = {s: _slot(prices, k, 0.5) for s, k in ref.items()}
        slots.setdefault("budget", None)
        slots.setdefault("upgrade", None)
        return slots
    return {"budget": None, "standard": _slot(prices, ref, 0.5), "upgrade": None}


def export_locked_artifacts(prices: dict, meta: dict) -> tuple:
    """Write the LOCKED master list to data/ (git-tracked, unlike output/).

    Only --lock ever writes these files: an ordinary re-pull refreshes the
    cache and the xlsx but cannot drift prices into estimates. Re-locking is
    an explicit, dated, git-diffable act. Consumed by src/sku_pricing.py and
    shipped verbatim inside rehab-estimator.skill.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    rows, unpriced = [], []
    for (cat, item, unit, driver, pricing, note) in MASTER:
        kind, ref = pricing
        slots = _grade_slots(prices, pricing)
        grades = {}
        for slot_name, s in slots.items():
            if s is None:
                grades[slot_name] = None
            elif s[0] is None:  # quote-only text slot (fabricated quartz)
                grades[slot_name] = {"quote": True, "text": s[1]}
            else:
                grades[slot_name] = {"unit_price": s[0], "title": s[1], "link": s[2]}
        if not grades.get("standard"):
            unpriced.append(item)
        row = {"key": _row_key(pricing, item), "category": cat, "item": item,
               "unit": unit, "qty_driver": driver, "kind": kind, "note": note,
               "grades": grades}
        if isinstance(ref, str) and ref in CASE_SQFT:
            row["per_case_sqft"] = CASE_SQFT[ref]
        rows.append(row)
    allowances = [{"key": _allow_key(item), "category": cat, "item": item,
                   "amount": amt, "basis": basis}
                  for (cat, item, amt, basis) in MASTER_ALLOWANCES]

    header = {"locked": True, "zip": meta["zip"],
              "market": market_for_zip(meta["zip"]),
              "pulled": meta["date"], "locked_on": date.today().isoformat(),
              "source": "SerpApi home_depot delivery_zip pricing",
              "rows": rows, "allowances": allowances}
    json_path = DATA_DIR / f"master_materials_locked_{meta['zip']}.json"
    json_path.write_text(json.dumps(header, indent=1), encoding="utf-8")

    csv_path = DATA_DIR / f"master_materials_locked_{meta['zip']}.csv"
    cols = ["kind", "category", "item", "unit", "qty_driver", "budget_price",
            "standard_price", "upgrade_price", "standard_product", "standard_link",
            "budget_product", "budget_link", "upgrade_product", "upgrade_link", "note"]
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        for row in rows:
            g = row["grades"]

            def part(slot, field):
                s = g.get(slot)
                if not s or s.get("quote"):
                    return ""
                return s.get(field, "")

            note = row["note"]
            quartz = g.get("upgrade") or {}
            if quartz.get("quote"):
                note = (note + " | " if note else "") + f"Upgrade: {quartz['text']}"
            w.writerow([f"sku-{row['kind']}", row["category"], row["item"], row["unit"],
                        row["qty_driver"], part("budget", "unit_price"),
                        part("standard", "unit_price"), part("upgrade", "unit_price"),
                        part("standard", "title"), part("standard", "link"),
                        part("budget", "title"), part("budget", "link"),
                        part("upgrade", "title"), part("upgrade", "link"), note])
        for a in allowances:
            w.writerow(["allowance", a["category"], a["item"], "", "", "",
                        a["amount"], "", "", "", "", "", "", "", a["basis"]])

    if unpriced:
        logger.warning("LOCK: %d rows have no standard price: %s",
                       len(unpriced), ", ".join(unpriced))
    logger.info("Locked artifacts: %s + %s (%d rows, %d allowances, pulled %s)",
                json_path, csv_path, len(rows), len(allowances), meta["date"])
    return json_path, csv_path


def build_master_workbook(prices: dict, meta: dict, out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Master List")
    _title(ws, "Master material list: every rehab, one standard",
           f"Home Depot pricing local to zip {meta['zip']} (SerpApi home_depot engine, "
           f"pulled {meta['date']}). Standard is the pick we buy by default; Budget and "
           "Upgrade are the approved swaps where a real grade choice exists.")
    cols = ["Category", "Material", "Unit", "Qty driver", "Budget $", "Standard $",
            "Upgrade $", "Standard product", "Budget / Upgrade options", "Note"]
    _sheet_header(ws, 4, cols)
    r, last_cat, unpriced = 5, None, []
    for (cat, item, unit, driver, pricing, note) in MASTER:
        kind, ref = pricing
        if kind == "graded":
            slots = {"budget": _slot(prices, ref, 0.25), "standard": _slot(prices, ref, 0.5),
                     "upgrade": _slot(prices, ref, 0.75)}
            if slots["standard"]:
                # A thin result set collapses the percentiles; only show a
                # spread that is real.
                if slots["budget"] and slots["budget"][0] == slots["standard"][0]:
                    slots["budget"] = None
                if slots["upgrade"] and slots["upgrade"][0] == slots["standard"][0]:
                    slots["upgrade"] = None
        elif kind == "classes":
            slots = {s: _slot(prices, k, 0.5) for s, k in ref.items()}
            slots.setdefault("budget", None)
            slots.setdefault("upgrade", None)
        else:
            slots = {"budget": None, "standard": _slot(prices, ref, 0.5), "upgrade": None}

        if cat != last_cat:
            ws.cell(row=r, column=1, value=cat).font = Font(bold=True, size=10, color=BLUE)
            last_cat = cat
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=driver).font = Font(size=9, color=GREY)
        std = slots.get("standard")
        if not std:
            ws.cell(row=r, column=8, value="not priced (no in-band API result)").font = \
                Font(size=9, color=GOLD)
            unpriced.append(item)
        else:
            for slot, col in (("budget", 5), ("standard", 6), ("upgrade", 7)):
                s = slots.get(slot)
                if s and s[0] is not None:
                    c = ws.cell(row=r, column=col, value=s[0])
                    c.number_format = "$#,##0.00"
            pc = ws.cell(row=r, column=8, value=std[1])
            pc.font = Font(size=9, color=NAVY)
            if std[2]:
                pc.hyperlink = std[2]
            opts = []
            for slot in ("budget", "upgrade"):
                s = slots.get(slot)
                if s and s[1] != std[1]:
                    opts.append(f"{slot.capitalize()}: {s[1][:60]}")
            if opts:
                ws.cell(row=r, column=9, value="  |  ".join(opts)).font = Font(size=8, color=GREY)
        if note:
            ws.cell(row=r, column=10, value=note).font = Font(size=9, color=GREY)
        for j in range(1, 11):
            ws.cell(row=r, column=j).border = THIN
            ws.cell(row=r, column=j).alignment = Alignment(
                vertical="top", wrap_text=(j in (4, 8, 9, 10)))
        r += 1
    if unpriced:
        r += 1
        ws.cell(row=r, column=2, value=f"Not priced: {', '.join(unpriced)}").font = \
            Font(size=9, color=GOLD)
    for j, w in enumerate([20, 32, 8, 26, 9, 10, 9, 46, 40, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    ws = wb.create_sheet("Allowances")
    _title(ws, "Standard allowances (no Home Depot SKU)",
           "Services, subs and quote-only scopes. Same standardization: one number "
           "everywhere until the PM revises it.")
    _sheet_header(ws, 4, ["Category", "Item", "Amount $", "Basis"])
    r = 5
    for (cat, item, amt, basis) in MASTER_ALLOWANCES:
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=item)
        c = ws.cell(row=r, column=3, value=amt)
        c.number_format = "$#,##0.00" if amt < 100 else "$#,##0"
        ws.cell(row=r, column=4, value=basis).font = Font(size=9, color=GREY)
        for j in range(1, 5):
            ws.cell(row=r, column=j).border = THIN
        r += 1
    for j, w in enumerate([22, 34, 10, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws = wb.create_sheet("Read Me", 0)
    _title(ws, "Master material list: how to use it")
    notes = [
        "PURPOSE: one standard materials basket for every rehab. The PM reviews this "
        "list, swaps any SKU they disagree with, and the result becomes the only "
        "materials we buy. Per-property lists then only decide QUANTITIES and which "
        "Budget/Upgrade swaps apply, so rehab costs stay stable across projects.",
        f"PRICING: every priced line is a live Home Depot SKU local to zip {meta['zip']} "
        f"(SerpApi home_depot engine, delivery_zip pricing, pulled {meta['date']}, "
        f"{meta['searches_used']} searches total). Product names link to the SKU.",
        "STANDARD / BUDGET / UPGRADE: Standard is the default buy (median in-band "
        "top-seller). Budget and Upgrade are the approved swaps: 25th/75th percentile "
        "products where grades exist on one search, or different product classes "
        "(3-tab vs architectural, laminate vs stocked slab vs quartz, unfinished vs "
        "shaker vs soft-close). Rows without a spread are commodity items where grade "
        "does not exist.",
        "QTY DRIVER: how each line scales to a real house, so any project's shopping "
        "list is this sheet times the driver.",
        "ALLOWANCES sheet: services and quote-only scopes with no HD SKU. Standardized "
        "the same way.",
        "MAINTENANCE: prices drift. Re-pull with 'python src/material_list.py --master "
        "--zip <zip>' before each project cycle (cached keys refresh only when evicted); "
        "a different market just needs its zip. A re-pull does NOT re-lock: the frozen "
        "prices the rehab engine and the rehab-estimator skill use only change when "
        "'--master --lock' is run explicitly (writes data/master_materials_locked_*).",
        "MATERIALS ONLY: labor, soft costs and scenario math stay in the rehab engine "
        "and the post-walkthrough workbook.",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=4 + i * 2, column=1, value=n)
        c.font = Font(size=10, color="374151")
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    wb.save(out_path)


# ── CLI ───────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="Tiered material list priced from Home Depot")
    ap.add_argument("--address", default="158 Old State Rd")
    ap.add_argument("--zip", dest="zip_code", required=True,
                    help="Target ZIP. sku_pricing consumes the locked list whose ZIP "
                         "matches its SKU_LOCKED_ZIP (default 78704, Austin); other "
                         "markets stay on the engine tables until their own list is "
                         "captured with --master --lock.")
    ap.add_argument("--sqft", type=int, default=1946)
    ap.add_argument("--beds", type=int, default=3, help="FINISHED config")
    ap.add_argument("--baths", type=float, default=2.0, help="FINISHED config")
    ap.add_argument("--year-built", type=int, default=1938)
    ap.add_argument("--profile", default="158", choices=["158", "sanland"],
                    help="Walk-driven line catalog to use")
    ap.add_argument("--master", action="store_true",
                    help="Build the property-agnostic master standardization list instead")
    ap.add_argument("--windows-done", type=int, default=11)
    ap.add_argument("--drywall-done-pct", type=int, default=60)
    ap.add_argument("--cached", action="store_true", help="Reuse the saved pricing pull")
    ap.add_argument("--lock", action="store_true",
                    help="With --master: export the LOCKED data/ artifacts the rehab "
                         "engine and rehab-estimator skill consume (explicit, dated act)")
    ap.add_argument("--out", default="")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s %(message)s")

    sqft = args.sqft
    roof_sqft = int(sqft * 1.1)
    window_openings = max(8, sqft // 100)          # rehab_estimator's estimate
    q = {
        "sqft": sqft,
        "floor_sqft": int(sqft * 1.1),             # 10% waste
        "roof_squares": round(roof_sqft / 100, 1),
        "shingle_bundles": -(-roof_sqft * 3 // 100),
        "drip_pieces": 19,
        "vapor_rolls": 6,
        "windows_done": args.windows_done,
        "windows_remaining": max(0, window_openings - args.windows_done),
        "drywall_done_pct": args.drywall_done_pct,
        "drywall_sheets": round(sqft * 3.4 * (100 - args.drywall_done_pct) / 100 / 32),
        "baseboard_lf": 720,
        "baseboard_sticks": 60,
    }
    subject = {"address": args.address, "beds": args.beds, "baths": args.baths}

    cache = OUTPUT_DIR / f"hd_materials_list_{args.zip_code}.json"
    if not args.master:
        lines = build_lines_sanland(q) if args.profile == "sanland" else build_lines(q)
        needed = set()
        for (_cat, _item, _qty, _unit, pricing, _note) in lines:
            kind, ref = pricing
            if kind in ("graded", "fixed"):
                needed.add(ref)
            elif kind == "per_tier":
                needed.update(ref.values())
    prices = pull_prices(args.zip_code, cache, args.cached,
                         None if args.master else needed)
    meta_file = json.loads(cache.read_text(encoding="utf-8"))

    if args.master:
        meta = {"zip": args.zip_code,
                "date": meta_file.get("date", date.today().isoformat()),
                "searches_used": meta_file.get("searches_used", 0)}
        out = args.out or "Master_Material_List.xlsx"
        build_master_workbook(prices, meta, out)
        print(f"\nMaster list: {len(MASTER)} materials + {len(MASTER_ALLOWANCES)} allowances")
        print(f"Workbook: {out}")
        if args.lock:
            jp, cp = export_locked_artifacts(prices, meta)
            print(f"LOCKED: {jp}\nLOCKED: {cp}")
        return 0
    # use_locked_materials=False on purpose: this column is the engine-table
    # baseline the SKU list is compared AGAINST. Letting it read the locked
    # list would compare the pull to itself.
    engine_mats = {t: estimate_rehab(sqft=sqft, bedrooms=args.beds, bathrooms=args.baths,
                                     year_built=args.year_built, tier=t, scope="full",
                                     region=DEFAULT_REGION,
                                     use_locked_materials=False).total_materials
                   for t in (1, 2, 3)}
    walk_notes = {
        "158": [
            "Quantities are scaled to the walk: 11 window openings already have new vinyl "
            f"(only {q['windows_remaining']} priced; verify the count onsite), "
            f"{q['drywall_done_pct']}% of drywall is already hung and taped (only the "
            "remaining hang is priced).",
            "Bath 2 (the 48 in shower kit, 24 in vanity, second toilet) exists only in the "
            "3/2 reconfig, which is still gated on layout verification.",
        ],
        "sanland": [
            "Quantities are scoped from the 2026-07-26 video walk: hoarding cleanout first, "
            "a FULL roof assumed until a roofer scopes it (multi-room ceiling failure), "
            "carport framing rot, and vines hiding the siding condition.",
            "Tier 1 (minimum viable) KEEPS the intact-but-dated kitchen, bath, windows and "
            "trim per the walk read; tiers 2-3 replace them and re-rock over the paneling.",
            "The egress window line is what legalizes the no-window rear bedroom; an "
            "investor exit can skip it, a retail relist needs the bed count.",
        ],
    }
    meta = {"zip": args.zip_code, "date": meta_file.get("date", date.today().isoformat()),
            "searches_used": meta_file.get("searches_used", 0),
            "engine_materials": engine_mats,
            "walk_notes": walk_notes[args.profile]}

    safe = "".join(c if c.isalnum() or c in " -" else "_" for c in args.address)[:40].strip().replace(" ", "_")
    out = args.out or f"{safe}_Material_List.xlsx"
    totals = build_workbook(subject, q, lines, prices, meta, out)
    print(f"\nMaterials: Tier 1 ${totals[1]:,} | Tier 2 ${totals[2]:,} | Tier 3 ${totals[3]:,}")
    print(f"Workbook: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
