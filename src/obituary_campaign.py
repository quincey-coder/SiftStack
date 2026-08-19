"""Consolidate the obituary run into one mail campaign: who we mail, and why.

Corrects a mistake in the first pass. That version treated "the owner is alive"
as a reason to HOLD the whole record. It is not. When an 89 year old widow owns
the house, her adult children are usually the ones handling her affairs, and
they belong on the mailer alongside her. The right output is not a hold, it is a
DIFFERENT letter to a DIFFERENT set of people.

So each house gets a situation, and the situation picks the roster:

  estate_verified    the owner died and research confirms the survivor set
                     -> mail the confirmed heirs
  estate_presumed    SmartSkip says the owner died, unverified
                     -> mail the inferred heirs, decedent still unconfirmed
  widowed_owner      the owner is alive, their spouse died
                     -> mail the OWNER plus adult children, never an estate letter
  owner_parent_died  the owner is alive, a parent died
                     -> mail the OWNER, they already control the property
  living_owner       the owner is alive, who died is unknown
                     -> mail the owner plus adult children, softly
  unresolved         the decedent could not be identified and the relatives do
                     not corroborate -> do not mail

Usage:
  python src/obituary_campaign.py --validate      # Smarty-check owner addresses too
  python src/obituary_campaign.py --out output/Obituary_Campaign.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from obituary_opportunity import CACHE, assemble, parse_date  # noqa: E402
from obituary_mail_validate import grade, load_env, norm  # noqa: E402

NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "CC0000"

SITUATION_LETTER = {
    "estate_verified": "Estate letter to the confirmed heirs. Name the property, not the death.",
    "estate_presumed": "Estate letter, but the decedent is unconfirmed. Softer wording, no condolence.",
    "widowed_owner": "NOT an estate letter. The owner is alive and recently widowed. "
                     "Write to the owner about the house; copy the adult children, who are "
                     "usually the ones handling her affairs.",
    "owner_parent_died": "Write to the OWNER. They already control the property and nothing "
                         "passed to anybody. Do not send condolences to the surviving parent.",
    "living_owner": "Owner is alive, we do not know who died. Plain we-buy-houses letter to "
                    "the owner plus adult children. No reference to a death.",
    "unresolved": "Do not mail. The decedent is unidentified and the relatives do not corroborate.",
    "entity_owner": "Do not mail. A trust or LLC holds title and was never name-traced. "
                    "Resolve the human trustee through Enformion BusinessV2 first.",
}


def situation_for(rec_ss, research, ss_deceased):
    """Classify one house. Verified research always beats the SmartSkip flag."""
    r = research or {}
    v = (r.get("verdict") or "").lower()
    if v.startswith("unresolved"):
        return "unresolved", r.get("evidence") or "decedent not identified"
    if "father" in v or "parent" in v:
        return "owner_parent_died", r.get("evidence") or ""
    if "wife" in v or "husband" in v or "spouse" in v:
        return "widowed_owner", r.get("evidence") or ""
    if "owner did die" in v:
        return "estate_verified", r.get("evidence") or ""
    return ("estate_presumed", "SmartSkip flags the owner deceased, unverified") if ss_deceased \
        else ("living_owner", "SmartSkip flags the owner alive, decedent unknown")


def adult(p):
    try:
        return int(str(p.get("age") or "").strip()) >= 25
    except ValueError:
        return True


def build_roster(crm, ss_recs, research_by_record):
    """One row per person we will mail, plus a per-house situation record."""
    ss_by_addr = {norm(r.get("property_address")): r for r in ss_recs}
    people, houses = [], []

    for m in crm:
        ss = ss_by_addr.get(norm(m["street"])) or {}
        res = research_by_record.get(ss.get("input_name") or "")

        # An LLC or trust was never name-traced, so it has no relatives and no
        # signer set. Mailing "Dr William Jerry Wilkins Trust" a we-buy-houses
        # letter is a wasted stamp: the human trustee is the target and only
        # Enformion BusinessV2 can name them.
        # owner_type is only "trust" when reisift parsed the deed cleanly. The
        # Luttrell revocable trust comes through as "incomplete" with the whole
        # trust name jammed into the company field, so test the words too.
        from obituary_dp_input import ENTITY_WORDS
        ent_words = set(re.findall(r"[a-z]+",
                                   (m.get("owner_company") or m.get("owner_name") or "").lower()))
        is_entity_owner = ((m.get("owner_type") or "").lower() in ("trust", "company", "entity")
                           or bool(ent_words & ENTITY_WORDS))
        if not ss and is_entity_owner:
            sit, evidence = "entity_owner", "trust or LLC on title, never name-traced"
        else:
            sit, evidence = situation_for(ss, res, bool(ss.get("deceased")))

        owner_alive = sit in ("widowed_owner", "owner_parent_died", "living_owner")
        drop = {n.strip().lower() for n in ((res or {}).get("drop_names") or [])}
        roster = []

        if owner_alive and sit != "entity_owner":
            roster.append({
                "who": "owner", "name": m["owner_name"], "relationship": "owner of record",
                "age": "", "street": m.get("mail_street") or m.get("street"),
                "city": m.get("mail_city") or m.get("city"),
                "state": m.get("mail_state") or m.get("state"),
                "zip": m.get("mail_zip") or m.get("zip"),
                "phones": "", "reason": "owns the property and can sign alone",
            })
        # The children go on the mailer in every situation except owner_parent_died,
        # where nothing passed to anybody, and unresolved, where we mail nobody.
        if sit not in ("owner_parent_died", "unresolved", "entity_owner"):
            for d in (ss.get("dms") or []):
                if not adult(d):
                    continue
                # Research can overturn a SmartSkip relative. Betty Burchell's
                # obituary names two children, not three; Edward Meek's "wife,
                # 41" is a grandson's wife. Those get dropped, not mailed.
                if (d.get("name") or "").strip().lower() in drop:
                    continue
                roster.append({
                    "who": "heir" if not owner_alive else "family",
                    "name": d.get("name"), "relationship": d.get("canon_rel"),
                    "age": d.get("age"), "street": d.get("mailing_street"),
                    "city": d.get("mailing_city"), "state": d.get("mailing_state"),
                    "zip": d.get("mailing_zip"),
                    "phones": ", ".join(p["number"] for p in (d.get("phones") or [])),
                    "reason": ("must sign as heir" if not owner_alive
                               else "likely handling the owner's affairs"),
                })

        for p in roster:
            people.append({**p, "record": ss.get("input_name") or m["owner_name"],
                           "property": m["street"], "property_city": m["city"],
                           "property_zip": m["zip"], "situation": sit,
                           "opportunity_score": m["score"], "value": m["estimate_value"],
                           "equity_pct": m["equity_pct"], "vacant": "yes" if m["vacant"] else "",
                           "months_since_death": m["months_since"]})

        houses.append({
            "record": ss.get("input_name") or m["owner_name"], "property": m["street"],
            "city": m["city"], "zip": m["zip"], "county": m["county"],
            "score": m["score"], "tier": m["tier"], "value": m["estimate_value"],
            "equity_pct": m["equity_pct"], "equity_dollars": m["equity_dollars"],
            "beds": m["beds"], "baths": m["baths"], "sqft": m["sqft"], "year_built": m["year_built"],
            "vacant": "yes" if m["vacant"] else "", "months_since_death": m["months_since"],
            "tax_owed": m["tax_amount"], "liens": m["lien_amount"],
            "owner_of_record": m["owner_name"],
            "situation": sit, "evidence": evidence,
            "verified": bool(res and not (res.get("verdict") or "").lower().startswith("unresolved")),
            "obituary_source": (res or {}).get("source") or "",
            "people_to_mail": len(roster),
            "relatives_found": len(ss.get("relatives") or []),
            "letter": SITUATION_LETTER[sit],
            "action": " ".join(x for x in [(res or {}).get("action"),
                                           (res or {}).get("roster_note")] if x),
            "corrected_roster": ", ".join(
                f"{c['name']} ({c['relationship']})"
                for c in ((res or {}).get("corrected_roster") or [])),
        })
    return people, houses


def validate_addresses(people, aid, tok, batch_size=100):
    from smartystreets_python_sdk import BasicAuthCredentials, Batch, ClientBuilder
    from smartystreets_python_sdk.us_street import Lookup
    from smartystreets_python_sdk.us_street.match_type import MatchType
    client = ClientBuilder(BasicAuthCredentials(aid, tok)).build_us_street_api_client()
    idx = [i for i, p in enumerate(people) if p.get("street")]
    for s in range(0, len(idx), batch_size):
        chunk = idx[s:s + batch_size]
        batch = Batch()
        for i in chunk:
            p = people[i]
            lk = Lookup()
            lk.street, lk.city = p["street"], p.get("city") or ""
            lk.state, lk.zipcode = p.get("state") or "TN", p.get("zip") or ""
            lk.candidates, lk.match = 1, MatchType.INVALID
            batch.add(lk)
        try:
            client.send_batch(batch)
        except Exception as e:
            for i in chunk:
                people[i]["mail_status"], people[i]["mail_reason"] = "error", str(e)[:120]
            continue
        for slot, i in enumerate(chunk):
            cands = batch[slot].result
            c = cands[0] if cands else None
            st, why = grade(c)
            people[i]["mail_status"], people[i]["mail_reason"] = st, why
            if c:
                people[i]["usps_address"] = f"{c.delivery_line_1}, {c.last_line}"
    for p in people:
        p.setdefault("mail_status", "unverified")
        p.setdefault("mail_reason", "no street address")
        p.setdefault("usps_address", "")
    return people


# ── workbook ──────────────────────────────────────────────────────────

def _title(ws, t, sub=""):
    ws.cell(row=1, column=1, value=t).font = Font(bold=True, size=15, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=10, color="666666")


def _hdr(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _w(ws, widths):
    for i, x in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def _rows(ws, start, rows, money=(), wrap=True):
    r = start
    for row in rows:
        for j, v in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(wrap_text=wrap, vertical="top")
            if j in money and isinstance(v, (int, float)):
                c.number_format = "$#,##0"
        r += 1
    return r


def _kv(ws, start, rows):
    r = start
    for pair in rows:
        a, b = (list(pair) + ["", ""])[:2]
        ws.cell(row=r, column=1, value=a)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if a and b == "":
            ws.cell(row=r, column=1).font = Font(bold=True, color=BLUE)
        r += 1
    return r


SIT_LABEL = {
    "estate_verified": "Estate, heirs confirmed by research",
    "estate_presumed": "Estate presumed, decedent unconfirmed",
    "widowed_owner": "Owner alive, spouse died",
    "owner_parent_died": "Owner alive, parent died",
    "living_owner": "Owner alive, decedent unknown",
    "unresolved": "Unresolved, do not mail",
    "entity_owner": "Trust or LLC on title, needs BusinessV2",
}
SIT_ORDER = ["estate_verified", "estate_presumed", "widowed_owner", "owner_parent_died",
             "living_owner", "unresolved", "entity_owner"]


def build_book(people, houses, out: Path, today):
    wb = Workbook()
    mailable = [p for p in people if p["mail_status"] == "deliverable"
                and p["situation"] not in ("unresolved", "entity_owner")]
    sit_people = Counter(p["situation"] for p in people)
    sit_house = Counter(h["situation"] for h in houses)

    ws = wb.active
    ws.title = "1 Summary"
    _title(ws, "Obituary mail campaign", f"ty+2, top 60 obituary opportunities, {today}")
    _w(ws, [40, 92])
    _kv(ws, 4, [
        ["The number", ""],
        ["Houses in the campaign", len(houses)],
        ["People we are mailing", len(mailable)],
        ["Average people per house", round(len(mailable) / max(len(houses), 1), 1)],
        ["Addresses USPS confirmed deliverable",
         f"{sum(1 for p in people if p['mail_status']=='deliverable')} of {len(people)}"],
        ["Held back (unresolved decedent)", sum(1 for p in people if p["situation"] == "unresolved")],
        ["", ""],
        ["What changed from the first pass", ""],
        ["", "The first pass held an entire record whenever the owner turned out to be alive. "
              "That was wrong. A living 89 year old widow still has adult children handling her "
              "affairs, and they belong on the mailer next to her. A living owner does not mean "
              "no lead, it means a different letter to a different set of people. Only the houses "
              "where the decedent could not be identified at all are held now."],
        ["", ""],
        ["The six situations, and who gets mailed", ""],
    ] + [[f"{SIT_LABEL[s]}  ({sit_house.get(s,0)} houses, {sit_people.get(s,0)} people)",
          SITUATION_LETTER[s]] for s in SIT_ORDER if sit_house.get(s)] + [
        ["", ""],
        ["What is verified and what is not", ""],
        ["Houses with a confirmed decedent",
         f"{sum(1 for h in houses if h['verified'])} of {len(houses)}"],
        ["", "dataflik carries no obituary id, url, decedent name or date of death on any of the "
              "740 records, so the source obituary has to be reconstructed from name plus date "
              "plus locality. That is reliable for a Helmboldt and useless for a Joyce Brown. "
              "Every unverified house is an inference from a relatives graph, not a fact."],
        ["Research hit rate so far",
         "3 of the 5 houses researched had a materially wrong signer set, including the "
         "top-ranked house. Treat the unverified rows accordingly."],
    ])

    # 2 Mail roster
    ws = wb.create_sheet("2 Mail roster")
    _title(ws, f"{len(mailable)} people to mail",
           "Sorted by opportunity score. Every address USPS validated.")
    cols = [("Score", 7), ("Property", 25), ("Situation", 24), ("Who", 8), ("Name", 22),
            ("Relationship", 13), ("Age", 5), ("Mailing address (USPS)", 40),
            ("Why this person", 30), ("Phones", 34)]
    _w(ws, [w for _, w in cols])
    _hdr(ws, 4, [h for h, _ in cols])
    _rows(ws, 5, [[p["opportunity_score"], p["property"], SIT_LABEL[p["situation"]], p["who"],
                   p["name"], p["relationship"], p["age"],
                   p.get("usps_address") or f"{p['street']}, {p['city']} {p['state']} {p['zip']}",
                   p["reason"], p["phones"]]
                  for p in sorted(mailable, key=lambda x: (-(x["opportunity_score"] or 0),
                                                           x["property"], x["who"]))])
    ws.freeze_panes = "C5"

    # 3 The houses
    ws = wb.create_sheet("3 The houses")
    _title(ws, f"{len(houses)} houses, one row each",
           "What is going on at each property and what to send.")
    cols = [("Score", 7), ("Tier", 5), ("Property", 25), ("City", 13), ("Zip", 7),
            ("Value", 11), ("Equity %", 8), ("Beds", 5), ("SqFt", 7), ("Built", 6),
            ("Vacant", 7), ("Months since death", 9), ("Tax owed", 10), ("Liens", 10),
            ("Owner of record", 20), ("Situation", 24), ("Verified", 8),
            ("People mailed", 8), ("Relatives found", 8),
            ("What to send", 46), ("Evidence / action", 52), ("Obituary source", 34)]
    _w(ws, [w for _, w in cols])
    _hdr(ws, 4, [h for h, _ in cols])
    _rows(ws, 5, [[h["score"], h["tier"], h["property"], h["city"], h["zip"], h["value"],
                   h["equity_pct"], h["beds"], h["sqft"], h["year_built"], h["vacant"],
                   h["months_since_death"], h["tax_owed"], h["liens"], h["owner_of_record"],
                   SIT_LABEL[h["situation"]], "yes" if h["verified"] else "",
                   h["people_to_mail"], h["relatives_found"], h["letter"],
                   " ".join(x for x in [h["evidence"], h["action"]] if x), h["obituary_source"]]
                  for h in sorted(houses, key=lambda x: -(x["score"] or 0))],
          money=(6, 13, 14))
    ws.freeze_panes = "D5"

    # 4 Research findings
    ws = wb.create_sheet("4 Research findings")
    _title(ws, "Stage C: what the obituary research actually found",
           "5 of 17 uncertain houses researched. Three of five overturned the inference.")
    _w(ws, [22, 26, 26, 60, 60, 34])
    _hdr(ws, 4, ["Record", "Property", "Verdict", "Evidence", "Action", "Source"])
    _rows(ws, 5, [[h["record"], h["property"], SIT_LABEL[h["situation"]], h["evidence"],
                   h["action"], h["obituary_source"]]
                  for h in houses if h["verified"] or h["situation"] == "unresolved"])

    # 5 Not mailed
    ws = wb.create_sheet("5 Not mailed")
    held = [p for p in people if p not in mailable]
    _title(ws, f"{len(held)} people not being mailed", "Every exclusion with its reason.")
    _w(ws, [25, 22, 14, 24, 40, 40])
    _hdr(ws, 4, ["Property", "Name", "Relationship", "Situation", "Why not mailed", "Address as given"])
    _rows(ws, 5, [[p["property"], p["name"], p["relationship"], SIT_LABEL[p["situation"]],
                   ("decedent unresolved, house held" if p["situation"] == "unresolved"
                    else f"{p['mail_status']}: {p['mail_reason']}"),
                   f"{p.get('street') or ''} {p.get('city') or ''} {p.get('zip') or ''}".strip()]
                  for p in held])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out, mailable, sit_house, sit_people


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--cache", default=str(CACHE))
    ap.add_argument("--ranked", default="output/dp/ranked_records.json")
    ap.add_argument("--research", default="output/dp/stage_c_research.json")
    ap.add_argument("--top", type=int, default=60)
    ap.add_argument("--out", default="output/Obituary_Campaign.xlsx")
    ap.add_argument("--validate", action="store_true", help="Smarty-validate every address")
    args = ap.parse_args()

    today = date.today()
    qualified, _, _ = assemble(Path(args.cache), 3, date(2026, 8, 7))
    crm = qualified[:args.top]
    ss_recs = json.loads(Path(args.ranked).read_text(encoding="utf-8"))
    research = json.loads(Path(args.research).read_text(encoding="utf-8"))
    rbr = {f["record"]: f for f in research["findings"]}

    people, houses = build_roster(crm, ss_recs, rbr)
    print(f"{len(houses)} houses, {len(people)} people on the roster")

    if args.validate:
        env = load_env()
        aid, tok = env.get("SMARTY_AUTH_ID", ""), env.get("SMARTY_AUTH_TOKEN", "")
        if aid and tok:
            validate_addresses(people, aid, tok)
            print(f"  validated: {dict(Counter(p['mail_status'] for p in people))}")
        else:
            print("  no Smarty creds, addresses unvalidated")
    else:
        for p in people:
            p.setdefault("mail_status", "unvalidated")
            p.setdefault("mail_reason", "")
            p.setdefault("usps_address", "")

    out, mailable, sh, sp = build_book(people, houses, Path(args.out), today)
    with Path("output/dp/mail_roster_final.csv").open("w", newline="", encoding="utf-8") as fh:
        cols = ["record", "property", "property_city", "property_zip", "situation", "who", "name",
                "relationship", "age", "street", "city", "state", "zip", "usps_address",
                "mail_status", "mail_reason", "reason", "phones", "opportunity_score", "value",
                "equity_pct", "vacant", "months_since_death"]
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(people)

    print(f"\n  MAILING {len(mailable)} people across "
          f"{len({p['property'] for p in mailable})} houses")
    for s in SIT_ORDER:
        if sh.get(s):
            print(f"    {SIT_LABEL[s]:42} {sh[s]:3} houses  {sp.get(s,0):3} people")
    print(f"\nwrote {out} and output/dp/mail_roster_final.csv")


if __name__ == "__main__":
    main()
