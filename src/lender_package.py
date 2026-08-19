"""The Private Lender Package (build 1.0.43, 2026-08).

Document 2 of the 8-piece lender set. The other 7 are Word documents built by
`lender_docs.py` from this same spec.

THE FRONT END IS A MIRROR OF THE REPAIR ESTIMATOR, block for block. That sheet
is what the team already reads, so the lender package uses its exact summary
layout rather than inventing one:

    Property header
    Property Values & Pricing        |  Holding Costs (Monthly)
    Financing Costs                  |  Buying Transaction Costs
                                     |  Selling Transaction Costs
    Estimated Net Profit and ROI Snapshot
    Purchase & Deal Analysis         |  Lender Coverage & Return

Six columns: label, percent, dollars on the left; label, percent, dollars on
the right. Banded section headers. Bold "Total X:" rows closing every block.
The one adaptation is the last right-hand block, which is the lender's
coverage instead of our cash on cash, because this is their document.

Repair Costs is the Estimator's detail grid on its own tab, Category / Y/N /
Repair Type / Qty / Unit / Unit Cost / Total / Notes, and it rolls up into
Estimated Repair Costs on the front page. On a straight relist every line is
zero and the front page shows zero, which is the correct answer rather than a
missing section.

Everything is a live Excel formula off workbook defined names, sentences
included. Blue cells are inputs. Grey cells are calculated.

Usage:
    python src/lender_package.py --spec deals/158_old_state_lender.json
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)

NAVY, BLUE, GREEN, GOLD, RED, GREY = "0A1130", "316AFF", "1B9E5A", "B8860B", "C00000", "666666"
INPUT_FILL = PatternFill("solid", fgColor="DCE6FA")
CALC_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="E4EAF4")
HAIR = Side(style="thin", color="C8D0E0")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

MONEY, MONEY2 = "$#,##0", "$#,##0.00"
PCT0, PCT1, PCT2 = "0%", "0.0%", "0.00%"
MULT, NUM = '0.00"x"', "#,##0"

TARGET_MAX_LTARV = 0.75
SHEET_ORDER = ["Deal Overview", "Your Investment", "Term Sheet", "Repair Costs",
               "Resale Value", "Comps", "Backups and Risk", "Next Steps"]


# ══ formatting ═════════════════════════════════════════════════════════

def _title(ws, text, sub="", note=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=15, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=10.5, color=NAVY)
    if note:
        ws.cell(row=3, column=1, value=note).font = Font(size=9, color=GREY)


def _band(ws, row, text, col=1, span=3, headers=()):
    """Banded section header. `headers` fills the percent / total columns."""
    for j in range(col, col + span):
        c = ws.cell(row=row, column=j)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BOX
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col + i, value=h)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    return row + 1


def _row(ws, row, col, label, pct=None, val=None, pct_fmt=PCT2, val_fmt=MONEY,
         bold=False, fill=CALC_FILL, pct_fill=None, note=None):
    """One Estimator-style row: label, percent column, dollar column."""
    c = ws.cell(row=row, column=col, value=label)
    c.font = Font(size=10, bold=bold)
    c.alignment = Alignment(wrap_text=False, vertical="center")
    if pct is not None:
        p = ws.cell(row=row, column=col + 1, value=pct)
        if pct_fmt:                       # a text cell has no number format
            p.number_format = pct_fmt
        p.fill = pct_fill or fill
        p.border = BOX
        p.font = Font(size=10, bold=bold)
        p.alignment = Alignment(horizontal="center")
    if val is not None:
        v = ws.cell(row=row, column=col + 2, value=val)
        if val_fmt:
            v.number_format = val_fmt
        v.fill = fill
        v.border = BOX
        v.font = Font(size=10, bold=bold)
    if note:
        ws.cell(row=row, column=col + 3, value=note).font = Font(size=8, color=GREY)
    return row + 1


def _para(ws, row, text, color=None, bold=False, size=10, col=1, span=6, height=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=size, bold=bold, color=color or "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row,
                       end_column=col + span - 1)
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


def _kv(ws, row, label, value, fmt=None, fill=CALC_FILL, col=1, bold=False,
        note=None, wrap=False, span=1):
    c = ws.cell(row=row, column=col, value=label)
    c.font = Font(size=10, bold=bold)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    v = ws.cell(row=row, column=col + 1, value=value)
    if fmt:
        v.number_format = fmt
    v.fill = fill
    v.border = BOX
    v.font = Font(size=10, bold=bold)
    v.alignment = Alignment(wrap_text=wrap, vertical="top")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col + 1, end_row=row,
                       end_column=col + span)
    if note is not None:
        n = ws.cell(row=row, column=col + 1 + max(span, 1), value=note)
        n.font = Font(size=9, color=GREY)
        n.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _hdr(ws, row, headers, col=1):
    for j, h in enumerate(headers, col):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="44506B")
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if j > col else "left")
        c.border = BOX
    return row + 1


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _money(v) -> str:
    if not isinstance(v, (int, float)):
        return str(v or "")
    return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"


# ── auto-fit ───────────────────────────────────────────────────────────
# Nobody should have to drag a column or a row to read this workbook. Widths
# come from the longest thing actually in each column, and wrapped rows get a
# height computed from the width they ended up with.

_FMT_ONLY = re.compile(r'^[\s$#,.0%x"\-]*$')


def _rendered_len(v) -> int:
    """How many characters this cell will SHOW, not how long its formula is.

    A formula cell holds '=..."&TEXT(Loan,"$#,##0")&"...' but displays a
    sentence. Sizing off the raw formula would blow every column out.
    """
    if v is None:
        return 0
    if not isinstance(v, str):
        return len(f"{v:,.0f}") if isinstance(v, (int, float)) else len(str(v))
    if not v.startswith("="):
        return len(v)
    literals = [x for x in re.findall(r'"([^"]*)"', v) if not _FMT_ONLY.match(x)]
    n = sum(len(x) for x in literals) + 12 * len(re.findall(r"TEXT\(", v))
    if "IF(" in v and literals:
        n = int(n * 0.62)          # only one branch of an IF ever renders
    return n or 12


def _autofit(ws, min_w=9.0, max_w=54.0, pad=2.4, line_h=12.8, base_h=15.5):
    merged_cells, anchors = set(), {}
    for rng in ws.merged_cells.ranges:
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                merged_cells.add((rr, cc))
        anchors[(rng.min_row, rng.min_col)] = rng

    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None or (c.row, c.column) in merged_cells:
                continue
            widths[c.column] = max(widths.get(c.column, 0), _rendered_len(c.value))
    for col, w in widths.items():
        cur = ws.column_dimensions[get_column_letter(col)].width or 0
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(w + pad, min_w, cur if cur and cur < max_w else 0), max_w)

    def eff_width(c):
        if (c.row, c.column) in merged_cells:
            rng = anchors.get((c.row, c.column))
            if rng is None:
                return None
            return sum(ws.column_dimensions[get_column_letter(x)].width or min_w
                       for x in range(rng.min_col, rng.max_col + 1))
        return ws.column_dimensions[get_column_letter(c.column)].width or min_w

    for row in ws.iter_rows():
        lines = 1
        for c in row:
            if c.value is None:
                continue
            if not (c.alignment and c.alignment.wrap_text):
                continue
            w = eff_width(c)
            if not w:
                continue
            lines = max(lines, math.ceil(_rendered_len(c.value) / max(w * 1.02, 8)))
        if lines > 1:
            ws.row_dimensions[row[0].row].height = base_h + (lines - 1) * line_h


def _polish(ws):
    """Everything that makes it read like a document on first open."""
    _autofit(ws)
    ws.sheet_view.zoomScale = 100
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def _t(name, fmt=MONEY):
    return f'TEXT({name},"{fmt}")'


def _say(parts):
    out = []
    for p in parts:
        raw = p.startswith(("TEXT(", "IF(", "ROUND("))
        out.append(p if raw else '"' + p.replace('"', '""') + '"')
    return "=" + "&".join(out)


# ══ 1. Deal Overview: mirror of the Repair Estimator summary page ══════

def sheet_overview(wb, spec):
    ws = wb.create_sheet("Deal Overview")
    p, plan, b = spec["property"], spec["plan"], spec["borrower"]
    hold = spec.get("holding", {})
    buy = spec.get("buying_costs", {})
    sell = spec.get("selling_costs", {})
    names = {}

    def name(nm, col, row):
        names[nm] = f"'Deal Overview'!${get_column_letter(col)}${row}"

    _title(ws, "The Private Lender Package", p["full_address"],
           "Blue cells are yours to change. Grey cells calculate off them. This page "
           "is laid out the same way as our Repair Estimator so the numbers sit where "
           "you expect them.")

    # ── property header ──
    r = 5
    r = _band(ws, r, "Property", col=1, span=6)
    _kv(ws, r, "Property Address", p["full_address"], span=2)
    _row(ws, r, 4, "Number of Units", val=spec.get("units", 1), val_fmt=NUM)
    r += 1
    _kv(ws, r, "Total Square Footage", p.get("sqft", 0), fmt=NUM)
    _row(ws, r, 4, "Occupied (Y/N)", val=spec.get("occupied", "N"), val_fmt=None)
    r += 1
    _kv(ws, r, "Beds / Baths", f"{p.get('beds','')} / {p.get('baths','')}")
    _row(ws, r, 4, "Year Built", val=p.get("year_built", ""), val_fmt=NUM)
    r += 1
    _kv(ws, r, "County / Parcel", f"{p.get('county','')} / {p.get('parcel_id','')}")
    _row(ws, r, 4, "Date Prepared", val=spec.get("date", ""), val_fmt=None)
    r += 1
    r = _para(ws, r, "Property Description: " + plan["strategy"] + ". " +
              p.get("condition", ""), color=GREY, size=9, height=40)
    r += 1

    # ── Property Values & Pricing  |  Holding Costs (Monthly) ──
    top = r
    _band(ws, r, "Property Values & Pricing", col=1, span=3)
    _band(ws, r, "Holding Costs (Monthly)", col=4, span=3,
          headers=("Annually", "Monthly"))
    r += 1
    lv = r
    _row(ws, lv, 1, "After Repair Value", val=spec["resale"]["point"], fill=INPUT_FILL)
    name("ARV", 3, lv); lv += 1
    _row(ws, lv, 1, "Current As-Is Value", val=(spec.get("as_is") or {}).get("point", 0),
         fill=INPUT_FILL)
    name("AsIs", 3, lv); lv += 1
    _row(ws, lv, 1, "Estimated Repair Costs", val="=RehabTotal")
    lv += 1
    _row(ws, lv, 1, "Purchase Price", val=spec["contract_price"], fill=INPUT_FILL)
    name("Purchase", 3, lv); lv += 1
    _row(ws, lv, 1, "Resale Band Low (stress test)", val=spec["resale"].get("lo", 0),
         fill=INPUT_FILL)
    name("ARVLo", 3, lv); lv += 1
    _row(ws, lv, 1, "Resale Band High", val=spec["resale"].get("hi", 0), fill=INPUT_FILL)
    name("ARVHi", 3, lv); lv += 1
    _row(ws, lv, 1, "Estimated Hold Time (months)",
         val=int(spec["loan"]["term_months"]), val_fmt="0", fill=INPUT_FILL)
    name("HoldMonths", 3, lv); lv += 1

    rv = r
    hold_rows = []
    for label, annual, monthly in [
        ("Property Taxes", hold.get("taxes_annual", 0), None),
        ("HOA and Condo Fees", None, hold.get("hoa_monthly", 0)),
        ("Insurance Costs", hold.get("insurance_annual", 0), None),
        ("Utilities: Gas", None, hold.get("gas_monthly", 0)),
        ("Utilities: Water", None, hold.get("water_monthly", 0)),
        ("Utilities: Electricity", None, hold.get("electric_monthly", 0)),
        ("Lawn and Other", None, hold.get("other_monthly", 0)),
        ("Miscellaneous Holding Costs", None, hold.get("misc_monthly", 0)),
    ]:
        c = ws.cell(row=rv, column=4, value=label)
        c.font = Font(size=10)
        if annual is not None:
            a = ws.cell(row=rv, column=5, value=annual)
            a.number_format = MONEY; a.fill = INPUT_FILL; a.border = BOX
            m = ws.cell(row=rv, column=6, value=f"=E{rv}/12")
            m.number_format = MONEY; m.fill = CALC_FILL; m.border = BOX
        else:
            m = ws.cell(row=rv, column=6, value=monthly)
            m.number_format = MONEY; m.fill = INPUT_FILL; m.border = BOX
        hold_rows.append(rv)
        rv += 1
    # totals close both columns, Estimator style
    end = max(lv, rv)
    _row(ws, end, 1, "Purchase and Repair Costs:", val="=Purchase+RehabTotal",
         bold=True, fill=TOTAL_FILL)
    c = ws.cell(row=end, column=4, value="Total Monthly Holding Costs:")
    c.font = Font(size=10, bold=True)
    m = ws.cell(row=end, column=6,
                value=f"=SUM(F{hold_rows[0]}:F{hold_rows[-1]})")
    m.number_format = MONEY; m.fill = TOTAL_FILL; m.border = BOX
    m.font = Font(bold=True)
    name("HoldMonthly", 6, end)
    end += 1
    c = ws.cell(row=end, column=4, value="Total Holding Costs (over the hold):")
    c.font = Font(size=10, bold=True)
    m = ws.cell(row=end, column=6, value="=HoldMonthly*HoldMonths")
    m.number_format = MONEY; m.fill = TOTAL_FILL; m.border = BOX
    m.font = Font(bold=True)
    name("HoldTotal", 6, end)
    r = end + 2

    # ── Financing Costs | Buying, then Selling Transaction Costs ──
    _band(ws, r, "Financing Costs", col=1, span=3, headers=("Rate", "Total"))
    _band(ws, r, "Buying Transaction Costs", col=4, span=3,
          headers=("Perc. Of Purch", "Total"))
    r += 1
    lf = r
    _row(ws, lf, 1, "Private Lender Loan Amount", val=spec["loan"]["amount"],
         fill=INPUT_FILL, note="The one number that sets the deal")
    name("Loan", 3, lf); lf += 1
    _row(ws, lf, 1, "Loan Points", pct=float(spec["loan"].get("points", 0)),
         pct_fmt="0.0", val="=Loan*Points/100", pct_fill=INPUT_FILL)
    name("Points", 2, lf); name("PointsAmt", 3, lf); lf += 1
    _row(ws, lf, 1, "Loan Interest", pct=spec["loan"]["rate"], pct_fmt=PCT2,
         val="=Loan*Rate*HoldMonths/12", pct_fill=INPUT_FILL)
    name("Rate", 2, lf); name("Interest", 3, lf); lf += 1
    _row(ws, lf, 1, "Guaranteed Minimum Interest",
         pct="=Rate*MinIntMonths/12", pct_fmt=PCT1,
         val="=Loan*Rate*MinIntMonths/12")
    name("MinIntPct", 2, lf); name("MinIntAmt", 3, lf); lf += 1
    _row(ws, lf, 1, "Minimum Interest (months)",
         val=int(spec["loan"].get("min_interest_months", 3)), val_fmt="0",
         fill=INPUT_FILL)
    name("MinIntMonths", 3, lf); lf += 1
    _row(ws, lf, 1, "Total Financing Costs:", val="=PointsAmt+Interest", bold=True,
         fill=TOTAL_FILL)
    name("FinCosts", 3, lf); lf += 1

    rf = r
    buy_rows = []
    for label, pct, val in [
        ("Escrow / Attorney Fees", None, buy.get("escrow", 0)),
        ("Title Insurance / Search", buy.get("title_pct", 0), "=Purchase*F_TITLEPCT"),
        ("Recording Fees", None, buy.get("recording", 0)),
        ("Miscellaneous Buying Costs", None, buy.get("misc", 0)),
    ]:
        c = ws.cell(row=rf, column=4, value=label); c.font = Font(size=10)
        if pct is not None:
            pc = ws.cell(row=rf, column=5, value=pct)
            pc.number_format = PCT2; pc.fill = INPUT_FILL; pc.border = BOX
            pc.alignment = Alignment(horizontal="center")
            names["TitlePct"] = f"'Deal Overview'!$E${rf}"
            v = ws.cell(row=rf, column=6, value="=Purchase*TitlePct")
            v.number_format = MONEY; v.fill = CALC_FILL; v.border = BOX
        else:
            v = ws.cell(row=rf, column=6, value=val)
            v.number_format = MONEY; v.fill = INPUT_FILL; v.border = BOX
        buy_rows.append(rf)
        rf += 1
    c = ws.cell(row=rf, column=4, value="Total Buying Transaction Costs:")
    c.font = Font(size=10, bold=True)
    v = ws.cell(row=rf, column=6, value=f"=SUM(F{buy_rows[0]}:F{buy_rows[-1]})")
    v.number_format = MONEY; v.fill = TOTAL_FILL; v.border = BOX; v.font = Font(bold=True)
    name("BuyCosts", 6, rf)
    rf += 2

    _band(ws, rf, "Selling Transaction Costs", col=4, span=3,
          headers=("Perc. Of ARV", "Total"))
    rf += 1
    sell_rows, var_rows = [], []
    for label, key, is_pct in [
        ("Escrow / Attorney Fees", "escrow", False),
        ("Selling Recording Fees", "recording", False),
        ("Realtor Fees", "realtor_pct", True),
        ("Transfer and Conveyance Fees", "transfer_pct", True),
        ("Home Warranty", "warranty", False),
        ("Staging Costs", "staging", False),
        ("Marketing Costs", "marketing", False),
        ("Miscellaneous Selling Costs", "misc", False),
    ]:
        c = ws.cell(row=rf, column=4, value=label); c.font = Font(size=10)
        if is_pct:
            pc = ws.cell(row=rf, column=5, value=sell.get(key, 0))
            pc.number_format = PCT2; pc.fill = INPUT_FILL; pc.border = BOX
            pc.alignment = Alignment(horizontal="center")
            v = ws.cell(row=rf, column=6, value=f"=ARV*E{rf}")
            v.number_format = MONEY; v.fill = CALC_FILL; v.border = BOX
            var_rows.append(rf)
        else:
            v = ws.cell(row=rf, column=6, value=sell.get(key, 0))
            v.number_format = MONEY; v.fill = INPUT_FILL; v.border = BOX
        sell_rows.append(rf)
        rf += 1
    c = ws.cell(row=rf, column=4, value="Total Selling Transaction Costs:")
    c.font = Font(size=10, bold=True)
    pc = ws.cell(row=rf, column=5, value="=IFERROR(SellCosts/ARV,0)")
    pc.number_format = PCT1; pc.fill = TOTAL_FILL; pc.border = BOX
    pc.alignment = Alignment(horizontal="center"); pc.font = Font(bold=True)
    v = ws.cell(row=rf, column=6, value=f"=SUM(F{sell_rows[0]}:F{sell_rows[-1]})")
    v.number_format = MONEY; v.fill = TOTAL_FILL; v.border = BOX; v.font = Font(bold=True)
    name("SellCosts", 6, rf)
    # fixed vs variable, so the band-floor case reprices commission correctly
    names["SellVarPct"] = f"'Deal Overview'!$E${rf}"
    fixed = "+".join(f"F{x}" for x in sell_rows if x not in var_rows)
    varpct = "+".join(f"E{x}" for x in var_rows)
    rf += 1
    ws.cell(row=rf, column=4, value="Fixed portion / variable rate").font = Font(
        size=8, color=GREY)
    v = ws.cell(row=rf, column=6, value=f"={fixed}")
    v.number_format = MONEY; v.fill = CALC_FILL; v.border = BOX
    name("SellFixed", 6, rf)
    pc = ws.cell(row=rf, column=5, value=f"={varpct}")
    pc.number_format = PCT2; pc.fill = CALC_FILL; pc.border = BOX
    names["SellVarPct"] = f"'Deal Overview'!$E${rf}"
    r = max(lf, rf) + 2

    # ── Net profit and ROI snapshot ──
    r = _band(ws, r, "Estimated Net Profit and ROI Snapshot", col=1, span=6)
    _row(ws, r, 1, "Estimated NET PROFIT",
         val="=ARV-SellCosts-Purchase-RehabTotal-BuyCosts-HoldTotal-FinCosts",
         bold=True, fill=TOTAL_FILL)
    names["NetProfit"] = f"'Deal Overview'!$C${r}"
    c = ws.cell(row=r, column=4, value="Total Costs Return on Investment (ROI)")
    c.font = Font(size=10, bold=True)
    v = ws.cell(row=r, column=6, value="=IFERROR(NetProfit/TotalCost,0)")
    v.number_format = PCT1; v.fill = TOTAL_FILL; v.border = BOX; v.font = Font(bold=True)
    r += 2

    # ── Purchase & Deal Analysis  |  Lender Coverage & Return ──
    _band(ws, r, "Purchase and Deal Analysis", col=1, span=3)
    _band(ws, r, "Lender Coverage and Return", col=4, span=3)
    r += 1
    left = [
        ("After Repair Value", "=ARV", MONEY, None),
        ("Purchase Price", "=Purchase", MONEY, None),
        ("Estimated Repair Costs", "=RehabTotal", MONEY, None),
        ("Total Financing Costs", "=FinCosts", MONEY, None),
        ("Total Holding Costs", "=HoldTotal", MONEY, None),
        ("Total Buying Transaction Costs", "=BuyCosts", MONEY, None),
        ("Total Selling Transaction Costs", "=SellCosts", MONEY, None),
        ("TOTAL PROJECT COST",
         "=Purchase+RehabTotal+BuyCosts+HoldTotal+FinCosts", MONEY, "TotalCost"),
        # Cash needed at and through closing. Interest is excluded on purpose:
        # it is paid out of the sale proceeds, not out of pocket.
        ("Borrower Cash Into The Deal",
         "=MAX(Purchase+RehabTotal+BuyCosts+HoldTotal-Loan,0)", MONEY, "BorrowerCash"),
    ]
    right = [
        ("Loan Amount", "=Loan", MONEY, None),
        ("Total Payoff To Lender", "=Loan+Interest+PointsAmt", MONEY, "Payoff"),
        ("Loan to After Repair Value", "=IFERROR(Loan/ARV,0)", PCT1, "LTV"),
        ("Equity Cushion", "=ARV-Loan", MONEY, "Equity"),
        ("Day One Coverage", "=IFERROR(AsIs/DayOne,0)", MULT, None),
        ("Coverage at the Ask", "=IFERROR(NetPoint/Payoff,0)", MULT, None),
        ("Coverage at the Band Floor", "=IFERROR(NetLo/Payoff,0)", MULT, None),
        ("Total Income To Lender", "=Interest+PointsAmt", MONEY, "LenderIncome"),
        ("Annualized Yield To Lender", "=Rate+(Points/100)*(12/HoldMonths)", PCT1,
         "AnnYield"),
    ]
    for i in range(max(len(left), len(right))):
        rr = r + i
        if i < len(left):
            lab, val, fmt, nm = left[i]
            bold = lab.isupper()
            _row(ws, rr, 1, lab, val=val, val_fmt=fmt, bold=bold,
                 fill=TOTAL_FILL if bold else CALC_FILL)
            if nm:
                name(nm, 3, rr)
        if i < len(right):
            lab, val, fmt, nm = right[i]
            _row(ws, rr, 4, lab, val=val, val_fmt=fmt, fill=CALC_FILL)
            if nm:
                name(nm, 6, rr)
    r += max(len(left), len(right)) + 1

    # ── derived values with no natural home on the Estimator layout ──
    r = _band(ws, r, "Supporting Calculations", col=1, span=6)
    helpers = [
        ("DayOne", "Advance at Closing", "=Loan-RehabTotal", MONEY),
        ("RehabTranche", "Held Back for Repairs", "=MAX(Loan-DayOne,0)", MONEY),
        ("Draws", "Number of Draws", int(spec["loan"].get("draws", 0)), "0"),
        ("PerDraw", "Size of Each Draw", "=IF(Draws>0,RehabTranche/Draws,0)", MONEY),
        ("NetPoint", "Net Proceeds at the Ask", "=ARV-SellCosts", MONEY),
        ("NetLo", "Net Proceeds at the Band Floor",
         "=ARVLo-(SellFixed+ARVLo*SellVarPct)", MONEY),
        ("NetHi", "Net Proceeds at the Band High",
         "=ARVHi-(SellFixed+ARVHi*SellVarPct)", MONEY),
        ("EquityPct", "Equity Cushion, % of ARV", "=IFERROR(Equity/ARV,0)", PCT0),
        ("CoverPoint", "Coverage at the Ask", "=IFERROR(NetPoint/Payoff,0)", MULT),
        ("CoverFloor", "Coverage at the Band Floor", "=IFERROR(NetLo/Payoff,0)", MULT),
        ("DayOneCover", "Day One Coverage", "=IFERROR(AsIs/DayOne,0)", MULT),
        ("ProfitLo", "Our Profit at the Band Floor",
         "=NetLo-Payoff-BorrowerCash", MONEY),
        ("ProfitHi", "Our Profit at the Band High",
         "=NetHi-Payoff-BorrowerCash", MONEY),
        ("TargetLoan", "Loan at the 75% Ceiling", f"=ARV*{TARGET_MAX_LTARV}", MONEY),
        ("ExtraCash", "Extra Cash to Reach 75%", "=MAX(Loan-TargetLoan,0)", MONEY),
    ]
    for i, (nm, label, formula, fmt) in enumerate(helpers):
        rr = r + i // 2
        col = 1 if i % 2 == 0 else 4
        _row(ws, rr, col, label, val=formula, val_fmt=fmt, fill=CALC_FILL)
        name(nm, col + 2, rr)
    r += (len(helpers) + 1) // 2 + 1

    for nm, ref in names.items():
        wb.defined_names[nm] = DefinedName(nm, attr_text=ref)
    _widths(ws, [32, 13, 15, 32, 13, 15])
    ws.freeze_panes = "A5"


# ══ 2. Repair Costs: the Estimator detail grid ═════════════════════════

DEFAULT_GRID = [
    ("EXTERIOR", [
        ("Roof", ["Roof, rip and replace, architectural shingle", "Roof repair or patch",
                  "Fascia, demo and install new", "Soffit, demo and install new"]),
        ("Gutters", ["Gutters and downspouts, demo and install new"]),
        ("Siding and Finish", ["Vinyl siding", "Fiber cement siding",
                               "Patch an exterior section", "Power wash exterior"]),
        ("Masonry", ["Tuckpoint brick", "Foundation parging and repair"]),
        ("Exterior Painting", ["Painting exterior only", "Paint trim only"]),
        ("Windows and Doors", ["Windows, vinyl, average size",
                               "Exterior door, demo and install new"]),
        ("Garage and Carport", ["Garage or carport structure repair",
                                "Garage door and opener"]),
        ("Landscaping", ["Clean up landscaping and yard only", "Tree removal, per tree"]),
        ("Concrete and Asphalt", ["Concrete driveway, patio or sidewalk",
                                  "Gravel driveway"]),
        ("Decks and Railings", ["Deck repair or replacement", "New railings"]),
        ("Fence", ["Wood fencing", "Chain link fence"]),
        ("Septic and Well", ["Septic inspection and repair allowance"]),
    ]),
    ("INTERIOR", [
        ("Interior Painting", ["Paint whole interior", "Prime and seal"]),
        ("Flooring", ["Luxury vinyl plank", "Carpet", "Refinish hardwood",
                      "Ceramic tile, bathrooms"]),
        ("Kitchen", ["Cabinets", "Countertops", "Sink and faucet", "Backsplash tile",
                     "Appliance package"]),
        ("Bathroom", ["Vanity and mirror", "Toilet", "Tub or shower surround",
                      "Fixtures and accessories"]),
        ("Framing and Structure", ["Sill plate and framing repair", "Subfloor",
                                   "Support beam or post"]),
        ("Insulation", ["Floor insulation", "Attic insulation, blown in"]),
        ("Walls and Ceilings", ["Drywall hang, finish and texture", "Ceiling repair"]),
        ("Doors and Trim", ["Interior doors", "Baseboard and casing trim"]),
        ("Crawlspace and Basement", ["Vapor barrier and moisture package", "Sump pump"]),
    ]),
    ("MECHANICALS", [
        ("HVAC", ["Heat pump or furnace and condenser", "Ductwork"]),
        ("Plumbing", ["Repipe supply lines", "Water heater", "Sewer lateral scope and repair"]),
        ("Electrical", ["Service and panel", "Rewire or partial rewire", "Fixtures and devices"]),
    ]),
    ("OTHER", [
        ("Cleanout and Demo", ["Contents removal and dumpsters", "Interior demo"]),
        ("Permits and Soft Costs", ["Permits and inspections", "Contingency"]),
        ("Flagged Items", ["Item pending a signed bid"]),
    ]),
]


def sheet_repairs(wb, spec):
    """Estimator detail grid. Rolls up into Estimated Repair Costs on the front
    page. On a straight relist every line is zero and the total is zero, which
    is the right answer rather than a missing tab."""
    ws = wb.create_sheet("Repair Costs")
    p = spec["property"]
    _title(ws, "Repair Costs", p["full_address"],
           "Same grid as the Repair Estimator. Put a Y in the Include column, a "
           "quantity and a unit cost, and the line totals itself. The grand total "
           "feeds Estimated Repair Costs on Deal Overview.")
    r = 5
    _kv(ws, r, "Address", p["full_address"], span=2)
    _row(ws, r, 5, "Beds / Baths", val=f"{p.get('beds','')} / {p.get('baths','')}",
         val_fmt=None)
    r += 1
    _kv(ws, r, "Square Footage", p.get("sqft", 0), fmt=NUM)
    _row(ws, r, 5, "Inspected By", val=spec.get("inspected_by", ""), val_fmt=None)
    r += 2

    if spec.get("no_rehab_note"):
        r = _para(ws, r, spec["no_rehab_note"], bold=True, span=8, height=46)
        r += 1

    r = _hdr(ws, r, ["Category", "Include", "Repair Type", "Qty", "Unit",
                     "Unit Cost", "Total", "Notes"])
    grid = spec.get("repair_grid") or DEFAULT_GRID
    total_rows = []
    for section, cats in grid:
        for j in range(1, 9):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
        c = ws.cell(row=r, column=1, value=section)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        r += 1
        for cat, items in cats:
            first_of_cat = True
            for item in items:
                # An item is either a plain string (blank template line) or a
                # dict carrying a real scope and price.
                d = item if isinstance(item, dict) else {"item": item}
                if first_of_cat:
                    c = ws.cell(row=r, column=1, value=cat)
                    c.font = Font(size=10, bold=True)
                    first_of_cat = False
                red = bool(d.get("placeholder"))
                inc = ws.cell(row=r, column=2, value=d.get("include", "N"))
                inc.fill = INPUT_FILL; inc.border = BOX
                inc.alignment = Alignment(horizontal="center")
                lab = ws.cell(row=r, column=3, value=d["item"])
                lab.font = Font(size=9.5, color=RED if red else "000000", bold=red)
                lab.alignment = Alignment(wrap_text=True, vertical="top")
                q = ws.cell(row=r, column=4, value=d.get("qty", 0))
                q.fill = INPUT_FILL; q.border = BOX; q.number_format = NUM
                u = ws.cell(row=r, column=5, value=d.get("unit", "ls"))
                u.fill = INPUT_FILL; u.border = BOX
                u.alignment = Alignment(horizontal="center")
                uc = ws.cell(row=r, column=6, value=d.get("cost", 0))
                uc.fill = INPUT_FILL; uc.border = BOX; uc.number_format = MONEY2
                uc.font = Font(size=10, color=RED if red else "000000", bold=red)
                tot = ws.cell(row=r, column=7, value=f'=IF(B{r}="Y",D{r}*F{r},0)')
                tot.fill = CALC_FILL; tot.border = BOX; tot.number_format = MONEY
                tot.font = Font(size=10, color=RED if red else "000000", bold=red)
                n = ws.cell(row=r, column=8, value=d.get("note", ""))
                n.font = Font(size=9, color=RED if red else GREY)
                n.alignment = Alignment(wrap_text=True, vertical="top")
                total_rows.append(r)
                r += 1
    r += 1
    c = ws.cell(row=r, column=1, value="TOTAL REPAIR COSTS")
    c.font = Font(bold=True, size=11, color=NAVY)
    v = ws.cell(row=r, column=7,
                value=f"=SUM(G{total_rows[0]}:G{total_rows[-1]})")
    v.number_format = MONEY; v.fill = TOTAL_FILL; v.border = BOX
    v.font = Font(bold=True, size=11)
    wb.defined_names["RehabTotal"] = DefinedName(
        "RehabTotal", attr_text=f"'Repair Costs'!$G${r}")
    r += 1
    c = ws.cell(row=r, column=1, value="Lines included")
    c.font = Font(size=9, color=GREY)
    v = ws.cell(row=r, column=7,
                value=f'=COUNTIF(B{total_rows[0]}:B{total_rows[-1]},"Y")')
    v.number_format = NUM; v.fill = CALC_FILL; v.border = BOX
    _widths(ws, [26, 10, 46, 8, 8, 13, 14, 34])
    ws.freeze_panes = "A9"


# ══ 3. Your Investment ═════════════════════════════════════════════════

def sheet_investment(wb, spec):
    ws = wb.create_sheet("Your Investment")
    p = spec["property"]
    _title(ws, "Your Investment",
           "What you put in, how it is secured, when it comes back, and what you earn.")
    r = 5
    r = _band(ws, r, "What You Are Funding", col=1, span=4)
    r = _kv(ws, r, "Loan amount", "=Loan", fmt=MONEY, bold=True, span=1,
            note="Set on Deal Overview")
    r = _kv(ws, r, "Advance at closing", "=DayOne", fmt=MONEY,
            note="Wired to the closing attorney's escrow, never to us")
    r = _kv(ws, r, "Held back for repairs", "=RehabTranche", fmt=MONEY,
            note="Zero when there is no repair scope")
    r = _kv(ws, r, "Interest rate", "=Rate", fmt=PCT1, note="Simple, on the balance")
    r = _kv(ws, r, "Term", '=HoldMonths&" months"',
            note="Payoff comes at the sale, earlier if the sale is earlier")
    r = _kv(ws, r, "Interest you are paid", "=Interest", fmt=MONEY,
            note="A flat figure, quoted for the full term")
    r = _kv(ws, r, "Guaranteed minimum",
            '=MinIntMonths&" months ("&TEXT(MinIntPct,"0.0%")&" of the loan, '
            '"&TEXT(MinIntAmt,"$#,##0")&")"',
            note="Earned even if it sells in thirty days")
    r += 1

    r = _band(ws, r, "How It Is Secured", col=1, span=4)
    for lab, val in [
        ("Collateral", f"{p['full_address']}, parcel {p.get('parcel_id','')}"),
        ("Lien position", spec["loan"].get("lien", "First lien, deed of trust")),
        ("Recorded", f"{p.get('county','Travis')} County County Clerk (Official Public Records), closing day"),
        ("Title insurance", "Lender's policy in your name for the full loan amount, "
                            "paid out of the loan"),
        ("Hazard insurance", "In force before funds disburse, with you named as "
                             "mortgagee and loss payee, so a claim check comes to you"),
        ("Personal guarantee", "Every member of the company guarantees the note. If we "
                               "ever default we liquidate immediately and the guarantee "
                               "covers any shortfall including the interest still owed."),
        ("Pooling", "None. Your money is secured against this one property."),
    ]:
        r = _kv(ws, r, lab, val, wrap=True, span=3)
        ws.row_dimensions[r - 1].height = 28
    r += 1

    r = _band(ws, r, "What You Get Paid At The Closing Table", col=1, span=4)
    r = _kv(ws, r, "Sale price at the ask", "=ARV", fmt=MONEY)
    r = _kv(ws, r, "Less selling transaction costs", "=-SellCosts", fmt=MONEY)
    r = _kv(ws, r, "Net proceeds", "=NetPoint", fmt=MONEY)
    r = _kv(ws, r, "Paid first: your principal", "=-Loan", fmt=MONEY)
    r = _kv(ws, r, "Paid first: your interest and points", "=-(Interest+PointsAmt)",
            fmt=MONEY)
    r = _kv(ws, r, "TOTAL TO YOU", "=Payoff", fmt=MONEY, fill=TOTAL_FILL, bold=True)
    r = _kv(ws, r, "What is left for us", "=NetPoint-Payoff", fmt=MONEY)
    r = _para(ws, r, _say([
        "Our profit swings from ", _t("ProfitLo"), " at the bottom of the band to ",
        _t("ProfitHi"), " at the top. Your ", _t("Payoff"), " does not move with it. "
        "You get paid the same in every case, and you get paid before we see a dollar."]),
        color=GREY, size=9, span=4)
    r += 1

    r = _band(ws, r, "What You Earn", col=1, span=4)
    r = _kv(ws, r, "Interest income", "=Interest", fmt=MONEY)
    r = _kv(ws, r, "Points income", "=PointsAmt", fmt=MONEY)
    r = _kv(ws, r, "TOTAL INCOME TO YOU", "=LenderIncome", fmt=MONEY, fill=TOTAL_FILL,
            bold=True)
    r = _kv(ws, r, "Annualized yield", "=AnnYield", fmt=PCT1)
    r = _kv(ws, r, "Same money in a 4% CD", "=Loan*0.04*HoldMonths/12", fmt=MONEY,
            note="Same window, so it is a fair comparison")
    r = _para(ws, r, _say([
        "Over the same ", _t("HoldMonths", "0"), " months this loan pays you ",
        _t("LenderIncome"), " against ", _t("Loan*0.04*HoldMonths/12"),
        " in a CD, and unlike the CD it is secured by a recorded first lien with ",
        _t("Equity"), " of equity behind you and a personal guarantee from every member "
        "of the company."]), color=GREY, size=9, span=4)
    _widths(ws, [34, 20, 46, 20])


# ══ 4. Term Sheet ══════════════════════════════════════════════════════

def sheet_terms(wb, spec):
    ws = wb.create_sheet("Term Sheet")
    b, p, ln = spec["borrower"], spec["property"], spec["loan"]
    members = b.get("members") or [b.get("signer") or "[SIGNER NAME]"]
    _title(ws, "Term Sheet",
           "Signature ready. Every dollar figure is live off Deal Overview.")
    r = 5
    r = _band(ws, r, "Parties", col=1, span=3)
    for lab, val in [
        ("Date", spec.get("date", "")),
        ("Lender", spec.get("lender_name", "[LENDER LEGAL NAME]")),
        ("Borrower", f"{b['entity']}, "
                     f"{b.get('formation','a Texas limited liability company')}"),
        ("Guarantors", ", ".join(members) + ", each individually"),
        ("Property", f"{p['full_address']} (parcel {p.get('parcel_id','')})"),
    ]:
        r = _kv(ws, r, lab, val, wrap=True, span=2)
    r += 1
    r = _band(ws, r, "Terms", col=1, span=3)
    for lab, val, fmt, bold in [
        ("Loan amount", "=Loan", MONEY, True),
        ("Purpose", spec["plan"].get("loan_purpose", ""), None, False),
        ("Interest rate", '=TEXT(Rate,"0.0%")&" per annum, simple interest on the '
                          'outstanding balance"', None, False),
        ("Points and fees", '=TEXT(Points,"0.0")&" points ("&TEXT(PointsAmt,"$#,##0")&'
                            '"). No origination fee, no processing fee, no exit fee."',
         None, False),
        ("Term", '=HoldMonths&" months from the date of funding"', None, False),
        ("Interest paid", '="A flat "&TEXT(Interest,"$#,##0")&", quoted for the full '
                          '"&HoldMonths&" months and payable in one payment at payoff."',
         None, False),
        ("Payments", ln.get("payment", "No monthly payments. Principal and all accrued "
                            "interest are due in one payment at the earlier of the sale "
                            "or the maturity date."), None, False),
        ("Guaranteed minimum interest",
         '=MinIntMonths&" months, which is "&TEXT(MinIntPct,"0.0%")&" of the loan ('
         '"&TEXT(MinIntAmt,"$#,##0")&"), earned even if we pay off sooner"', None, False),
        ("Prepayment", "Allowed at any time with no penalty, subject to the guaranteed "
                       "minimum above", None, False),
        ("Security", f"{ln.get('lien','First lien, deed of trust')} on "
                     f"{p['short_address']}, recorded in {p.get('county','Travis')} "
                     f"County, Texas", None, False),
        ("Additional security", "Personal guarantee from every member of the company, "
                                "plus the hazard policy naming the lender as mortgagee "
                                "and loss payee", None, False),
        ("Funding", '=IF(Draws=0,"One advance of "&TEXT(Loan,"$#,##0")&" at closing.",'
                    'TEXT(DayOne,"$#,##0")&" at closing, then "&'
                    'TEXT(RehabTranche,"$#,##0")&" in "&Draws&" draws against completed '
                    'work.")', None, False),
        ("Costs", "The loan amount covers all closing costs, document preparation, "
                  "recording fees and the lender's title policy premium, repaid with "
                  "interest and backed by the personal guarantee.", None, False),
        ("Default", ln.get("default_terms", ""), None, False),
        ("Reporting", "Written update every two weeks, and you may inspect the property "
                      "at any time without notice", None, False),
        ("Governing law", "State of Texas", None, False),
        ("Open until", spec.get("term_sheet_expiry", "14 days from the date above"),
         None, False),
    ]:
        r = _kv(ws, r, lab, val, fmt=fmt, bold=bold, wrap=True, span=2,
                fill=TOTAL_FILL if bold else CALC_FILL)
        ws.row_dimensions[r - 1].height = 30
    r += 1
    r = _para(ws, r, "This sets out the commercial terms both sides intend to document. "
                     "It is not a loan commitment on its own and neither side is "
                     "obligated until the promissory note, the guarantee and the deed of "
                     "trust are signed.", color=GREY, size=9, span=3)
    r += 1
    r = _band(ws, r, "Agreed", col=1, span=3)
    for role, nm in ([("LENDER", spec.get("lender_name", "[LENDER LEGAL NAME]")),
                      ("BORROWER", f"{b['entity']}, by "
                                   f"{b.get('signer') or '[SIGNER NAME]'}, "
                                   f"{b.get('title','Managing Member')}")]
                     + [("GUARANTOR", f"{m}, individually") for m in members]):
        ws.cell(row=r, column=1, value=role).font = Font(bold=True, size=9, color=NAVY)
        ws.cell(row=r, column=2, value=nm).font = Font(size=10)
        r += 1
        ws.cell(row=r, column=1, value="Signature").font = Font(size=9, color=GREY)
        ws.cell(row=r, column=2,
                value="__________________________________    Date: ____________"
                ).font = Font(size=10)
        r += 2
    _widths(ws, [28, 40, 34])


# ══ 5. Resale Value ════════════════════════════════════════════════════

def sheet_resale(wb, spec):
    ws = wb.create_sheet("Resale Value")
    _title(ws, "Resale Value",
           "Where the number we are asking comes from, and what the stress tests run on.")
    r = 5
    r = _band(ws, r, "The Numbers", col=1, span=3)
    r = _kv(ws, r, "Resale band low, used by every stress test", "=ARVLo", fmt=MONEY)
    r = _kv(ws, r, "WORKING FIGURE, THE ASK", "=ARV", fmt=MONEY, fill=TOTAL_FILL,
            bold=True)
    r = _kv(ws, r, "Resale band high", "=ARVHi", fmt=MONEY)
    r = _kv(ws, r, "Working figure per square foot",
            f"=IFERROR(ARV/{max(spec['property'].get('sqft', 0), 1)},0)", fmt=MONEY2)
    r = _kv(ws, r, "Band width", "=IFERROR((ARVHi-ARVLo)/ARVLo,0)", fmt=PCT0)
    r = _kv(ws, r, "Selling costs at the ask", "=SellCosts", fmt=MONEY)
    r = _kv(ws, r, "Net proceeds at the ask", "=NetPoint", fmt=MONEY, bold=True,
            fill=TOTAL_FILL)
    r += 1
    r = _para(ws, r, spec["resale"].get("basis", ""), span=4, height=110)
    r += 1
    if (spec.get("as_is") or {}).get("basis"):
        r = _band(ws, r, "As-Is Value Today", col=1, span=3)
        r = _kv(ws, r, "As-is value", "=AsIs", fmt=MONEY, fill=TOTAL_FILL, bold=True)
        r = _kv(ws, r, "Day one coverage against it", "=DayOneCover", fmt=MULT)
        r = _para(ws, r, spec["as_is"]["basis"], span=4, height=76)
    _widths(ws, [40, 18, 26, 34])


# ══ 6. Comps ═══════════════════════════════════════════════════════════

def sheet_comps(wb, spec):
    ws = wb.create_sheet("Comps")
    _title(ws, "Comps", "The closed sales behind the resale figure.",
           "Closed sales only. No listings and no automated estimates. The ones we "
           "threw out are on here too so you can see what we threw out and why.")
    r = 5
    comps = spec.get("comps", [])
    if not comps:
        _para(ws, r, "No comparable sales supplied.", color=RED, bold=True)
        _widths(ws, [110])
        return
    r = _hdr(ws, r, ["Address", "Dist", "SF", "Bd", "Ba", "Sold", "Price", "$/SF",
                     "Why it is here"])
    first = r
    for c in comps:
        ws.cell(row=r, column=1, value=c["address"]).font = Font(size=10)
        ws.cell(row=r, column=2, value=c.get("dist", ""))
        cell = ws.cell(row=r, column=3, value=c.get("sqft", "")); cell.number_format = NUM
        ws.cell(row=r, column=4, value=c.get("beds", ""))
        ws.cell(row=r, column=5, value=c.get("baths", ""))
        ws.cell(row=r, column=6, value=c.get("sold", ""))
        cell = ws.cell(row=r, column=7, value=c.get("price", ""))
        cell.number_format = MONEY; cell.fill = INPUT_FILL; cell.border = BOX
        cell = ws.cell(row=r, column=8, value=f'=IFERROR(G{r}/C{r},"")')
        cell.number_format = MONEY; cell.fill = CALC_FILL; cell.border = BOX
        n = ws.cell(row=r, column=9, value=c.get("role", ""))
        n.alignment = Alignment(wrap_text=True, vertical="top"); n.font = Font(size=9)
        if str(c.get("role", "")).upper().startswith("EXCLUDED"):
            for col in range(1, 9):
                ws.cell(row=r, column=col).font = Font(size=10, color=GREY, italic=True)
            n.font = Font(size=9, color=GREY, italic=True)
        r += 1
    last = r - 1
    for lab, fn in [("Average", "AVERAGE"), ("Median", "MEDIAN")]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        for cc, letter, fmt in ((3, "C", NUM), (7, "G", MONEY), (8, "H", MONEY)):
            cell = ws.cell(row=r, column=cc,
                           value=f'=IFERROR({fn}({letter}{first}:{letter}{last}),"")')
            cell.number_format = fmt; cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL; cell.border = BOX
        r += 1
    r += 1
    r = _kv(ws, r, "Our ask", "=ARV", fmt=MONEY, fill=TOTAL_FILL, bold=True)
    for warn in spec.get("comp_warnings", []):
        r += 1
        r = _para(ws, r, warn, color=GOLD, span=9, height=44)
    _widths(ws, [26, 8, 9, 5, 5, 12, 13, 10, 46])


# ══ 7. Backups and Risk ════════════════════════════════════════════════

def sheet_risk(wb, spec):
    ws = wb.create_sheet("Backups and Risk")
    _title(ws, "Backups and Risk",
           "What happens if we are wrong, which is the only question that really "
           "matters on a private loan.")
    r = 5
    r = _band(ws, r, "Stress Tests", col=1, span=4)
    r = _hdr(ws, r, ["Scenario", "You net", "Coverage", "Are you whole"])
    for label, net, cov, verdict in [
        ("It sells at the ask", "=NetPoint", "=CoverPoint",
         '=IF(CoverPoint>=1,"Yes, paid in full","No")'),
        ("It sells at the band floor", "=NetLo", "=CoverFloor",
         '=IF(CoverFloor>=1,"Yes, paid in full","Short by "&'
         'TEXT(Payoff-NetLo,"$#,##0")&", covered by the personal guarantee")'),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(size=10)
        v = ws.cell(row=r, column=2, value=net)
        v.number_format = MONEY; v.fill = CALC_FILL; v.border = BOX
        v = ws.cell(row=r, column=3, value=cov)
        v.number_format = MULT; v.fill = CALC_FILL; v.border = BOX
        v.font = Font(bold=True)
        c = ws.cell(row=r, column=4, value=verdict)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    for s in spec.get("stress_cases", []):
        ws.cell(row=r, column=1, value=s["scenario"]).alignment = Alignment(
            wrap_text=True, vertical="top")
        c = ws.cell(row=r, column=2, value=s["outcome"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=4, value=s["repaid"])
        c.font = Font(bold=True, size=10,
                      color=GREEN if str(s["repaid"]).startswith("Yes") else GOLD)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32
        r += 1
    r += 1

    r = _band(ws, r, "Loan To Value, Said Out Loud", col=1, span=4)
    r = _kv(ws, r, "This loan against the ask", "=LTV", fmt=PCT1, bold=True)
    r = _kv(ws, r, "The usual private money ceiling", TARGET_MAX_LTARV, fmt=PCT0)
    r = _kv(ws, r, "Loan if we held to that ceiling", "=TargetLoan", fmt=MONEY)
    r = _kv(ws, r, "Extra cash we would bring to get there", "=ExtraCash", fmt=MONEY)
    r = _para(ws, r, _say([
        'IF(LTV>0.75,"This loan is at "&TEXT(LTV,"0.0%")&" of the ask, which is above '
        'the 75% a conservative private lender usually wants, and we are not going to '
        'bury that. Your day one advance is "&TEXT(DayOne,"$#,##0")&" against a house '
        'worth "&TEXT(AsIs,"$#,##0")&" today"&IF(Draws>0,", and the balance only funds '
        'against work that is already in the building","")&". On top of that, every '
        'member of the company personally guarantees the note, so a shortfall is covered '
        'by us and not by you. If you would still rather sit at 75%, we will bring '
        '"&TEXT(ExtraCash,"$#,##0")&" to closing and write it at '
        '"&TEXT(TargetLoan,"$#,##0")&". Your call.","This loan is at '
        '"&TEXT(LTV,"0.0%")&" of the ask, comfortably inside the 75% ceiling a '
        'conservative private lender wants, and there is nothing to explain away.")']),
        span=4, height=76)
    r += 1

    r = _band(ws, r, "Backup Exits, In The Order We Would Use Them", col=1, span=4)
    r = _hdr(ws, r, ["Backup", "Proceeds", "Pays you off", "Why it works"])
    for e in spec.get("exit_backups", []):
        ws.cell(row=r, column=1, value=e["name"]).alignment = Alignment(
            wrap_text=True, vertical="top")
        c = ws.cell(row=r, column=2, value=e.get("proceeds", ""))
        if isinstance(e.get("proceeds"), (int, float)):
            c.number_format = MONEY
        c.fill = CALC_FILL; c.border = BOX
        c = ws.cell(row=r, column=3, value=e.get("repays", ""))
        c.font = Font(bold=True, size=10,
                      color=GREEN if str(e.get("repays", "")).startswith("Yes") else GOLD)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c = ws.cell(row=r, column=4, value=e.get("why", ""))
        c.alignment = Alignment(wrap_text=True, vertical="top"); c.font = Font(size=9)
        ws.row_dimensions[r].height = 34
        r += 1
    r += 1
    r = _band(ws, r, "What Could Go Sideways", col=1, span=4)
    r = _hdr(ws, r, ["Risk", "How likely", "What it costs", "How we handle it"])
    for k in spec.get("risks", []):
        for col, key in ((1, "risk"), (2, "likelihood"), (3, "cost"), (4, "mitigation")):
            c = ws.cell(row=r, column=col, value=k.get(key, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=9 if col == 4 else 10)
        ws.row_dimensions[r].height = 44
        r += 1
    r += 1
    r = _band(ws, r, "The Worst Case, Plainly", col=1, span=4)
    r = _para(ws, r, "If we default, we do not make you chase us. We move immediately to "
                     "liquidate the property for as much as the market will pay, and the "
                     "personal guarantee from every member of the company covers whatever "
                     "the sale does not, including the interest still owed. You also hold "
                     "a recorded first lien and the original note, so you have every "
                     "ordinary remedy on top of that.", span=4, height=58)
    _widths(ws, [34, 22, 22, 58])


# ══ 8. Next Steps ══════════════════════════════════════════════════════

def sheet_next(wb, spec):
    ws = wb.create_sheet("Next Steps")
    _title(ws, "Next Steps",
           "Fill in the blue cells and send this tab back. We hand it straight to the "
           "closing attorney.")
    r = 5
    r = _band(ws, r, "Your Details", col=1, span=4)
    for f in ["Legal name exactly as it should read on the note and deed of trust",
              "Entity, trust or individual", "Mailing address", "City, state, ZIP",
              "Phone", "Email", "Date of birth or entity EIN, for the closing file",
              "Custodian and account number, if lending from a retirement account",
              "Your attorney, if you want one reviewing"]:
        c = ws.cell(row=r, column=1, value=f)
        c.alignment = Alignment(wrap_text=True, vertical="top"); c.font = Font(size=10)
        v = ws.cell(row=r, column=2, value="")
        v.fill = INPUT_FILL; v.border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1
    r = _band(ws, r, "What You Are Agreeing To", col=1, span=4)
    r = _kv(ws, r, "Amount", "=Loan", fmt=MONEY, bold=True)
    r = _kv(ws, r, "Property", spec["property"]["full_address"], span=2)
    r = _kv(ws, r, "Rate and term",
            '=TEXT(Rate,"0.0%")&" for "&HoldMonths&" months"')
    r = _kv(ws, r, "Interest you are paid", "=Interest", fmt=MONEY)
    r = _kv(ws, r, "Estimated payoff to you", "=Payoff", fmt=MONEY, fill=TOTAL_FILL,
            bold=True)
    r = _kv(ws, r, "Closing attorney",
            spec.get("closing_agent", "[CLOSING ATTORNEY / TITLE COMPANY]"), span=2)
    r = _kv(ws, r, "Target closing date",
            spec["plan"].get("closing_date", "[CLOSING DATE]"))
    r += 1
    r = _para(ws, r, "Before you wire anything, call the closing attorney's office at a "
                     "number you looked up yourself and confirm the instructions out "
                     "loud. Do not rely on wire instructions that only arrived by email, "
                     "including from us.", color=RED, bold=True, span=4, height=40)
    r += 1
    r = _band(ws, r, "Closing Checklist", col=1, span=4)
    r = _hdr(ws, r, ["Done", "Item", "Who", "When"])
    for item, owner, when in [
        ("Term sheet signed by both sides", "Both", "Before docs are ordered"),
        ("Closing attorney engaged and sent the file", "Us", "14 days out"),
        ("Draft note, guarantee and deed of trust sent to you", "Us", "14 days out"),
        ("Your own attorney reviews, if you want one", "You", "14 to 7 days out"),
        ("Title searched and cleared", "Attorney", "10 days out"),
        ("Lender's title policy ordered for the loan amount", "Attorney", "10 days out"),
        ("Insurance bound, you named mortgagee and loss payee", "Us", "7 days out"),
        ("Certificate delivered to the closing attorney", "Insurance agent", "7 days out"),
        ("Wire instructions confirmed by phone", "You", "3 days out"),
        ("Funds wired to the attorney's escrow account", "You", "1 to 2 days out"),
        ("Attorney confirms receipt in writing", "Attorney", "On receipt"),
        ("Note, guarantee and deed of trust signed", "Us", "Closing day"),
        ("Deed of trust recorded", "Attorney", "Closing day"),
        ("Signed copies and the original note delivered to you", "Us", "Closing day"),
        ("Investor Information Sheet completed and sent to you", "Us", "Within 48 hours"),
        ("You are paid from the settlement before we are", "Attorney", "Sale day"),
        ("Note marked paid in full, release recorded", "Both", "Within 10 days of payoff"),
    ]:
        c = ws.cell(row=r, column=1, value="[  ]")
        c.font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=item).alignment = Alignment(
            wrap_text=True, vertical="top")
        ws.cell(row=r, column=3, value=owner).font = Font(size=9)
        ws.cell(row=r, column=4, value=when).font = Font(size=9)
        r += 1
    _widths(ws, [56, 30, 18, 24])


# Repair Costs builds first so RehabTotal exists; sheets re-sort afterwards.
BUILDERS = [sheet_repairs, sheet_overview, sheet_investment, sheet_terms,
            sheet_resale, sheet_comps, sheet_risk, sheet_next]


def build_workbook(spec: dict, out_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    for fn in BUILDERS:
        try:
            fn(wb, spec)
        except Exception as exc:
            logger.warning("sheet %s failed: %s", fn.__name__, exc)
            ws = wb.create_sheet(fn.__name__.replace("sheet_", "")[:28])
            _title(ws, "Section unavailable", str(exc))
    for ws in wb:
        try:
            _polish(ws)
        except Exception as exc:
            logger.warning("autofit failed on %s: %s", ws.title, exc)
    wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title)
                    if s.title in SHEET_ORDER else 99)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def compute(spec: dict) -> dict:
    """Mirror of the sheet formulas, for logging, tests and lender_docs."""
    ln, res, ai = spec["loan"], spec["resale"], spec.get("as_is") or {}
    hold, buy, sell = (spec.get("holding", {}), spec.get("buying_costs", {}),
                       spec.get("selling_costs", {}))
    rate, points = float(ln["rate"]), float(ln.get("points", 0))
    term = int(ln["term_months"])
    mim = int(ln.get("min_interest_months", 3))
    draws = int(ln.get("draws", 0))
    rehab = float(spec.get("rehab_total", 0))
    purchase = spec["contract_price"]
    arv, arv_lo, arv_hi = res["point"], res.get("lo", res["point"]), res.get("hi", res["point"])

    hold_monthly = (hold.get("taxes_annual", 0) / 12 + hold.get("hoa_monthly", 0)
                    + hold.get("insurance_annual", 0) / 12 + hold.get("gas_monthly", 0)
                    + hold.get("water_monthly", 0) + hold.get("electric_monthly", 0)
                    + hold.get("other_monthly", 0) + hold.get("misc_monthly", 0))
    hold_total = hold_monthly * term
    buy_costs = (buy.get("escrow", 0) + purchase * buy.get("title_pct", 0)
                 + buy.get("recording", 0) + buy.get("misc", 0))
    sell_fixed = sum(sell.get(k, 0) for k in
                     ("escrow", "recording", "warranty", "staging", "marketing", "misc"))
    sell_var = sell.get("realtor_pct", 0) + sell.get("transfer_pct", 0)
    sell_costs = sell_fixed + arv * sell_var

    loan = ln["amount"]
    points_d = loan * points / 100
    interest = loan * rate * term / 12
    payoff = loan + interest + points_d
    fin_costs = points_d + interest
    total_cost = purchase + rehab + buy_costs + hold_total + fin_costs
    borrower_cash = max(purchase + rehab + buy_costs + hold_total - loan, 0)
    day_one = loan - rehab
    as_is = ai.get("point", 0)
    net_p = arv - sell_costs
    net_lo = arv_lo - (sell_fixed + arv_lo * sell_var)
    net_hi = arv_hi - (sell_fixed + arv_hi * sell_var)
    return {
        "rate": rate, "points": points, "term": term, "draws": draws,
        "min_int_months": mim, "min_int_amt": round(loan * rate * mim / 12),
        "min_int_pct": rate * mim / 12,
        "purchase": round(purchase), "rehab": round(rehab),
        "hold_monthly": round(hold_monthly), "hold_total": round(hold_total),
        "buy_costs": round(buy_costs), "sell_costs": round(sell_costs),
        "sell_pct": sell_costs / arv if arv else 0,
        "fin_costs": round(fin_costs), "total_cost": round(total_cost),
        "day_one": round(day_one), "as_is": round(as_is), "arv": round(arv),
        "arv_lo": round(arv_lo), "arv_hi": round(arv_hi),
        "rehab_tranche": round(max(loan - day_one, 0)),
        "loan": round(loan), "interest": round(interest), "payoff": round(payoff),
        "points_d": round(points_d), "borrower_cash": round(borrower_cash),
        "ltv": loan / arv if arv else 0, "equity": round(arv - loan),
        "day_one_cover": as_is / day_one if day_one else 0,
        "cover_point": net_p / payoff if payoff else 0,
        "cover_floor": net_lo / payoff if payoff else 0,
        "net_point": round(net_p), "net_lo": round(net_lo), "net_hi": round(net_hi),
        "net_profit": round(arv - sell_costs - purchase - rehab - buy_costs
                            - hold_total - fin_costs),
        "per_draw": round((loan - day_one) / draws) if draws else 0,
        "ann_yield": rate + (points / 100) * (12 / term) if term else rate,
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description="The Private Lender Package")
    ap.add_argument("--spec")
    ap.add_argument("--out")
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if not args.spec:
        ap.error("--spec is required")
    spec = json.loads(Path(args.spec).read_text(encoding="utf-8"))
    out = args.out or (f"output/lender/{spec['deal_name'].replace(' ', '_')}"
                       f"/2. The Private Lender Package.xlsx")
    build_workbook(spec, out)
    m = compute(spec)
    logger.info("Loan %s | interest %s | payoff %s | LTV %.1f%% | ask cover %.2fx | "
                "floor %.2fx | our profit %s",
                _money(m["loan"]), _money(m["interest"]), _money(m["payoff"]),
                m["ltv"] * 100, m["cover_point"], m["cover_floor"],
                _money(m["net_profit"]))
    logger.info("Wrote %s", out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
