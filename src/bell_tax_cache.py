"""Bell County tax data cache — owner name → property + delinquency lookup.

Mirrors `travis_tax_cache` for Travis County. BellCAD has no live API, so the
master cross-reference comes from the Appraisal Data Condensed XLSX (~190K
rows, full county). The Delinquent Roll Condensed XLSX (~24K rows) overlays
authoritative tax-delinquency fields onto matching properties.

Both XLSX files have rotating date stamps in their filenames; this module
discovers the latest URL by fetching the BellCAD data portal index on each
download.

Indices (lazy, process-wide):
  _INDEX[normalized_last_name] → list[property dict]
  _ADDR_INDEX[(normalized_street, zip5)] → property dict

Each property dict carries `source` = "delinquent_situs" (authoritative) or
"appraisal_master" (heuristic). Delinquent records win on address lookup.
Mirrors the pattern in `travis_tax_cache._load_delq` / `_load_cur`.
"""

from __future__ import annotations

import logging
import os
import re
import time
from pathlib import Path
from typing import Iterable

import requests

import config

logger = logging.getLogger(__name__)


_INDEX: dict[str, list[dict]] | None = None
_ADDR_INDEX: dict[tuple[str, str], dict] | None = None


# ── URL discovery ────────────────────────────────────────────────────
def _fetch_portal_html() -> str:
    r = requests.get(config.BELLCAD_DATA_PORTAL_URL, timeout=30)
    r.raise_for_status()
    return r.text


def discover_latest_url(pattern: str) -> str | None:
    """Find the latest matching XLSX URL on the BellCAD data portal page.

    The portal lists multiple historical files for some categories (e.g.,
    2024 + 2025 appraisal exports). When multiple match, prefer the one
    with the highest YYYYMMDD date stamp at the end of the filename.
    """
    try:
        html = _fetch_portal_html()
    except Exception as e:
        logger.warning("BellCAD portal fetch failed: %s", e)
        return None

    # Anchor URLs in the page have absolute hrefs
    candidates = re.findall(
        r'https://[^\s"\'<>]+?' + pattern,
        html,
    )
    if not candidates:
        # Pattern may bleed past the URL terminator; broaden and re-filter
        candidates = re.findall(r'https://[^\s"\'<>]+\.xlsx', html)
        candidates = [c for c in candidates if re.search(pattern, c)]
    if not candidates:
        logger.warning("BellCAD portal had no match for %s", pattern)
        return None

    def _date_key(url: str) -> str:
        m = re.search(r"(\d{8})\.xlsx", url)
        return m.group(1) if m else ""

    candidates.sort(key=_date_key, reverse=True)
    return candidates[0]


# ── Cache I/O ────────────────────────────────────────────────────────
def _cache_dir() -> Path:
    base = Path(config.BELL_TAX_CACHE_DIR)
    base.mkdir(parents=True, exist_ok=True)
    return base


def _is_fresh(path: Path, ttl_hours: int) -> bool:
    if not path.exists():
        return False
    age = time.time() - path.stat().st_mtime
    return age < ttl_hours * 3600


def _download(url: str, dest: Path) -> int:
    logger.info("Downloading %s → %s", url, dest)
    r = requests.get(url, timeout=600, stream=True)
    r.raise_for_status()
    tmp = dest.with_suffix(dest.suffix + ".tmp")
    total = 0
    with open(tmp, "wb") as f:
        for chunk in r.iter_content(chunk_size=1 << 20):
            f.write(chunk)
            total += len(chunk)
    os.replace(tmp, dest)
    logger.info("  Saved %.1f MB to %s", total / (1 << 20), dest)
    return total


def download_if_stale(force: bool = False) -> tuple[Path | None, Path | None]:
    """Download both Bell XLSX files if missing or stale.

    Returns (delinquent_path, appraisal_path). Either may be None if the
    portal didn't surface a matching link (network error or layout change).
    """
    ttl = config.BELL_TAX_CACHE_TTL_HOURS
    cache = _cache_dir()
    delq_path = cache / "BellCAD_Delinquent_Roll_Condensed_latest.xlsx"
    appr_path = cache / "BellCAD_Appraisal_Data_Condensed_latest.xlsx"

    delq_url = appr_url = None
    if force or not _is_fresh(delq_path, ttl):
        delq_url = discover_latest_url(config.BELLCAD_DELINQUENT_ROLL_PATTERN)
    if force or not _is_fresh(appr_path, ttl):
        appr_url = discover_latest_url(config.BELLCAD_APPRAISAL_ROLL_PATTERN)

    if delq_url:
        try:
            _download(delq_url, delq_path)
        except Exception as e:
            logger.warning("Bell delinquent download failed: %s", e)
            if not delq_path.exists():
                delq_path = None
    if appr_url:
        try:
            _download(appr_url, appr_path)
        except Exception as e:
            logger.warning("Bell appraisal download failed: %s", e)
            if not appr_path.exists():
                appr_path = None

    return (
        delq_path if delq_path and delq_path.exists() else None,
        appr_path if appr_path and appr_path.exists() else None,
    )


# ── Index helpers ────────────────────────────────────────────────────
def _normalize_street(street: str) -> str:
    if not street:
        return ""
    s = street.upper()
    s = re.sub(r"[.,]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_last(name: str) -> str:
    """Pull the last-name token used as the index key.

    Bell delinquent owner name is `LAST, FIRST [ETUX/ETAL ...]` — so the
    pre-comma chunk is the surname. Fall back to the first whitespace token
    when there's no comma (matches the Travis behavior when format drifts).
    """
    if not name:
        return ""
    s = name.strip().upper()
    if "," in s:
        return s.split(",", 1)[0].strip()
    for suf in (" JR", " SR", " II", " III", " IV", " MD", " DDS", " ESQ"):
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
    parts = s.split()
    return parts[0] if parts else ""


def _clean_mailing(addr: str) -> str:
    """Bell's `Ownr_Addr` packs street + city/state/zip into one field with
    `_x000d_\\n` literal newline markers. Collapse to a single line."""
    if not addr:
        return ""
    s = addr.replace("_x000d_", "").replace("\r", " ").replace("\n", ", ")
    s = re.sub(r"\s+", " ", s).strip().strip(",").strip()
    return s


def _split_situs_addr(addr: str) -> tuple[str, str, str]:
    """Bell appraisal `Situs_Addr` packs street + city/state/zip on one line:
        ' FM 2410, HARKER HEIGHTS TX  76548'
    Returns (street, city, zip5).
    """
    if not addr:
        return "", "", ""
    s = re.sub(r"\s+", " ", addr).strip().strip(",")
    # Strategy: split on the last comma first (street vs city/state/zip)
    if "," in s:
        street, tail = s.rsplit(",", 1)
    else:
        # Some rows don't use a comma — try to split before "TX" suffix
        m = re.match(r"^(.*?)\s+([A-Z][A-Z ]+?\s+TX\s+\d{5}.*)$", s)
        if m:
            street, tail = m.group(1), m.group(2)
        else:
            return s, "", ""
    street = street.strip()
    tail = tail.strip()
    m2 = re.match(r"^(.*?)\s+TX\s+(\d{5})", tail)
    if m2:
        return street, m2.group(1).strip().title(), m2.group(2)
    return street, tail.title(), ""


# ── Loaders ──────────────────────────────────────────────────────────
def _parse_years_owed(years_owed_raw: str | int | None) -> tuple[str, int]:
    """Bell's `years_owed` is either a single year ('2025') or a range
    ('2014-2025'). Returns (first_year_str, count_of_years).
    """
    if years_owed_raw in (None, ""):
        return "", 0
    s = str(years_owed_raw).strip()
    m = re.match(r"^(\d{4})(?:-(\d{4}))?$", s)
    if not m:
        return "", 0
    first = int(m.group(1))
    last = int(m.group(2)) if m.group(2) else first
    if last < first:
        first, last = last, first
    return str(first), (last - first + 1)


def _load_delinquent(
    path: Path,
    idx: dict[str, list[dict]],
    addr_idx: dict[tuple[str, str], dict],
) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        wb.close()
        return 0
    col = {h: i for i, h in enumerate(header) if h is not None}
    count = 0
    for row in rows:
        prop_type = (row[col.get("prop_type", -1)] if col.get("prop_type") is not None else "") or ""
        if str(prop_type).strip() != "Real Property":
            continue
        owner = (row[col["Ownr_Name"]] if "Ownr_Name" in col else "") or ""
        owner = str(owner).strip()
        if not owner:
            continue
        key = _normalize_last(owner)
        if not key:
            continue
        situs = (row[col["situs_address"]] if "situs_address" in col else "") or ""
        situs = str(situs).strip()
        if not situs:
            continue
        city = str(row[col.get("situs_city", -1)] or "").strip().title() if "situs_city" in col else ""
        zip5 = str(row[col.get("situs_zip", -1)] or "").strip()[:5] if "situs_zip" in col else ""
        years_owed_raw = row[col.get("years_owed", -1)] if "years_owed" in col else ""
        first_year, years_count = _parse_years_owed(years_owed_raw)
        delq_total = str(row[col.get("Base_Tax_Due", -1)] or "").strip() if "Base_Tax_Due" in col else ""
        prop_id = str(row[col["prop_id"]] if "prop_id" in col else "").strip()
        state_cd = str(row[col.get("state_cd", -1)] or "").strip() if "state_cd" in col else ""
        legal = str(row[col.get("legal_desc", -1)] or "").strip() if "legal_desc" in col else ""
        ownr_addr = _clean_mailing(str(row[col.get("Ownr_Addr", -1)] or "")) if "Ownr_Addr" in col else ""
        record = {
            "fullname": owner,
            "situsaddress": situs,
            "scity": city.upper() if city else "",
            "szip": zip5,
            "quickrefid": prop_id,
            "totalpropmktvalue": "",
            "propertytypedesc": state_cd,
            "delinquent_total": delq_total,
            "years_delinquent": str(years_count) if years_count else "",
            "first_year_delinquent": first_year,
            "mailing": ownr_addr,
            "legal_desc": legal,
            "source": "delinquent_situs",
        }
        idx.setdefault(key, []).append(record)
        addr_idx[(_normalize_street(situs), zip5)] = record
        count += 1
    wb.close()
    return count


def _load_appraisal(
    path: Path,
    idx: dict[str, list[dict]],
    addr_idx: dict[tuple[str, str], dict],
) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        wb.close()
        return 0
    col = {h: i for i, h in enumerate(header) if h is not None}
    count = 0
    for row in rows:
        prop_type_cd = str(row[col.get("prop_type_cd", -1)] or "").strip() if "prop_type_cd" in col else ""
        if prop_type_cd != "R":
            continue  # skip personal property, autos, mobile homes
        # BellCAD's export header for the owner column has drifted between
        # "Owner_name" (current) and "Ownr_Name" (older). Match either so a
        # header rename can't silently zero out the whole appraisal index.
        owner_col = next(
            (c for c in ("Owner_name", "Ownr_Name", "owner_name", "Owner_Name", "OwnerName")
             if c in col),
            None,
        )
        owner = str(row[col[owner_col]] if owner_col else "").strip()
        if not owner:
            continue
        key = _normalize_last(owner)
        if not key:
            continue
        situs_raw = str(row[col["Situs_Addr"]] if "Situs_Addr" in col else "").strip()
        if not situs_raw:
            continue
        street, city, zip5 = _split_situs_addr(situs_raw)
        if not street:
            continue
        prop_id = str(row[col["prop_id"]] if "prop_id" in col else "").strip()
        market = str(row[col.get("market", -1)] or "").strip() if "market" in col else ""
        state_cd = str(row[col.get("state_cd", -1)] or "").strip() if "state_cd" in col else ""
        legal = str(row[col.get("legal_desc", -1)] or "").strip() if "legal_desc" in col else ""
        ownr_addr = _clean_mailing(str(row[col.get("Ownr_Addr", -1)] or "")) if "Ownr_Addr" in col else ""
        record = {
            "fullname": owner,
            "situsaddress": street,
            "scity": city.upper() if city else "",
            "szip": zip5,
            "quickrefid": prop_id,
            "totalpropmktvalue": market,
            "propertytypedesc": state_cd,
            "delinquent_total": "",
            "years_delinquent": "",
            "mailing": ownr_addr,
            "legal_desc": legal,
            "source": "appraisal_master",
        }
        idx.setdefault(key, []).append(record)
        # Only fill address index when no delinquent record already owns this key
        addr_idx.setdefault((_normalize_street(street), zip5), record)
        count += 1
    wb.close()
    return count


def load_index(force_download: bool = False) -> dict[str, list[dict]]:
    global _INDEX, _ADDR_INDEX
    if _INDEX is not None and not force_download:
        return _INDEX

    delq_path, appr_path = download_if_stale(force=force_download)
    idx: dict[str, list[dict]] = {}
    addr_idx: dict[tuple[str, str], dict] = {}

    if delq_path:
        logger.info("Loading Bell delinquent XLSX into index…")
        n = _load_delinquent(delq_path, idx, addr_idx)
        logger.info("  %d delinquent-situs records indexed", n)
    else:
        logger.warning("Bell delinquent file unavailable — name lookups limited to appraisal master")

    if appr_path:
        logger.info("Loading Bell appraisal master XLSX into index (large file, ~30s)…")
        n = _load_appraisal(appr_path, idx, addr_idx)
        logger.info("  %d appraisal-master records indexed", n)
    else:
        logger.warning("Bell appraisal master unavailable — probate/foreclosure backfill disabled for Bell")

    _INDEX = idx
    _ADDR_INDEX = addr_idx
    logger.info(
        "Bell tax cache ready: %d name keys, %d address keys, %d total records",
        len(idx), len(addr_idx), sum(len(v) for v in idx.values()),
    )
    return idx


def search_by_name(last_name: str, first_name: str = "") -> list[dict]:
    """Return candidate property dicts matching `last_name` (exact key match).

    Matches the WCAD-shaped result dicts so `cad_lookup.lookup_property_by_name`
    can score them with the same logic.
    """
    idx = load_index()
    key = (last_name or "").strip().upper()
    if not key:
        return []
    return idx.get(key, [])


def search_by_address(street: str, zip5: str) -> dict | None:
    global _ADDR_INDEX
    if _ADDR_INDEX is None:
        load_index()
    if _ADDR_INDEX is None:
        return None
    zip5 = (zip5 or "").strip()[:5]
    if not street or not zip5:
        return None
    return _ADDR_INDEX.get((_normalize_street(street), zip5))


def iter_delinquent_path() -> Path | None:
    """Return the cached delinquent XLSX path (downloading if stale).

    Used by `tax_delinquent_bell.py` for the primary scrape — keeps the
    download in one place so the cache is shared with master-roll lookups.
    """
    delq_path, _ = download_if_stale(force=False)
    return delq_path
