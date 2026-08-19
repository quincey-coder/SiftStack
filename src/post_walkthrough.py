"""Post-walkthrough deal package: the one workbook you build the hour after you
walk a house.

This is the join point for the whole deal-analysis stack. It spins together:

  * the COMP engine      (zillow_market_api /search band pull -> boundary clip ->
                          condition bucketing -> dual-track ARV, from comp_package)
  * the REHAB engine     (rehab_estimator 4-tier room-by-room, expanded to a
                          4-scenario matrix with a real itemized breakdown)
  * the WALKTHROUGH      (what you actually saw: work already done = credits,
                          work still open = debits, team-walk flags = priced)
  * the EXIT engine      (six exits scored off the conservative ARV track, with
                          the reason each one is recommended or rejected)
  * the DISPO stack      (buyer_sweep ranked buyers + dispo_skiptrace contacts)
  * the LIVE SIFT LEAD   (CRM record, status, lists, tags, owner, motivation,
                          message board, SiftMap deed history)

Output matches "Post Walkthrough Template.xlsx" exactly, sheet for sheet:

  Overview | Exit Strats | Comps | Active-Pending | Repair Logic |
  Repair Numbers | Buyer Targets | Outreach Sheet

Every input is optional. With no walkthrough file it renders the pre-walk
numbers; with no API key it renders whatever the spec supplies. Nothing hard
fails, so you always get a workbook you can hand to the team.

Usage:
  # full auto: Sift lead + API comps + rehab matrix + walkthrough findings
  python src/post_walkthrough.py --address "158 Old State Rd" --city Knoxville --zip 37914 \
      --beds 2 --baths 1 --sqft 1946 --year-built 1938 \
      --bbox "35.996,36.016,-83.895,-83.840" \
      --streets "old state|nash rd|seahorn|holston dr|bona|grata|silva|chilhowee" \
      --walkthrough walk_158.json \
      --buyers output/buyer_sweep_37914_2026-07-22.json \
      --outreach output/dispo_skiptrace_158.json

  python src/post_walkthrough.py --walkthrough-template   # writes walkthrough_template.json
  python src/post_walkthrough.py --spec pack.json --out X.xlsx   # render a saved pack
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
import statistics
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from comp_analyzer import (ADJ_PER_BATHROOM, ADJ_PER_BEDROOM, ADJ_PER_SQFT,
                           ADJ_PER_YEAR_BUILT, fetch_subject_property)
from comp_package import GUT_ALLOWANCE_PER_SQFT, SOFT_COST_PCT, classify, dual_track_arv
from rehab_estimator import estimate_rehab
from zillow_market_api import MarketListing, ZillowMarketAPI, filter_bbox, filter_streets

logger = logging.getLogger(__name__)

NAVY, BLUE, GREEN, GOLD, RED, GREY = "0A1130", "316AFF", "1B9E5A", "B8860B", "C0392B", "666666"

_API_CLIENTS = Path(r"C:\Users\Tyrus\OneDrive\Desktop\Deal Room Coaching Call\_api\clients")

# Deal-math constants. Kept here (not buried in the sheet builders) so the
# Sources block can print them and the team can argue with them in one place.
SELF_PERFORM_LABOR_FACTOR = 0.55   # self-performer pays ~55% of GC labor
SELLING_COST_PCT = 0.08            # commissions + seller closing on a retail exit
HOLDING_COST_PCT = 0.04            # carry, insurance, utilities over the hold
BUY_CLOSE_FLAT = 900               # escrow/attorney on the buy (TN)
BUY_TITLE_PCT = 0.0077             # buy-side title insurance + fees (rehab-estimator contract)
MIN_WHOLESALE_FEE = 10_000         # below this an assignment is not worth the file
MIN_FLIP_ROI = 0.15                # below this a flip is not worth the capital
REFI_LTV = 0.75                    # BRRRR refinance loan-to-value
SHELL_WORK_RATIO = 0.25            # rehab/ARV above this = cash-buyer-only shell


# ══ small helpers ══════════════════════════════════════════════════════

def _money(v) -> str:
    """Sign goes outside the dollar sign: -$5,000, never $-5,000."""
    if not isinstance(v, (int, float)):
        return str(v or "")
    return f"-${abs(v):,.0f}" if v < 0 else f"${v:,.0f}"


def _ba(v: float) -> str:
    """1.0 -> '1', 1.5 -> '1.5' (bath counts read badly with a trailing zero)."""
    return str(int(v)) if float(v or 0) == int(v or 0) else str(v)


def _num(v) -> float:
    if isinstance(v, str):
        v = re.sub(r"[^\d.\-]", "", v) or 0
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ── Money that knows how sure it is ───────────────────────────────────
# Team rule (Marwan, 158 review): any price we are not confident in ships as a
# RANGE, never a single number. A point number is a promise; a range is the
# truth when the inputs are a model output rather than a signed bid.

def _round_money(v: float) -> int:
    v = float(v or 0)
    step = 1000 if abs(v) >= 10_000 else 100
    return int(round(v / step) * step)


def rng(lo, hi=None, confident: bool = False, point=None) -> dict:
    """A price. Confident ones print as one number, the rest print as a range."""
    if hi is None:
        hi = lo
    if confident and lo == hi:
        # A signed figure is a fact: $81,500 on the contract prints as $81,500,
        # never rounded to $82,000. Rounding is for model outputs only.
        v = int(round(float(lo or 0)))
        return {"lo": v, "hi": v, "point": v, "confident": True}
    lo, hi = _round_money(min(lo, hi)), _round_money(max(lo, hi))
    mid = point if point is not None else (lo + hi) / 2
    return {"lo": lo, "hi": hi, "point": _round_money(mid),
            "confident": bool(confident) or lo == hi}


# Ty, 112 Milligan review: "we need an exact amount for each of these."
# A wide band is not an answer you can take to a seller or a buyer, so every
# figure prints as ONE number. The lo/hi still drives the downside test
# internally and surfaces as a single sensitivity line, not in the cells.
EXACT_NUMBERS = True


def fmt_rng(r):
    """One exact number for the cell."""
    if not r:
        return ""
    if EXACT_NUMBERS or r["confident"] or r["lo"] == r["hi"]:
        return r["point"]
    sep = " to " if min(r["lo"], r["hi"]) < 0 else " - "
    return f"{_money(r['lo'])}{sep}{_money(r['hi'])}"


def swing(r) -> str:
    """The downside/upside behind an exact number, for a sensitivity line."""
    if not r or r["lo"] == r["hi"]:
        return ""
    return f"{_money(r['lo'])} downside / {_money(r['hi'])} upside"


def _mid(r) -> float:
    return (r["lo"] + r["hi"]) / 2 if r else 0.0


def _miles(lat1, lon1, lat2, lon2) -> float:
    """Great-circle distance, used for the comp Dist column."""
    if not all([lat1, lon1, lat2, lon2]):
        return 0.0
    r = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return round(2 * r * math.asin(math.sqrt(a)), 2)


# ══ styling ════════════════════════════════════════════════════════════

def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=NAVY)
    if sub:
        c = ws.cell(row=2, column=1, value=sub)
        c.font = Font(size=10, color=GREY)
        c.alignment = Alignment(wrap_text=False, vertical="top")


def _header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _subhead(ws, row, text, col=1, color=BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=11, color=color)
    return row + 1


def _widths(ws, widths, start=1):
    for i, w in enumerate(widths, start):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kv(ws, start, rows, wrap_col=2):
    """rows: [label, value]; a label with no value renders as a blue subhead."""
    r = start
    for pair in rows:
        label, value = (list(pair) + ["", ""])[:2]
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=wrap_col, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if label and value == "":
            ws.cell(row=r, column=1).font = Font(bold=True, size=11, color=BLUE)
        r += 1
    return r


# ══ 1. Live Sift lead context ══════════════════════════════════════════

DEFAULT_SIFT_ACCOUNT = "datasift-apikey"   # Open API key: no expiry, no refresh dance


def load_lead(address: str, city: str = "", zip_code: str = "",
              account: str = DEFAULT_SIFT_ACCOUNT) -> dict:
    """Pull the live CRM record so the package is about a LEAD, not an address.

    Auth goes through the shared Deal Room auth store. It defaults to the
    Api-Key account (no expiry) via REISIFT_ACCOUNT, which reisift_auth honors
    ahead of the configured active_account; pass account= to use a JWT account
    when the lead lives outside the key's account.

    Returns a flat dict of everything the Overview sheet needs. Never raises:
    no creds, no record, or no _api checkout all degrade to {"available": False}
    with the reason, and the workbook still builds.
    """
    import os

    out = {"available": False, "reason": "", "address": address}
    if not _API_CLIENTS.exists():
        out["reason"] = f"Deal Room _api clients not found at {_API_CLIENTS}"
        return out
    sys.path.insert(0, str(_API_CLIENTS))
    try:
        from dossier import build_dossier  # noqa: E402
    except Exception as exc:  # noqa: BLE001 - any import problem is non-fatal
        out["reason"] = f"could not import the CRM clients: {exc}"
        return out

    prior = os.environ.get("REISIFT_ACCOUNT")
    if account:
        os.environ["REISIFT_ACCOUNT"] = account
    try:
        d = build_dossier(address, flow="A")
    except Exception as exc:  # noqa: BLE001 - auth/network are expected failures
        out["reason"] = f"CRM lookup failed on account '{account}': {exc}"
        return out
    finally:
        if prior is None:
            os.environ.pop("REISIFT_ACCOUNT", None)
        else:
            os.environ["REISIFT_ACCOUNT"] = prior

    det = d.get("detail") or {}
    if not d.get("uuid"):
        out["reason"] = f"no Sift record matched '{address}' (flow {d.get('flow')})"
        return out

    addr = det.get("address") or {}
    owner = det.get("owner") or {}
    owner_addr = owner.get("address") or {}
    assignee = det.get("assignee") or det.get("assigned_to") or {}
    if isinstance(assignee, str):                # the record stores a bare uuid
        assignee = {}
    summary = d.get("activity_summary") or {}
    smap = d.get("siftmap_detail") or {}
    owner_info = (smap.get("owner_info") or {}) if isinstance(smap, dict) else {}

    cf = {}
    for item in d.get("custom_fields") or []:
        label = ((item.get("custom_field") or {}).get("label") or "").strip()
        if label and item.get("value") not in (None, ""):
            cf[label] = item["value"]

    messages = []
    for m in (d.get("messages") or [])[:12]:
        a = m.get("author") or {}
        who = f"{a.get('first_name', '')} {a.get('last_name', '')}".strip() or m.get("source", "")
        messages.append({"date": str(m.get("created", ""))[:10], "who": who,
                         "text": (m.get("message") or "").strip()})

    out.update({
        "available": True,
        "uuid": d["uuid"],
        "dataflik_id": d.get("dataflik_id"),
        "url": f"https://app.reisift.io/records/{d['uuid']}",
        "street": addr.get("street") or address,
        "city": addr.get("city") or city,
        "state": addr.get("state") or "",
        "zip": (addr.get("postal_code") or zip_code or "").split("-")[0],
        "county": addr.get("county") or cf.get("County") or "",
        "lat": _num(addr.get("latitude")),
        "lon": _num(addr.get("longitude")),
        "vacant": addr.get("vacant"),
        "owner_name": (f"{owner.get('first_name', '') or ''} {owner.get('last_name', '') or ''}".strip()
                       or owner.get("company") or ""),
        "owner_deceased": bool(owner.get("deceased")),
        "owner_age": owner.get("age"),
        "mailing": ", ".join(str(x) for x in [owner_addr.get("street"), owner_addr.get("city"),
                                              f"{owner_addr.get('state', '')} {owner_addr.get('postal_code', '')}".strip()]
                             if x) or owner_info.get("mailing_address") or "",
        "status": det.get("status") or (d.get("chosen") or {}).get("status") or "",
        "lists": det.get("lists") or [],
        "tags": det.get("tags") or [],
        "assignee": f"{assignee.get('first_name', '') or ''} {assignee.get('last_name', '') or ''}".strip(),
        "last_contact": summary.get("last_event_date") or "",
        "last_contact_type": summary.get("last_event_type") or "",
        "cards": [f"{(c.get('board') or {}).get('title')} / {(c.get('column') or {}).get('title')}"
                  for c in d.get("cards") or []],
        "custom_fields": cf,
        "messages": messages,
        # Subject facts the record already knows. These beat a Zillow lookup
        # (the CRM carries the county-sourced numbers) but still lose to an
        # explicit CLI override off the county card.
        "estimated_value": _num(det.get("estimate_value") or owner_info.get("estimated_value")),
        "equity_percent": det.get("equity_percent") or owner_info.get("equity_percent"),
        "last_sale_date": det.get("last_sold") or "",
        "last_sale_price": _num(det.get("last_sale_price")),
        "rental_value": _num(det.get("rental_value")),
        "beds": int(_num(det.get("bedrooms"))),
        "baths": _num(det.get("bathrooms")),
        "sqft": int(_num(det.get("sqft") or det.get("building_sqft"))),
        "year_built": int(_num(det.get("year"))),
        "lot_acres": _num(det.get("lot_size")),
        "parcel_id": det.get("parcel_id") or det.get("apn") or "",
        "structure_type": det.get("structure_type") or "",
        "investor_score": det.get("investor_score"),
        "tax_delinquent_value": det.get("tax_delinquent_value"),
        "tax_delinquent_year": det.get("tax_delinquent_year"),
        "tax_auction_date": det.get("tax_auction_date"),
        "foreclosure_date": det.get("foreclosure_date"),
        "probate_open_date": det.get("probate_open_date"),
        "personal_representative": det.get("personal_representative") or "",
        "portfolio_size": owner_info.get("portfolio_size") or owner_info.get("property_count"),
        "siftmap": smap,
    })
    return out


def lead_motivation(lead: dict, walk: dict) -> str:
    """Best available answer to 'why would this person sell'.

    Walkthrough note wins (you just talked to them), then the CRM custom field,
    then the distress signals the record carries.
    """
    if walk.get("motivation"):
        return walk["motivation"]
    cf = lead.get("custom_fields") or {}
    for key in ("Motivation", "Seller Motivation", "Notice Type", "Lead Source"):
        if cf.get(key):
            return str(cf[key])
    signals = distress_signals(lead)
    if lead.get("lists"):
        signals += [str(x) for x in lead["lists"][:4]]
    return ", ".join(signals[:6]) or "not captured on the record yet"


def acquired_line(lead: dict) -> str:
    """How and when the current owner got it, with the equity position."""
    if not lead.get("last_sale_price"):
        # Old family transfers often carry no recorded price. "Held since 1980"
        # is still a motivation signal, so print the date rather than nothing.
        date = lead.get("last_sale_date")
        return f"Held since {date}, no recorded sale price" if date else ""
    line = f"{lead.get('last_sale_date', '')} for {_money(lead['last_sale_price'])}"
    equity = _num(lead.get("equity_percent"))
    if equity:
        line += f", equity {equity:.0f}%"
    return line.strip()


def distress_signals(lead: dict) -> list[str]:
    """The hard distress facts the record carries, not the soft tags."""
    out = []
    if lead.get("owner_deceased"):
        out.append("owner deceased on record")
    if lead.get("vacant"):
        out.append("flagged vacant")
    if lead.get("tax_delinquent_value") or lead.get("tax_delinquent_year"):
        out.append(f"tax delinquent {lead.get('tax_delinquent_year') or ''} "
                   f"{_money(_num(lead.get('tax_delinquent_value'))) if lead.get('tax_delinquent_value') else ''}".strip())
    if lead.get("tax_auction_date"):
        out.append(f"tax auction {lead['tax_auction_date']}")
    if lead.get("foreclosure_date"):
        out.append(f"foreclosure {lead['foreclosure_date']}")
    if lead.get("probate_open_date"):
        out.append(f"probate opened {lead['probate_open_date']}")
    if _num(lead.get("equity_percent")) >= 90:
        out.append(f"{_num(lead['equity_percent']):.0f}% equity, free and clear")
    return out


# ══ 2. Comps ═══════════════════════════════════════════════════════════

def adjusted_price(comp: MarketListing, subject: dict) -> float:
    """Line-item adjust a comp to the subject (SiftStack comp_analyzer rates)."""
    if not comp.price:
        return 0.0
    adj = comp.price
    adj += (subject.get("beds", 0) - comp.beds) * ADJ_PER_BEDROOM
    adj += (subject.get("baths", 0) - comp.baths) * ADJ_PER_BATHROOM
    adj += (subject.get("sqft", 0) - comp.sqft) * ADJ_PER_SQFT
    if subject.get("year_built") and comp.raw.get("yearBuilt"):
        adj += (subject["year_built"] - _num(comp.raw["yearBuilt"])) * ADJ_PER_YEAR_BUILT
    return round(adj)


# A Zestimate that has already absorbed the sale carries no condition signal:
# the ratio snaps to ~1.0 whether the house was a gut job or a retail sale.
ZEST_ABSORBED = (0.97, 1.03)

# Below this share of the Zestimate a "sale" is almost never arm's length:
# it is a family deed, an estate transfer or a quitclaim. Kept out of the
# as-is band, which is otherwise trivially dragged to zero by one bad row.
NON_ARMS_LENGTH_RATIO = 0.30


def _deed_investor(deed: dict | None) -> bool:
    """Deed shows an entity or a cash buyer: an investor trade, not retail."""
    buyer = (deed or {}).get("buyer") or ""
    return bool(re.search(r"\bLLC\b|\bINC\b|\bTRUST\b|PROPERT|HOLDINGS|HOMES|CAPITAL|INVEST|COUNTY",
                          buyer, re.I)) or bool((deed or {}).get("cash"))


def refine_bucket(comp: MarketListing, deed: dict, median_retail_ppsf: float,
                  retail_floor: float = 0.0) -> str:
    """Correct the sold-vs-Zestimate bucket using $/sf and the deed.

    `comp_package.classify` reads sold price against Zestimate, which fails on
    recent sales because Zillow re-anchors the Zestimate to the sale: a $105K
    investor buy on a shell shows a 1.00 ratio and reads RENOVATED. Left alone
    it drags the same-bed retail median down and understates the ARV.

    Two tells override it: the ratio sitting inside the absorbed band, and a
    deed-verified entity or cash buyer. Investors do not pay retail, so when
    either holds and the $/sf is well under the retail median, rebucket it.
    """
    base = classify(comp)
    if base != "RENOVATED/RETAIL" or not (median_retail_ppsf and comp.ppsf):
        return base
    ratio = comp.price / comp.zestimate if comp.zestimate else 0
    absorbed = ZEST_ABSORBED[0] <= ratio <= ZEST_ABSORBED[1]
    investor = _deed_investor(deed)
    if not (absorbed or investor):
        return base
    # Size guard: $/sf falls as houses get bigger, so a low $/sf alone would
    # demote every large renovated comp. A sale priced inside the retail band
    # is a retail sale no matter how many square feet it spreads over, so the
    # total price has to be under the band's floor before the $/sf can demote it.
    if retail_floor and comp.price >= retail_floor:
        return base
    relative = comp.ppsf / median_retail_ppsf
    if relative < 0.65:
        return "DISTRESSED"
    if relative < 0.85:
        return "AVERAGE"
    return base


def bucket_comps(sold: list, deeds: dict) -> dict:
    """zpid -> refined bucket, plus the retail median the refinement used.

    The median and floor that decide a demotion come from CLEAN retail sales
    only: deed-verified investor buys are definitionally not retail, and in a
    small pocket they can be a quarter of the classify-retail set. Left in,
    the lower quartile lands on their own price and the floor guard protects
    the exact comps it exists to demote (158 Old State: two $105K absorbed
    investor buys set the floor at $105K and survived as RENOVATED/RETAIL).
    """
    retail = [l for l in sold if isinstance(l, MarketListing)
              and classify(l) == "RENOVATED/RETAIL" and l.ppsf
              and not _deed_investor(deeds.get(_norm_addr(l.address)))]
    median_ppsf = statistics.median(l.ppsf for l in retail) if retail else 0.0
    prices = sorted(l.price for l in retail if l.price)
    floor = prices[len(prices) // 4] if prices else 0.0   # retail band lower quartile
    out = {}
    for l in sold:
        if isinstance(l, MarketListing):
            out[l.zpid or l.address] = refine_bucket(
                l, deeds.get(_norm_addr(l.address)), median_ppsf, floor)
    return {"buckets": out, "median_ppsf": median_ppsf, "retail_floor": floor}


# ── Comp similarity ───────────────────────────────────────────────────
# A transparent, weighted model. Every dimension scores 0-100 on its own
# stated decay curve, then the weighted average is the comp's similarity.
# Nothing is a black box: the per-dimension scores print next to the comp so
# a low score can always be traced to the reason for it.
#
# Weights reflect what has actually moved money on this deal set: distance and
# size first (a $/sf figure does not survive a 2x size gap or a hop across a
# highway), then the bed and bath bands (the bath band alone measured $42K-$75K
# on 112 Milligan), then condition, recency and lot.
#
# yearBuilt is deliberately absent: the /search feed carries it 0% of the time,
# so scoring age would be scoring a guess.
SIM_WEIGHTS = {
    "dist": 22,      # proximity to the subject
    "size": 20,      # living area ratio
    "bed": 15,       # bedroom band
    "bath": 15,      # bath band
    "cond": 13,      # condition bucket vs a renovated exit
    "recency": 10,   # months since sale
    "lot": 5,        # lot acreage ratio
}
SIM_RULES = {
    "dist": "100 within 0.25 mi, straight-line decay to 0 at 2.0 mi",
    "size": "100 at equal sqft, 50 at a 25% gap, 0 at a 50% gap",
    "bed": "100 exact, 45 within one bedroom, 0 beyond",
    "bath": "100 same band and exact, 85 same band, 30 across the 1.5 bath line",
    "cond": "renovated 100, average 55, unknown 40, distressed 12",
    "recency": "100 inside 6 months, decay to 0 at 36 months",
    "lot": "100 within 40% of subject acreage, decay to 0 at 150%",
}


def _clamp01(v: float) -> float:
    return max(0.0, min(1.0, v))


def comp_similarity(comp: MarketListing, subject: dict, bucket: str, dist: float,
                    target: dict | None = None) -> dict:
    """Score one comp against the FINISHED subject. Missing data is dropped,
    not guessed: the weight is redistributed and coverage is reported."""
    tgt = target or subject
    parts: dict[str, float] = {}

    if comp.latitude and comp.longitude and subject.get("lat"):
        parts["dist"] = 100 * _clamp01(1 - max(0.0, dist - 0.25) / 1.75)
    if comp.sqft and subject.get("sqft"):
        parts["size"] = 100 * _clamp01(1 - abs(comp.sqft / subject["sqft"] - 1) / 0.50)
    if comp.beds and tgt.get("beds"):
        gap = abs(comp.beds - tgt["beds"])
        parts["bed"] = 100.0 if gap == 0 else (45.0 if gap == 1 else 0.0)
    if comp.baths and tgt.get("baths"):
        same_band = (comp.baths <= 1.5) == (float(tgt["baths"]) <= 1.5)
        parts["bath"] = (100.0 if comp.baths == float(tgt["baths"]) else 85.0) if same_band else 30.0
    parts["cond"] = {"RENOVATED/RETAIL": 100.0, "AVERAGE": 55.0,
                     "UNKNOWN": 40.0, "DISTRESSED": 12.0}.get(bucket, 40.0)
    if comp.sold_date:
        months = (datetime.now() - datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")).days / 30.4
        parts["recency"] = 100.0 if months <= 6 else 100 * _clamp01(1 - (months - 6) / 30)
    if comp.lot_acres and subject.get("lot_acres"):
        r = abs(comp.lot_acres / subject["lot_acres"] - 1)
        parts["lot"] = 100 * _clamp01(1 - max(0.0, r - 0.40) / 1.10)

    used = sum(SIM_WEIGHTS[k] for k in parts)
    total = sum(parts[k] * SIM_WEIGHTS[k] for k in parts) / used if used else 0.0
    grade = "A" if total >= 80 else "B" if total >= 65 else "C" if total >= 50 else "D"
    detail = " ".join(f"{k} {parts[k]:.0f}" for k in SIM_WEIGHTS if k in parts)
    missing = [k for k in SIM_WEIGHTS if k not in parts]
    if missing:
        detail += f"  (no {', '.join(missing)})"
    return {"score": round(total), "grade": grade, "parts": parts,
            "detail": detail, "coverage": round(100 * used / sum(SIM_WEIGHTS.values()))}


def comp_role(comp: MarketListing, subject: dict, bucket: str, dist: float) -> str:
    """The one line that says why this comp is in the package.

    A comp with no stated role is a comp the buyer will argue with, so every
    row gets one: anchor, ceiling, bed-band cap, or as-is floor.
    """
    bits = []
    same_bed = comp.beds == subject.get("beds")
    sqft_delta = abs(comp.sqft - subject.get("sqft", 0)) / max(subject.get("sqft", 1), 1)

    if (comp.zestimate and comp.price
            and comp.price / comp.zestimate < NON_ARMS_LENGTH_RATIO):
        return ("NOT ARM'S LENGTH: sold at "
                f"{comp.price / comp.zestimate:.0%} of Zestimate. Reads as a family deed, estate "
                "transfer or quitclaim, so it is excluded from the as-is band. Verify at the "
                "Register of Deeds before quoting it to anyone.")
    if bucket == "RENOVATED/RETAIL":
        if same_bed and dist and dist <= 0.5 and sqft_delta <= 0.25:
            bits.append("Best anchor: same bed count, same pocket, closest size")
        elif same_bed:
            bits.append(f"Same-bed retail comp: sets the {comp.beds}-bed ceiling")
        elif comp.beds > subject.get("beds", 0):
            bits.append(f"Higher-bed band ({comp.beds}bd): upside track only, not base ARV")
        else:
            bits.append("Lower-bed retail sale: floor reference")
    elif bucket == "DISTRESSED":
        bits.append("As-is data point: what an investor actually paid here")
    elif bucket == "AVERAGE":
        bits.append("Dated or partial-update sale: the wholetail lane")
    if comp.home_type == "LOT":
        bits.append("LOT type: verify against the county card, may be a house at land value")
    if comp.sqft and subject.get("sqft") and comp.sqft > subject["sqft"] * 1.25:
        bits.append("size ceiling comp")
    if dist:
        bits.append(f"{dist} mi out")
    return ". ".join(bits)


def load_comps(subject: dict, args, spec: dict) -> tuple[list, list, dict]:
    """Boundary-filtered sold + active comps and the dual-track ARV.

    Falls back to comps carried in the spec so a saved pack re-renders without
    burning another API pull.
    """
    sold_spec, active_spec = spec.get("sold_comps"), spec.get("active_comps")
    if sold_spec is not None:
        revive = [MarketListing(**c) for c in sold_spec] if sold_spec and isinstance(sold_spec[0], dict) else sold_spec
        revive_a = [MarketListing(**c) for c in active_spec] if active_spec and isinstance(active_spec[0], dict) else (active_spec or [])
        return revive, revive_a, spec.get("arv", {})
    if args.sold_json and Path(args.sold_json).exists():
        # Re-use a saved band pull (the 50-80 call partition) instead of paying
        # for it again. Same shape buyer_sweep --sold-json accepts.
        from zillow_market_api import normalize
        items = json.load(open(args.sold_json, encoding="utf-8"))
        if isinstance(items, dict):
            items = items.get("sold") or items.get("records") or []
        sold = [normalize(i) for i in items if isinstance(i, dict) and "streetAddress" in i]
        sold = [l for l in sold if l.home_type == "SINGLE_FAMILY" or not l.home_type]
        # Recency is a filter, not a preference. A saved pull spans years, so
        # the lookback has to be enforced here or stale sales quietly set the ARV.
        cutoff = (datetime.now() - timedelta(days=args.months * 30)).strftime("%Y-%m-%d")
        before = len(sold)
        sold = [l for l in sold if l.sold_date >= cutoff]
        logger.info("Recency filter: %d of %d sold are within %d months (since %s)",
                    len(sold), before, args.months, cutoff)
        active = []
        logger.info("Loaded %d sold comps from %s", len(sold), args.sold_json)
        return _clip(sold, args), _clip(active, args), {}

    if getattr(args, "no_api", False):
        logger.info("Comp pull skipped (--no-api)")
        return [], [], {}

    try:
        api = ZillowMarketAPI()
    except ValueError as exc:
        logger.warning("Comp pull skipped: %s", exc)
        return [], [], {}

    location = f"{subject['city']}, {subject['state']} {subject['zip']}"
    sold = _clip(api.pull_sold(location, months_back=args.months), args)
    active = _clip(api.pull_active(location), args)
    logger.info("Boundary-filtered: %d sold, %d active", len(sold), len(active))
    # The ARV is computed in run() AFTER bucket refinement, so demoted investor
    # buys are withheld from it. Computing it here would skip the refinement.
    return sold, active, {}


def point_in_polygon(lat: float, lon: float, poly: list) -> bool:
    """Ray casting. poly is [(lat, lon), ...], implicitly closed."""
    inside = False
    n = len(poly)
    for i in range(n):
        lat1, lon1 = poly[i]
        lat2, lon2 = poly[(i + 1) % n]
        if (lon1 > lon) != (lon2 > lon):
            t = (lon - lon1) / (lon2 - lon1)
            if lat < lat1 + t * (lat2 - lat1):
                inside = not inside
    return inside


def load_polygon(spec: str) -> list:
    """A drawn boundary as 'lat,lon lat,lon ...' or a JSON file of pairs.

    This is how the boundary actually gets decided: someone traces the pocket
    on a map along the roads that separate one micro-market from the next. A
    bounding box cannot express that, and every box wide enough to hold the
    pocket also holds the subdivisions on the far side of the highway.
    """
    if not spec:
        return []
    path = Path(spec)
    if path.exists():
        data = json.load(open(path, encoding="utf-8"))
        pts = data.get("polygon", data) if isinstance(data, dict) else data
        return [(float(p[0]), float(p[1])) for p in pts]
    out = []
    for pair in spec.replace(";", " ").split():
        lat, lon = pair.split(",")
        out.append((float(lat), float(lon)))
    return out


def _clip(listings: list, args) -> list:
    """Apply the drawn boundary: polygon, then bbox, then street regex.

    A polygon beats a box whenever one exists, because the pocket boundary is a
    set of roads, not a rectangle. bbox and streets still apply on top so a
    coarse polygon can be tightened without redrawing it.
    """
    poly = getattr(args, "_polygon", None) or load_polygon(getattr(args, "polygon", "") or "")
    if poly:
        args._polygon = poly
        before = len(listings)
        listings = [l for l in listings
                    if l.latitude and l.longitude and point_in_polygon(l.latitude, l.longitude, poly)]
        logger.info("Polygon boundary (%d vertices): kept %d of %d", len(poly), len(listings), before)
    if args.bbox:
        lat_min, lat_max, lon_min, lon_max = (float(x) for x in args.bbox.split(","))
        listings = filter_bbox(listings, lat_min, lat_max, lon_min, lon_max)
    if args.streets:
        listings = filter_streets(listings, re.compile(args.streets, re.I))
    return listings


def _bucket_of(l, buckets: dict | None) -> str:
    if buckets:
        return buckets.get(l.zpid or l.address) or classify(l)
    return classify(l)


ARV_SIZE_LO, ARV_SIZE_HI = 0.70, 1.35   # widest a comp can be and still be the same house
ARV_TIGHT_LO, ARV_TIGHT_HI = 0.85, 1.15  # preferred size band
ARV_PREFERRED_MONTHS = 12               # recent first, widen only if forced
MIN_ARV_COMPS = 3                       # below this the ARV is indicative, not underwritable


def tight_arv(subject: dict, sold: list, buckets: dict | None = None,
              months: int = 18) -> dict:
    """One exact ARV off the tightest comp set that still holds up.

    An escalation ladder, tightest first, widening only when a step leaves
    fewer than MIN_ARV_COMPS. Every widen is recorded in notes so the number
    is auditable:

      1. tight size (0.85-1.15x), 12 months, same bed, same bath band
      2. tight size, full lookback
      3. wide size (0.70-1.35x), full lookback
      4. wide size, any bath, discounted by the MEASURED bath gap

    Size and bath both matter because $/sf does not carry across a 2x size gap,
    and at this size a single bath is a priced deficiency, not a rounding error.
    """
    sqft = subject.get("sqft") or 0
    beds = subject.get("beds") or 0
    baths = float(subject.get("baths") or 0)

    def pool(size_lo: float, size_hi: float, window_days: int, bath_match: bool) -> list:
        cutoff = (datetime.now() - timedelta(days=window_days)).strftime("%Y-%m-%d")
        out = []
        for l in sold:
            if not isinstance(l, MarketListing) or not l.ppsf or not l.sqft:
                continue
            if _bucket_of(l, buckets) != "RENOVATED/RETAIL":
                continue
            if l.sold_date < cutoff or l.beds != beds:
                continue
            if sqft and not (size_lo <= l.sqft / sqft <= size_hi):
                continue
            if bath_match and baths and (l.baths <= 1.5) != (baths <= 1.5):
                continue
            out.append(l)
        return out

    ladder = [
        (ARV_TIGHT_LO, ARV_TIGHT_HI, ARV_PREFERRED_MONTHS * 30, True, ""),
        (ARV_TIGHT_LO, ARV_TIGHT_HI, months * 30, True,
         f"Widened past {ARV_PREFERRED_MONTHS} months to {months}"),
        (ARV_SIZE_LO, ARV_SIZE_HI, months * 30, True,
         f"Widened the size band to {ARV_SIZE_LO:.2f}-{ARV_SIZE_HI:.2f}x subject sqft"),
    ]

    notes, comps, bath_adj = [], [], 1.0
    for lo, hi, days, bmatch, why in ladder:
        comps = pool(lo, hi, days, bmatch)
        if len(comps) >= MIN_ARV_COMPS:
            if why:
                notes.append(f"{why} to reach {len(comps)} comps")
            break

    # Last resort: price off any-bath and discount by the gap this pocket shows.
    if len(comps) < MIN_ARV_COMPS and baths:
        allb = pool(ARV_SIZE_LO, ARV_SIZE_HI, months * 30, False)
        mine = [l.ppsf for l in allb if (l.baths <= 1.5) == (baths <= 1.5)]
        other = [l.ppsf for l in allb if (l.baths <= 1.5) != (baths <= 1.5)]
        if len(allb) >= MIN_ARV_COMPS and len(mine) >= 1 and len(other) >= 2:
            if len(mine) >= 2:
                bath_adj = statistics.median(mine) / statistics.median(other)
            else:
                bath_adj = mine[0] / statistics.median(other)
            notes.append(f"Only {len(mine)} same-bath comps, too thin to price off alone. "
                         f"Priced off all {len(allb)} same-bed sales and adjusted "
                         f"{(bath_adj - 1):+.0%} for the measured bath gap in this pocket")
            comps = allb

    if not comps:
        return {"arv": 0, "n": 0, "ppsf": 0, "comps": [],
                "notes": ["No retail comps survived the recency, size and bath filters"],
                "basis": "no qualifying comps"}
    if len(comps) < MIN_ARV_COMPS:
        notes.append(f"THIN: only {len(comps)} comp(s) qualify. Treat the ARV as indicative and "
                     "confirm against the county card before contracting")

    ppsf = statistics.median(l.ppsf for l in comps) * bath_adj
    est = ppsf * sqft
    prices = sorted(l.price * bath_adj for l in comps)
    # Clamp to the band median ONLY when the comps are the same size of house.
    # Against a set of much smaller homes the median price is a different
    # product, and clamping to it would understate a larger subject.
    med_sqft = statistics.median(l.sqft for l in comps)
    if sqft and 0.88 <= med_sqft / sqft <= 1.12:
        band_median = statistics.median(prices)
        if est > band_median:
            est = band_median
            notes.append(f"Clamped to the band median sale ({_money(band_median)}): extra sqft "
                         "does not lift a house out of its bed and bath band")

    newest = max(l.sold_date for l in comps)
    oldest = min(l.sold_date for l in comps)
    bath_lbl = ("1-1.5 bath" if baths <= 1.5 else "2 plus bath") if baths else "any bath"
    basis = (f"{len(comps)} comps, {beds}bd {bath_lbl}, "
             f"{int(min(l.sqft for l in comps)):,}-{int(max(l.sqft for l in comps)):,} sqft, "
             f"sold {oldest} to {newest}, median ${round(ppsf)}/sf"
             + ("  |  " + "; ".join(notes) if notes else ""))
    return {"arv": _round_money(est), "n": len(comps), "ppsf": round(ppsf),
            "band": (prices[0], prices[-1]), "comps": comps, "notes": notes, "basis": basis}


def as_is_value(sold: list, subject: dict, walk: dict, buckets: dict | None = None,
                deeds: dict | None = None, retail_floor: float = 0.0) -> float:
    """What the house is worth today, in its current condition.

    Walkthrough number wins. Otherwise: median distressed $/sf times subject
    sqft, which is the number the local cash buyers have actually been paying.
    """
    if walk.get("as_is_value"):
        return float(walk["as_is_value"])
    sqft = subject.get("sqft") or 0

    def comparable(l) -> bool:
        """Size band. $/sf is not transferable across a 2x size gap, so a
        2,800 sqft sale cannot set the as-is number on a 1,400 sqft house."""
        return bool(l.sqft) and (not sqft or 0.65 <= l.sqft / sqft <= 1.40)

    # BEST SOURCE: what investors actually paid at the deed. "Sold below
    # Zestimate" is not the same thing as an as-is trade, and conflating them
    # is how a $365,000 sale at a 0.60 ratio ends up setting the as-is band.
    if deeds:
        investor = []
        for l in sold:
            if not isinstance(l, MarketListing) or not l.ppsf or not comparable(l):
                continue
            d = deeds.get(_norm_addr(l.address))
            if not d:
                continue
            entity = re.search(r"\bLLC\b|\bINC\b|\bTRUST\b|PROPERT|HOLDINGS|HOMES|CAPITAL|INVEST",
                               d.get("buyer") or "", re.I)
            if d.get("cash") or entity:
                investor.append(l)
        if len(investor) >= 3:
            logger.info("As-is band from %d deed-verified investor buys (median $%.0f/sf)",
                        len(investor), statistics.median(l.ppsf for l in investor))
            return round(statistics.median(l.ppsf for l in investor) * sqft / 1000) * 1000

    # FALLBACK: the distressed bucket, size-filtered, arm's-length only, and
    # capped below the retail band. "Sold under Zestimate" catches plenty of
    # ordinary sales at ordinary prices; a $300,000 trade is not the as-is
    # number for a house that needs six figures of work.
    distressed = [l for l in sold if isinstance(l, MarketListing)
                  and _bucket_of(l, buckets) == "DISTRESSED" and l.ppsf and comparable(l)
                  and (not l.zestimate or l.price / l.zestimate >= NON_ARMS_LENGTH_RATIO)
                  and (not retail_floor or l.price < retail_floor)]
    if distressed and sqft:
        return round(statistics.median(l.ppsf for l in distressed) * sqft / 1000) * 1000
    return 0.0


def wholetail_value(sold: list, subject: dict, base_arv: float,
                    buckets: dict | None = None) -> float:
    """Resale after a clean-and-list, no renovation: the AVERAGE bucket."""
    average = [l for l in sold if isinstance(l, MarketListing)
               and _bucket_of(l, buckets) == "AVERAGE" and l.ppsf and l.beds == subject.get("beds")]
    if not average:
        average = [l for l in sold if isinstance(l, MarketListing)
                   and _bucket_of(l, buckets) == "AVERAGE" and l.ppsf]
    if average and subject.get("sqft"):
        return round(statistics.median(l.ppsf for l in average) * subject["sqft"] / 1000) * 1000
    return round(base_arv * 0.85 / 1000) * 1000


# ══ 3. Rehab matrix (the 4-scenario Repair Numbers engine) ═════════════

# rehab_estimator category -> the label the template uses
CATEGORY_LABELS = [
    ("Kitchen", "Kitchen"),
    ("Master Bathroom", "Master bath"),
    ("Secondary Bathroom(s)", "Secondary bath (added)"),
    ("Flooring", "Flooring"),
    ("Paint (Interior)", "Interior paint"),
    ("Exterior", "Exterior"),
    ("Windows", "Windows"),
    ("Roof", "Roof"),
    ("HVAC", "HVAC"),
    ("Electrical", "Electrical"),
    ("Plumbing", "Plumbing"),
    ("Foundation/Structural", "Foundation/structural"),
]
GUT_LABEL = "Gut allowance (demo/drywall/insul)"
CREDIT_LABEL = "Walkthrough credits (work already done)"
FLAG_LABEL = "Team walk flags (priced)"

# Envelope work a mid-reno leaves alone
MID_RENO_DROP = ("Roof", "Windows", "Foundation/Structural")


def scenario_defs(subject: dict, walk: dict) -> list[dict]:
    """The four columns of the Repair Numbers sheet.

    Cosmetic keeps the house as configured; the other three price the
    reconfigured target (usually adding the bedroom/bath that moves the house
    into the next value band).
    """
    sb, sba = subject.get("beds", 3), subject.get("baths", 1.0)
    target = walk.get("target_config") or {}
    tb = int(target.get("beds") or max(sb, 3))
    tba = float(target.get("baths") or max(sba, 2.0))
    # A walk that has consolidated onto ONE plan (comps say what the finish
    # is, the contract is signed, the menu phase is over) collapses the matrix
    # to that single column. Exits then price every lane off this one number.
    ss = walk.get("single_scenario")
    if ss:
        b = int(ss.get("beds") or tb)
        ba = float(ss.get("baths") or tba)
        return [{
            "key": ss.get("key") or "reno",
            "label": ss.get("label") or f"Comp-Match Reno ({b}/{_ba(ba)})",
            "tier": int(ss.get("tier") or 2), "scope": ss.get("scope") or "full",
            "beds": b, "baths": ba, "gut": bool(ss.get("gut")),
            "drop": tuple(ss.get("drop") or ()),
        }]
    return [
        {"key": "cosmetic", "label": f"Cosmetic ({sb}/{_ba(sba)})", "tier": 2,
         "scope": "wholetail", "beds": sb, "baths": sba, "gut": False, "drop": ()},
        {"key": "mid", "label": f"Mid Reno ({tb}/{_ba(tba)})", "tier": 2,
         "scope": "full", "beds": tb, "baths": tba, "gut": False, "drop": MID_RENO_DROP},
        {"key": "gut_t2", "label": f"Full Gut T2 ({tb}/{_ba(tba)})", "tier": 2,
         "scope": "full", "beds": tb, "baths": tba, "gut": True, "drop": ()},
        {"key": "gut_t3", "label": f"Full Gut T3 ({tb}/{_ba(tba)})", "tier": 3,
         "scope": "full", "beds": tb, "baths": tba, "gut": True, "drop": ()},
    ]


def _line_rows(room) -> list[tuple[str, float]]:
    """Turn one RoomEstimate into readable (line item, cost) pairs.

    The estimator stores three different line_items shapes (component table,
    per-sqft rates, fixed installed cost); this normalizes all of them to money.
    """
    li = room.line_items or {}
    # Labels must NOT carry the tier-dependent unit rate, or the same line
    # splits into one row per tier and the breakdown stops lining up.
    if "materials_per_sqft" in li:
        sq = int(li.get("sqft") or 0)
        return [(f"Materials ({sq:,} sf)", room.materials),
                (f"Labor ({sq:,} sf)", room.labor)]
    if "per_window" in li:
        return [(f"{int(li.get('count') or 0)} windows, installed", room.total)]
    if "per_sqft" in li:
        return [(f"{int(li.get('roof_sqft') or 0):,} roof sf, tear off and reshingle", room.total)]
    if "total_installed" in li:
        return [("Installed, 60/40 labor to materials", room.total)]
    return [(k.replace("_", " ").capitalize(), v) for k, v in li.items() if v]


def _walk_items(walk: dict, key: str, amount_key: str) -> list[dict]:
    return [i for i in (walk.get(key) or []) if _num(i.get(amount_key))]


def build_rehab_matrix(subject: dict, walk: dict, soft_pct: float = SOFT_COST_PCT) -> dict:
    """Run the rehab engine once per scenario and fold in the walkthrough.

    Credits (work the seller already paid for) and team-walk flags (what the
    walk found that the model cannot see) are kept as their OWN rows rather
    than smeared into categories, so every dollar stays auditable back to the
    Repair Logic sheet.
    """
    sqft = subject.get("sqft") or 0
    scenarios = []
    cat_present: list[str] = []
    line_index: list[tuple[str, str]] = []
    placeholder_lines: set = set()

    credits = _walk_items(walk, "work_done", "credit")
    flags = _walk_items(walk, "flags", "cost")

    # Labor model: default is the own-crew self-perform factor; a walk that is
    # being PM'd with subs overrides the factor (and names the model) so the
    # second budget line is the one the operator actually runs.
    sp_factor = _num(walk.get("self_perform_factor")) or SELF_PERFORM_LABOR_FACTOR
    sp_label = walk.get("labor_model_label") or "self-perform (own crew)"

    for sd in scenario_defs(subject, walk):
        est = estimate_rehab("", sqft, sd["beds"], sd["baths"], subject.get("year_built", 0),
                             tier=sd["tier"], scope=sd["scope"], region=walk.get("region", "knoxville"))
        rooms = [r for r in est.rooms if r.category not in sd["drop"]]

        cats: dict[str, float] = {}
        lines: dict[tuple[str, str], float] = {}
        materials = labor = 0.0
        for room in rooms:
            label = dict(CATEGORY_LABELS).get(room.category, room.category)
            cats[label] = cats.get(label, 0) + room.total
            materials += room.materials
            labor += room.labor
            for item, cost in _line_rows(room):
                lines[(label, item)] = lines.get((label, item), 0) + cost

        if sd["gut"] and sqft:
            gut = round(GUT_ALLOWANCE_PER_SQFT * sqft)
            cats[GUT_LABEL] = gut
            lines[(GUT_LABEL, f"Demo, drywall, insulation at ${GUT_ALLOWANCE_PER_SQFT:.0f}/sf x {sqft:,} sf")] = gut
            materials += gut * 0.35          # allowance is mostly labor
            labor += gut * 0.65

        # Walkthrough credits: subtract, but never below zero on the category.
        credit_total = 0.0
        for c in credits:
            if sd["key"] in (c.get("scenarios") or [s["key"] for s in scenario_defs(subject, walk)]):
                amt = _num(c["credit"])
                credit_total += amt
                lines[(CREDIT_LABEL, f"{c.get('item', 'Credit')}: {c.get('detail', '')}".strip(": "))] = -amt
        if credit_total:
            cats[CREDIT_LABEL] = -credit_total
            materials -= credit_total * 0.4
            labor -= credit_total * 0.6

        flag_total = 0.0
        for f in flags:
            if sd["key"] in (f.get("scenarios") or [s["key"] for s in scenario_defs(subject, walk)]):
                amt = _num(f["cost"])
                flag_total += amt
                line_key = (FLAG_LABEL, f"{f.get('item', 'Flag')}: {f.get('note', '')}".strip(": "))
                lines[line_key] = amt
                # Ty's rule: a pending quote never renders blank. It carries a
                # placeholder dollar figure, painted red on the sheet, so the
                # total is always a true number. Replace with the real bid.
                if f.get("placeholder"):
                    placeholder_lines.add(line_key)
        if flag_total:
            cats[FLAG_LABEL] = flag_total
            materials += flag_total * 0.4
            labor += flag_total * 0.6

        net = max(sum(cats.values()), 0.0)
        soft = round(net * soft_pct)
        grand = round(net + soft)
        self_perform = round((max(materials, 0) + max(labor, 0) * sp_factor) * (1 + soft_pct))

        scenarios.append({
            "key": sd["key"], "label": sd["label"], "tier": sd["tier"],
            "categories": {k: round(v) for k, v in cats.items()},
            "lines": {k: round(v) for k, v in lines.items()},
            "materials": round(max(materials, 0)), "labor": round(max(labor, 0)),
            "subtotal": round(net), "soft": soft, "grand": grand,
            "self_perform": self_perform,
            "ppsf": round(grand / sqft, 2) if sqft else 0,
            # Weeks off the RETAINED rooms only (a mid reno skips the envelope),
            # with the estimator's parallel-work factor.
            "weeks": round(sum(r.weeks for r in rooms) * 0.6, 1),
        })
        for k in cats:
            if k not in cat_present:
                cat_present.append(k)
        for k in lines:
            if k not in line_index:
                line_index.append(k)

    # Category rows in template order, then the synthetic rows, then anything new
    ordered = [lbl for _, lbl in CATEGORY_LABELS if lbl in cat_present]
    for extra in (GUT_LABEL, CREDIT_LABEL, FLAG_LABEL):
        if extra in cat_present:
            ordered.append(extra)
    ordered += [c for c in cat_present if c not in ordered]
    line_index.sort(key=lambda kv: (ordered.index(kv[0]) if kv[0] in ordered else 99, kv[1]))

    return {"scenarios": scenarios, "categories": ordered, "lines": line_index,
            "soft_pct": soft_pct, "sp_factor": sp_factor, "sp_label": sp_label,
            "placeholders": sorted("|".join(k) for k in placeholder_lines),
            "totals": {s["key"]: s["grand"] for s in scenarios}}


# ══ 4. Exit strategies ═════════════════════════════════════════════════

def build_exits(subject: dict, arv: dict, rehab: dict, walk: dict,
                as_is: float, wholetail_arv: float) -> dict:
    """Score the exits that are actually on the table for THIS house.

    Two team rules drive the shape of this:
      * Only real candidates get a suggestion block. The template's six slots
        were placeholders (Rami), so an exit that cannot happen here is
        explained in the logic block instead of taking up a slot.
      * Every price we are not confident in ships as a RANGE (Marwan). Rehab is
        a model output until a bid lands, so it stays a range; ARV tightens to
        one number only when the comp band is deep and narrow.
    """
    base_t, up_t = (arv.get("base") or {}), (arv.get("upside") or {})
    base = base_t.get("arv", 0) or _num(walk.get("arv_override"))
    upside = up_t.get("arv", 0)
    totals = rehab.get("totals", {})

    def arv_rng(track: dict, point: float) -> dict:
        """ARV range = the band the comps actually drew, never tighter."""
        band = track.get("band") or ()
        n = track.get("n", 0)
        if not point:
            return rng(0, 0, confident=True)
        if len(band) == 2 and band[0] and band[1]:
            lo, hi = min(band[0], point), max(band[1], point)
            tight = n >= 5 and (hi - lo) / max(point, 1) <= 0.30
            return rng(lo, hi, confident=tight, point=point)
        return rng(point * 0.90, point * 1.10, point=point)

    base_r = arv_rng(base_t, base)
    upside_r = arv_rng(up_t, upside) if upside else None

    # Rehab overruns skew high, so the band is asymmetric. A signed bid in the
    # walkthrough file collapses it to a single number.
    bids = walk.get("bids") or {}

    def work_rng(key: str) -> dict:
        if _num(bids.get(key)):
            return rng(_num(bids[key]), confident=True)
        v = totals.get(key)
        if v is None and len(totals) == 1:
            # Single-scenario walk: every lane prices off the one plan.
            v = next(iter(totals.values()))
        v = v or 0
        return rng(v * 0.90, v * 1.15, point=v)

    cosmetic, mid_w = work_rng("cosmetic"), work_rng("mid")
    gut2, gut3 = work_rng("gut_t2"), work_rng("gut_t3")

    # A signed contract is a fact; a derived MAO is a range.
    if _num(walk.get("contract_price")):
        purchase = rng(_num(walk["contract_price"]), confident=True)
    elif base:
        purchase = rng(base * 0.70 - gut2["hi"], base * 0.75 - gut2["lo"])
    else:
        purchase = rng(0, 0, confident=True)
    contract = purchase["point"]

    as_is_r = (rng(as_is, confident=True) if walk.get("as_is_value")
               else rng(as_is * 0.90, as_is * 1.10, point=as_is))
    if _num(walk.get("assignment_price")):
        assign = rng(_num(walk["assignment_price"]), confident=True)
    else:
        assign = rng(as_is_r["lo"] * 0.90, as_is_r["hi"] * 0.98, point=as_is * 0.95)

    wholetail_r = (rng(wholetail_arv * 0.92, wholetail_arv * 1.08, point=wholetail_arv)
                   if wholetail_arv else rng(0, 0, confident=True))
    rent = _num(walk.get("monthly_rent"))

    work_ratio = (_mid(gut2) / base) if base else 0
    shell = walk.get("unfinanceable")
    if shell is None:
        shell = work_ratio > SHELL_WORK_RATIO
    reconfig_verified = bool(walk.get("reconfig_verified"))

    # ── Financing (walk "financing" block): private money baked into profit ──
    # Without the block every resale lane stays cash-basis (the old behavior).
    # With it, each lane's profit is NET OF DEBT: points + interest on the
    # committed balance over that lane's hold, plus buy-side closing/title.
    # Interest accrues on the FULL balance (a draw schedule makes reality
    # cheaper), so the number is conservative.
    fin = walk.get("financing") or {}
    fin_on = bool(fin)
    # Explicit zero is a real term (a no-points lender), so defaults apply
    # only when the key is absent, never when it is 0.
    fin_rate = _num(fin["rate"]) if fin.get("rate") is not None else 0.12
    fin_points = _num(fin["points"]) if fin.get("points") is not None else 2.0
    fin_term = int(_num(fin["term_months"])) if fin.get("term_months") is not None else 9
    fin_ltc = _num(fin["ltc"]) if fin.get("ltc") is not None else 1.0

    def buy_close(p: float) -> float:
        return BUY_CLOSE_FLAT + p * BUY_TITLE_PCT

    def fin_cost(p: float, w: float, months: int) -> float:
        loan = (p + w) * fin_ltc
        return loan * fin_points / 100 + loan * fin_rate * months / 12

    def resale_profit(sell: dict, work: dict, hold_months: int) -> dict:
        """Worst case pairs the low sale with the high rehab, and the reverse."""
        hold = contract * HOLDING_COST_PCT * (hold_months / 6)
        debt_lo = debt_hi = 0.0
        if fin_on:
            debt_lo = fin_cost(purchase["lo"], work["lo"], hold_months) + buy_close(purchase["lo"])
            debt_hi = fin_cost(purchase["hi"], work["hi"], hold_months) + buy_close(purchase["hi"])
        lo = sell["lo"] * (1 - SELLING_COST_PCT) - purchase["hi"] - work["hi"] - hold - debt_hi
        hi = sell["hi"] * (1 - SELLING_COST_PCT) - purchase["lo"] - work["lo"] - hold - debt_lo
        return rng(lo, hi, confident=(sell["confident"] and work["confident"]
                                      and purchase["confident"]))

    def roi(profit: dict, work: dict) -> float:
        if fin_on:
            # Financed deal: the honest lens is cash-on-cash. Cash in = the
            # cost the loan does not cover, plus points and closing paid at
            # the table.
            total = contract + _mid(work)
            loan = total * fin_ltc
            cash_in = max(total - loan, 0) + buy_close(contract) + loan * fin_points / 100
            return _mid(profit) / max(cash_in, 1)
        return _mid(profit) / max(contract + _mid(work), 1)

    def flip_viable(profit: dict, work: dict) -> bool:
        """Cash-on-cash goes huge at high LTC, so a financed flip must ALSO
        beat what wholesaling the same file pays, or the debt is not worth
        carrying."""
        ok = roi(profit, work) >= MIN_FLIP_ROI
        if fin_on:
            ok = ok and _mid(profit) >= MIN_WHOLESALE_FEE
        return ok

    def flip_clears_at(sell: dict, work: dict) -> float:
        """Entry price where a flip hits the ROI floor, solved off the same
        mids the gate uses (sell band mid, rehab band mid, work point in the
        denominator). Six-month hold assumed, matching the flip exits."""
        smid = (sell["lo"] + sell["hi"]) / 2
        wmid = (work["lo"] + work["hi"]) / 2
        wpt = work.get("point") or wmid
        return max(((1 - SELLING_COST_PCT) * smid - wmid - MIN_FLIP_ROI * wpt)
                   / (1 + HOLDING_COST_PCT + MIN_FLIP_ROI), 0.0)

    exits = []

    p = rng(assign["lo"] - purchase["hi"], assign["hi"] - purchase["lo"],
            confident=assign["confident"] and purchase["confident"])
    exits.append({
        "name": "Wholesale assignment (as-is)", "kind": "assign",
        "arv": base_r, "work": rng(0, 0, confident=True), "purchase": purchase,
        "sell": assign, "profit": p,
        "viable": _mid(p) >= MIN_WHOLESALE_FEE,
        "why": "No capital, no rehab risk, closes in days. The pool that pays this number is "
               "self-performers and landlords, not GC-model flippers.",
        "why_not": (f"Spread mids at {_money(_mid(p))}, under the {_money(MIN_WHOLESALE_FEE)} "
                    "floor: the fee has to be won on the buyer side, not the buy side."
                    if _mid(p) < MIN_WHOLESALE_FEE else ""),
    })

    p = resale_profit(wholetail_r, cosmetic, 3)
    exits.append({
        "name": "Wholetail (clean, light cosmetic, list as-is)", "kind": "resale",
        "arv": wholetail_r, "work": cosmetic, "purchase": purchase, "sell": wholetail_r,
        "profit": p,
        "viable": bool(wholetail_arv) and not shell and flip_viable(p, cosmetic),
        # A shell blocks wholetail for ANY owner, so it can never be the exit we
        # pitch the buyer either (price-gated exits can: the buyer's basis differs).
        "property_gated": shell or not wholetail_arv,
        "why": "Fastest retail dollar: no permits, no systems, sells into the dated-but-livable band.",
        "why_not": ("House is an unfinanceable shell, so there is no retail buyer for it in this "
                    "condition. Wholetail needs a house someone can move into." if shell else
                    ("No dated-condition comps inside the boundary to price a wholetail against."
                     if not wholetail_arv else
                     f"ROI mids at {roi(p, cosmetic):.0%}, under the {MIN_FLIP_ROI:.0%} floor.")),
    })

    p = resale_profit(base_r, gut2, 6)
    exits.append({
        "name": "Fix and flip, same config (Tier 2)", "kind": "resale",
        "arv": base_r, "work": gut2, "purchase": purchase, "sell": base_r, "profit": p,
        "viable": bool(base) and flip_viable(p, gut2),
        "why": "Sells into the same-bed retail band the comps actually support. The honest base "
               "case: no layout bet.",
        "why_not": ("No same-bed retail comps to underwrite against." if not base else
                    f"ROI mids at {roi(p, gut2):.0%}, under the {MIN_FLIP_ROI:.0%} floor at this "
                    f"rehab number. Clears at a {_money(_round_money(flip_clears_at(base_r, gut2)))} buy."),
    })

    # Same flip, our own crew on the labor. A separate lane, not a discount on
    # the GC number: the self-perform basis is what the heavy-rehab buyer pool
    # underwrites with, so this block is also the buyer's math made visible.
    scen_list = rehab.get("scenarios") or []
    sp = next((s.get("self_perform") for s in scen_list if s.get("key") == "gut_t2"),
              scen_list[0].get("self_perform", 0) if len(scen_list) == 1 else 0)
    if sp and base:
        sp_factor = rehab.get("sp_factor", SELF_PERFORM_LABOR_FACTOR)
        sp_label = rehab.get("sp_label", "self-perform (own crew)")
        sp_r = rng(sp * 0.90, sp * 1.15, point=sp)
        p = resale_profit(base_r, sp_r, 6)
        sp_clears = flip_clears_at(base_r, sp_r)
        exits.append({
            "name": f"Fix and flip, {sp_label}", "kind": "resale",
            "arv": base_r, "work": sp_r, "purchase": purchase, "sell": base_r, "profit": p,
            "viable": flip_viable(p, sp_r),
            "why": f"Labor at the {sp_factor:.0%} factor instead of GC retail: the {sp_label} "
                   "basis. The lane that actually buys heavy-rehab houses: the buyers we dial run "
                   "this same math.",
            "why_not": ("" if roi(p, sp_r) >= MIN_FLIP_ROI else
                        f"ROI mids at {roi(p, sp_r):.0%}, under the {MIN_FLIP_ROI:.0%} floor at this "
                        f"buy. Clears at a {_money(_round_money(sp_clears))} buy."),
        })

    if upside_r:
        p = resale_profit(upside_r, gut3, 7)
        exits.append({
            "name": "Fix and flip, reconfig to target beds (Tier 3)", "kind": "resale",
            "arv": upside_r, "work": gut3, "purchase": purchase, "sell": upside_r, "profit": p,
            "viable": reconfig_verified and flip_viable(p, gut3),
            "property_gated": not reconfig_verified,
            "why": "Moves the house into the higher-bed value band, which is where the real "
                   "spread on this street lives.",
            "why_not": ("UPSIDE TRACK, not underwritable yet: the walk has to confirm the layout "
                        "converts (bearing walls, egress, ceiling height) before this ARV is "
                        "allowed into the math." if not reconfig_verified else
                        f"ROI mids at {roi(p, gut3):.0%}, under the {MIN_FLIP_ROI:.0%} floor."),
        })

    # BRRRR profit is CASH OUT AT REFI, not a sale profit. Kept because it is a
    # real use of the house; the logic block states that the unit differs.
    refi = rng(base_r["lo"] * REFI_LTV, base_r["hi"] * REFI_LTV,
               confident=base_r["confident"], point=base * REFI_LTV)
    all_in_lo = purchase["lo"] + mid_w["lo"] + refi["lo"] * 0.03
    all_in_hi = purchase["hi"] + mid_w["hi"] + refi["hi"] * 0.03
    cash_out = rng(refi["lo"] - all_in_hi, refi["hi"] - all_in_lo,
                   confident=(refi["confident"] and mid_w["confident"] and purchase["confident"]))
    annual_cf = round(rent * 12 * 0.55 - _mid(refi) * 0.075) if rent else 0
    exits.append({
        "name": "Rental hold (BRRRR, refi cash out)", "kind": "hold",
        "arv": base_r, "work": mid_w, "purchase": purchase, "sell": refi, "profit": cash_out,
        "viable": bool(rent) and bool(base) and (_mid(cash_out) >= 0 or annual_cf > 0),
        "why": f"Refi at {REFI_LTV:.0%} returns {fmt_money_rng(refi)} against "
               f"{_money(all_in_lo)} to {_money(all_in_hi)} all-in"
               + (f", cash flow about {_money(annual_cf)}/yr after debt service." if annual_cf
                  else ", with no rent comp to prove the cash flow."),
        "why_not": ("No rent estimate on the record or in the walkthrough file, so the hold cannot "
                    "be underwritten: a refi you cannot service is not an exit." if not rent else
                    f"{_money(abs(_mid(cash_out)))} stays trapped at this rehab number."),
    })

    ranked = sorted(exits, key=lambda e: (e["viable"], _mid(e["profit"])), reverse=True)
    # Only real candidates get a block. If nothing clears its gate, show the two
    # closest misses so the sheet is never blank, and say exactly that.
    suggested = [e for e in ranked if e["viable"]]
    forced = not suggested
    if forced:
        suggested = ranked[:2]
    # The walkthrough can pin an exit onto the sheet even when it does not
    # clear (walk "show_exits": name substrings). It renders as a full block
    # tagged "Shown on request", never as a suggestion.
    for want in (walk.get("show_exits") or []):
        for e in ranked:
            if want.lower() in e["name"].lower() and e not in suggested:
                e["forced_show"] = True
                suggested.append(e)
    main = suggested[0] if suggested else None
    resales = ([e for e in suggested if e.get("kind") == "resale"]
               or [e for e in ranked
                   if e.get("kind") == "resale" and not e.get("property_gated")]
               or [e for e in ranked if e.get("kind") == "resale"])
    buyer_exit = resales[0] if resales else main

    logic = []
    if base:
        logic.append(f"Base ARV {fmt_money_rng(base_r)} is the same-bed track and is the ONLY "
                     "number underwriting is allowed to use."
                     + (f" Upside {fmt_money_rng(upside_r)} rides on a verified reconfig."
                        if upside_r else ""))
    logic.append(f"Rehab against ARV at full gut mids at {work_ratio:.0%}. "
                 + ("Above the shell line, so the buyer pool is CASH self-performers and "
                    "landlords: a GC-model flipper's MAO collapses at this rehab number."
                    if shell else "Inside the range a GC-model flipper can still underwrite."))
    if fin_on:
        logic.append(f"FINANCING IS IN THE PROFIT MATH: {fin.get('kind', 'private money')} at "
                     f"{fin_rate:.0%} + {fin_points:g} pts on {fin_ltc:.0%} of cost, interest on the "
                     f"full balance over each lane's hold (a draw schedule beats this), plus buy-side "
                     f"closing ({_money(BUY_CLOSE_FLAT)} + {BUY_TITLE_PCT:.2%} title). Every resale "
                     "profit is net of debt; ROI reads as cash-on-cash, and a financed flip must also "
                     f"beat the {_money(MIN_WHOLESALE_FEE)} wholesale floor to stay suggested."
                     + (" TERMS ARE ASSUMED: replace with the signed term sheet (Lender Analysis "
                        "sheet)." if fin.get("assumed") else ""))
    if as_is and contract >= as_is * 0.95:
        logic.append(f"Contract {fmt_money_rng(purchase)} sits at or above the as-is band "
                     f"({fmt_money_rng(as_is_r)}). That converts this from a discount wholesale "
                     "into a dispo-EXECUTION play: the fee is won on the buyer side, so price per "
                     "buyer model instead of blasting one number.")
    logic.append("Every figure is ONE number, taken off the tight comp set (recent first, same "
                 "bed count, same size of house). The sensitivity line under each block shows "
                 "what the profit becomes if the sale comes in low and the rehab runs high.")
    logic.append("Profit is net cash at closing on every sale exit (after selling costs, carry, "
                 "rehab and the buy). The BRRRR line is the odd unit: that number is cash out at "
                 "refinance and you still own the house, so do not read it as a check.")
    if forced:
        logic.append("NOTHING CLEARS ITS GATE at these numbers. The blocks below are the closest "
                     "misses, shown so the sheet is not blank. Re-price the buy or get a real "
                     "rehab bid before taking this to a buyer.")
    for e in ranked:
        verdict = ("RECOMMENDED" if e is main else
                   ("on the table" if e["viable"] else "not recommended"))
        reason = e["why"] if e["viable"] else (e["why_not"] or e["why"])
        logic.append(f"{e['name']}: {verdict}. {reason}")

    # Lanes we actually run: wholesale, wholetail, fix and flip, rental.
    # Novation is deliberately not modelled.
    return {"exits": ranked, "suggested": suggested, "main": main, "buyer_exit": buyer_exit,
            "logic": logic, "forced": forced,
            "inputs": {"arv_basis": arv.get("tight", {}).get("basis", ""),
                       "base_arv": base_r, "upside_arv": upside_r, "as_is": as_is_r,
                       "wholetail": wholetail_r, "contract": purchase, "assign": assign,
                       "work_ratio": work_ratio, "shell": shell,
                       "reconfig_verified": reconfig_verified,
                       "contract_point": contract, "assign_point": assign["point"],
                       "as_is_point": as_is,
                       "financing": ({"kind": fin.get("kind", "private money"),
                                      "rate": fin_rate, "points": fin_points,
                                      "term_months": fin_term, "ltc": fin_ltc,
                                      "draws": int(_num(fin.get("draws")) or 4),
                                      "lender": fin.get("lender", ""),
                                      "assumed": bool(fin.get("assumed"))}
                                     if fin_on else None)}}


def fmt_money_rng(r) -> str:
    """Always a string, for prose. fmt_rng is what goes in a cell."""
    v = fmt_rng(r)
    return _money(v) if isinstance(v, (int, float)) else v



# ══ 5. Buyers + outreach (dispo stack inputs) ══════════════════════════

def load_buyers(path: str | None) -> list[dict]:
    """Ranked buyers from a buyer_sweep run (its `ranked` list, or a CSV)."""
    if not path or not Path(path).exists():
        return []
    if path.lower().endswith(".json"):
        data = json.load(open(path, encoding="utf-8"))
        rows = (data.get("ranked") or data.get("buyers") or data.get("targets") or []) \
            if isinstance(data, dict) else data
    else:
        import csv as _csv
        rows = list(_csv.DictReader(open(path, newline="", encoding="utf-8-sig")))
    return rows


def _norm_addr(a: str) -> str:
    """Loose street key for joining Zillow addresses to deed records."""
    a = re.sub(r"[^a-z0-9 ]", "", str(a or "").lower())
    a = re.sub(r"\b(rd|road|st|street|ave|avenue|dr|drive|ln|lane|ct|court|pike|way|blvd|"
               r"cir|circle|pl|place|ne|nw|se|sw|n|s|e|w)\b", "", a)
    return re.sub(r"\s+", " ", a).strip()


def load_deeds(path: str | None) -> dict:
    """Address -> {buyer, cash} from the buyer_sweep `records` block.

    This is what turns the comp table into a dispo list: the buyer of the
    distressed comp two streets over is the person to call about this house.
    """
    if not path or not Path(path).exists() or not path.lower().endswith(".json"):
        return {}
    data = json.load(open(path, encoding="utf-8"))
    records = data.get("records") if isinstance(data, dict) else None
    if not records:
        return {}
    out = {}
    for rec in records:
        key = _norm_addr(rec.get("address"))
        if key and rec.get("buyer"):
            out[key] = {"buyer": rec["buyer"], "cash": bool(rec.get("cash")),
                        "sale_date": rec.get("sale_date"), "sale_price": rec.get("sale_price")}
    return out


def buyer_type(b: dict) -> str:
    """Read the buyer's MODEL off their deed record, because the ask price is
    tuned per model: a self-performer pays more than an out-of-state landlord."""
    n = int(_num(b.get("n_buys") or b.get("purchase_count")))
    cash = int(_num(b.get("cash_n")))
    portfolio = int(_num(b.get("portfolio_n") or b.get("portfolio_size")))
    entity = b.get("is_entity") or re.search(r"LLC|INC|TRUST|PROPERTIES|HOLDINGS",
                                             str(b.get("buyer") or b.get("buyer_name") or ""), re.I)
    if portfolio >= 10 and n <= portfolio / 3:
        return "Landlord / portfolio holder"
    if n >= 3 and cash >= max(1, n * 0.5):
        return "Volume cash flipper"
    if cash >= 1:
        return "Cash flipper"
    return "Investor entity" if entity else "Individual investor"


def load_outreach(path: str | None) -> list[dict]:
    """Skip-traced contacts from a dispo_skiptrace run."""
    if not path or not Path(path).exists():
        return []
    data = json.load(open(path, encoding="utf-8"))
    return data.get("contacts") or (data if isinstance(data, list) else [])


def buyer_fit(buyer: dict, as_is: float, contract: float, ask: float = 0.0) -> str:
    """Rank a buyer against THIS deal's band, not against the market generally.

    A 40-flip machine that only buys at $40K is a worse call than a 3-flip
    operator whose average buy is exactly your ask.
    """
    avg = _num(buyer.get("avg_price") or buyer.get("avg_buy") or buyer.get("median_price"))
    count = int(_num(buyer.get("n_buys") or buyer.get("purchase_count") or 0))
    cash = int(_num(buyer.get("cash_n")))
    target = ask or max(as_is, contract)
    band_lo, band_hi = target * 0.55, target * 1.8
    in_band = bool(avg and band_lo <= avg <= band_hi)
    if in_band and count >= 3:
        return "1 - hottest"
    if in_band and (count >= 1 or cash):
        return "2"
    if count >= 5:
        return "3"
    return "4 - fallback"


def _phone_cell(contact: dict) -> str:
    """All dialable numbers with Trestle score and tier, best first."""
    phones = contact.get("phones") or []
    out = []
    for p in phones[:4]:
        num = p.get("formatted") or p.get("phone") or ""
        score = p.get("score")
        t = (p.get("tier") or "").upper()
        conf = f"x{p['confirm_count']}" if p.get("confirm_count", 0) > 1 else ""
        bits = ", ".join(str(b) for b in [score, t, conf] if b not in (None, ""))
        out.append(f"{num} ({bits})" if bits else num)
    return " | ".join(out)


# ══ 6. Sheet builders (template order) ═════════════════════════════════

def sheet_overview(wb, pack):
    subject, lead, walk = pack["subject"], pack["lead"], pack["walk"]
    arv, exits = pack["arv"], pack["exits"]
    ws = wb.create_sheet("Overview")

    ws.cell(row=1, column=1, value="Property Address").font = Font(bold=True, size=11, color=NAVY)
    ws.cell(row=1, column=2, value=subject["full_address"])
    ws.cell(row=2, column=1, value="County").font = Font(bold=True, size=11, color=NAVY)
    ws.cell(row=2, column=2, value=lead.get("county") or subject.get("county", ""))
    ws.cell(row=3, column=2, value=f"Post-walkthrough package built {datetime.now():%m/%d/%Y}"
            + (f". Walked {walk['walk_date']}." if walk.get("walk_date") else ".")).font = Font(size=10, color=GREY)

    base = (arv.get("base") or {})
    upside = (arv.get("upside") or {})
    source = walk.get("subject_source") or ("Sift record + county card" if lead.get("available") else "Zillow property details")

    rows = [
        [f"SUBJECT ({source})", ""],
        ["Motivation", lead_motivation(lead, walk)],
        ["Estimated ARV", (f"{fmt_money_rng(exits['inputs']['base_arv'])} base, same-bed track"
                           + (f" | {fmt_money_rng(exits['inputs']['upside_arv'])} upside if the "
                              "reconfig verifies" if exits["inputs"].get("upside_arv") else ""))],
        ["Sqft", f"{subject['sqft']:,}" if subject.get("sqft") else ""],
        ["Bed/Bath", f"{subject.get('beds', '')}/{_ba(subject.get('baths', 0))}"],
        ["Lot Size", subject.get("lot", "")],
        ["Year Built", subject.get("year_built", "")],
        ["Zestimate", (_money(subject.get("zestimate")) + (" (Sift estimate)" if not walk.get("zestimate_source")
                                                           else "")) if subject.get("zestimate") else ""],
        ["Parcel", subject.get("parcel_id", "")],
        ["As-is value (distressed band)", fmt_money_rng(exits["inputs"]["as_is"])
         if exits["inputs"]["as_is_point"] else "no distressed comps inside the boundary"],
        ["", ""],
        ["OWNERSHIP / SELLER ANGLE", ""],
        ["Current owner", (lead.get("owner_name") or walk.get("owner") or "")
         + (f" (mailing: {lead['mailing']})" if lead.get("mailing") else "")
         + (" [DECEASED on record]" if lead.get("owner_deceased") else "")],
        ["Acquired", walk.get("acquired") or acquired_line(lead)],
        ["Distress signal", walk.get("distress") or ", ".join(distress_signals(lead))
         or ", ".join(str(t) for t in (lead.get("tags") or [])[:6])],
        ["Note", walk.get("note") or ""],
        ["", ""],
        ["THE SIFT LEAD", ""],
    ]
    if lead.get("available"):
        cf = lead.get("custom_fields") or {}
        rows += [
            ["Record", lead["url"]],
            ["Status", lead.get("status") or "(none)"],
            ["Lists", ", ".join(str(x) for x in (lead.get("lists") or [])) or "(none)"],
            ["Tags", ", ".join(str(x) for x in (lead.get("tags") or [])) or "(none)"],
            ["SIFTline", ", ".join(lead.get("cards") or []) or "(not on a board)"],
            ["Assigned to", lead.get("assignee") or "unassigned"],
            ["Last contact", f"{lead.get('last_contact', '')} ({lead.get('last_contact_type', '')})".strip()],
            ["Sift estimated value", _money(lead["estimated_value"]) if lead.get("estimated_value") else ""],
            ["Rent estimate", _money(lead["rental_value"]) + "/mo" if lead.get("rental_value") else ""],
            ["Investor score", str(lead.get("investor_score") or "")],
            ["Decision maker", cf.get("Decision Maker") or lead.get("personal_representative", "")],
            ["DM relationship", cf.get("DM Relationship", "")],
        ]
        for m in (lead.get("messages") or [])[:3]:
            rows.append([f"Note {m['date']}", f"{m['who']}: {m['text'][:500]}"])
    else:
        rows.append(["Not linked", f"No live Sift context: {lead.get('reason', 'lookup skipped')}"])

    _kv(ws, 5, rows)
    _widths(ws, [32, 110])


def sheet_exits(wb, pack):
    ex = pack["exits"]
    ws = wb.create_sheet("Exit Strats")
    ws.cell(row=1, column=1, value="Exit Strategy").font = Font(bold=True, size=14, color=NAVY)

    suggested, main = ex["suggested"], ex["main"]
    headline = (f"MAIN: {main['name']} at {fmt_money_rng(main['profit'])}." if main
                else "No exit scored.")
    also = [e for e in suggested[1:] if not e.get("forced_show")]
    shown = [e for e in suggested if e.get("forced_show")]
    if also:
        headline += "  Also live: " + " | ".join(
            f"{e['name']} {fmt_money_rng(e['profit'])}" for e in also)
    if shown:
        headline += "  Shown on request: " + " | ".join(
            f"{e['name']} {fmt_money_rng(e['profit'])}" for e in shown)
    rejected = [e for e in ex["exits"] if e not in suggested]
    if rejected:
        headline += (f"  Ruled out ({len(rejected)}): " + ", ".join(e["name"] for e in rejected)
                     + ". Reasons below.")

    ws.cell(row=2, column=1, value="Deal Suggestions").font = Font(bold=True, size=11, color=BLUE)
    ws.cell(row=2, column=2, value=headline)
    ws.cell(row=3, column=1, value="Deal Type Suggestion Logic").font = Font(bold=True, size=11, color=BLUE)
    ws.cell(row=3, column=2, value=(ex["logic"][0] if ex["logic"] else ""))
    for r in (2, 3):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Blocks flow two per band, left then right, for as many exits as actually
    # apply. The template's six slots were placeholders, not a quota.
    row = 5
    for i, e in enumerate(suggested):
        col = 1 if i % 2 == 0 else 4
        if i and col == 1:
            row += 8
        tag = ("Shown on request" if e.get("forced_show")
               else ("Suggestion 1 - main" if i == 0 else f"Suggestion {i + 1}"))
        head = ws.cell(row=row, column=col, value=f"{tag}: {e['name']}")
        head.font = Font(bold=True, size=11,
                         color=GREEN if i == 0 else (NAVY if e["viable"] else GOLD))
        for j, (label, key) in enumerate([("ARV", "arv"), ("Work Needed", "work"),
                                          ("Purchase Price", "purchase"), ("Sell Price", "sell"),
                                          ("Our Profit", "profit")], start=1):
            ws.cell(row=row + j, column=col, value=label)
            v = fmt_rng(e[key])
            if key == "work" and v == 0:
                v = "none, sold as-is"
            c = ws.cell(row=row + j, column=col + 1, value=v)
            if isinstance(v, (int, float)):
                c.number_format = "$#,##0"
            if label == "Our Profit":
                c.font = Font(bold=True, color=GREEN if _mid(e["profit"]) > 0 else RED)
        sens = swing(e["profit"])
        if sens:
            ws.cell(row=row + 6, column=col, value="If it moves")
            sc = ws.cell(row=row + 6, column=col + 1, value=sens)
            sc.font = Font(size=9, color=GREY)
            sc.alignment = Alignment(wrap_text=True, vertical="top")

    r = row + 9
    if ex.get("forced"):
        c = ws.cell(row=r, column=1, value="Nothing cleared its gate at these numbers. The blocks "
                                           "above are the closest misses, not recommendations.")
        c.font = Font(bold=True, color=RED)
        r += 2

    r = _subhead(ws, r, "SUGGESTION LOGIC: why this one, why not the others")
    for line in ex["logic"]:
        ws.cell(row=r, column=1, value=line).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    inp = ex["inputs"]
    r = _subhead(ws, r, "INPUTS BEHIND EVERY BLOCK")
    rows = [
        ("Base ARV (tight set)", fmt_money_rng(inp["base_arv"])),
        ("ARV comp set", inp.get("arv_basis", "")),
        ("Upside ARV (reconfig)", (fmt_money_rng(inp["upside_arv"]) + (
            "" if inp["reconfig_verified"] else "   NOT VERIFIED: excluded from underwriting"))
         if inp.get("upside_arv") else (
             "reconfig ruled out on the walk: target config equals the existing layout"
             if (pack["walk"].get("target_config")
                 and int(pack["walk"]["target_config"].get("beds") or 0) == int(pack["subject"]["beds"])
                 and float(pack["walk"]["target_config"].get("baths") or 0) == float(pack["subject"]["baths"]))
             else "no higher-bed comps in the boundary")),
        ("As-is value", fmt_money_rng(inp["as_is"])),
        ("Wholetail resale", fmt_money_rng(inp["wholetail"])),
        ("Purchase price used", fmt_money_rng(inp["contract"])),
        ("Assignment price used", fmt_money_rng(inp["assign"])),
        ("Selling / holding costs", f"{SELLING_COST_PCT:.0%} of the sale, "
                                    f"{HOLDING_COST_PCT:.0%} of the buy per 6 months held"),
        ("Buyer pool", "cash self-performers and landlords" if inp["shell"]
         else "GC-model flippers still work"),
        ("Why a range", "A single number is a promise. Anything derived from the model (rehab, "
                        "a thin comp band, a derived MAO) prints as a range until a signed bid or "
                        "a signed contract collapses it."),
    ]
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    _widths(ws, [36, 26, 4, 36, 26])



def sheet_comps(wb, pack):
    """The comp table, ranked by a transparent similarity score.

    Columns earn their place. "Date"/"Cash?" from the template were two
    ambiguous date columns; they are replaced by vs Zest (the signal the Bucket
    is derived from, so the condition call is auditable) and Buyer (deed), which
    wires the table into the dispo list. Sim/Grade/In ARV sit up front because
    the first question about any comp is how comparable it actually is.
    """
    subject, sold = pack["subject"], pack["sold"]
    target = pack.get("finished") or subject
    ws = wb.create_sheet("Comps")
    _title(ws, "Comps ranked by similarity to the finished subject",
           f"Scored against the FINISHED {target.get('beds')}bd/{_ba(target.get('baths', 0))}ba at "
           f"{subject.get('sqft', 0):,} sqft. Sorted most similar first. Method and weights are at "
           "the bottom of this sheet.")
    hdr = ["Address", "Sim", "Grade", "In ARV", "Sold", "Sold Price", "Adj. Price", "Bd", "Ba",
           "SqFt", "$/SF", "Dist", "vs Zest", "Buyer (deed)", "Bucket", "Score detail",
           "Role in the package"]
    _header(ws, 3, hdr)

    deeds = pack.get("deeds") or {}
    buckets = pack.get("buckets") or {}
    raw_ids = pack.get("arv_comp_ids") or []
    arv_ids = {str(i) for i in raw_ids} if isinstance(raw_ids, (list, set, tuple)) else set()
    rows, corrected = [], 0
    for l in sold:
        if not isinstance(l, MarketListing) or not l.sqft:
            continue
        bucket = _bucket_of(l, buckets)
        if bucket != classify(l):
            corrected += 1
        dist = _miles(subject.get("lat"), subject.get("lon"), l.latitude, l.longitude)
        sim = comp_similarity(l, subject, bucket, dist, target)
        ratio = round(l.price / l.zestimate, 2) if l.zestimate and l.price else None
        d = deeds.get(_norm_addr(l.address)) or {}
        buyer = d.get("buyer") or ""
        if buyer and d.get("cash"):
            buyer = f"CASH: {buyer}"
        rows.append([l.address, sim["score"], sim["grade"],
                     "YES" if str(l.zpid or l.address) in arv_ids else "",
                     l.sold_date, l.price, adjusted_price(l, subject), l.beds, l.baths,
                     l.sqft, round(l.ppsf) if l.ppsf else None, dist or None, ratio, buyer,
                     bucket, sim["detail"], comp_role(l, subject, bucket, dist)])
    rows.sort(key=lambda r: -r[1])

    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=4 + i, column=j, value=v)
            if j in (6, 7):
                c.number_format = "$#,##0"
            elif j == 12 and v:
                c.number_format = '0.00"mi"'
            elif j == 13 and v:
                c.number_format = "0.00"
            elif j in (16, 17):
                c.alignment = Alignment(wrap_text=True, vertical="top")
        grade = row[2]
        ws.cell(row=4 + i, column=2).font = Font(
            bold=True, color={"A": GREEN, "B": NAVY, "C": GOLD}.get(grade, GREY), size=10)
        ws.cell(row=4 + i, column=3).font = Font(bold=True, size=10,
                                                 color={"A": GREEN, "B": NAVY, "C": GOLD}.get(grade, GREY))
        if row[3] == "YES":
            ws.cell(row=4 + i, column=4).font = Font(bold=True, color=GREEN, size=10)
        color = {"RENOVATED/RETAIL": GREEN, "DISTRESSED": GOLD}.get(row[14])
        if color:
            ws.cell(row=4 + i, column=15).font = Font(bold=True, color=color, size=10)
        if row[13]:
            ws.cell(row=4 + i, column=14).font = Font(color=BLUE, size=10)

    r = 5 + len(rows)
    r = _subhead(ws, r, "HOW THE SIMILARITY SCORE IS BUILT")
    ws.cell(row=r, column=1, value=(
        "Each comp is scored 0-100 on every dimension below, then those are combined as a "
        "weighted average. A dimension with no data is DROPPED and its weight is spread over the "
        "rest, so a comp is never scored on a guess. The per-comp breakdown in Score detail shows "
        "the dimensions that were used and names any that were missing."))
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    r += 2

    _header(ws, r, ["Dimension", "Weight", "Scoring rule", "Why it carries this weight"])
    why = {
        "dist": "Location is the one thing a rehab cannot fix. A sale across a highway is a "
                "different micro-market no matter how well it matches on paper.",
        "size": "A $/sf figure does not survive a large size gap. Pricing a 1,400 sqft house off "
                "a 2,400 sqft sale is the most common way an ARV gets inflated.",
        "bed": "Bedroom count sets the value band, and extra sqft does not lift a house out of "
               "its band.",
        "bath": "Measured on this deal: 3/2s sold a median $345,000 against $280,000-$290,000 for "
                "3/1s at the same size inside the boundary.",
        "cond": "A distressed sale prices the as-is trade, not the renovated exit. Mixing the two "
                "is what pulls an ARV down.",
        "recency": "Recent sales carry current conditions. Older sales still inform, at less weight.",
        "lot": "Real but secondary at this price point, and acreage varies widely on rural parcels.",
    }
    r += 1
    for k, w in sorted(SIM_WEIGHTS.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=w / 100).number_format = "0%"
        ws.cell(row=r, column=3, value=SIM_RULES[k]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4, value=why[k]).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    for line in [
        "Grades: A is 80 and above, B is 65-79, C is 50-64, D is under 50.",
        "In ARV marks the comps that actually set the ARV. Similarity ranks every sale in the "
        "boundary; the ARV uses only those clearing the hard recency, size, bed and bath gates, "
        "so a high-scoring comp can still sit outside it.",
        "Year built is NOT scored. The comp feed carries it 0% of the time, and scoring a field "
        "that is never populated would be scoring a guess.",
        "Adj. Price applies the SiftStack line-item adjustments (bed $5,000, bath $7,500, "
        "$85/sqft) to bring each comp onto the subject. It is a cross-check on the score, not a "
        "substitute for it.",
    ]:
        ws.cell(row=r, column=1, value=line).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    if not deeds:
        ws.cell(row=r, column=1, value="Buyer (deed) is empty because no buyer sweep was passed. "
                                       "Run src/buyer_sweep.py for the zip and pass --buyers.")
    else:
        matched = sum(1 for row in rows if row[13])
        ws.cell(row=r, column=1, value=f"Deed match on {matched} of {len(rows)} comps. The sweep "
                                       "only covers the investor price band, so retail sales above "
                                       "it are blank by design.")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    if corrected:
        ws.cell(row=r + 1, column=1, value=(
            f"{corrected} comp(s) rebucketed off the raw Zestimate ratio: Zillow re-anchors the "
            "Zestimate to a recent sale, so an investor buy reads 1.00 and looks retail. Where the "
            "deed shows an entity or cash buyer and the $/sf sits well under the retail median, "
            "the comp is rebucketed and kept OUT of the ARV set."
        )).alignment = Alignment(wrap_text=True, vertical="top")

    _widths(ws, [26, 6, 7, 7, 11, 12, 12, 5, 5, 7, 7, 8, 8, 28, 18, 40, 58])


def sheet_active(wb, pack):
    ws = wb.create_sheet("Active-Pending")
    _title(ws, "Active listings inside boundary", f"As of {datetime.now():%m/%d/%Y}. "
                                                  "This is what your buyer's exit competes against today.")
    _header(ws, 4, ["Address", "List Price", "Bd", "Ba", "SqFt", "$/SF", "Notes"])
    for i, l in enumerate(pack["active"]):
        if not isinstance(l, MarketListing):
            continue
        note = []
        if l.days_on_zillow:
            note.append(f"{l.days_on_zillow} days on market")
        if l.zestimate and l.price > l.zestimate * 1.1:
            note.append("listed above Zestimate: likely a finished flip resale")
        if l.beds == pack["subject"].get("beds"):
            note.append("same bed count: direct competition for our exit")
        vals = [l.address, l.price, l.beds, l.baths, l.sqft,
                round(l.ppsf) if l.ppsf else None, ". ".join(note)]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=5 + i, column=j, value=v)
            if j == 2:
                c.number_format = "$#,##0"
            if j == 7:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    _widths(ws, [26, 13, 5, 5, 8, 7, 74])


def sheet_repair_logic(wb, pack):
    walk = pack["walk"]
    ws = wb.create_sheet("Repair Logic")
    _title(ws, f"Walkthrough findings{' - ' + walk['walk_date'] if walk.get('walk_date') else ''}",
           walk.get("media", "") + (f" Layout: {walk['layout']}" if walk.get("layout") else ""))

    r = 4
    r = _subhead(ws, r, "WORK ALREADY DONE (real credits)", color=GREEN)
    for item in walk.get("work_done") or []:
        ws.cell(row=r, column=1, value=item.get("item", ""))
        detail = item.get("detail", "")
        if _num(item.get("credit")):
            detail = f"{detail} [credit {_money(_num(item['credit']))}]".strip()
        c = ws.cell(row=r, column=2, value=detail)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if not walk.get("work_done"):
        ws.cell(row=r, column=2, value="Nothing logged. Fill work_done in the walkthrough JSON.")
        r += 1

    r += 1
    r = _subhead(ws, r, "STILL OPEN (debits)", color=GOLD)
    for item in walk.get("still_open") or []:
        ws.cell(row=r, column=1, value=item.get("item", ""))
        c = ws.cell(row=r, column=2, value=item.get("detail", ""))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    flags = walk.get("flags") or []
    r = _subhead(ws, r, f"TEAM WALK FLAGS (priced into the numbers): "
                        f"{_money(sum(_num(f.get('cost')) for f in flags))} total", color=RED)
    for item in flags:
        ws.cell(row=r, column=1, value=item.get("item", ""))
        c = ws.cell(row=r, column=2, value=f"{_money(_num(item.get('cost')))} - {item.get('note', '')}")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    gates = walk.get("gates") or []
    if gates:
        r += 1
        r = _subhead(ws, r, "GATE ITEMS: verify before the contract goes hard", color=RED)
        for g in gates:
            ws.cell(row=r, column=2, value=g).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    _widths(ws, [42, 108])


def sheet_repair_numbers(wb, pack):
    rehab, subject = pack["rehab"], pack["subject"]
    scen = rehab["scenarios"]
    ws = wb.create_sheet("Repair Numbers")
    pct = f"{rehab['soft_pct']:.0%}"

    # ── Left block: category x scenario overview ──
    ws.cell(row=1, column=1, value="Overview").font = Font(bold=True, size=12, color=NAVY)
    _header(ws, 2, ["Category"] + [s["label"] for s in scen])
    r = 3
    for cat in rehab["categories"]:
        ws.cell(row=r, column=1, value=cat)
        for j, s in enumerate(scen, 2):
            v = s["categories"].get(cat)
            if v is not None:
                c = ws.cell(row=r, column=j, value=v)
                c.number_format = "$#,##0;[Red]-$#,##0"
        r += 1
    ws.cell(row=r, column=1, value=f"Permits + contingency ({pct})").font = Font(italic=True)
    for j, s in enumerate(scen, 2):
        ws.cell(row=r, column=j, value=s["soft"]).number_format = "$#,##0"
    r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, color=NAVY)
    for j, s in enumerate(scen, 2):
        c = ws.cell(row=r, column=j, value=s["grand"])
        c.number_format = "$#,##0"
        c.font = Font(bold=True, color=NAVY)
    r += 1
    ws.cell(row=r, column=1, value="$/sqft").font = Font(bold=True)
    for j, s in enumerate(scen, 2):
        ws.cell(row=r, column=j, value=s["ppsf"]).number_format = "$#,##0.00"
    r += 1
    ws.cell(row=r, column=1, value="Timeline (weeks)")
    for j, s in enumerate(scen, 2):
        ws.cell(row=r, column=j, value=s["weeks"])

    # ── Right block: itemized breakdown ──
    ws.cell(row=1, column=7, value="Breakdown").font = Font(bold=True, size=12, color=NAVY)
    for j, h in enumerate(["Category", "Itemized List"] + [s["label"] for s in scen], 7):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)

    br = 3
    placeholders = set(rehab.get("placeholders") or [])
    for cat, item in rehab["lines"]:
        is_ph = f"{cat}|{item}" in placeholders
        ws.cell(row=br, column=7, value=cat)
        ic = ws.cell(row=br, column=8, value=item)
        ic.alignment = Alignment(wrap_text=True, vertical="top")
        if is_ph:
            ic.font = Font(color="C00000")
        for j, s in enumerate(scen, 9):
            v = s["lines"].get((cat, item))
            if v is not None:
                c = ws.cell(row=br, column=j, value=v)
                c.number_format = "$#,##0;[Red]-$#,##0"
                if is_ph:
                    c.font = Font(bold=True, color="C00000")
        br += 1

    br += 1
    sp_factor = rehab.get("sp_factor", SELF_PERFORM_LABOR_FACTOR)
    sp_label = rehab.get("sp_label", "self-perform")
    for label, key, bold in [("Materials", "materials", False), ("Labor", "labor", False),
                             ("Subtotal", "subtotal", True), (f"Soft costs {pct}", "soft", False),
                             ("GRAND TOTAL (GC model)", "grand", True),
                             (f"{sp_label[:1].upper()}{sp_label[1:]} estimate (labor at {sp_factor:.0%})",
                              "self_perform", True)]:
        ws.cell(row=br, column=8, value=label).font = Font(bold=bold, color=NAVY if bold else "000000")
        for j, s in enumerate(scen, 9):
            c = ws.cell(row=br, column=j, value=s[key])
            c.number_format = "$#,##0"
            if bold:
                c.font = Font(bold=True, color=NAVY)
        br += 1

    br += 1
    if rehab.get("placeholders"):
        ws.cell(row=br, column=8, value=(
            "RED = PLACEHOLDER pending a sub bid. A real assumed number, never a blank, so the "
            "total stays true; replace each with the signed quote and re-render."
        )).font = Font(bold=True, color="C00000")
        br += 2
    ws.cell(row=br, column=8, value=(
        f"Engine: SiftStack rehab_estimator, {subject.get('sqft', 0):,} sqft, built "
        f"{subject.get('year_built', '')}, Knoxville regional multiplier. Credits are the work "
        "the seller already paid for; flags are what the team walk found and the model cannot "
        "see. Both are separate rows so every dollar traces back to Repair Logic."
    )).alignment = Alignment(wrap_text=True, vertical="top")

    _widths(ws, [34, 16, 16, 16, 16, 3])
    _widths(ws, [26, 52, 15, 15, 15, 15], start=7)


def sheet_buyers(wb, pack):
    ws = wb.create_sheet("Buyer Targets")
    inp = pack["exits"]["inputs"]
    pool = "cash self-performers and landlords" if inp["shell"] else "flippers first, landlords second"
    _title(ws, "Wholesale buyer targets",
           f"Profile fit: {pool}. Deed-verified from the SiftMap buyer sweep, ranked "
           f"against THIS deal's band ({fmt_money_rng(inp['contract'])} buy to "
           f"{fmt_money_rng(inp['assign'])} ask).")
    _header(ws, 4, ["Buyer", "Type", "Property 1", "Buy/Sell Price", "Property 2", "Buy/Sell Price",
                    "Property 3", "Buy/Sell Price", "Property 4", "Buy/Sell Price", "Fit"])

    rows = pack["buyers"]
    if not rows:
        ws.cell(row=5, column=1, value="No buyer sweep supplied. Run: python src/buyer_sweep.py "
                                       f"--zip {pack['subject'].get('zip', '')} --months 18, then pass --buyers.")
        _widths(ws, [34, 22, 24, 16, 24, 16, 24, 16, 24, 16, 14])
        return

    scored = []
    for b in rows:
        fit = b.get("fit") or buyer_fit(b, inp["as_is_point"], inp["contract_point"],
                                        inp["assign_point"])
        scored.append((fit, -_num(b.get("n_buys") or b.get("purchase_count")), b))
    scored.sort(key=lambda t: (t[0], t[1]))
    cap = pack.get("max_buyers") or len(scored)
    dropped = max(0, len(scored) - cap)
    scored = scored[:cap]

    for i, (fit, _, b) in enumerate(scored):
        name = b.get("buyer") or b.get("buyer_name") or b.get("name", "")
        principal = b.get("principal") or b.get("principal_name")
        if principal:
            name = f"{name} ({principal})"
        vals = [name, buyer_type(b)]
        props = b.get("buys") or b.get("properties") or b.get("purchases") or []
        for k in range(4):
            p = props[k] if k < len(props) else None
            if p is None:
                vals += ["", ""]
            elif isinstance(p, (list, tuple)):          # buyer_sweep: [addr, date, price]
                price = _num(p[2]) if len(p) > 2 else 0
                vals += [p[0] if p else "",
                         f"{_money(price)} {str(p[1])[:7]}" if price else
                         (f"undisclosed {str(p[1])[:7]}" if len(p) > 1 else "")]
            elif isinstance(p, dict):
                buy, sell = _num(p.get("price") or p.get("buy_price")), _num(p.get("sell_price"))
                vals += [p.get("address", ""),
                         f"{_money(buy)}{'/' + _money(sell) if sell else ''}" if buy else ""]
            else:
                vals += [str(p), ""]
        vals.append(fit)
        for j, v in enumerate(vals, 1):
            ws.cell(row=5 + i, column=j, value=v)
        if str(fit).startswith("1"):
            ws.cell(row=5 + i, column=1).font = Font(bold=True, color=GREEN)

    r = 6 + len(scored)
    note = ("Buy prices are deed-verified from the SiftMap sale history, so this is what each "
            "buyer ACTUALLY paid, not what they say they pay. Price per buyer model: a "
            "self-performer clears more rehab than a GC-model flipper and can pay more.")
    if dropped:
        note += (f" Showing the top {cap} of {cap + dropped} swept buyers by deal fit; "
                 f"{dropped} lower-fit buyers are in the buyer_sweep file.")
    ws.cell(row=r, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    _widths(ws, [40, 24, 24, 18, 24, 18, 24, 18, 24, 18, 14])


def _buyer_exit_line(bx, inp, gut, ask) -> str:
    """The buyer's math at the BUYER'S basis: they buy at our ask, and a shell
    buyer self-performs the labor. Quoting our contract-basis profit here would
    show the buyer a number that is not theirs (the 158 rebuild printed a
    negative net off our basis while the self-perform math cleared $30K+)."""
    if not bx:
        return ""
    sell = bx["sell"]["point"]
    if inp.get("shell") and gut:
        work, work_lbl = gut["self_perform"], f"{_money(gut['self_perform'])} self-performed"
    else:
        work, work_lbl = bx["work"]["point"], f"{_money(bx['work']['point'])} on a GC model"
    net = sell - sell * SELLING_COST_PCT - ask - work - ask * HOLDING_COST_PCT
    return (f"{bx['name']} on the BUYER'S basis: buy at {_money(ask)}, {work_lbl}, "
            f"resale {_money(sell)} nets them about {_money(_round_money(net))} after selling "
            "and carry. That is the math to walk them through, not ours.")


def sheet_outreach(wb, pack):
    ws = wb.create_sheet("Outreach Sheet")
    subject, inp, walk = pack["subject"], pack["exits"]["inputs"], pack["walk"]
    _title(ws, f"{subject['short_address']} - Dispo Outreach Sheet (ready to dial)",
           f"Built {datetime.now():%m/%d/%Y}. Sources: Enformion BusinessV2 + Person Search, "
           "Tracerfy batch, web cross-check. Every number Trestle scored.")
    _header(ws, 4, ["Buyer / Entity", "Contact person", "Phone (Trestle score, tier)", "Email",
                    "Mailing address", "Why them / angle", "Confidence"])

    contacts = pack["outreach"]
    r = 5
    for c in contacts:
        vals = [c.get("entity") or c.get("buyer") or "",
                (f"{c.get('name', '')}" + (f" ({c['age']})" if c.get("age") else "")).strip(),
                _phone_cell(c) or "no phone: direct mail / driving contact",
                " / ".join(c.get("emails") or ([c["email"]] if c.get("email") else [])),
                c.get("mailing") or c.get("address", ""),
                c.get("why") or c.get("angle", ""),
                c.get("confidence") or ("HIGH" if c.get("single_source_flag") is False else "MEDIUM")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if not contacts:
        ws.cell(row=r, column=1, value="No skip trace supplied. Run: python src/dispo_skiptrace.py "
                                       "with the buyer contacts, then pass --outreach.")
        r += 1

    r += 1
    bx = pack["exits"].get("buyer_exit") or pack["exits"]["main"]
    ask, floor = inp["assign_point"], _round_money(inp["contract_point"] * 1.15)
    ladder = walk.get("price_ladder") or (
        f"open {_money(ask)} firm-ish | concede to {_money(_round_money(ask * 0.92))} for a 7-day "
        f"close | floor {_money(floor)}, below that we novate or list")
    _scen = pack["rehab"]["scenarios"]
    gut = next((s for s in _scen if s["key"] == "gut_t2"),
               _scen[0] if len(_scen) == 1 else None)
    credits = sum(_num(w.get("credit")) for w in walk.get("work_done") or [])
    pitch = walk.get("pitch") or (
        f"{subject['sqft']:,} sqft {subject.get('beds', '')}/{_ba(subject.get('baths', 0))} on "
        f"{subject['short_address']}. "
        + (f"Seller already put in about {_money(credits)} of work. " if credits else "")
        + f"Retail on this street is {fmt_money_rng(inp['base_arv'])}, as-is trades run "
        f"{fmt_money_rng(inp['as_is'])}. "
        + (f"Full gut runs {fmt_money_rng(bx['work']) if bx else ''} on a GC model and about "
           f"{_money(gut['self_perform'])} if you self-perform." if gut else ""))

    for label, value in [
            ("PITCH SCRIPT ANCHOR", pitch),
            ("PRICE LADDER", ladder),
            ("SEQUENCE", walk.get("sequence") or
             "Day 1 dial every Trestle 80+ line in fit rank 1. Day 2 text follow-up plus the "
             "walkthrough photo set. Day 3 dial rank 2 and 3. Day 5 mail the entity-only buyers. "
             "Day 7 re-price off whoever engaged."),
            ("EXIT WE ARE SELLING INTO", _buyer_exit_line(bx, inp, gut, ask)),
            ("DEPENDENCY", walk.get("dependency") or
             ("Not under contract yet. Lock the seller before the blast goes out."
              if not walk.get("under_contract") else "Under contract. Clock is running."))]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=11, color=BLUE)
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    _widths(ws, [32, 40, 44, 30, 40, 66, 22])


def sheet_lender(wb, pack):
    """Private lender package: sources and uses, exposure vs value, payoff
    waterfall, and the lender's own return. Renders only when the walk carries
    a financing block; the borrower-side numbers tie to Exit Strats."""
    inp = pack["exits"]["inputs"]
    fin = inp.get("financing")
    if not fin:
        return
    rehab = pack["rehab"]
    scen = rehab["scenarios"]
    op = next((s for s in scen if s["key"] == "gut_t2"), scen[0] if scen else None)
    if op is None:
        return

    contract = inp["contract_point"]
    arv = inp["base_arv"]["point"]
    arv_lo = inp["base_arv"]["lo"]
    as_is = inp["as_is_point"]
    sp_label = rehab.get("sp_label", "self-perform")
    budget = op.get("self_perform") or op["grand"]      # the budget the operator runs
    stress_budget = op["grand"]                          # GC-model stress case
    rate, points, term, ltc = fin["rate"], fin["points"], fin["term_months"], fin["ltc"]
    draws = fin.get("draws", 4)

    loan = round((contract + budget) * ltc)
    points_d = round(loan * points / 100)
    interest_full = round(loan * rate * term / 12)
    closing = round(BUY_CLOSE_FLAT + contract * BUY_TITLE_PCT)
    initial_draw = round(contract * ltc)
    per_draw = round((loan - initial_draw) / max(draws, 1))
    total_project = contract + closing + budget + points_d + interest_full
    cash_in = points_d + closing + max(round(contract + budget - loan), 0)

    sell_net = arv * (1 - SELLING_COST_PCT)
    payoff = loan + interest_full
    borrower_net = round(sell_net - payoff - cash_in - round(contract * HOLDING_COST_PCT * term / 6)
                         + (loan - contract - budget if ltc > 1 else 0))
    ann_yield = rate + (points / 100) * (12 / term)

    ws = wb.create_sheet("Lender Analysis")
    _title(ws, "Private lender package",
           f"{pack['subject']['full_address']} | contract {_money(contract)} | ARV {_money(arv)} "
           f"(tight comp set) | budget = {sp_label} plan")
    r = 4
    if fin.get("assumed"):
        c = ws.cell(row=r, column=1, value="TERMS ARE ASSUMED PLACEHOLDERS (red rule): rate, points, "
                                           "term and LTC below are modeling assumptions. Replace with "
                                           "the signed term sheet and re-render.")
        c.font = Font(bold=True, color="C00000")
        r += 2

    r = _subhead(ws, r, "TERMS" + (f" ({fin['lender']})" if fin.get("lender") else ""))
    money_rows: set = set()

    def block(rows):
        nonlocal r
        for label, val, is_money in rows:
            ws.cell(row=r, column=1, value=label)
            c = ws.cell(row=r, column=2, value=val)
            if is_money:
                c.number_format = "$#,##0"
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

    block([
        ("Structure", f"{fin['kind']}, {rate:.0%} interest + {points:g} points, "
                      f"{term}-month term, {ltc:.0%} of cost funded", False),
        ("Draw schedule", f"Initial draw {_money(initial_draw)} at close (purchase), then "
                          f"{draws} rehab draws of ~{_money(per_draw)} released against completed "
                          "work, so exposure tracks value creation", False),
    ])
    r += 1
    r = _subhead(ws, r, "SOURCES AND USES")
    block([
        ("Purchase (contract, signed)", contract, True),
        ("Buy-side closing + title", closing, True),
        (f"Renovation budget ({sp_label})", budget, True),
        ("Points at close", points_d, True),
        (f"Interest, full balance x {term} mo (draws make this cheaper)", interest_full, True),
        ("TOTAL PROJECT COST", total_project, True),
        ("Loan amount", loan, True),
        ("Borrower cash in (points + closing + any cost above the loan)", cash_in, True),
    ])
    r += 1
    r = _subhead(ws, r, "LENDER SECURITY")
    block([
        ("Loan to ARV", f"{loan / arv:.0%}  ({_money(loan)} against {_money(arv)})", False),
        ("Equity cushion at ARV", f"{_money(round(arv - loan))}  ({(arv - loan) / arv:.0%})", False),
        ("Day-one exposure vs as-is value",
         f"{_money(initial_draw)} initial draw against {_money(round(as_is))} as-is: "
         f"{as_is / max(initial_draw, 1):.2f}x covered" if as_is else "no as-is band priced", False),
        ("Full-loan coverage at completion", f"{arv / max(loan, 1):.2f}x at the {_money(arv)} ARV", False),
        ("Stress: sale at the band floor",
         f"{_money(round(arv_lo * (1 - SELLING_COST_PCT)))} net of selling still repays "
         f"{_money(payoff)} principal + interest: "
         f"{arv_lo * (1 - SELLING_COST_PCT) / max(payoff, 1):.2f}x", False),
        ("Stress: GC-model budget", f"Budget at {_money(stress_budget)} (GC retail) moves loan to "
                                    f"{_money(round((contract + stress_budget) * ltc))}: "
                                    f"{(contract + stress_budget) * ltc / arv:.0%} of ARV", False),
    ])
    r += 1
    r = _subhead(ws, r, "PAYOFF WATERFALL AT SALE")
    block([
        ("Sale at base ARV", arv, True),
        (f"Less selling costs ({SELLING_COST_PCT:.0%})", -round(arv * SELLING_COST_PCT), True),
        ("Less loan principal", -loan, True),
        (f"Less accrued interest ({term} mo)", -interest_full, True),
        ("Less borrower cash in + carry", -(cash_in + round(contract * HOLDING_COST_PCT * term / 6)), True),
        ("BORROWER NET (ties to Exit Strats, term-length hold)", borrower_net, True),
        ("", "This waterfall holds the POINT ARV for the FULL term: the conservative case. "
             "Exit Strats mids the comp band and a faster execution, so its profit reads "
             "higher; the truth lives between the two and improves with every week saved.", False),
    ])
    r += 1
    r = _subhead(ws, r, "LENDER RETURN")
    block([
        ("Interest income over term", interest_full, True),
        ("Points income", points_d, True),
        ("Total lender income", interest_full + points_d, True),
        ("Annualized yield on the balance", f"{ann_yield:.1%}", False),
    ])
    _widths(ws, [46, 64])


BUILDERS = [sheet_overview, sheet_exits, sheet_comps, sheet_active,
            sheet_repair_logic, sheet_repair_numbers, sheet_buyers, sheet_outreach,
            sheet_lender]


def build_workbook(pack: dict, out_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    for fn in BUILDERS:
        fn(wb, pack)
    wb.save(out_path)
    return out_path


# ══ 7. Assembly ════════════════════════════════════════════════════════

def assemble(args, spec: dict, walk: dict) -> dict:
    """Pull every source and hand the sheet builders one finished pack."""
    lead = {"available": False, "reason": "--no-sift"}
    if not args.no_sift:
        logger.info("Pulling the live Sift lead (account %s) ...", args.sift_account)
        lead = load_lead(args.address, args.city, args.zip_code, account=args.sift_account)
        logger.info("Sift lead: %s", "linked" if lead["available"] else lead["reason"])

    zdata = None
    if not args.no_api:
        zdata = fetch_subject_property(args.address, args.city, args.state, args.zip_code)

    cf = lead.get("custom_fields") or {}
    subject = {
        "address": args.address,
        "short_address": args.address,
        "full_address": f"{args.address}, {args.city}, {args.state} {args.zip_code}".strip(", "),
        "city": args.city, "state": args.state, "zip": args.zip_code,
        "county": lead.get("county") or cf.get("County", ""),
        # Source order: county-card override on the CLI, then the Sift record
        # (county-sourced), then Zillow. Aggregators get bed counts wrong, so
        # the human override always wins.
        "beds": args.beds or lead.get("beds") or _num(cf.get("Bedrooms")) or (zdata.bedrooms if zdata else 0) or 3,
        "baths": args.baths or lead.get("baths") or _num(cf.get("Bathrooms")) or (zdata.bathrooms if zdata else 0) or 1.0,
        "sqft": int(args.sqft or lead.get("sqft") or _num(cf.get("Living SqFt")) or (zdata.sqft if zdata else 0) or 0),
        "year_built": int(args.year_built or lead.get("year_built") or _num(cf.get("Year Built"))
                          or (zdata.year_built if zdata else 0) or 0),
        "lot": walk.get("lot") or (f"{lead['lot_acres']:.2f} ac" if lead.get("lot_acres")
                                   else (f"{zdata.lot_sqft / 43560:.2f} ac" if zdata and zdata.lot_sqft else "")),
        "zestimate": (zdata.zestimate if zdata else 0) or lead.get("estimated_value", 0),
        "parcel_id": lead.get("parcel_id", ""),
        "lot_acres": _num(lead.get("lot_acres")) or ((zdata.lot_sqft / 43560) if zdata and zdata.lot_sqft else 0.0),
        "lat": (zdata.latitude if zdata else 0.0) or lead.get("lat", 0.0),
        "lon": (zdata.longitude if zdata else 0.0) or lead.get("lon", 0.0),
    }
    # The record already knows the rent; do not make the walker retype it.
    if lead.get("rental_value") and not walk.get("monthly_rent"):
        walk["monthly_rent"] = lead["rental_value"]
    subject["beds"] = int(subject["beds"])
    logger.info("Subject: %s | %s sqft %sbd/%sba built %s", subject["full_address"],
                f"{subject['sqft']:,}", subject["beds"], _ba(subject["baths"]), subject["year_built"])

    deeds = load_deeds(args.buyers)
    sold, active, arv = load_comps(subject, args, spec)

    # Refine the condition buckets BEFORE the ARV runs. dual_track_arv calls
    # classify() internally, so the way to correct it without duplicating its
    # logic is to withhold the comps whose retail label the deed and the $/sf
    # both contradict. They still appear on the Comps sheet, correctly labeled.
    bucketing = bucket_comps(sold, deeds)
    buckets = bucketing["buckets"]
    demoted = [l for l in sold if isinstance(l, MarketListing)
               and classify(l) == "RENOVATED/RETAIL"
               and buckets.get(l.zpid or l.address) != "RENOVATED/RETAIL"]
    if demoted:
        logger.info("Rebucketed %d comps out of the retail set (Zestimate had absorbed the sale): %s",
                    len(demoted), ", ".join(f"{l.address} {_money(l.price)}" for l in demoted[:6]))
    arv_sold = [l for l in sold if l not in demoted]
    # Recompute even when a spec supplied a saved ARV: the refinement above may
    # have demoted comps the saved number still leans on. Comps are already in
    # hand, so the recompute is free; the spec ARV only stands when no comps
    # revived at all.
    if arv_sold:
        arv = dual_track_arv(subject["beds"], subject["sqft"], arv_sold)
    # In ARV marks must survive the fallback path too: when tight_arv finds no
    # qualifying set and the dual-track base stands, its comps are the ARV set.
    if arv:
        arv["comp_ids"] = {str(l.zpid or l.address) for l in arv.pop("base_comps", [])
                           if isinstance(l, MarketListing)}

    # The dual track gives the wide market picture; underwriting uses the TIGHT
    # set (recent, same bed, same size). Overwrite the base track with it so
    # every downstream number comes off one defensible ARV.
    # ARV values the FINISHED product, so it keys off the TARGET config, not
    # what the house is today. Renovating a 3/1 into a 3/2 sells as a 3/2.
    target = walk.get("target_config") or {}
    finished = dict(subject)
    finished["beds"] = int(target.get("beds") or subject["beds"])
    finished["baths"] = float(target.get("baths") or subject["baths"])
    if (finished["beds"], finished["baths"]) != (subject["beds"], subject["baths"]):
        logger.info("ARV priced as the FINISHED %s bd/%s ba, not the current %s/%s",
                    finished["beds"], _ba(finished["baths"]), subject["beds"], _ba(subject["baths"]))
    tight = tight_arv(finished, arv_sold, buckets, months=args.months)
    if tight.get("arv"):
        arv = dict(arv or {})
        arv["base"] = {"arv": tight["arv"], "band": tight.get("band"),
                       "ppsf": tight["ppsf"], "n": tight["n"]}
        arv["tight"] = tight
        arv["comp_ids"] = {str(c.zpid or c.address) for c in tight.get("comps", [])}
        logger.info("Tight ARV %s from %d comps at $%s/sf median%s",
                    _money(tight["arv"]), tight["n"], tight["ppsf"],
                    (" | " + "; ".join(tight["notes"])) if tight["notes"] else "")
    # A walkthrough that explicitly sets the target AT the existing config has
    # ruled the reconfig out (no footprint for it), so the upside track is not
    # a labeled maybe, it is off. Only an explicit target says this; an absent
    # target_config leaves the dual-track upside alone.
    if (target and arv.get("upside")
            and (finished["beds"], finished["baths"]) == (int(subject["beds"]), float(subject["baths"]))):
        arv = dict(arv)
        arv["upside"] = {}
        logger.info("Reconfig track OFF: walk target config equals the existing %s/%s",
                    subject["beds"], _ba(subject["baths"]))

    as_is = as_is_value(sold, subject, walk, buckets, deeds, bucketing.get("retail_floor", 0.0))
    base_arv = (arv.get("base") or {}).get("arv", 0)
    wt = wholetail_value(sold, subject, base_arv, buckets)

    rehab = build_rehab_matrix(subject, walk, soft_pct=args.soft_pct)
    exits = build_exits(subject, arv, rehab, walk, as_is, wt)

    spec_ids = spec.get("arv_comp_ids")
    arv_ids = (arv or {}).get("comp_ids") or (spec_ids if isinstance(spec_ids, list) else [])
    return {"subject": subject, "lead": lead, "walk": walk, "sold": sold, "active": active,
            "arv": arv, "rehab": rehab, "exits": exits,
            "buyers": load_buyers(args.buyers), "outreach": load_outreach(args.outreach),
            "deeds": deeds, "buckets": buckets, "max_buyers": args.max_buyers,
            "finished": finished, "arv_comp_ids": sorted(str(i) for i in arv_ids)}


WALKTHROUGH_TEMPLATE = {
    "_comment": ("Anything you fill in here OVERRIDES the live Sift record, so leave a field "
                 "empty unless the walk proved the record wrong. work_done credits and flags "
                 "flow straight into the Repair Numbers matrix; scenarios defaults to all four."),
    "walk_date": "2026-07-22",
    "media": "24 stills + 7 walkthrough videos.",
    "layout": "main house + attached rear wing",
    "motivation": "",
    "owner": "",
    "acquired": "",
    "distress": "",
    "note": "",
    "lot": "",
    "target_config": {"beds": 3, "baths": 2},
    "reconfig_verified": False,
    "unfinanceable": True,
    "under_contract": False,
    "contract_price": 35000,
    "assignment_price": 75000,
    "as_is_value": 90000,
    "seller_net": 0,
    "monthly_rent": 1400,
    "work_done": [
        {"item": "Drywall", "detail": "Roughly 60% of main house hung and taped",
         "credit": 6000, "scenarios": ["mid", "gut_t2", "gut_t3"]},
        {"item": "Windows", "detail": "11 openings already have new vinyl single-hungs",
         "credit": 4000, "scenarios": ["gut_t2", "gut_t3"]}
    ],
    "still_open": [
        {"item": "Structural", "detail": "Rotted sill plate over masonry, failing parging"},
        {"item": "Plumbing", "detail": "Galvanized supply in crawl: repipe"}
    ],
    "flags": [
        {"item": "Termite: WDO inspection, treatment, framing allowance", "cost": 4000,
         "note": "Suspected at rotted sill, verify before contract",
         "scenarios": ["cosmetic", "mid", "gut_t2", "gut_t3"]},
        {"item": "Crawl moisture package: 20-mil barrier plus sump", "cost": 4000,
         "note": "Crawl wet at the new CMU wall"}
    ],
    "gates": ["Structural engineer eval before the contract goes hard"],
    "price_ladder": "",
    "sequence": "",
    "dependency": "",
    "pitch": ""
}


def main() -> int:
    ap = argparse.ArgumentParser(description="Build the post-walkthrough deal workbook")
    ap.add_argument("--address", help="Subject street address")
    ap.add_argument("--city", default="Knoxville")
    ap.add_argument("--state", default="TN")
    ap.add_argument("--zip", dest="zip_code", default="")
    ap.add_argument("--beds", type=int, help="County card override (beats Zillow and Sift)")
    ap.add_argument("--baths", type=float)
    ap.add_argument("--sqft", type=int)
    ap.add_argument("--year-built", type=int)
    ap.add_argument("--months", type=int, default=18, help="Sold comp lookback (default 18)")
    ap.add_argument("--bbox", help="lat_min,lat_max,lon_min,lon_max boundary box")
    ap.add_argument("--polygon", help="Drawn boundary: 'lat,lon lat,lon ...' or a JSON file of "
                                      "[lat, lon] pairs. Beats a bbox: the pocket is roads, not a box")
    ap.add_argument("--streets", help="Regex of in-boundary street names")
    ap.add_argument("--sold-json", help="Re-use a saved /search sold pull instead of a new API pull")
    ap.add_argument("--walkthrough", help="Walkthrough findings JSON (see --walkthrough-template)")
    ap.add_argument("--buyers", help="buyer_sweep .json/.csv output")
    ap.add_argument("--outreach", help="dispo_skiptrace .json output")
    ap.add_argument("--max-buyers", type=int, default=25,
                    help="Buyer Targets rows, best deal-fit first (default 25, 0 = all)")
    ap.add_argument("--spec", help="Saved pack JSON (re-render without new API pulls)")
    ap.add_argument("--soft-pct", type=float, default=SOFT_COST_PCT,
                    help=f"Permits + contingency rate (default {SOFT_COST_PCT})")
    ap.add_argument("--no-sift", action="store_true", help="Skip the live CRM lookup")
    ap.add_argument("--sift-account", default=DEFAULT_SIFT_ACCOUNT,
                    help=f"reisift_auth account for the lead pull (default {DEFAULT_SIFT_ACCOUNT})")
    ap.add_argument("--no-api", action="store_true", help="Skip Zillow property details")
    ap.add_argument("--save-pack", help="Write the assembled pack JSON here")
    ap.add_argument("--out", help="Output .xlsx path")
    ap.add_argument("--walkthrough-template", action="store_true",
                    help="Write walkthrough_template.json and exit")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s %(message)s")

    if args.walkthrough_template:
        Path("walkthrough_template.json").write_text(
            json.dumps(WALKTHROUGH_TEMPLATE, indent=1), encoding="utf-8")
        print("wrote walkthrough_template.json")
        return 0

    spec = json.load(open(args.spec, encoding="utf-8")) if args.spec else {}
    walk = json.load(open(args.walkthrough, encoding="utf-8")) if args.walkthrough else spec.get("walk", {})
    if not args.address:
        args.address = spec.get("address") or ""
    if not args.address:
        ap.error("--address is required (or supply it in --spec)")

    pack = assemble(args, spec, walk)

    safe = "".join(c if c.isalnum() or c in " -" else "_" for c in args.address)[:44].strip().replace(" ", "_")
    out = args.out or str(Path.cwd() / f"{safe}_Post_Walkthrough.xlsx")
    build_workbook(pack, out)

    if args.save_pack:
        def _json_safe(obj):
            if isinstance(obj, dict):
                return {("|".join(str(p) for p in k) if isinstance(k, tuple) else k): _json_safe(v)
                        for k, v in obj.items()}
            if isinstance(obj, list):
                return [_json_safe(v) for v in obj]
            return obj

        serial = {k: v for k, v in pack.items() if k not in ("sold", "active")}
        serial["sold_comps"] = [vars(l) | {"raw": {}} for l in pack["sold"] if isinstance(l, MarketListing)]
        serial["active_comps"] = [vars(l) | {"raw": {}} for l in pack["active"] if isinstance(l, MarketListing)]
        serial["lead"] = {k: v for k, v in pack["lead"].items() if k != "siftmap"}
        Path(args.save_pack).write_text(
            json.dumps(_json_safe(serial), indent=1, default=str), encoding="utf-8")
        print(f"pack saved: {args.save_pack}")

    e, inp = pack["exits"], pack["exits"]["inputs"]
    print(f"\nARV base {fmt_money_rng(inp['base_arv'])} | upside "
          f"{fmt_money_rng(inp['upside_arv']) or 'n/a'} | as-is {fmt_money_rng(inp['as_is'])}")
    print("Rehab: " + " | ".join(f"{s['label']} {_money(s['grand'])}" for s in pack["rehab"]["scenarios"]))
    if e["main"]:
        print(f"Suggested exits ({len(e['suggested'])} of {len(e['exits'])} scored): "
              + " | ".join(f"{x['name']} {fmt_money_rng(x['profit'])}" for x in e["suggested"]))
    print(f"Workbook: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
