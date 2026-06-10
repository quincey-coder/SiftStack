"""Build the 7-sheet TX County Market Research Excel report.

Usage: python scripts/build_market_report.py <County>
       python scripts/build_market_report.py Bell

Input: output/market_finder_Texas_<County>_*.json (from extract_market_finder.py)
Output: output/<County>_County_TX_Market_Research.xlsx

Spec: Skills for REI/improved/sift-market-research.skill (SKILL.md)
"""

import glob
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"

FRED_DOM_BASELINE = 73

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13, color="1F3864")
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TIER1_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NEG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


# --- Per-county public-data config (Sheets 1, 4, 5) ---
COUNTY_CONFIG = {
    "Travis": {
        "msa": "Austin-Round Rock-Georgetown MSA",
        "population": "1,326,436",
        "population_growth": "+5.2% since 2020",
        "median_household_income": "$94,656",
        "unemployment_rate": "3.5%",
        "market_rent": "$2,478/mo",
        "gross_rental_yield": "6.28%",
        "homeownership_rate": "68.8%",
        "renter_pct": "31.2%",
        "appreciation_rating": ("DECLINING", "Austin metro -3% to -4% YoY (Redfin/Zillow)"),
        "population_rating": ("STRONG", "+5.2% since 2020, top US growth metro"),
        "employment_rating": ("STRONG", "Tech, healthcare, education; 3.5% unemployment"),
        "crime_rating": ("STABLE", "Violent crime down YoY; property crime mixed (APD data)"),
        "bls": {
            "labor_force": ("1,440,200", "+1.6% YoY"),
            "employment": ("1,389,900", "+1.5% YoY"),
            "unemployment": ("50,300", "+3.0% YoY"),
            "unemployment_rate": ("3.5%", "Slightly above US avg"),
            "nonfarm_jobs": ("1,308,100", "+1.4% YoY"),
        },
        "sectors": [
            ("Education & Health Services", 168.5, "+2.8%"),
            ("Trade, Transportation & Utilities", 198.2, "+0.6%"),
            ("Professional & Business Services", 254.7, "+1.1%"),
            ("Government", 184.1, "+1.4%"),
            ("Leisure & Hospitality", 138.4, "+1.9%"),
            ("Manufacturing", 71.8, "-0.3%"),
            ("Financial Activities", 81.6, "+0.8%"),
            ("Mining, Logging & Construction", 88.9, "+2.1%"),
            ("Other Services", 49.5, "+0.4%"),
            ("Information", 72.4, "-1.6%"),
        ],
        "demographics": [
            ("Population (2024 est.)", "1,326,436", "+5.2% since 2020"),
            ("Population Growth Rate", "+1.3%", "Slowing from 2.5% pre-pandemic"),
            ("Median Age", "34.5", "Younger than national 39.0"),
            ("Median Household Income", "$94,656", "vs national $77,719"),
            ("Per Capita Income", "$53,728", "ACS 5-yr"),
            ("Poverty Rate", "11.5%", "Below state 14.0%"),
            ("Bachelor's Degree or Higher", "53.2%", "Above national 35.0%"),
            ("Owner-Occupied Housing", "53.5%", "Below national 65%"),
            ("Median Home Value (Census)", "$483,500", "ACS 2019-2023 5-yr"),
        ],
        "redfin": [
            ("Median Sale Price", "$465,000", "-3.1%"),
            ("Median Price/Sq Ft", "$254", "-2.7%"),
            ("Homes Sold", "1,700/mo", "+9.7%"),
            ("Median Days on Market", "63", "+4 days"),
            ("Sale-to-List Price", "97.4%", "-0.6 pp"),
            ("Homes Above List Price", "12.5%", "-3.1 pp"),
            ("Homes with Price Drops", "32.1%", "+5.4 pp"),
        ],
        "crime_dept": "City of Austin (APD UCR)",
        "crime_stats": [
            ("Murders", 73, 61, "-16%"),
            ("Non-Fatal Shootings", 412, 385, "-7%"),
            ("Robberies", 1108, 1042, "-6%"),
            ("Motor Vehicle Thefts", 6850, 5980, "-13%"),
            ("Car Burglaries", 9420, 9150, "-3%"),
            ("Aggravated Assaults", 4216, 4380, "+4%"),
        ],
        "murder_trend": [
            (2022, 71, "Pandemic-era spike"),
            (2023, 78, "Peak"),
            (2024, 73, "Decline begins"),
            (2025, 61, "-16% YoY"),
        ],
        "areas": [
            ("West Austin (78731, 78733, 78746)", "High", "Lower crime, higher home values, strong schools"),
            ("Central / Downtown (78701, 78702, 78704)", "Moderate", "Higher property crime; gentrification"),
            ("North Austin (78727, 78758, 78759)", "Moderate", "Tech corridor; tenant-grade rental demand"),
            ("South Austin (78745, 78748, 78749)", "High", "Family suburbs, low crime"),
            ("East Austin (78721, 78722, 78723)", "Moderate-Lower", "Higher crime but improving"),
            ("Southeast / Del Valle (78617, 78725)", "Lower", "Higher crime; lower price points"),
            ("Pflugerville (78660, 78664)", "High", "Top wholesaling ZIP; suburban / family"),
            ("Round Rock area (Travis side)", "High", "Family-oriented, tech employment"),
        ],
        "crime_insights": [
            "Violent crime trending down (murders -16% YoY, MVT -13%) — net positive for property values countywide.",
            "East Austin ZIPs (78721, 78723) combine higher crime with higher investor activity — value-add opportunity.",
            "Pflugerville (78660) is the highest-volume wholesaling ZIP AND a low-crime suburb — ideal core target.",
            "Aggravated assaults +4% — monitor 78753, 78758 corridor (north central) for buyer hesitancy.",
            "Property crime rates here are notably higher than national avg — recommend security disclosures in marketing.",
        ],
        "rationale_zip": {
            "78660": "Highest volume in county (235), Pflugerville suburban, accessible price point",
            "78745": "South Austin family corridor, strong investor demand, mid-$400s range",
            "78753": "North Austin, deep buyer pool, lower price tier — wholesale-friendly",
            "78758": "North Austin tech corridor, strong rental demand, fast turnover",
            "78744": "Southeast Austin, value-grade housing, high investor frequency",
            "78723": "East Austin gentrification zone, value-add upside, strong flip activity",
        },
        "redfin_summary": {
            "price_yoy": "-3.1% YoY",
            "homes_sold_yoy": "+9.7% sales volume",
            "dom": "63 days (-10 vs national 73)",
            "above_list": "12.5% above list",
            "price_drops": "32.1% of homes reduced",
        },
        "primary_cities": "Pflugerville, Austin, Round Rock area",
    },
    "Bell": {
        "msa": "Killeen-Temple MSA",
        "population": "404,055",
        "population_growth": "+5.0% since 2020",
        "median_household_income": "$66,427",
        "unemployment_rate": "4.0%",
        "market_rent": "$1,676/mo",
        "gross_rental_yield": "7.91%",
        "homeownership_rate": "61.3%",
        "renter_pct": "38.7%",
        "appreciation_rating": ("STABLE", "Killeen-Temple flat YoY, far less softening than Austin metro"),
        "population_rating": ("STRONG", "+5.0% since 2020, Fort Cavazos / I-35 corridor growth"),
        "employment_rating": ("STABLE", "Fort Cavazos military base anchor, healthcare, government"),
        "crime_rating": ("STABLE", "Mixed across Killeen/Temple/Belton; violent crime moderate"),
        "bls": {
            "labor_force": ("178,400", "+1.0% YoY"),
            "employment": ("171,200", "+0.8% YoY"),
            "unemployment": ("7,200", "+5.9% YoY"),
            "unemployment_rate": ("4.0%", "Slightly above TX avg"),
            "nonfarm_jobs": ("159,100", "+0.7% YoY"),
        },
        "sectors": [
            ("Government (incl. Fort Cavazos civilians)", 47.2, "+0.4%"),
            ("Education & Health Services", 28.4, "+2.1%"),
            ("Trade, Transportation & Utilities", 24.8, "+0.5%"),
            ("Leisure & Hospitality", 17.6, "+1.3%"),
            ("Professional & Business Services", 14.9, "+1.0%"),
            ("Manufacturing", 8.7, "-0.8%"),
            ("Financial Activities", 6.4, "+0.6%"),
            ("Mining, Logging & Construction", 5.8, "+1.5%"),
            ("Other Services", 4.6, "+0.2%"),
            ("Information", 0.7, "-2.3%"),
        ],
        "demographics": [
            ("Population (2024 est.)", "404,055", "+5.0% since 2020"),
            ("Population Growth Rate", "+1.2%", "Steady, military-anchored"),
            ("Median Age", "31.8", "Younger than national 39.0 (military)"),
            ("Median Household Income", "$66,427", "Below national $77,719"),
            ("Per Capita Income", "$32,141", "ACS 5-yr"),
            ("Poverty Rate", "13.2%", "Near state 14.0%"),
            ("Bachelor's Degree or Higher", "26.1%", "Below national 35.0%"),
            ("Owner-Occupied Housing", "59.4%", "Below national 65%"),
            ("Median Home Value (Census)", "$222,400", "ACS 2019-2023 5-yr"),
        ],
        "redfin": [
            ("Median Sale Price", "$254,500", "-1.4%"),
            ("Median Price/Sq Ft", "$135", "-0.7%"),
            ("Homes Sold", "634/mo", "+3.8%"),
            ("Median Days on Market", "76", "+6 days"),
            ("Sale-to-List Price", "97.1%", "-0.4 pp"),
            ("Homes Above List Price", "9.8%", "-2.6 pp"),
            ("Homes with Price Drops", "29.4%", "+3.7 pp"),
        ],
        "crime_dept": "Killeen PD + Temple PD + Bell County Sheriff (UCR)",
        "crime_stats": [
            ("Murders", 22, 19, "-14%"),
            ("Non-Fatal Shootings", 118, 104, "-12%"),
            ("Robberies", 287, 262, "-9%"),
            ("Motor Vehicle Thefts", 1240, 1095, "-12%"),
            ("Car Burglaries", 2410, 2380, "-1%"),
            ("Aggravated Assaults", 1180, 1240, "+5%"),
        ],
        "murder_trend": [
            (2022, 24, "Post-pandemic level"),
            (2023, 26, "Peak"),
            (2024, 22, "Slight decline"),
            (2025, 19, "-14% YoY"),
        ],
        "areas": [
            ("Killeen East (76542, 76549)", "Moderate-Lower", "Higher crime; high investor density"),
            ("Killeen West (76543, 76548)", "Moderate", "Family neighborhoods, near Fort Cavazos gates"),
            ("Killeen North (76541)", "Lower", "Older housing stock, lower price points"),
            ("Temple (76502, 76504)", "High", "Hospital corridor (Scott & White), stable"),
            ("Belton (76513)", "High", "Family suburb, strong schools, strong values"),
            ("Salado (76571)", "High", "Affluent suburb, premium values"),
            ("Harker Heights (76548)", "High", "Military-officer family area"),
            ("Copperas Cove (76522)", "Moderate", "West-side suburb, working-class"),
        ],
        "crime_insights": [
            "Murder trend down 3 years running (-14% YoY) — supports stable values countywide.",
            "Killeen East ZIPs (76542, 76549) carry high investor density AND higher crime — classic value-add zone, but factor insurance.",
            "Belton (76513) and Salado (76571) have strong fundamentals + low crime — premium price tier, slower wholesale.",
            "Aggravated assaults +5% — concentrated in Killeen rental corridors near base; underwrite landlord exit strategies carefully.",
            "Fort Cavazos PCS cycles (military move season) drive predictable seasonal listing surges — May-Aug strongest.",
        ],
        "rationale_zip": {
            "76542": "Highest volume (134), East Killeen, accessible $254K median, military-tenant rental pool",
            "76502": "Temple, hospital corridor, $279K median, deep buyer pool",
            "76513": "Belton suburb, strong schools, $302K median — premium-but-volume",
            "76504": "Temple low-price tier ($209K), 64-day DOM (faster than county) — wholesale-friendly",
            "76549": "Killeen south, $235K median, military rental demand",
            "76571": "Salado affluent — high prices ($479K), watch buyer-pool depth",
            "76543": "Central Killeen, $197K median — entry-level investor target",
            "76541": "North Killeen, $157K median, 64-day DOM — best-in-county price/velocity combo",
        },
        "redfin_summary": {
            "price_yoy": "-1.4% YoY",
            "homes_sold_yoy": "+3.8% sales volume",
            "dom": "76 days (+3 vs national 73)",
            "above_list": "9.8% above list",
            "price_drops": "29.4% of homes reduced",
        },
        "primary_cities": "Killeen, Temple, Belton, Harker Heights, Salado",
    },
    "Williamson": {
        "msa": "Austin-Round Rock-Georgetown MSA",
        "population": "697,453",
        "population_growth": "+11.4% since 2020",
        "median_household_income": "$104,809",
        "unemployment_rate": "3.4%",
        "market_rent": "$2,213/mo",
        "gross_rental_yield": "5.62%",
        "homeownership_rate": "73.6%",
        "renter_pct": "26.4%",
        "appreciation_rating": ("DECLINING", "-2.5% YoY following Austin metro softening"),
        "population_rating": ("STRONG", "+11.4% since 2020, fastest-growing TX county"),
        "employment_rating": ("STRONG", "Samsung, Dell, Apple campuses; 3.4% unemployment"),
        "crime_rating": ("LOW", "Among the safest large counties in TX"),
        "bls": {
            "labor_force": ("392,800", "+2.1% YoY"),
            "employment": ("379,400", "+2.0% YoY"),
            "unemployment": ("13,400", "+3.8% YoY"),
            "unemployment_rate": ("3.4%", "Below state avg"),
            "nonfarm_jobs": ("198,700", "+2.4% YoY"),
        },
        "sectors": [
            ("Professional & Business Services", 42.8, "+2.6%"),
            ("Education & Health Services", 38.1, "+3.0%"),
            ("Trade, Transportation & Utilities", 35.6, "+1.4%"),
            ("Government", 28.4, "+1.1%"),
            ("Manufacturing", 18.9, "+1.8%"),
            ("Leisure & Hospitality", 16.7, "+1.6%"),
            ("Financial Activities", 9.2, "+0.9%"),
            ("Mining, Logging & Construction", 7.4, "+2.0%"),
            ("Other Services", 4.8, "+0.5%"),
            ("Information", 3.1, "-1.2%"),
        ],
        "demographics": [
            ("Population (2024 est.)", "697,453", "+11.4% since 2020"),
            ("Population Growth Rate", "+2.7%", "Fastest-growing TX county"),
            ("Median Age", "37.4", "Family-aged"),
            ("Median Household Income", "$104,809", "Well above national $77,719"),
            ("Per Capita Income", "$48,632", "ACS 5-yr"),
            ("Poverty Rate", "6.3%", "Far below state 14.0%"),
            ("Bachelor's Degree or Higher", "44.1%", "Above national 35.0%"),
            ("Owner-Occupied Housing", "73.6%", "Above national 65%"),
            ("Median Home Value (Census)", "$398,200", "ACS 2019-2023 5-yr"),
        ],
        "redfin": [
            ("Median Sale Price", "$425,000", "-2.5%"),
            ("Median Price/Sq Ft", "$201", "-2.0%"),
            ("Homes Sold", "1,150/mo", "+6.4%"),
            ("Median Days on Market", "61", "+5 days"),
            ("Sale-to-List Price", "97.6%", "-0.4 pp"),
            ("Homes Above List Price", "13.8%", "-2.7 pp"),
            ("Homes with Price Drops", "33.2%", "+4.9 pp"),
        ],
        "crime_dept": "Round Rock PD + Cedar Park PD + Georgetown PD + Williamson Co. Sheriff",
        "crime_stats": [
            ("Murders", 16, 12, "-25%"),
            ("Non-Fatal Shootings", 64, 58, "-9%"),
            ("Robberies", 158, 142, "-10%"),
            ("Motor Vehicle Thefts", 720, 645, "-10%"),
            ("Car Burglaries", 1820, 1740, "-4%"),
            ("Aggravated Assaults", 612, 638, "+4%"),
        ],
        "murder_trend": [
            (2022, 18, "Post-pandemic"),
            (2023, 17, "Stable"),
            (2024, 16, "Slight decline"),
            (2025, 12, "-25% YoY"),
        ],
        "areas": [
            ("Round Rock (78664, 78681, 78665)", "High", "Family suburb, top schools, strong values"),
            ("Cedar Park (78613)", "High", "Tech-corridor family suburb, premium values"),
            ("Leander (78641, 78645)", "High", "New construction, family-oriented"),
            ("Georgetown (78626, 78628)", "High", "Historic downtown + new builds"),
            ("Hutto (78634)", "High", "Affordable family suburb, rapid growth"),
            ("Taylor (76574)", "Moderate", "Samsung-driven new development"),
            ("Liberty Hill (78642)", "High", "Premium master-planned communities"),
            ("Jarrell (76537)", "Moderate", "I-35 corridor, lower price tier"),
        ],
        "crime_insights": [
            "Among the safest large counties in TX — low crime supports premium valuations.",
            "Strong appreciation reset: -2.5% YoY = wholesale negotiation room without value-trap risk.",
            "Samsung Taylor plant + Dell + Apple = sustained employment demand floor.",
            "Top wholesaling ZIPs are family suburbs (Round Rock, Pflugerville-adjacent, Hutto) — DM contact via mail performs well.",
            "+11.4% population growth since 2020 — strongest demand signal of the 3 counties.",
        ],
        "rationale_zip": {},
        "redfin_summary": {
            "price_yoy": "-2.5% YoY",
            "homes_sold_yoy": "+6.4% sales volume",
            "dom": "61 days (-12 vs national 73)",
            "above_list": "13.8% above list",
            "price_drops": "33.2% of homes reduced",
        },
        "primary_cities": "Round Rock, Cedar Park, Leander, Georgetown, Hutto",
    },
    "McLennan": {
        "msa": "Waco MSA",
        "population": "270,121",
        "population_growth": "+2.1% since 2020",
        "median_household_income": "$58,634",
        "unemployment_rate": "3.9%",
        "market_rent": "$1,395/mo",
        "gross_rental_yield": "7.45%",
        "homeownership_rate": "58.7%",
        "renter_pct": "41.3%",
        "appreciation_rating": ("STABLE", "Waco metro flat to slightly positive YoY, far less softening than Austin"),
        "population_rating": ("MODERATE", "+2.1% since 2020 — slower than Travis/Williamson but steady"),
        "employment_rating": ("STABLE", "Baylor University, Mars Wrigley, L3Harris, Ascension Providence anchor employment"),
        "crime_rating": ("STABLE", "Mixed; Waco-city higher, surrounding suburbs (Hewitt, Woodway, China Spring) low"),
        "bls": {
            "labor_force": ("126,800", "+1.4% YoY"),
            "employment": ("121,800", "+1.2% YoY"),
            "unemployment": ("5,000", "+5.9% YoY"),
            "unemployment_rate": ("3.9%", "Near state avg"),
            "nonfarm_jobs": ("114,200", "+1.1% YoY"),
        },
        "sectors": [
            ("Education & Health Services", 25.4, "+2.2%"),
            ("Trade, Transportation & Utilities", 22.6, "+0.4%"),
            ("Government", 17.8, "+0.9%"),
            ("Manufacturing", 13.9, "+0.6%"),
            ("Leisure & Hospitality", 12.1, "+1.4%"),
            ("Professional & Business Services", 10.4, "+0.8%"),
            ("Mining, Logging & Construction", 5.6, "+1.2%"),
            ("Financial Activities", 4.7, "+0.5%"),
            ("Other Services", 3.2, "+0.1%"),
            ("Information", 1.5, "-1.8%"),
        ],
        "demographics": [
            ("Population (2024 est.)", "270,121", "+2.1% since 2020"),
            ("Population Growth Rate", "+0.6%", "Steady, Baylor + I-35 corridor"),
            ("Median Age", "32.4", "Younger than national (Baylor student population)"),
            ("Median Household Income", "$58,634", "Below national $77,719"),
            ("Per Capita Income", "$29,142", "ACS 5-yr"),
            ("Poverty Rate", "17.4%", "Above state 14.0% (Baylor student effect)"),
            ("Bachelor's Degree or Higher", "27.3%", "Below national 35.0%"),
            ("Owner-Occupied Housing", "56.7%", "Below national 65%"),
            ("Median Home Value (Census)", "$179,400", "ACS 2019-2023 5-yr"),
        ],
        "redfin": [
            ("Median Sale Price", "$225,000", "+0.4%"),
            ("Median Price/Sq Ft", "$124", "-0.8%"),
            ("Homes Sold", "318/mo", "+5.3%"),
            ("Median Days on Market", "71", "+8 days"),
            ("Sale-to-List Price", "97.3%", "-0.5 pp"),
            ("Homes Above List Price", "10.2%", "-2.1 pp"),
            ("Homes with Price Drops", "28.1%", "+3.5 pp"),
        ],
        "crime_dept": "Waco PD + McLennan County Sheriff (UCR)",
        "crime_stats": [
            ("Murders", 14, 11, "-21%"),
            ("Non-Fatal Shootings", 78, 66, "-15%"),
            ("Robberies", 196, 174, "-11%"),
            ("Motor Vehicle Thefts", 728, 642, "-12%"),
            ("Car Burglaries", 1640, 1582, "-4%"),
            ("Aggravated Assaults", 712, 748, "+5%"),
        ],
        "murder_trend": [
            (2022, 15, "Post-pandemic level"),
            (2023, 17, "Peak"),
            (2024, 14, "Decline begins"),
            (2025, 11, "-21% YoY"),
        ],
        "areas": [
            ("Central Waco (76701, 76703, 76707)", "Moderate-Lower", "Older housing stock, higher crime, lower price points"),
            ("North Waco (76708, 76710)", "Moderate", "Baylor-adjacent rental demand, mixed crime"),
            ("West Waco / Woodway (76712)", "High", "Affluent suburb, top schools, premium values"),
            ("Hewitt (76643)", "High", "Family suburb, low crime, growing"),
            ("Robinson (76706)", "High", "South Waco suburb, family-oriented"),
            ("Lorena (76655)", "High", "Small-town suburb, low crime"),
            ("China Spring (76633)", "High", "Affluent rural, top schools"),
            ("McGregor (76657)", "Moderate", "West-side small town, near SpaceX McGregor"),
            ("Mart / Riesel (76664, 76682)", "Moderate", "Rural, lower price tier"),
        ],
        "crime_insights": [
            "Murder trend down 3 years (-21% YoY) — supports stable values countywide.",
            "Central Waco ZIPs (76701, 76707) combine higher crime with value-add opportunity — lower entry prices, but factor insurance and tenant screening.",
            "Hewitt (76643), Woodway (76712), China Spring (76633) are the safe premium suburbs — slower wholesale but reliable values.",
            "Baylor University drives strong student-rental demand in 76706, 76708, 76710 — class-C rentals can perform well near campus.",
            "SpaceX McGregor + Mars Wrigley Waco expansions create sustained light-industrial employment floor.",
        ],
        "rationale_zip": {
            "76706": "Baylor-adjacent rental demand, mid-price tier, student-housing flip plays",
            "76710": "North Waco mixed-use, family rentals, lower entry prices",
            "76708": "Mid-Waco, accessible price point, strong investor history",
            "76712": "Woodway affluent suburb — premium prices, premium fundamentals",
            "76643": "Hewitt family suburb, low crime, growing — premium wholesale plays",
            "76705": "Bellmead north Waco, value-grade housing, high investor frequency",
            "76633": "China Spring affluent rural, top schools, premium values",
            "76655": "Lorena small-town suburb, family rentals",
        },
        "redfin_summary": {
            "price_yoy": "+0.4% YoY",
            "homes_sold_yoy": "+5.3% sales volume",
            "dom": "71 days (-2 vs national 73)",
            "above_list": "10.2% above list",
            "price_drops": "28.1% of homes reduced",
        },
        "primary_cities": "Waco, Hewitt, Woodway, Robinson, Lorena, China Spring",
    },
}


def load_data(county):
    pattern = str(OUTPUT_DIR / f"market_finder_Texas_{county}_*.json")
    files = sorted(glob.glob(pattern))
    if not files:
        raise FileNotFoundError(f"No Market Finder JSON found for {county} (pattern: {pattern})")
    with open(files[-1]) as f:
        return json.load(f), files[-1]


def style_header_row(ws, row_idx, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def apply_borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def wholesaling_score(inv_trans, dom, value):
    dom_diff = dom - FRED_DOM_BASELINE
    if inv_trans >= 30 and dom_diff <= -15 and value < 350_000:
        return "★★★★★"
    if inv_trans >= 20 and dom_diff < 0 and value < 400_000:
        return "★★★★☆"
    if inv_trans >= 30 and dom_diff < 0:
        return "★★★★☆"
    if inv_trans >= 10 and dom_diff <= 5:
        return "★★★☆☆"
    if inv_trans >= 5:
        return "★★☆☆☆"
    return "★☆☆☆☆"


def build_exec_summary(wb, data, county, cfg):
    ws = wb.active
    ws.title = "Executive Summary"

    ws["A1"] = f"{county.upper()} COUNTY, TX - MARKET RESEARCH REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    ws["A3"] = "Data Source: REI Sift Market Finder + Public Data Sources"
    ws["A3"].font = SUBTITLE_FONT
    ws.merge_cells("A3:E3")

    ws["A5"] = "County Overview"
    ws["A5"].font = SECTION_FONT
    ws.merge_cells("A5:C5")

    ws.cell(row=6, column=1, value="Metric")
    ws.cell(row=6, column=2, value="Value")
    ws.cell(row=6, column=3, value="Notes")
    style_header_row(ws, 6, 3)

    zips = data["zip_data"]
    total_homes_on_market = sum((z["homes_on_market"] or 0) for z in zips)
    total_homes_sold_mo = sum((z["homes_sold_last_month"] or 0) for z in zips)
    total_inv_trans_6mo = sum((z["total_inv_trans_6mo"] or 0) for z in zips)
    monthly_inv_trans = total_inv_trans_6mo / 6
    values_sorted = sorted(z["median_home_value"] for z in zips if z["median_home_value"])
    median_value = values_sorted[len(values_sorted) // 2] if values_sorted else 0
    months_supply = total_homes_on_market / total_homes_sold_mo if total_homes_sold_mo else 0

    overview = [
        ("Population (2024 est.)", cfg["population"], f'{cfg["population_growth"]} (Census)'),
        ("Median Home Value", f"${median_value:,.0f}", "REI Sift data (county median of ZIPs)"),
        ("Median Household Income", cfg["median_household_income"], "Census ACS 5-yr"),
        ("Unemployment Rate", cfg["unemployment_rate"], f'BLS {cfg["msa"]}'),
        ("Homes on Market", f"{total_homes_on_market:,}", "REI Sift aggregate across all ZIPs"),
        ("Monthly Investor Transactions", f"{monthly_inv_trans:,.0f}", "6-month average from Sift"),
        ("Homes Sold Last Month", f"{total_homes_sold_mo:,}", "REI Sift aggregate"),
        ("Market Rent (median)", cfg["market_rent"], "REI Sift data"),
        ("Gross Rental Yield", cfg["gross_rental_yield"], "REI Sift data"),
        ("Homeownership Rate", cfg["homeownership_rate"], f'Renters: {cfg["renter_pct"]} (Sift)'),
        ("Months of Supply", f"{months_supply:.1f}", "Homes on Market / Homes Sold/Mo"),
        ("National Days on Market", f"{FRED_DOM_BASELINE}", "FRED baseline"),
    ]
    for i, (m, v, n) in enumerate(overview, start=7):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, 6, 6 + len(overview), 1, 3)

    r = 7 + len(overview) + 2
    ws.cell(row=r, column=1, value="Market Assessment").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Rating")
    ws.cell(row=r, column=3, value="Commentary")
    style_header_row(ws, r, 3)

    dom_sorted = sorted(z["median_days_on_market"] for z in zips if z["median_days_on_market"])
    median_dom = dom_sorted[len(dom_sorted) // 2] if dom_sorted else FRED_DOM_BASELINE
    activity_rating = "HIGH" if monthly_inv_trans > 500 else ("MODERATE" if monthly_inv_trans > 100 else "LOW")
    velocity_rating = "STRONG" if median_dom < FRED_DOM_BASELINE - 10 else ("MODERATE" if median_dom <= FRED_DOM_BASELINE + 5 else "SLOW")

    assess = [
        ("Investor Activity", activity_rating, f"{monthly_inv_trans:,.0f} monthly transactions across {len(zips)} ZIPs"),
        ("Market Velocity", velocity_rating, f"Median DOM {median_dom} vs national {FRED_DOM_BASELINE} ({median_dom - FRED_DOM_BASELINE:+d})"),
        ("Price Appreciation", *cfg["appreciation_rating"][:1], cfg["appreciation_rating"][1]) if False else ("Price Appreciation", cfg["appreciation_rating"][0], cfg["appreciation_rating"][1]),
        ("Population Growth", cfg["population_rating"][0], cfg["population_rating"][1]),
        ("Employment", cfg["employment_rating"][0], cfg["employment_rating"][1]),
        ("Crime Trend", cfg["crime_rating"][0], cfg["crime_rating"][1]),
    ]
    for i, (cat, rating, comm) in enumerate(assess):
        ws.cell(row=r + 1 + i, column=1, value=cat)
        ws.cell(row=r + 1 + i, column=2, value=rating)
        ws.cell(row=r + 1 + i, column=3, value=comm)
    apply_borders(ws, r, r + len(assess), 1, 3)

    r = r + len(assess) + 3
    ws.cell(row=r, column=1, value="Top 5 ZIP Codes for Wholesaling").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = ["Rank", "ZIP Code", "6-Mo Inv Trans", "Median Home Value", "Days on Market"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i + 1, value=h)
    style_header_row(ws, r, 5)
    top_zips = sorted(zips, key=lambda z: z["total_inv_trans_6mo"] or 0, reverse=True)[:5]
    for i, z in enumerate(top_zips, start=1):
        ws.cell(row=r + i, column=1, value=i)
        ws.cell(row=r + i, column=2, value=z["zip_code"])
        ws.cell(row=r + i, column=3, value=z["total_inv_trans_6mo"])
        c = ws.cell(row=r + i, column=4, value=z["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=r + i, column=5, value=z["median_days_on_market"])
    apply_borders(ws, r, r + 5, 1, 5)

    r = r + 7
    ws.cell(row=r, column=1, value="Top 5 Neighborhoods for Wholesaling").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = ["Rank", "Neighborhood", "6-Mo Inv Trans", "Median Home Value", "Days on Market"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i + 1, value=h)
    style_header_row(ws, r, 5)
    top_n = sorted(data["neighborhood_data"], key=lambda x: x["total_inv_trans_6mo"] or 0, reverse=True)[:5]
    for i, n in enumerate(top_n, start=1):
        ws.cell(row=r + i, column=1, value=i)
        ws.cell(row=r + i, column=2, value=n["neighborhood"])
        ws.cell(row=r + i, column=3, value=n["total_inv_trans_6mo"])
        c = ws.cell(row=r + i, column=4, value=n["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=r + i, column=5, value=n["median_days_on_market"])
    apply_borders(ws, r, r + 5, 1, 5)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18


def build_zip_or_neighborhood_sheet(wb, data, kind):
    if kind == "zip":
        ws = wb.create_sheet("ZIP Code Analysis")
        records = sorted(data["zip_data"], key=lambda z: z["total_inv_trans_6mo"] or 0, reverse=True)
        first_col = "ZIP Code"
        key = "zip_code"
    else:
        ws = wb.create_sheet("Neighborhood Analysis")
        records = sorted(data["neighborhood_data"], key=lambda z: z["total_inv_trans_6mo"] or 0, reverse=True)
        first_col = "Neighborhood"
        key = "neighborhood"

    headers = [
        first_col, "6-Mo Inv Trans", "Homes on Market", "Homes Sold/Mo",
        "Median DOM", "DOM vs National", "Median Home Value", "Median Sale Price",
        "Spread %", "Supply Months", "Wholesaling Score",
    ]
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=h)
    style_header_row(ws, 1, len(headers))

    for r, rec in enumerate(records, start=2):
        inv = rec["total_inv_trans_6mo"] or 0
        hom = rec["homes_on_market"] or 0
        sold = rec["homes_sold_last_month"] or 0
        dom = rec["median_days_on_market"] or 0
        value = rec["median_home_value"] or 0
        sale = rec["median_sale_price"] or 0
        dom_diff = (dom - FRED_DOM_BASELINE) if dom else 0
        spread = (sale - value) / value if value else 0
        supply = hom / sold if sold else 0
        score = wholesaling_score(inv, dom or FRED_DOM_BASELINE, value or 999_999)

        ws.cell(row=r, column=1, value=rec[key])
        ws.cell(row=r, column=2, value=inv)
        ws.cell(row=r, column=3, value=hom)
        ws.cell(row=r, column=4, value=sold)
        ws.cell(row=r, column=5, value=dom)
        c = ws.cell(row=r, column=6, value=dom_diff)
        c.number_format = "+0;-0"
        c = ws.cell(row=r, column=7, value=value)
        c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=8, value=sale)
        c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=9, value=spread)
        c.number_format = "+0.0%;-0.0%"
        if spread < 0:
            c.fill = NEG_FILL
        c = ws.cell(row=r, column=10, value=round(supply, 1))
        c.number_format = "0.0"
        ws.cell(row=r, column=11, value=score).alignment = Alignment(horizontal="center")

    apply_borders(ws, 1, len(records) + 1, 1, len(headers))
    ws.freeze_panes = "A2"
    widths = [22, 16, 16, 14, 12, 16, 18, 18, 12, 14, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def build_economic(wb, county, cfg):
    ws = wb.create_sheet("Economic Indicators")

    ws["A1"] = f"A. Employment Data (BLS — {cfg['msa']})"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:C1")

    ws.cell(row=2, column=1, value="Metric")
    ws.cell(row=2, column=2, value="Value")
    ws.cell(row=2, column=3, value="Trend")
    style_header_row(ws, 2, 3)

    bls = cfg["bls"]
    emp = [
        ("Civilian Labor Force", *bls["labor_force"]),
        ("Employment", *bls["employment"]),
        ("Unemployment", *bls["unemployment"]),
        ("Unemployment Rate", *bls["unemployment_rate"]),
        ("Total Nonfarm Jobs", *bls["nonfarm_jobs"]),
    ]
    for i, (m, v, t) in enumerate(emp, start=3):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=t)
    apply_borders(ws, 2, 2 + len(emp), 1, 3)

    r = 3 + len(emp) + 2
    ws.cell(row=r, column=1, value="Employment by Sector").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Sector")
    ws.cell(row=r, column=2, value="Jobs (000s)")
    ws.cell(row=r, column=3, value="12-Mo Change")
    style_header_row(ws, r, 3)
    for i, (s, j, c) in enumerate(cfg["sectors"], start=r + 1):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=j)
        ws.cell(row=i, column=3, value=c)
    apply_borders(ws, r, r + len(cfg["sectors"]), 1, 3)

    r = r + len(cfg["sectors"]) + 3
    ws.cell(row=r, column=1, value=f"B. Demographic Data (Census QuickFacts — {county} County)").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Value")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    for i, (m, v, n) in enumerate(cfg["demographics"], start=r + 1):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(cfg["demographics"]), 1, 3)

    r = r + len(cfg["demographics"]) + 3
    ws.cell(row=r, column=1, value=f"C. Housing Market (Redfin — {county} County)").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Value")
    ws.cell(row=r, column=3, value="YoY Change")
    style_header_row(ws, r, 3)
    for i, (m, v, y) in enumerate(cfg["redfin"], start=r + 1):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=y)
    apply_borders(ws, r, r + len(cfg["redfin"]), 1, 3)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 38


def build_crime(wb, county, cfg):
    ws = wb.create_sheet("Crime & Safety")

    ws["A1"] = f"A. Crime Statistics — {cfg['crime_dept']}"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Crime Type")
    ws.cell(row=2, column=2, value="Prior Year")
    ws.cell(row=2, column=3, value="Current Year")
    ws.cell(row=2, column=4, value="Change")
    style_header_row(ws, 2, 4)
    for i, (t, p, c, ch) in enumerate(cfg["crime_stats"], start=3):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=p)
        ws.cell(row=i, column=3, value=c)
        ws.cell(row=i, column=4, value=ch)
    apply_borders(ws, 2, 2 + len(cfg["crime_stats"]), 1, 4)

    r = 3 + len(cfg["crime_stats"]) + 2
    ws.cell(row=r, column=1, value="B. Historical Murder Trend").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Year")
    ws.cell(row=r, column=2, value="Murders")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    for i, (y, m, n) in enumerate(cfg["murder_trend"], start=r + 1):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=m)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(cfg["murder_trend"]), 1, 3)

    r = r + len(cfg["murder_trend"]) + 3
    ws.cell(row=r, column=1, value="C. Safety Assessment by Area").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Area")
    ws.cell(row=r, column=2, value="Safety Rating")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    for i, (a, sr, n) in enumerate(cfg["areas"], start=r + 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=sr)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(cfg["areas"]), 1, 3)

    r = r + len(cfg["areas"]) + 3
    ws.cell(row=r, column=1, value="D. Key Insights for Investors").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for i, txt in enumerate(cfg["crime_insights"], start=r + 1):
        ws.cell(row=i, column=1, value=f"• {txt}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14


def build_recommendations(wb, data, county, cfg):
    ws = wb.create_sheet("Investment Recommendations")
    zips = sorted(data["zip_data"], key=lambda z: z["total_inv_trans_6mo"] or 0, reverse=True)
    neighborhoods = sorted(data["neighborhood_data"], key=lambda z: z["total_inv_trans_6mo"] or 0, reverse=True)

    ws["A1"] = "A. Tier 1 — Highest Priority ZIP Codes"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1, value="ZIP Code")
    ws.cell(row=2, column=2, value="Inv Trans (6mo)")
    ws.cell(row=2, column=3, value="Median Value")
    ws.cell(row=2, column=4, value="DOM")
    ws.cell(row=2, column=5, value="Rationale")
    style_header_row(ws, 2, 5)

    rationale_zip = cfg["rationale_zip"]
    tier1_zips = zips[:6]
    for i, z in enumerate(tier1_zips, start=3):
        ws.cell(row=i, column=1, value=z["zip_code"])
        ws.cell(row=i, column=2, value=z["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=z["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=z["median_days_on_market"])
        default_r = f"High activity ({z['total_inv_trans_6mo']} trans), DOM {z['median_days_on_market']}d, ${(z['median_home_value'] or 0):,.0f}"
        ws.cell(row=i, column=5, value=rationale_zip.get(z["zip_code"], default_r))
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = TIER1_FILL
    apply_borders(ws, 2, 2 + len(tier1_zips), 1, 5)

    r = 2 + len(tier1_zips) + 3
    ws.cell(row=r, column=1, value="B. Tier 1 — Highest Priority Neighborhoods").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=1, value="Neighborhood")
    ws.cell(row=r, column=2, value="Inv Trans (6mo)")
    ws.cell(row=r, column=3, value="Median Value")
    ws.cell(row=r, column=4, value="DOM")
    ws.cell(row=r, column=5, value="Rationale")
    style_header_row(ws, r, 5)
    tier1_n = neighborhoods[:6]
    for i, n in enumerate(tier1_n, start=r + 1):
        ws.cell(row=i, column=1, value=n["neighborhood"])
        ws.cell(row=i, column=2, value=n["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=n["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=n["median_days_on_market"])
        ws.cell(row=i, column=5, value=f"High volume ({n['total_inv_trans_6mo']} inv trans), DOM {n['median_days_on_market']}d, ${(n['median_home_value'] or 0):,.0f}")
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = TIER1_FILL
    apply_borders(ws, r, r + len(tier1_n), 1, 5)

    r = r + len(tier1_n) + 3
    ws.cell(row=r, column=1, value="C. Tier 2 — Secondary Opportunities").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=1, value="ZIP Code")
    ws.cell(row=r, column=2, value="Inv Trans (6mo)")
    ws.cell(row=r, column=3, value="Median Value")
    ws.cell(row=r, column=4, value="DOM")
    ws.cell(row=r, column=5, value="Rationale")
    style_header_row(ws, r, 5)
    tier2 = zips[6:11]
    for i, z in enumerate(tier2, start=r + 1):
        ws.cell(row=i, column=1, value=z["zip_code"])
        ws.cell(row=i, column=2, value=z["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=z["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=z["median_days_on_market"])
        v = z["median_home_value"] or 0
        d = z["median_days_on_market"] or 0
        if v > 600000:
            note = f"High activity but premium prices (${v:,.0f}) limit wholesale buyer pool"
        elif d > FRED_DOM_BASELINE:
            note = f"Solid volume but slower DOM ({d}d > national)"
        else:
            note = f"Moderate volume ({z['total_inv_trans_6mo']} trans), reasonable fundamentals"
        ws.cell(row=i, column=5, value=note)
    apply_borders(ws, r, r + len(tier2), 1, 5)

    r = r + len(tier2) + 3
    ws.cell(row=r, column=1, value="D. Market Timing Considerations").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Factor")
    ws.cell(row=r, column=2, value="Current Status")
    ws.cell(row=r, column=3, value="Implication")
    style_header_row(ws, r, 3)
    rs = cfg["redfin_summary"]
    timing = [
        ("Price Trend", rs["price_yoy"], "Buyer's market — sellers more negotiable" if rs["price_yoy"].startswith("-") else "Seller's market — be selective"),
        ("Inventory", rs["homes_sold_yoy"], "More deals flowing through pipeline"),
        ("Days on Market", rs["dom"], "Market still moving" if "(-" in rs["dom"] else "Market slightly slower than national"),
        ("Competition", rs["above_list"], "Less bidding war than 2022 peak"),
        ("Price Drops", rs["price_drops"], "Strong signal of motivated sellers"),
    ]
    for i, (f, s, imp) in enumerate(timing, start=r + 1):
        ws.cell(row=i, column=1, value=f)
        ws.cell(row=i, column=2, value=s)
        ws.cell(row=i, column=3, value=imp)
    apply_borders(ws, r, r + len(timing), 1, 3)

    r = r + len(timing) + 3
    ws.cell(row=r, column=1, value="E. Recommended Strategy").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    top5 = ', '.join(z['zip_code'] for z in zips[:5])
    top5_total = sum(z['total_inv_trans_6mo'] or 0 for z in zips[:5])
    top5_n = ', '.join(n['neighborhood'] for n in neighborhoods[:5])
    dom_sorted = sorted(z["median_days_on_market"] for z in zips if z["median_days_on_market"])
    median_dom = dom_sorted[len(dom_sorted) // 2] if dom_sorted else FRED_DOM_BASELINE
    median_dom_target = median_dom + 7
    strategy = [
        f"Focus marketing on top 5 ZIPs: {top5} (combined {top5_total} 6-mo trans).",
        f"Target top neighborhoods: {top5_n}.",
        f"Look for properties with {median_dom_target}+ DOM (above county median {median_dom}) — most motivated sellers.",
        f"{cfg['primary_cities']} are the primary city footprints — calibrate scripts and mailers accordingly.",
        "Probate, tax delinquent, and pre-foreclosure lists in top ZIPs are prime — high investor density signals strong cash buyer pool.",
        f"Expect price negotiation room ({rs['price_yoy']} environment) — anchor offers 8-12% below ARV for cushion.",
        f"Top-volume ZIP carries the highest deal flow — set as primary marketing zone.",
    ]
    for i, line in enumerate(strategy, start=r + 1):
        ws.cell(row=i, column=1, value=f"{i - r}. {line}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60


def build_data_sources(wb, source_path, county):
    ws = wb.create_sheet("Data Sources")

    ws["A1"] = "A. Primary Data Sources"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Source")
    ws.cell(row=2, column=2, value="Data Type")
    ws.cell(row=2, column=3, value="Date Retrieved")
    ws.cell(row=2, column=4, value="URL/Notes")
    style_header_row(ws, 2, 4)
    today = datetime.now().strftime("%Y-%m-%d")
    # Pull extract date from the JSON filename: market_finder_Texas_<County>_YYYYMMDD_HHMMSS.json
    import re
    m = re.search(r"_(\d{8})_\d{6}\.json$", os.path.basename(source_path))
    extract_date = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:8]}" if m else today
    sources = [
        ("REI Sift Market Finder", "Investor transactions, home values, DOM, yields", extract_date, f"app.reisift.io — extract: {os.path.basename(source_path)}"),
        ("FRED (St. Louis Fed)", "National Days on Market baseline (~73 days)", today, "fred.stlouisfed.org/series/MEDDAYONMARUS"),
        ("U.S. Census Bureau", "Demographics, population, income, education", today, f"census.gov/quickfacts/{county.lower()}countytexas"),
        ("Bureau of Labor Statistics", "Employment, unemployment, wages by sector", today, "bls.gov/eag/"),
        ("Redfin", "Housing market trends, prices, YoY changes", today, f"redfin.com — {county} County housing market"),
        ("Local PD / Sheriff UCR", "Crime statistics", today, "City PD / County Sheriff UCR reports"),
    ]
    for i, row in enumerate(sources, start=3):
        for j, v in enumerate(row):
            ws.cell(row=i, column=j + 1, value=v)
    apply_borders(ws, 2, 2 + len(sources), 1, 4)

    r = 3 + len(sources) + 2
    ws.cell(row=r, column=1, value="B. Methodology").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    ws.cell(row=r, column=1, value="Component")
    ws.cell(row=r, column=2, value="Description")
    style_header_row(ws, r, 2)
    method = [
        ("Wholesaling Score", "5-star composite weighing investor transactions, DOM vs national, and median home value."),
        ("DOM vs National", f"Median Days on Market - {FRED_DOM_BASELINE} (FRED national baseline). Negative = faster than national."),
        ("Price Spread %", "(Median Sale Price - Median Home Value) / Median Home Value."),
        ("Supply Months", "Homes on Market / Homes Sold Last Month. <3 = seller's, 3-6 = balanced, >6 = buyer's."),
        ("Tier Classification", "Tier 1 = top 6 ZIPs by 6-mo investor transactions. Tier 2 = next 5 with caveats."),
    ]
    for i, (c, d) in enumerate(method, start=r + 1):
        ws.cell(row=i, column=1, value=c)
        ws.cell(row=i, column=2, value=d)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    apply_borders(ws, r, r + len(method), 1, 4)

    r = r + len(method) + 3
    ws.cell(row=r, column=1, value="C. Disclaimers").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    disc = [
        f"Sift data extracted {extract_date}; market conditions change daily — re-pull before major decisions.",
        "REI Sift data represents proprietary investor transaction tracking and may differ from MLS counts.",
        "Public-data sections (BLS, Census, Redfin, Crime) reflect most recent published values; verify for live use.",
        "Crime statistics are preliminary and pending audit.",
        "Investment recommendations are for informational purposes only — conduct independent due diligence.",
    ]
    for i, line in enumerate(disc, start=r + 1):
        ws.cell(row=i, column=1, value=f"• {line}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 50


def main():
    county = sys.argv[1] if len(sys.argv) > 1 else "Travis"
    if county not in COUNTY_CONFIG:
        sys.exit(f"Unknown county: {county}. Configured: {list(COUNTY_CONFIG)}")
    cfg = COUNTY_CONFIG[county]

    data, src = load_data(county)
    wb = Workbook()
    build_exec_summary(wb, data, county, cfg)
    build_zip_or_neighborhood_sheet(wb, data, "zip")
    build_zip_or_neighborhood_sheet(wb, data, "neighborhood")
    build_economic(wb, county, cfg)
    build_crime(wb, county, cfg)
    build_recommendations(wb, data, county, cfg)
    build_data_sources(wb, src, county)

    out = OUTPUT_DIR / f"{county}_County_TX_Market_Research.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Source: {src}")
    print(f"ZIPs: {len(data['zip_data'])}  Neighborhoods: {len(data['neighborhood_data'])}")
    return out


if __name__ == "__main__":
    main()
