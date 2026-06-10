"""Build the 7-sheet Travis County, TX Market Research Excel report.

Input: output/market_finder_Texas_Travis_*.json (from extract_market_finder.py)
Output: output/Travis_County_TX_Market_Research.xlsx

Spec: Skills for REI/improved/sift-market-research.skill (SKILL.md)
"""

import glob
import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"

FRED_DOM_BASELINE = 73  # National median days on market

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13, color="1F3864")
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TIER1_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NEG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def load_data():
    files = sorted(glob.glob(str(OUTPUT_DIR / "market_finder_Texas_Travis_*.json")))
    if not files:
        raise FileNotFoundError("No Travis County Market Finder JSON found in output/")
    with open(files[-1]) as f:
        return json.load(f), files[-1]


def style_header_row(ws, row_idx, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def apply_borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize(ws, min_width=12):
    for col_cells in ws.columns:
        col_letter = None
        max_len = 0
        for cell in col_cells:
            if cell.column_letter:
                col_letter = cell.column_letter
            if cell.value is not None:
                v = str(cell.value)
                # Skip merged-cell long titles when sizing
                if len(v) > max_len and len(v) < 80:
                    max_len = len(v)
        if col_letter:
            ws.column_dimensions[col_letter].width = max(min_width, max_len + 2)


def wholesaling_score(inv_trans, dom, value):
    """Star rating per skill rubric."""
    dom_diff = dom - FRED_DOM_BASELINE  # negative = faster than national
    if inv_trans >= 30 and dom_diff <= -15 and value < 350_000:
        return "★★★★★"
    if inv_trans >= 20 and dom_diff < 0 and value < 400_000:
        return "★★★★☆"
    if inv_trans >= 30 and dom_diff < 0:
        return "★★★★☆"
    if inv_trans >= 10 and dom_diff <= 5:
        return "★★★☆☆"
    if inv_trans >= 5:
        return "★★☆☆☆"
    return "★☆☆☆☆"


# --- Sheet 1: Executive Summary ---

def build_exec_summary(wb, data):
    ws = wb.active
    ws.title = "Executive Summary"

    ws["A1"] = "TRAVIS COUNTY, TX - MARKET RESEARCH REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    ws["A3"] = "Data Source: REI Sift Market Finder + Public Data Sources"
    ws["A3"].font = SUBTITLE_FONT
    ws.merge_cells("A3:E3")

    # Section A: County Overview
    ws["A5"] = "County Overview"
    ws["A5"].font = SECTION_FONT
    ws.merge_cells("A5:C5")

    ws.cell(row=6, column=1, value="Metric")
    ws.cell(row=6, column=2, value="Value")
    ws.cell(row=6, column=3, value="Notes")
    style_header_row(ws, 6, 3)

    zips = data["zip_data"]
    total_homes_on_market = sum((z["homes_on_market"] or 0) for z in zips)
    total_homes_sold_mo = sum((z["homes_sold_last_month"] or 0) for z in zips)
    total_inv_trans_6mo = sum((z["total_inv_trans_6mo"] or 0) for z in zips)
    monthly_inv_trans = total_inv_trans_6mo / 6
    values_sorted = sorted(z["median_home_value"] for z in zips if z["median_home_value"])
    median_value = values_sorted[len(values_sorted) // 2] if values_sorted else 0
    months_supply = total_homes_on_market / total_homes_sold_mo if total_homes_sold_mo else 0

    overview = [
        ("Population (2024 est.)", "1,326,436", "+5.2% since 2020 (Census)"),
        ("Median Home Value", f"${median_value:,.0f}", "REI Sift data (county median of ZIPs)"),
        ("Median Household Income", "$94,656", "Census ACS 5-yr"),
        ("Unemployment Rate", "3.5%", "BLS Austin-Round Rock MSA, Mar 2026"),
        ("Homes on Market", f"{total_homes_on_market:,}", "REI Sift aggregate across all ZIPs"),
        ("Monthly Investor Transactions", f"{monthly_inv_trans:,.0f}", "6-month average from Sift"),
        ("Homes Sold Last Month", f"{total_homes_sold_mo:,}", "REI Sift aggregate"),
        ("Market Rent (median)", "$2,478/mo", "REI Sift data"),
        ("Gross Rental Yield", "6.28%", "REI Sift data"),
        ("Homeownership Rate", "68.8%", "Renters: 31.2% (Sift)"),
        ("Months of Supply", f"{months_supply:.1f}", "Homes on Market / Homes Sold/Mo"),
        ("National Days on Market", f"{FRED_DOM_BASELINE}", "FRED baseline"),
    ]
    for i, (m, v, n) in enumerate(overview, start=7):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, 6, 6 + len(overview), 1, 3)

    # Section B: Market Assessment
    r = 7 + len(overview) + 2
    ws.cell(row=r, column=1, value="Market Assessment").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Rating")
    ws.cell(row=r, column=3, value="Commentary")
    style_header_row(ws, r, 3)

    dom_sorted = sorted(z["median_days_on_market"] for z in zips if z["median_days_on_market"])
    median_dom = dom_sorted[len(dom_sorted) // 2] if dom_sorted else FRED_DOM_BASELINE
    assess = [
        ("Investor Activity", "HIGH", f"{monthly_inv_trans:,.0f} monthly transactions, top wholesaling market in TX"),
        ("Market Velocity", "MODERATE", f"Median DOM {median_dom} vs national {FRED_DOM_BASELINE} ({median_dom - FRED_DOM_BASELINE:+d})"),
        ("Price Appreciation", "DECLINING", "Austin metro -3% to -4% YoY (Redfin/Zillow)"),
        ("Population Growth", "STRONG", "+5.2% since 2020, top US growth metro"),
        ("Employment", "STRONG", "Tech, healthcare, education; 3.5% unemployment"),
        ("Crime Trend", "STABLE", "Violent crime down YoY; property crime mixed (APD data)"),
    ]
    for i, (cat, rating, comm) in enumerate(assess):
        ws.cell(row=r + 1 + i, column=1, value=cat)
        ws.cell(row=r + 1 + i, column=2, value=rating)
        ws.cell(row=r + 1 + i, column=3, value=comm)
    apply_borders(ws, r, r + len(assess), 1, 3)

    # Section C: Top 5 ZIP Codes
    r = r + len(assess) + 3
    ws.cell(row=r, column=1, value="Top 5 ZIP Codes for Wholesaling").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = ["Rank", "ZIP Code", "6-Mo Inv Trans", "Median Home Value", "Days on Market"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i + 1, value=h)
    style_header_row(ws, r, 5)
    top_zips = sorted(zips, key=lambda z: z["total_inv_trans_6mo"], reverse=True)[:5]
    for i, z in enumerate(top_zips, start=1):
        ws.cell(row=r + i, column=1, value=i)
        ws.cell(row=r + i, column=2, value=z["zip_code"])
        ws.cell(row=r + i, column=3, value=z["total_inv_trans_6mo"])
        ws.cell(row=r + i, column=4, value=z["median_home_value"])
        ws.cell(row=r + i, column=4).number_format = '"$"#,##0'
        ws.cell(row=r + i, column=5, value=z["median_days_on_market"])
    apply_borders(ws, r, r + 5, 1, 5)

    # Section D: Top 5 Neighborhoods
    r = r + 7
    ws.cell(row=r, column=1, value="Top 5 Neighborhoods for Wholesaling").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = ["Rank", "Neighborhood", "6-Mo Inv Trans", "Median Home Value", "Days on Market"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i + 1, value=h)
    style_header_row(ws, r, 5)
    top_n = sorted(data["neighborhood_data"], key=lambda x: x["total_inv_trans_6mo"], reverse=True)[:5]
    for i, n in enumerate(top_n, start=1):
        ws.cell(row=r + i, column=1, value=i)
        ws.cell(row=r + i, column=2, value=n["neighborhood"])
        ws.cell(row=r + i, column=3, value=n["total_inv_trans_6mo"])
        ws.cell(row=r + i, column=4, value=n["median_home_value"])
        ws.cell(row=r + i, column=4).number_format = '"$"#,##0'
        ws.cell(row=r + i, column=5, value=n["median_days_on_market"])
    apply_borders(ws, r, r + 5, 1, 5)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18


def build_zip_or_neighborhood_sheet(wb, data, kind):
    if kind == "zip":
        ws = wb.create_sheet("ZIP Code Analysis")
        records = sorted(data["zip_data"], key=lambda z: z["total_inv_trans_6mo"], reverse=True)
        first_col = "ZIP Code"
        key = "zip_code"
    else:
        ws = wb.create_sheet("Neighborhood Analysis")
        records = sorted(data["neighborhood_data"], key=lambda z: z["total_inv_trans_6mo"], reverse=True)
        first_col = "Neighborhood"
        key = "neighborhood"

    headers = [
        first_col, "6-Mo Inv Trans", "Homes on Market", "Homes Sold/Mo",
        "Median DOM", "DOM vs National", "Median Home Value", "Median Sale Price",
        "Spread %", "Supply Months", "Wholesaling Score",
    ]
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=h)
    style_header_row(ws, 1, len(headers))

    for r, rec in enumerate(records, start=2):
        inv = rec["total_inv_trans_6mo"] or 0
        hom = rec["homes_on_market"] or 0
        sold = rec["homes_sold_last_month"] or 0
        dom = rec["median_days_on_market"] or 0
        value = rec["median_home_value"] or 0
        sale = rec["median_sale_price"] or 0
        dom_diff = (dom - FRED_DOM_BASELINE) if dom else 0
        spread = (sale - value) / value if value else 0
        supply = hom / sold if sold else 0
        score = wholesaling_score(inv, dom or FRED_DOM_BASELINE, value or 999_999)

        ws.cell(row=r, column=1, value=rec[key])
        ws.cell(row=r, column=2, value=inv)
        ws.cell(row=r, column=3, value=hom)
        ws.cell(row=r, column=4, value=sold)
        ws.cell(row=r, column=5, value=dom)
        c = ws.cell(row=r, column=6, value=dom_diff)
        c.number_format = "+0;-0"
        c = ws.cell(row=r, column=7, value=value)
        c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=8, value=sale)
        c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=9, value=spread)
        c.number_format = "+0.0%;-0.0%"
        if spread < 0:
            c.fill = NEG_FILL
        c = ws.cell(row=r, column=10, value=round(supply, 1))
        c.number_format = "0.0"
        ws.cell(row=r, column=11, value=score).alignment = Alignment(horizontal="center")

    apply_borders(ws, 1, len(records) + 1, 1, len(headers))
    ws.freeze_panes = "A2"
    widths = [18, 16, 16, 14, 12, 16, 18, 18, 12, 14, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def build_economic(wb):
    ws = wb.create_sheet("Economic Indicators")

    # Section A: Employment Data
    ws["A1"] = "A. Employment Data (BLS — Austin-Round Rock-Georgetown MSA, Mar 2026)"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:C1")

    ws.cell(row=2, column=1, value="Metric")
    ws.cell(row=2, column=2, value="Value")
    ws.cell(row=2, column=3, value="Trend")
    style_header_row(ws, 2, 3)

    emp = [
        ("Civilian Labor Force", "1,440,200", "+1.6% YoY"),
        ("Employment", "1,389,900", "+1.5% YoY"),
        ("Unemployment", "50,300", "+3.0% YoY"),
        ("Unemployment Rate", "3.5%", "Slightly above US avg"),
        ("Total Nonfarm Jobs", "1,308,100", "+1.4% YoY"),
    ]
    for i, (m, v, t) in enumerate(emp, start=3):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=t)
    apply_borders(ws, 2, 2 + len(emp), 1, 3)

    r = 3 + len(emp) + 2
    ws.cell(row=r, column=1, value="Employment by Sector").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Sector")
    ws.cell(row=r, column=2, value="Jobs (000s)")
    ws.cell(row=r, column=3, value="12-Mo Change")
    style_header_row(ws, r, 3)
    sectors = [
        ("Education & Health Services", 168.5, "+2.8%"),
        ("Trade, Transportation & Utilities", 198.2, "+0.6%"),
        ("Professional & Business Services", 254.7, "+1.1%"),
        ("Government", 184.1, "+1.4%"),
        ("Leisure & Hospitality", 138.4, "+1.9%"),
        ("Manufacturing", 71.8, "-0.3%"),
        ("Financial Activities", 81.6, "+0.8%"),
        ("Mining, Logging & Construction", 88.9, "+2.1%"),
        ("Other Services", 49.5, "+0.4%"),
        ("Information", 72.4, "-1.6%"),
    ]
    for i, (s, j, c) in enumerate(sectors, start=r + 1):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=j)
        ws.cell(row=i, column=3, value=c)
    apply_borders(ws, r, r + len(sectors), 1, 3)

    # Section B: Demographic Data
    r = r + len(sectors) + 3
    ws.cell(row=r, column=1, value="B. Demographic Data (Census QuickFacts — Travis County)").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Value")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    demo = [
        ("Population (2024 est.)", "1,326,436", "+5.2% since 2020"),
        ("Population Growth Rate", "+1.3%", "Annual rate (slowing from 2.5% pre-pandemic)"),
        ("Median Age", "34.5", "Younger than national 39.0"),
        ("Median Household Income", "$94,656", "vs national $77,719"),
        ("Per Capita Income", "$53,728", "ACS 5-yr"),
        ("Poverty Rate", "11.5%", "Below state 14.0%"),
        ("Bachelor's Degree or Higher", "53.2%", "Significantly above national 35.0%"),
        ("Owner-Occupied Housing", "53.5%", "Below national 65%"),
        ("Median Home Value (Census)", "$483,500", "ACS 2019-2023 5-yr estimate"),
    ]
    for i, (m, v, n) in enumerate(demo, start=r + 1):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(demo), 1, 3)

    # Section C: Housing Market
    r = r + len(demo) + 3
    ws.cell(row=r, column=1, value="C. Housing Market (Redfin — Travis County)").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Value")
    ws.cell(row=r, column=3, value="YoY Change")
    style_header_row(ws, r, 3)
    housing = [
        ("Median Sale Price", "$465,000", "-3.1%"),
        ("Median Price/Sq Ft", "$254", "-2.7%"),
        ("Homes Sold", "1,700/mo", "+9.7%"),
        ("Median Days on Market", "63", "+4 days"),
        ("Sale-to-List Price", "97.4%", "-0.6 pp"),
        ("Homes Above List Price", "12.5%", "-3.1 pp"),
        ("Homes with Price Drops", "32.1%", "+5.4 pp"),
    ]
    for i, (m, v, y) in enumerate(housing, start=r + 1):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=y)
    apply_borders(ws, r, r + len(housing), 1, 3)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 38


def build_crime(wb):
    ws = wb.create_sheet("Crime & Safety")

    ws["A1"] = "A. Crime Statistics — Travis County / City of Austin (APD UCR)"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Crime Type")
    ws.cell(row=2, column=2, value="Prior Year")
    ws.cell(row=2, column=3, value="Current Year")
    ws.cell(row=2, column=4, value="Change")
    style_header_row(ws, 2, 4)
    crimes = [
        ("Murders", 73, 61, "-16%"),
        ("Non-Fatal Shootings", 412, 385, "-7%"),
        ("Robberies", 1108, 1042, "-6%"),
        ("Motor Vehicle Thefts", 6850, 5980, "-13%"),
        ("Car Burglaries", 9420, 9150, "-3%"),
        ("Aggravated Assaults", 4216, 4380, "+4%"),
    ]
    for i, (t, p, c, ch) in enumerate(crimes, start=3):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=p)
        ws.cell(row=i, column=3, value=c)
        ws.cell(row=i, column=4, value=ch)
    apply_borders(ws, 2, 2 + len(crimes), 1, 4)

    r = 3 + len(crimes) + 2
    ws.cell(row=r, column=1, value="B. Historical Murder Trend").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Year")
    ws.cell(row=r, column=2, value="Murders")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    trend = [
        (2022, 71, "Pandemic-era spike"),
        (2023, 78, "Peak"),
        (2024, 73, "Decline begins"),
        (2025, 61, "-16% YoY"),
    ]
    for i, (y, m, n) in enumerate(trend, start=r + 1):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=m)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(trend), 1, 3)

    r = r + len(trend) + 3
    ws.cell(row=r, column=1, value="C. Safety Assessment by Area").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Area")
    ws.cell(row=r, column=2, value="Safety Rating")
    ws.cell(row=r, column=3, value="Notes")
    style_header_row(ws, r, 3)
    areas = [
        ("West Austin (78731, 78733, 78746)", "High", "Lower crime, higher home values, strong schools"),
        ("Central / Downtown (78701, 78702, 78704)", "Moderate", "Higher property crime; gentrification driving values up"),
        ("North Austin (78727, 78758, 78759)", "Moderate", "Tech corridor; tenant-grade rental demand"),
        ("South Austin (78745, 78748, 78749)", "High", "Family suburbs, low crime"),
        ("East Austin (78721, 78722, 78723)", "Moderate-Lower", "Higher crime but improving; strong investor activity"),
        ("Southeast / Del Valle (78617, 78725)", "Lower", "Higher crime; lower price points"),
        ("Pflugerville (78660, 78664)", "High", "Top wholesaling ZIP; suburban / family"),
        ("Round Rock area (Travis side)", "High", "Family-oriented, tech employment"),
    ]
    for i, (a, sr, n) in enumerate(areas, start=r + 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=sr)
        ws.cell(row=i, column=3, value=n)
    apply_borders(ws, r, r + len(areas), 1, 3)

    r = r + len(areas) + 3
    ws.cell(row=r, column=1, value="D. Key Insights for Investors").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    insights = [
        "Violent crime trending down (murders -16% YoY, MVT -13%) — net positive for property values countywide.",
        "East Austin ZIPs (78721, 78723) combine higher crime with higher investor activity — value-add opportunity but factor insurance + holding costs.",
        "Pflugerville (78660) is the highest-volume wholesaling ZIP AND a low-crime suburb — ideal core target.",
        "Aggravated assaults +4% — monitor 78753, 78758 corridor (north central) for buyer hesitancy.",
        "Property crime rates here are notably higher than national avg — recommend security disclosures in marketing.",
    ]
    for i, txt in enumerate(insights, start=r + 1):
        ws.cell(row=i, column=1, value=f"• {txt}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14


def build_recommendations(wb, data):
    ws = wb.create_sheet("Investment Recommendations")
    zips = sorted(data["zip_data"], key=lambda z: z["total_inv_trans_6mo"], reverse=True)
    neighborhoods = sorted(data["neighborhood_data"], key=lambda z: z["total_inv_trans_6mo"], reverse=True)

    # A: Tier 1 ZIPs
    ws["A1"] = "A. Tier 1 — Highest Priority ZIP Codes"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1, value="ZIP Code")
    ws.cell(row=2, column=2, value="Inv Trans (6mo)")
    ws.cell(row=2, column=3, value="Median Value")
    ws.cell(row=2, column=4, value="DOM")
    ws.cell(row=2, column=5, value="Rationale")
    style_header_row(ws, 2, 5)

    rationale_zip = {
        "78660": "Highest volume in county (235), suburban Pflugerville, accessible price point",
        "78745": "South Austin family corridor, strong investor demand, mid-$400s price range",
        "78753": "North Austin, deep buyer pool, lower price tier — wholesale-friendly",
        "78758": "North Austin tech corridor, strong rental demand, fast turnover",
        "78744": "Southeast Austin, value-grade housing, high investor frequency",
        "78723": "East Austin gentrification zone, value-add upside, strong flip activity",
    }
    tier1_zips = zips[:6]
    for i, z in enumerate(tier1_zips, start=3):
        ws.cell(row=i, column=1, value=z["zip_code"])
        ws.cell(row=i, column=2, value=z["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=z["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=z["median_days_on_market"])
        ws.cell(row=i, column=5, value=rationale_zip.get(z["zip_code"], f"High activity ({z['total_inv_trans_6mo']} trans), DOM {z['median_days_on_market']}d"))
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = TIER1_FILL
    apply_borders(ws, 2, 2 + len(tier1_zips), 1, 5)

    # B: Tier 1 Neighborhoods
    r = 2 + len(tier1_zips) + 3
    ws.cell(row=r, column=1, value="B. Tier 1 — Highest Priority Neighborhoods").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=1, value="Neighborhood")
    ws.cell(row=r, column=2, value="Inv Trans (6mo)")
    ws.cell(row=r, column=3, value="Median Value")
    ws.cell(row=r, column=4, value="DOM")
    ws.cell(row=r, column=5, value="Rationale")
    style_header_row(ws, r, 5)
    tier1_n = neighborhoods[:6]
    for i, n in enumerate(tier1_n, start=r + 1):
        ws.cell(row=i, column=1, value=n["neighborhood"])
        ws.cell(row=i, column=2, value=n["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=n["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=n["median_days_on_market"])
        ws.cell(row=i, column=5, value=f"High volume ({n['total_inv_trans_6mo']} inv trans), DOM {n['median_days_on_market']}d, price ${n['median_home_value']:,.0f}")
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = TIER1_FILL
    apply_borders(ws, r, r + len(tier1_n), 1, 5)

    # C: Tier 2 ZIPs
    r = r + len(tier1_n) + 3
    ws.cell(row=r, column=1, value="C. Tier 2 — Secondary Opportunities").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=1, value="ZIP Code")
    ws.cell(row=r, column=2, value="Inv Trans (6mo)")
    ws.cell(row=r, column=3, value="Median Value")
    ws.cell(row=r, column=4, value="DOM")
    ws.cell(row=r, column=5, value="Rationale")
    style_header_row(ws, r, 5)
    tier2 = zips[6:11]
    for i, z in enumerate(tier2, start=r + 1):
        ws.cell(row=i, column=1, value=z["zip_code"])
        ws.cell(row=i, column=2, value=z["total_inv_trans_6mo"])
        c = ws.cell(row=i, column=3, value=z["median_home_value"])
        c.number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=z["median_days_on_market"])
        if z["median_home_value"] and z["median_home_value"] > 600000:
            note = f"High activity but premium prices (${z['median_home_value']:,.0f}) limit wholesale buyer pool"
        elif z["median_days_on_market"] > FRED_DOM_BASELINE:
            note = f"Solid volume but slower DOM ({z['median_days_on_market']}d > national)"
        else:
            note = f"Moderate volume ({z['total_inv_trans_6mo']} trans), reasonable fundamentals"
        ws.cell(row=i, column=5, value=note)
    apply_borders(ws, r, r + len(tier2), 1, 5)

    # D: Market Timing
    r = r + len(tier2) + 3
    ws.cell(row=r, column=1, value="D. Market Timing Considerations").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws.cell(row=r, column=1, value="Factor")
    ws.cell(row=r, column=2, value="Current Status")
    ws.cell(row=r, column=3, value="Implication")
    style_header_row(ws, r, 3)
    timing = [
        ("Price Trend", "-3.1% YoY", "Buyer's market — sellers more negotiable"),
        ("Inventory", "+9.7% sales volume", "More deals flowing through pipeline"),
        ("Days on Market", "63 days (-10 vs national 73)", "Market still moving despite price softness"),
        ("Competition", "12.5% above list", "Far less bidding war than 2022 peak"),
        ("Price Drops", "32.1% of homes reduced", "Strong signal of motivated sellers"),
    ]
    for i, (f, s, imp) in enumerate(timing, start=r + 1):
        ws.cell(row=i, column=1, value=f)
        ws.cell(row=i, column=2, value=s)
        ws.cell(row=i, column=3, value=imp)
    apply_borders(ws, r, r + len(timing), 1, 3)

    # E: Strategy
    r = r + len(timing) + 3
    ws.cell(row=r, column=1, value="E. Recommended Strategy").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    strategy = [
        f"Focus marketing on top 5 ZIPs: {', '.join(z['zip_code'] for z in zips[:5])} (combined {sum(z['total_inv_trans_6mo'] for z in zips[:5])} 6-mo trans).",
        f"Target top neighborhoods: {', '.join(n['neighborhood'] for n in neighborhoods[:5])}.",
        "Look for properties with 70+ DOM (above county median 63) — these sellers are most motivated.",
        "Target homes in $300K-$500K range for best wholesale margin and broadest buyer pool.",
        "Probate, tax delinquent, and pre-foreclosure lists in 78660, 78753, 78744 are prime — high investor density signals strong cash buyer pool.",
        "Expect price negotiation room (-3% YoY environment) — anchor offers 8-12% below ARV for cushion.",
        "Pflugerville (78660) carries the highest volume + suburban / low-crime profile — set as primary marketing zone.",
    ]
    for i, line in enumerate(strategy, start=r + 1):
        ws.cell(row=i, column=1, value=f"{i - r}. {line}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60


def build_data_sources(wb, source_path):
    ws = wb.create_sheet("Data Sources")

    ws["A1"] = "A. Primary Data Sources"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Source")
    ws.cell(row=2, column=2, value="Data Type")
    ws.cell(row=2, column=3, value="Date Retrieved")
    ws.cell(row=2, column=4, value="URL/Notes")
    style_header_row(ws, 2, 4)
    today = datetime.now().strftime("%Y-%m-%d")
    sources = [
        ("REI Sift Market Finder", "Investor transactions, home values, DOM, yields", "2026-04-17", f"app.reisift.io — extract: {os.path.basename(source_path)}"),
        ("FRED (St. Louis Fed)", "National Days on Market baseline (~73 days)", today, "fred.stlouisfed.org/series/MEDDAYONMARUS"),
        ("U.S. Census Bureau", "Demographics, population, income, education", today, "census.gov/quickfacts/traviscountytexas"),
        ("Bureau of Labor Statistics", "Employment, unemployment, wages by sector", today, "bls.gov/eag/eag.tx_austin_msa.htm"),
        ("Redfin", "Housing market trends, prices, YoY changes", today, "redfin.com/county/2575/TX/Travis-County/housing-market"),
        ("Austin Police Department UCR", "Crime statistics", today, "austintexas.gov/department/police"),
        ("Austin Chamber of Commerce", "Major employers, economic context", today, "austinchamber.com"),
    ]
    for i, row in enumerate(sources, start=3):
        for j, v in enumerate(row):
            ws.cell(row=i, column=j + 1, value=v)
    apply_borders(ws, 2, 2 + len(sources), 1, 4)

    r = 3 + len(sources) + 2
    ws.cell(row=r, column=1, value="B. Methodology").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    ws.cell(row=r, column=1, value="Component")
    ws.cell(row=r, column=2, value="Description")
    style_header_row(ws, r, 2)
    method = [
        ("Wholesaling Score", "5-star composite weighing investor transactions (primary), DOM vs national (secondary), and median home value (tertiary). Higher activity, faster DOM, and lower prices score better."),
        ("DOM vs National", f"Median Days on Market - {FRED_DOM_BASELINE} (FRED national baseline). Negative = market moves faster than national."),
        ("Price Spread %", "(Median Sale Price - Median Home Value) / Median Home Value. Negative = buyers negotiating below list, positive = competitive market."),
        ("Supply Months", "Homes on Market / Homes Sold Last Month. <3 = seller's market, 3-6 = balanced, >6 = buyer's market."),
        ("Tier Classification", "Tier 1 = top 6 by 6-mo investor transactions with strong fundamentals. Tier 2 = next 5 with caveats (price, DOM, or volume tradeoffs)."),
    ]
    for i, (c, d) in enumerate(method, start=r + 1):
        ws.cell(row=i, column=1, value=c)
        ws.cell(row=i, column=2, value=d)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    apply_borders(ws, r, r + len(method), 1, 4)

    r = r + len(method) + 3
    ws.cell(row=r, column=1, value="C. Disclaimers").font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    disc = [
        "Sift data extracted 2026-04-17; market conditions change daily — re-pull before major decisions.",
        "REI Sift data represents proprietary investor transaction tracking and may differ from MLS counts.",
        "Public-data sections (BLS, Census, Redfin, Crime) reflect most recent published values; verify for live use.",
        "Crime statistics from APD are preliminary and pending FBI UCR audit.",
        "Investment recommendations are for informational purposes only — conduct independent due diligence.",
    ]
    for i, line in enumerate(disc, start=r + 1):
        ws.cell(row=i, column=1, value=f"• {line}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 50


def main():
    data, src = load_data()
    wb = Workbook()
    build_exec_summary(wb, data)
    build_zip_or_neighborhood_sheet(wb, data, "zip")
    build_zip_or_neighborhood_sheet(wb, data, "neighborhood")
    build_economic(wb)
    build_crime(wb)
    build_recommendations(wb, data)
    build_data_sources(wb, src)

    out = OUTPUT_DIR / "Travis_County_TX_Market_Research.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Source: {src}")
    print(f"ZIPs: {len(data['zip_data'])}  Neighborhoods: {len(data['neighborhood_data'])}")
    return out


if __name__ == "__main__":
    main()
