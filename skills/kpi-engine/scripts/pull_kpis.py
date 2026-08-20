"""
pull_kpis.py - self-contained DataSift (REISift) KPI puller for the kpi-engine skill.

Reads YOUR account's per-record activity log and produces a graded KPI report:
dials, answer/conversation/contact rates, correct numbers, dispositions, leads
(including new_lead statuses), talk time, per caller and per day, with funnel
pacing. Markdown + CSV always; Excel with --xlsx (needs openpyxl); optional
Slack digest with --slack <webhook-url>.

AUTH (your own login, ~48h token):
  app.reisift.io -> DevTools (F12) -> Network -> any apiv2.reisift.io request ->
  copy the authorization Bearer JWT. Then either:
    set REISIFT_TOKEN=<jwt>          (env var), or
    save it in reisift_token.txt next to this script.

USAGE
  python pull_kpis.py --days 7
  python pull_kpis.py --from 2026-07-06 --to 2026-07-16 --xlsx
  python pull_kpis.py --days 1 --slack https://hooks.slack.com/services/...
  python pull_kpis.py --show-benchmarks
  python pull_kpis.py --days 7 --tz America/Chicago

Benchmarks: defaults below; override any key in a benchmarks.json next to this
script (add your admin logins to excluded_callers, custom lead statuses, etc.).
Read-only against DataSift. Standard library only.
"""
from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from zoneinfo import ZoneInfo

HERE = Path(__file__).resolve().parent
API = "https://apiv2.reisift.io"
HEADERS_BASE = {
    "accept": "application/json, text/plain, */*",
    "origin": "https://app.reisift.io",
    "referer": "https://app.reisift.io/",
    "x-reisift-ui-version": "2022.02.01.7",
    "user-agent": "Mozilla/5.0",
}

DEFAULT_BENCHMARKS = {
    "dials_floor_per_caller": 150,
    "dials_target_per_caller": 200,
    "conversations_floor_per_caller": 5,
    "conversation_min_seconds": 60,
    "meaningful_conversation_min_seconds": 120,
    "voicemail_max_seconds": 30,
    "dials_per_correct_scored": 9,
    "dials_per_correct_blind": 32,
    "correct_numbers_per_deal": 100,
    "leads_per_caller_day": [2, 3],
    "leads_per_contract": [15, 20],
    "appointment_take_rate": 0.25,
    "lead_statuses": ["Cold Lead", "Warm Lead", "Hot Lead", "new_lead", "New Lead",
                      "No Contact New Lead", "Nurture New Lead"],
    "excluded_callers": [],
}

CORRECT_STATES = {"CORRECT", "CORRECT_DNC"}
WRONG_STATES = {"WRONG", "WRONG_DNC"}
CALL_EVENTS = {"owner.call.made", "owner.call.answered", "owner.call.noanswer",
               "owner.call.received", "owner.call.missed"}


def log(msg: str) -> None:
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}", file=sys.stderr, flush=True)


def load_benchmarks() -> dict:
    bench = dict(DEFAULT_BENCHMARKS)
    f = HERE / "benchmarks.json"
    if f.exists():
        bench.update(json.loads(f.read_text(encoding="utf-8")))
    return bench


def get_token() -> str:
    tok = os.environ.get("REISIFT_TOKEN", "").strip()
    if not tok:
        f = HERE / "reisift_token.txt"
        if f.exists():
            tok = f.read_text(encoding="utf-8").strip()
    if not tok:
        sys.exit("No token. Set REISIFT_TOKEN or create reisift_token.txt (see script docstring).")
    return tok.removeprefix("Bearer ").strip()


def req(token: str, path: str, *, method="GET", body=None, method_override=None, params=None):
    url = API + path
    if params:
        url += ("&" if "?" in url else "?") + urllib.parse.urlencode(params)
    headers = dict(HEADERS_BASE)
    headers["authorization"] = f"Bearer {token}"
    if body is not None:
        headers["content-type"] = "application/json"
    if method_override:
        headers["x-http-method-override"] = method_override
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(url, data=data, headers=headers, method=method)
    time.sleep(0.03)
    try:
        with urllib.request.urlopen(r, timeout=30) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        if e.code == 401:
            sys.exit("Token expired or invalid (HTTP 401). Grab a fresh one from DevTools.")
        raise


def search_updated(token: str, day_from: str, day_to_excl: str) -> list[dict]:
    out, offset, limit = [], 0, 200
    while True:
        body = {"limit": limit, "offset": offset, "ordering": "-list_count",
                "query": {"must": {"property_type": "clean", "updated": [day_from, day_to_excl]}}}
        r = req(token, "/api/internal/property/", method="POST", body=body, method_override="GET")
        rows = r.get("results") or r.get("data") or []
        out.extend(rows)
        total = r.get("count", len(out))
        offset += limit
        if offset >= total or not rows or offset > 20000:
            break
    return out


def get_logs(token: str, uuid: str) -> list[dict]:
    r = req(token, f"/api/internal/property/{uuid}/logs/", params={"limit": 250, "offset": 0})
    return r.get("results") or r.get("data") or []


# ---- event helpers ----
def local_dt(ts: str, tz):
    try:
        return (datetime.datetime.strptime(ts[:19], "%Y-%m-%d %H:%M:%S")
                .replace(tzinfo=ZoneInfo("UTC")).astimezone(tz))
    except Exception:
        return None


def caller_of(ev):
    call = (ev.get("payload") or {}).get("call") or {}
    if call.get("direction") == "inbound":
        return ("inbound", "Inbound")
    eu = call.get("external_user") or {}
    return (eu.get("email") or "unknown", eu.get("name") or eu.get("email") or "Unknown")


def author_of(ev):
    info = ev.get("author_extra_info") or {}
    email = info.get("email") or ev.get("author") or "system"
    return (email, (f"{info.get('first_name', '')} {info.get('last_name', '')}".strip() or email))


def detect_new_index(events, key):
    """Status payloads carry [a,b] with no guaranteed order; vote by chaining."""
    by_tgt = defaultdict(list)
    for ev in events:
        obj = (ev.get("payload") or {}).get("owner" if key == "phone" else "property")
        if not isinstance(obj, dict):
            continue
        pair = obj.get("status")
        tgt = obj.get("phone") if key == "phone" else ev.get("resource")
        if isinstance(pair, list) and len(pair) == 2 and tgt:
            by_tgt[tgt].append((ev.get("timestamp", ""), pair))
    old_new = new_old = 0
    for seq in by_tgt.values():
        seq.sort()
        for (_, e1), (_, e2) in zip(seq, seq[1:]):
            if e1[1] == e2[0] and e1[1] != e2[1]:
                old_new += 1
            elif e1[0] == e2[1] and e1[0] != e2[0]:
                new_old += 1
    return 0 if new_old > old_new else 1


def new_status(ev, key, idx):
    pair = ((ev.get("payload") or {}).get("owner" if key == "phone" else "property") or {}).get("status")
    if isinstance(pair, list) and len(pair) == 2:
        return pair[idx]
    return pair[-1] if isinstance(pair, list) and pair else None


def blank():
    return {"dials": 0, "answered": 0, "noanswer": 0, "talk_seconds": 0, "sms_sent": 0,
            "sms_received": 0, "conversations": 0, "meaningful_conversations": 0,
            "band_vm": 0, "band_brief": 0, "correct_numbers": 0, "wrong_numbers": 0,
            "dead_numbers": 0, "dnc_numbers": 0, "leads": 0, "not_interested": 0,
            "follow_ups": 0, "appointments": 0, "records": set(), "days": set(),
            "first_call": None, "last_call": None}


def pull(token: str, day_from: str, day_to: str, tz, bench: dict) -> dict:
    lead_set = {s.lower() for s in bench["lead_statuses"]}  # case-insensitive: reisift mixes casings
    excluded = {e.lower() for e in bench["excluded_callers"]}
    conv_s, mean_s, vm_s = (bench["conversation_min_seconds"],
                            bench["meaningful_conversation_min_seconds"],
                            bench["voicemail_max_seconds"])
    to_excl = (datetime.date.fromisoformat(day_to) + datetime.timedelta(days=1)).isoformat()
    cand = search_updated(token, day_from, to_excl)
    log(f"{len(cand)} candidate records updated in window; fetching activity logs...")

    rec_events, rec_meta = {}, {}
    for i, r in enumerate(cand, 1):
        uuid = r.get("uuid")
        if not uuid:
            continue
        try:
            evs = get_logs(token, uuid)
        except Exception:
            continue
        keep = []
        for e in evs:
            dt = local_dt(e.get("timestamp", ""), tz)
            if dt and day_from <= dt.date().isoformat() <= day_to:
                keep.append((dt, e))
        if keep:
            rec_events[uuid] = keep
            rec_meta[uuid] = {"address": r.get("address") or "", "status": r.get("status") or ""}
        if i % 50 == 0:
            log(f"  {i}/{len(cand)} scanned...")
    log(f"{len(rec_events)} records had activity in window")

    all_phone = [e for evs in rec_events.values() for _, e in evs
                 if e.get("event_type") == "owner.phone.status.updated"]
    all_prop = [e for evs in rec_events.values() for _, e in evs
                if e.get("event_type") == "property.status.updated"]
    p_idx = detect_new_index(all_phone, "phone")
    s_idx = detect_new_index(all_prop, "property")

    acct, per, daily = blank(), defaultdict(blank), defaultdict(blank)
    names, phone_final, prop_final, seen = {}, {}, {}, set()

    def bump(email, day, field, amt=1):
        acct[field] += amt
        per[email][field] += amt
        daily[day][field] += amt

    for uuid, evs in rec_events.items():
        for dt, ev in evs:
            et, day = ev.get("event_type"), dt.date().isoformat()
            if et in CALL_EVENTS or et in ("owner.sms.sent", "owner.sms.received"):
                o = (ev.get("payload") or {}).get("call") or (ev.get("payload") or {}).get("sms") or {}
                k = o.get("uuid") or o.get("external_id")
                if k is not None:
                    if (et, k) in seen:
                        continue
                    seen.add((et, k))
            if et in CALL_EVENTS:
                email, name = caller_of(ev)
                names[email] = name
                call = (ev.get("payload") or {}).get("call") or {}
                if et == "owner.call.made":
                    bump(email, day, "dials")
                    for scope in (acct, per[email]):
                        scope["records"].add(uuid)
                        scope["days"].add(day)
                        if scope["first_call"] is None or dt < scope["first_call"]:
                            scope["first_call"] = dt
                        if scope["last_call"] is None or dt > scope["last_call"]:
                            scope["last_call"] = dt
                elif et == "owner.call.answered":
                    bump(email, day, "answered")
                    dur = int(call.get("duration") or 0)
                    bump(email, day, "talk_seconds", dur)
                    if dur >= mean_s:
                        bump(email, day, "meaningful_conversations")
                        bump(email, day, "conversations")
                    elif dur >= conv_s:
                        bump(email, day, "conversations")
                    elif dur >= vm_s:
                        bump(email, day, "band_brief")
                    else:
                        bump(email, day, "band_vm")
                elif et == "owner.call.noanswer":
                    bump(email, day, "noanswer")
            elif et == "owner.sms.sent":
                eu = ((ev.get("payload") or {}).get("sms") or {}).get("external_user") or {}
                bump(eu.get("email") or author_of(ev)[0], day, "sms_sent")
            elif et == "owner.sms.received":
                acct["sms_received"] += 1
            elif et == "owner.phone.status.updated":
                phone = ((ev.get("payload") or {}).get("owner") or {}).get("phone")
                ns = new_status(ev, "phone", p_idx)
                if phone and ns:
                    prev = phone_final.get(phone)
                    if prev is None or dt >= prev[0]:
                        phone_final[phone] = (dt, ns, author_of(ev)[0])
            elif et == "property.status.updated":
                ns = new_status(ev, "property", s_idx)
                if ns:
                    prev = prop_final.get(uuid)
                    if prev is None or dt >= prev[0]:
                        prop_final[uuid] = (dt, ns, author_of(ev)[0])
            elif et == "task.completed":
                title = (((ev.get("payload") or {}).get("task") or {}).get("title") or "").lower()
                if any(k in title for k in ("appoint", "appt", "meeting", "consult")):
                    bump(author_of(ev)[0], day, "appointments")

    for phone, (dt, ns, email) in phone_final.items():
        day = dt.date().isoformat()
        if ns in CORRECT_STATES:
            bump(email, day, "correct_numbers")
        elif ns in WRONG_STATES:
            bump(email, day, "wrong_numbers")
        elif ns == "DEAD":
            bump(email, day, "dead_numbers")
        elif ns == "DNC":
            bump(email, day, "dnc_numbers")
    for uuid, (dt, ns, email) in prop_final.items():
        day = dt.date().isoformat()
        if ns.lower() in lead_set:
            bump(email, day, "leads")
        elif ns == "not_interested":
            bump(email, day, "not_interested")
        elif ns == "follow_up":
            bump(email, day, "follow_ups")

    # pull excluded/admin callers out of rollups
    for email in list(per):
        if email.lower() in excluded or email in ("inbound", "system", "unknown"):
            c = per.pop(email)
            for k, v in c.items():
                if isinstance(v, int):
                    acct[k] = max(0, acct[k] - v)

    # record-level detail (one row per worked record)
    details = []
    for uuid, evs in rec_events.items():
        d = {"address": rec_meta[uuid]["address"], "status": rec_meta[uuid]["status"],
             "dials": 0, "answered": 0, "talk_seconds": 0, "callers": set(),
             "final_phone_dispo": "", "status_set_to": "", "status_set_by": "",
             "is_lead": False, "last_activity": ""}
        for dt, ev in sorted(evs, key=lambda x: x[0]):
            et = ev.get("event_type")
            d["last_activity"] = dt.strftime("%Y-%m-%d %H:%M")
            if et == "owner.call.made":
                d["dials"] += 1
                d["callers"].add(caller_of(ev)[1])
            elif et == "owner.call.answered":
                d["answered"] += 1
                d["talk_seconds"] += int(((ev.get("payload") or {}).get("call") or {}).get("duration") or 0)
        if uuid in prop_final:
            _, ns, email = prop_final[uuid]
            d["status_set_to"] = ns
            d["status_set_by"] = names.get(email, email)
            d["is_lead"] = ns.lower() in lead_set
        details.append(d)
    details.sort(key=lambda d: (-d["is_lead"], -d["dials"]))
    return {"from": day_from, "to": day_to, "account_totals": acct,
            "callers": dict(per), "daily": dict(daily), "names": names,
            "records_detail": details}


# ---- rendering ----
def fmt_hms(sec):
    h, r = divmod(int(sec), 3600)
    m, s = divmod(r, 60)
    return f"{h}h{m:02d}m" if h else f"{m}m{s:02d}s"


def pct(n, d):
    return f"{n / d * 100:.1f}%" if d else "0.0%"


def render_md(res: dict, bench: dict) -> str:
    a = res["account_totals"]
    lead_lo, lead_hi = bench["leads_per_caller_day"]
    c_lo, c_hi = bench["leads_per_contract"]
    take = bench["appointment_take_rate"]
    dials = a["dials"]
    out = [f"# KPI Report, {res['from']} to {res['to']}", "",
           f"- Dials: {dials}  |  Answered: {a['answered']} ({pct(a['answered'], dials)})",
           f"- Conversations 60s+: {a['conversations']} ({pct(a['conversations'], dials)})  |  "
           f"Meaningful 120s+: {a['meaningful_conversations']}",
           f"- Correct numbers: {a['correct_numbers']} ({pct(a['correct_numbers'], dials)} right-party)"
           f"  |  Wrong {a['wrong_numbers']}  Dead {a['dead_numbers']}  DNC {a['dnc_numbers']}",
           f"- Leads: {a['leads']}  |  Not interested: {a['not_interested']}  |  "
           f"Follow-ups: {a['follow_ups']}  |  Appointments logged: {a['appointments']}",
           f"- Talk time: {fmt_hms(a['talk_seconds'])}  |  Texts: {a['sms_sent']} out / {a['sms_received']} in",
           "", "## By caller", "",
           "| Caller | Days | Dials | Dials/day | Ans% | Convos | Correct | NI | Leads | Floor |",
           "|---|---|---|---|---|---|---|---|---|---|"]
    for email, c in sorted(res["callers"].items(), key=lambda kv: -kv[1]["dials"]):
        if not any(c[k] for k in ("dials", "correct_numbers", "leads", "not_interested")):
            continue
        nd = len(c["days"]) or 1
        floor = bench["dials_floor_per_caller"] * len(c["days"])
        out.append(f"| {res['names'].get(email, email)} | {len(c['days'])} | {c['dials']} | "
                   f"{c['dials'] / nd:.0f} | {pct(c['answered'], c['dials'])} | {c['conversations']} | "
                   f"{c['correct_numbers']} | {c['not_interested']} | {c['leads']} | "
                   f"{'MET' if c['dials'] >= floor and floor else 'BELOW'} |")
    out += ["", "## Daily trend", "", "| Day | Dials | Answered | Convos | Correct | NI | Leads |",
            "|---|---|---|---|---|---|---|"]
    for day in sorted(res["daily"]):
        d = res["daily"][day]
        out.append(f"| {day} | {d['dials']} | {d['answered']} | {d['conversations']} | "
                   f"{d['correct_numbers']} | {d['not_interested']} | {d['leads']} |")
    dpc = f"{dials / a['correct_numbers']:.1f}" if a["correct_numbers"] else "n/a"
    out += ["", "## Funnel and pacing", "",
            f"- Dials per correct number: {dpc} (scored target ~{bench['dials_per_correct_scored']}, "
            f"blind ~{bench['dials_per_correct_blind']})",
            f"- Correct toward next deal: {a['correct_numbers']} / {bench['correct_numbers_per_deal']}",
            f"- Leads: {a['leads']} (baseline {lead_lo}-{lead_hi}/caller/day)  ->  "
            f"~{round(a['leads'] * take)} appointments at {take * 100:.0f}% take  ->  "
            f"{a['leads'] / c_hi:.1f}-{a['leads'] / c_lo:.1f} contracts at 1 per {c_lo}-{c_hi} leads"]
    return "\n".join(out)


def write_csv(res: dict, path: Path) -> None:
    fields = ["caller", "days", "dials", "answered", "conversations",
              "meaningful_conversations", "correct_numbers", "wrong_numbers",
              "not_interested", "leads", "appointments", "talk_seconds", "sms_sent"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(fields)
        for email, c in sorted(res["callers"].items(), key=lambda kv: -kv[1]["dials"]):
            w.writerow([res["names"].get(email, email), len(c["days"])] +
                       [c[f] for f in fields[2:]])


def write_xlsx(res: dict, bench: dict, path: Path) -> bool:
    try:
        from openpyxl import Workbook
    except ImportError:
        log("openpyxl not installed; skipping Excel (pip install openpyxl)")
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "By Caller"
    ws.append(["Caller", "Days", "Dials", "Answered", "Convos", "Correct", "NI", "Leads", "Talk"])
    for email, c in sorted(res["callers"].items(), key=lambda kv: -kv[1]["dials"]):
        ws.append([res["names"].get(email, email), len(c["days"]), c["dials"], c["answered"],
                   c["conversations"], c["correct_numbers"], c["not_interested"], c["leads"],
                   fmt_hms(c["talk_seconds"])])
    ws2 = wb.create_sheet("Daily")
    ws2.append(["Day", "Dials", "Answered", "Convos", "Correct", "NI", "Leads"])
    for day in sorted(res["daily"]):
        d = res["daily"][day]
        ws2.append([day, d["dials"], d["answered"], d["conversations"],
                    d["correct_numbers"], d["not_interested"], d["leads"]])
    if res.get("records_detail"):
        wsr = wb.create_sheet("Records Detail")
        wsr.append(["Address", "Current status", "Dials", "Answered", "Talk", "Callers",
                    "Status set to", "Set by", "Lead", "Last activity"])
        for d in res["records_detail"]:
            wsr.append([d["address"], d["status"], d["dials"], d["answered"],
                        fmt_hms(d["talk_seconds"]), ", ".join(sorted(d["callers"])),
                        d["status_set_to"], d["status_set_by"],
                        "YES" if d["is_lead"] else "", d["last_activity"]])
    ws3 = wb.create_sheet("Benchmarks")
    for k, v in bench.items():
        ws3.append([k, json.dumps(v) if isinstance(v, (list, dict)) else v])
    wb.save(path)
    return True


def post_slack(res: dict, webhook: str) -> None:
    a = res["account_totals"]
    text = (f"KPI Report {res['from']} to {res['to']}: {a['dials']} dials, "
            f"{a['conversations']} conversations, {a['correct_numbers']} correct numbers, "
            f"{a['not_interested']} not interested, {a['leads']} leads, "
            f"talk {fmt_hms(a['talk_seconds'])}")
    r = urllib.request.Request(webhook, data=json.dumps({"text": text}).encode(),
                               headers={"content-type": "application/json"}, method="POST")
    with urllib.request.urlopen(r, timeout=15):
        pass
    log("Slack digest posted")


def main() -> int:
    ap = argparse.ArgumentParser(description="DataSift KPI puller (kpi-engine skill)")
    ap.add_argument("--from", dest="day_from")
    ap.add_argument("--to", dest="day_to")
    ap.add_argument("--days", type=int, help="trailing N days ending today")
    ap.add_argument("--tz", default="America/New_York")
    ap.add_argument("--xlsx", action="store_true")
    ap.add_argument("--detail", action="store_true",
                    help="also write a record-level detail CSV (one row per worked record)")
    ap.add_argument("--slack", metavar="WEBHOOK_URL")
    ap.add_argument("--show-benchmarks", action="store_true")
    args = ap.parse_args()

    bench = load_benchmarks()
    if args.show_benchmarks:
        print(json.dumps(bench, indent=2))
        return 0
    tz = ZoneInfo(args.tz)
    today = datetime.datetime.now(tz).date()
    if args.days:
        day_from = (today - datetime.timedelta(days=args.days - 1)).isoformat()
        day_to = today.isoformat()
    elif args.day_from:
        day_from, day_to = args.day_from, args.day_to or today.isoformat()
    else:
        day_from = day_to = today.isoformat()

    token = get_token()
    res = pull(token, day_from, day_to, tz, bench)
    md = render_md(res, bench)
    print("\n" + md)
    stem = f"kpi_{day_from}_{day_to}"
    (HERE / f"{stem}.md").write_text(md, encoding="utf-8")
    write_csv(res, HERE / f"{stem}.csv")
    log(f"wrote {stem}.md and {stem}.csv")
    if args.detail:
        with open(HERE / f"{stem}_records.csv", "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["address", "current_status", "dials", "answered", "talk",
                        "callers", "status_set_to", "status_set_by", "is_lead", "last_activity"])
            for d in res["records_detail"]:
                w.writerow([d["address"], d["status"], d["dials"], d["answered"],
                            fmt_hms(d["talk_seconds"]), ", ".join(sorted(d["callers"])),
                            d["status_set_to"], d["status_set_by"],
                            "YES" if d["is_lead"] else "", d["last_activity"]])
        log(f"wrote {stem}_records.csv ({len(res['records_detail'])} records)")
    if args.xlsx and write_xlsx(res, bench, HERE / f"{stem}.xlsx"):
        log(f"wrote {stem}.xlsx")
    if args.slack:
        post_slack(res, args.slack)
    return 0


if __name__ == "__main__":
    sys.exit(main())
