"""Generate 7-sheet Market Research Excel workbook from extracted JSON data.

Takes the JSON output from extract_market_finder.py and produces a
professionally formatted Excel workbook following the SKILL.md specification.

Usage:
    python generate_market_excel.py --input market_finder_Tennessee_Knox.json
    python generate_market_excel.py --input data.json --fred-dom 71 --output Knox_County_TN_Market_Research.xlsx

Works in both Claude Code CLI and Co-Work (only requires openpyxl).
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ── Style Definitions ─────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13)
SUBSECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MONEY_FORMAT = '#,##0'
CURRENCY_FORMAT = '$#,##0'
PERCENT_FORMAT = '0.0%'


def _apply_header_row(ws, row_num: int, headers: list[str]):
    """Apply header styling to a row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_border(ws, row_num: int, col_count: int):
    """Apply thin border to a data row."""
    for col in range(1, col_count + 1):
        ws.cell(row=row_num, column=col).border = THIN_BORDER


def _auto_fit_columns(ws, min_width: int = 12, max_width: int = 40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ── Scoring ───────────────────────────────────────────────────────────

def _wholesaling_score(inv_trans: int, dom: int, home_value: float,
                       fred_dom: int) -> str:
    """Calculate wholesaling star rating per SKILL.md rubric."""
    dom_diff = dom - fred_dom if dom and fred_dom else 0

    if (inv_trans and inv_trans >= 30 and dom_diff <= -15
            and home_value and home_value < 350000):
        return "★★★★★"
    if (inv_trans and inv_trans >= 20 and dom_diff < 0
            and home_value and home_value < 400000):
        return "★★★★☆"
    if inv_trans and inv_trans >= 30 and dom_diff < 0:
        return "★★★★☆"  # Volume overrides price
    if inv_trans and inv_trans >= 10 and dom_diff <= 0:
        return "★★★☆☆"
    if inv_trans and inv_trans >= 5:
        return "★★☆☆☆"
    return "★☆☆☆☆"


def _supply_months(homes_on_market: int, homes_sold: int) -> float | None:
    """Calculate months of supply."""
    if not homes_sold or homes_sold == 0:
        return None
    return round(homes_on_market / homes_sold, 1)


def _spread_pct(sale_price: float, home_value: float) -> float | None:
    """Calculate price spread percentage."""
    if not home_value or home_value == 0:
        return None
    return round((sale_price - home_value) / home_value, 4)


# ── Sheet Builders ────────────────────────────────────────────────────

def _build_executive_summary(wb, data: dict, fred_dom: int):
    """Sheet 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"

    county = data.get("county", "Unknown")
    state = data.get("state", "Unknown")
    summary = data.get("summary", {})

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"{county.upper()} COUNTY, {state.upper()} — MARKET RESEARCH REPORT"
    title_cell.font = Font(bold=True, size=16)

    ws["A2"] = f"Generated: {datetime.now():%B %d, %Y}"
    ws["A3"] = "Data Source: REI Sift Market Finder + Public Data Sources"

    # Section A: County Overview
    row = 5
    ws.cell(row=row, column=1, value="COUNTY OVERVIEW").font = SECTION_FONT
    row += 1
    _apply_header_row(ws, row, ["Metric", "Value", "Notes"])
    row += 1

    overview_metrics = [
        ("Median Home Value", summary.get("median home value", "N/A"), "REI Sift data"),
        ("Homes on Market", summary.get("homes on market", "N/A"), "REI Sift data"),
        ("Mo. Investor Transactions", summary.get("mo investor transactions", "N/A"), "6-month average"),
        ("Homes Sold Last Month", summary.get("homes sold last month", "N/A"), "REI Sift data"),
        ("Market Rent", summary.get("market_rent", summary.get("market rent", "N/A")), "REI Sift data"),
        ("Gross Rental Yield", summary.get("gross_rental_yield", summary.get("gross rental yield", "N/A")), "REI Sift data"),
        ("Homeownership Rate", summary.get("homeownership_rate", summary.get("homeownership rate", "N/A")), "REI Sift / Census"),
        ("National Days on Market", str(fred_dom), f"FRED {datetime.now():%b %Y}"),
    ]

    for metric, value, notes in overview_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=notes)
        _apply_border(ws, row, 3)
        row += 1

    # Section B: Top 5 ZIP Codes
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 ZIP CODES FOR WHOLESALING").font = SECTION_FONT
    row += 1
    _apply_header_row(ws, row, ["Rank", "ZIP Code", "6-Mo Inv Trans", "Median Home Value", "Days on Market"])
    row += 1

    zip_data = sorted(
        data.get("zip_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )
    for i, z in enumerate(zip_data[:5], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=z.get("zip_code", ""))
        ws.cell(row=row, column=3, value=z.get("total_inv_trans_6mo"))
        ws.cell(row=row, column=4, value=z.get("median_home_value_raw", ""))
        ws.cell(row=row, column=5, value=z.get("median_days_on_market"))
        _apply_border(ws, row, 5)
        row += 1

    # Section C: Top 5 Neighborhoods
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 NEIGHBORHOODS FOR WHOLESALING").font = SECTION_FONT
    row += 1
    _apply_header_row(ws, row, ["Rank", "Neighborhood", "6-Mo Inv Trans", "Median Home Value", "Days on Market"])
    row += 1

    nbr_data = sorted(
        data.get("neighborhood_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )
    for i, n in enumerate(nbr_data[:5], 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=n.get("neighborhood", ""))
        ws.cell(row=row, column=3, value=n.get("total_inv_trans_6mo"))
        ws.cell(row=row, column=4, value=n.get("median_home_value_raw", ""))
        ws.cell(row=row, column=5, value=n.get("median_days_on_market"))
        _apply_border(ws, row, 5)
        row += 1

    _auto_fit_columns(ws)


def _build_zip_analysis(wb, data: dict, fred_dom: int):
    """Sheet 2: ZIP Code Analysis — all ZIPs with calculated columns."""
    ws = wb.create_sheet("ZIP Code Analysis")

    headers = [
        "ZIP Code", "6-Mo Inv Trans", "Homes on Market", "Homes Sold/Mo",
        "Median DOM", "DOM vs National", "Median Home Value", "Median Sale Price",
        "Spread %", "Supply Months", "Wholesaling Score",
    ]
    _apply_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    zip_data = sorted(
        data.get("zip_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )

    for i, z in enumerate(zip_data, 2):
        inv = z.get("total_inv_trans_6mo") or 0
        dom = z.get("median_days_on_market") or 0
        hom = z.get("homes_on_market") or 0
        sold = z.get("homes_sold_last_month") or 0
        value = z.get("median_home_value") or 0
        sale = z.get("median_sale_price") or 0

        dom_diff = dom - fred_dom if dom else None
        spread = _spread_pct(sale, value)
        supply = _supply_months(hom, sold)
        score = _wholesaling_score(inv, dom, value, fred_dom)

        ws.cell(row=i, column=1, value=z.get("zip_code", ""))
        ws.cell(row=i, column=2, value=inv)
        ws.cell(row=i, column=3, value=hom)
        ws.cell(row=i, column=4, value=sold)
        ws.cell(row=i, column=5, value=dom)

        dom_cell = ws.cell(row=i, column=6, value=dom_diff)
        if dom_diff is not None:
            if dom_diff < 0:
                dom_cell.fill = GREEN_FILL
            elif dom_diff <= fred_dom * 0.15:
                dom_cell.fill = YELLOW_FILL
            else:
                dom_cell.fill = RED_FILL

        ws.cell(row=i, column=7, value=z.get("median_home_value_raw", ""))
        ws.cell(row=i, column=8, value=z.get("median_sale_price_raw", ""))

        spread_cell = ws.cell(row=i, column=9)
        if spread is not None:
            spread_cell.value = spread
            spread_cell.number_format = '0.0%'
            if spread < 0:
                spread_cell.fill = GREEN_FILL

        supply_cell = ws.cell(row=i, column=10, value=supply)
        if supply is not None:
            if supply < 3:
                supply_cell.fill = GREEN_FILL
            elif supply <= 6:
                supply_cell.fill = YELLOW_FILL
            else:
                supply_cell.fill = RED_FILL

        ws.cell(row=i, column=11, value=score)

        # Bold Tier 1 rows (10+ monthly inv trans)
        if inv >= 60:  # 60 in 6 months ≈ 10/month
            for col in range(1, 12):
                ws.cell(row=i, column=col).font = Font(bold=True)

        _apply_border(ws, i, 11)

    _auto_fit_columns(ws)


def _build_neighborhood_analysis(wb, data: dict, fred_dom: int):
    """Sheet 3: Neighborhood Analysis — same structure as ZIP analysis."""
    ws = wb.create_sheet("Neighborhood Analysis")

    headers = [
        "Neighborhood", "6-Mo Inv Trans", "Homes on Market", "Homes Sold/Mo",
        "Median DOM", "DOM vs National", "Median Home Value", "Median Sale Price",
        "Spread %", "Supply Months", "Wholesaling Score",
    ]
    _apply_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    nbr_data = sorted(
        data.get("neighborhood_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )

    for i, n in enumerate(nbr_data, 2):
        inv = n.get("total_inv_trans_6mo") or 0
        dom = n.get("median_days_on_market") or 0
        hom = n.get("homes_on_market") or 0
        sold = n.get("homes_sold_last_month") or 0
        value = n.get("median_home_value") or 0
        sale = n.get("median_sale_price") or 0

        dom_diff = dom - fred_dom if dom else None
        spread = _spread_pct(sale, value)
        supply = _supply_months(hom, sold)
        score = _wholesaling_score(inv, dom, value, fred_dom)

        ws.cell(row=i, column=1, value=n.get("neighborhood", ""))
        ws.cell(row=i, column=2, value=inv)
        ws.cell(row=i, column=3, value=hom)
        ws.cell(row=i, column=4, value=sold)
        ws.cell(row=i, column=5, value=dom)

        dom_cell = ws.cell(row=i, column=6, value=dom_diff)
        if dom_diff is not None:
            if dom_diff < 0:
                dom_cell.fill = GREEN_FILL
            elif dom_diff <= fred_dom * 0.15:
                dom_cell.fill = YELLOW_FILL
            else:
                dom_cell.fill = RED_FILL

        ws.cell(row=i, column=7, value=n.get("median_home_value_raw", ""))
        ws.cell(row=i, column=8, value=n.get("median_sale_price_raw", ""))

        spread_cell = ws.cell(row=i, column=9)
        if spread is not None:
            spread_cell.value = spread
            spread_cell.number_format = '0.0%'
            if spread < 0:
                spread_cell.fill = GREEN_FILL

        supply_cell = ws.cell(row=i, column=10, value=supply)
        if supply is not None:
            if supply < 3:
                supply_cell.fill = GREEN_FILL
            elif supply <= 6:
                supply_cell.fill = YELLOW_FILL
            else:
                supply_cell.fill = RED_FILL

        ws.cell(row=i, column=11, value=score)

        if inv >= 60:
            for col in range(1, 12):
                ws.cell(row=i, column=col).font = Font(bold=True)

        _apply_border(ws, i, 11)

    _auto_fit_columns(ws)


def _build_economic_indicators(wb, data: dict):
    """Sheet 4: Economic Indicators — placeholder for public data blending."""
    ws = wb.create_sheet("Economic Indicators")

    county = data.get("county", "Unknown")
    state = data.get("state", "Unknown")

    ws["A1"] = f"ECONOMIC INDICATORS — {county.upper()} COUNTY, {state.upper()}"
    ws["A1"].font = SECTION_FONT

    ws["A3"] = "Section A: Employment Data (BLS)"
    ws["A3"].font = SUBSECTION_FONT
    _apply_header_row(ws, 4, ["Metric", "Value", "Trend"])
    for i, metric in enumerate([
        "Civilian Labor Force", "Employment", "Unemployment",
        "Unemployment Rate", "Total Nonfarm Jobs",
    ], 5):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value="[To be populated from BLS]")
        _apply_border(ws, i, 3)

    ws["A11"] = "Section B: Demographic Data (Census)"
    ws["A11"].font = SUBSECTION_FONT
    _apply_header_row(ws, 12, ["Metric", "Value", "Notes"])
    for i, metric in enumerate([
        "Population", "Population Growth Rate", "Median Age",
        "Median Household Income", "Per Capita Income", "Poverty Rate",
        "Bachelor's Degree or Higher", "Owner-Occupied Housing",
        "Median Home Value (Census)",
    ], 13):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value="[To be populated from Census]")
        _apply_border(ws, i, 3)

    ws["A23"] = "Section C: Housing Market (Redfin)"
    ws["A23"].font = SUBSECTION_FONT
    _apply_header_row(ws, 24, ["Metric", "Value", "YoY Change"])
    for i, metric in enumerate([
        "Median Sale Price", "Median Price/Sq Ft", "Homes Sold",
        "Median Days on Market", "Sale-to-List Price",
        "Homes Above List Price", "Homes with Price Drops",
    ], 25):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value="[To be populated from Redfin]")
        _apply_border(ws, i, 3)

    _auto_fit_columns(ws)


def _build_crime_safety(wb, data: dict):
    """Sheet 5: Crime & Safety — placeholder for crime data."""
    ws = wb.create_sheet("Crime & Safety")

    county = data.get("county", "Unknown")
    state = data.get("state", "Unknown")

    ws["A1"] = f"CRIME & SAFETY — {county.upper()} COUNTY, {state.upper()}"
    ws["A1"].font = SECTION_FONT

    ws["A3"] = "Section A: Crime Statistics"
    ws["A3"].font = SUBSECTION_FONT
    _apply_header_row(ws, 4, ["Crime Type", "Prior Year", "Current Year", "Change"])
    for i, crime in enumerate([
        "Murders", "Non-Fatal Shootings", "Robberies",
        "Motor Vehicle Thefts", "Car Burglaries", "Aggravated Assaults",
    ], 5):
        ws.cell(row=i, column=1, value=crime)
        ws.cell(row=i, column=2, value="[Data needed]")
        _apply_border(ws, i, 4)

    ws["A12"] = "Section B: Safety Assessment by Area"
    ws["A12"].font = SUBSECTION_FONT
    _apply_header_row(ws, 13, ["Area", "Safety Rating", "Notes"])
    ws.cell(row=14, column=1, value="[Populate with local area assessments]")
    _apply_border(ws, 14, 3)

    _auto_fit_columns(ws)


def _build_recommendations(wb, data: dict, fred_dom: int):
    """Sheet 6: Investment Recommendations."""
    ws = wb.create_sheet("Investment Recommendations")

    county = data.get("county", "Unknown")

    ws["A1"] = f"INVESTMENT RECOMMENDATIONS — {county.upper()} COUNTY"
    ws["A1"].font = SECTION_FONT

    # Tier 1 ZIPs
    zip_data = sorted(
        data.get("zip_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )

    ws["A3"] = "Section A: Tier 1 — Highest Priority ZIP Codes"
    ws["A3"].font = SUBSECTION_FONT
    _apply_header_row(ws, 4, [
        "ZIP Code", "Inv Trans", "Median Value", "DOM", "Supply Months", "Rationale",
    ])

    row = 5
    for z in zip_data:
        inv = z.get("total_inv_trans_6mo") or 0
        dom = z.get("median_days_on_market") or 0
        hom = z.get("homes_on_market") or 0
        sold = z.get("homes_sold_last_month") or 0
        value = z.get("median_home_value") or 0
        supply = _supply_months(hom, sold)

        # Tier 1: 10+ monthly transactions (60+ in 6 months)
        if inv < 60:
            continue
        dom_label = "below avg" if dom < fred_dom else "at avg"
        rationale = f"{inv} transactions, {dom_label} DOM, moderate prices"

        ws.cell(row=row, column=1, value=z.get("zip_code", ""))
        ws.cell(row=row, column=2, value=inv)
        ws.cell(row=row, column=3, value=z.get("median_home_value_raw", ""))
        ws.cell(row=row, column=4, value=dom)
        ws.cell(row=row, column=5, value=supply)
        ws.cell(row=row, column=6, value=rationale)
        _apply_border(ws, row, 6)
        row += 1

    if row == 5:  # No Tier 1 ZIPs found, lower threshold
        for z in zip_data[:3]:
            inv = z.get("total_inv_trans_6mo") or 0
            dom = z.get("median_days_on_market") or 0
            hom = z.get("homes_on_market") or 0
            sold = z.get("homes_sold_last_month") or 0
            supply = _supply_months(hom, sold)
            dom_label = "below avg" if dom < fred_dom else "at avg"
            rationale = f"Top activity ({inv} trans), {dom_label} DOM"

            ws.cell(row=row, column=1, value=z.get("zip_code", ""))
            ws.cell(row=row, column=2, value=inv)
            ws.cell(row=row, column=3, value=z.get("median_home_value_raw", ""))
            ws.cell(row=row, column=4, value=dom)
            ws.cell(row=row, column=5, value=supply)
            ws.cell(row=row, column=6, value=rationale)
            _apply_border(ws, row, 6)
            row += 1

    # Tier 1 Neighborhoods
    row += 1
    ws.cell(row=row, column=1, value="Section B: Tier 1 — Highest Priority Neighborhoods").font = SUBSECTION_FONT
    row += 1
    _apply_header_row(ws, row, [
        "Neighborhood", "Inv Trans", "Median Value", "DOM", "Supply Months", "Rationale",
    ])
    row += 1

    nbr_data = sorted(
        data.get("neighborhood_data", []),
        key=lambda x: x.get("total_inv_trans_6mo") or 0,
        reverse=True,
    )

    for n in nbr_data[:5]:
        inv = n.get("total_inv_trans_6mo") or 0
        dom = n.get("median_days_on_market") or 0
        hom = n.get("homes_on_market") or 0
        sold = n.get("homes_sold_last_month") or 0
        supply = _supply_months(hom, sold)
        dom_label = "below avg" if dom < fred_dom else "at avg"

        ws.cell(row=row, column=1, value=n.get("neighborhood", ""))
        ws.cell(row=row, column=2, value=inv)
        ws.cell(row=row, column=3, value=n.get("median_home_value_raw", ""))
        ws.cell(row=row, column=4, value=dom)
        ws.cell(row=row, column=5, value=supply)
        ws.cell(row=row, column=6, value=f"Top neighborhood, {inv} trans, {dom_label} DOM")
        _apply_border(ws, row, 6)
        row += 1

    # Market Timing
    row += 1
    ws.cell(row=row, column=1, value="Section D: Market Timing Considerations").font = SUBSECTION_FONT
    row += 1
    _apply_header_row(ws, row, ["Factor", "Current Status", "Implication"])
    row += 1
    timing = [
        ("National DOM", f"{fred_dom} days (FRED)", "Benchmark for market velocity comparison"),
        ("Supply", "[Calculate from data]", "Under 3 mo = seller's market"),
        ("Investor Activity", f"{len(zip_data)} active ZIP codes", "Market has proven buyer demand"),
    ]
    for factor, status, impl in timing:
        ws.cell(row=row, column=1, value=factor)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=impl)
        _apply_border(ws, row, 3)
        row += 1

    _auto_fit_columns(ws)


def _build_data_sources(wb, data: dict, fred_dom: int):
    """Sheet 7: Data Sources — methodology and disclaimers."""
    ws = wb.create_sheet("Data Sources")

    ws["A1"] = "DATA SOURCES & METHODOLOGY"
    ws["A1"].font = SECTION_FONT

    # Section A: Sources
    ws["A3"] = "Section A: Primary Data Sources"
    ws["A3"].font = SUBSECTION_FONT
    _apply_header_row(ws, 4, ["Source", "Data Type", "Date Retrieved", "URL/Notes"])

    sources = [
        ("REI Sift Market Finder", "Investor transactions, home values, DOM",
         datetime.now().strftime("%Y-%m-%d"), "app.reisift.io/market-finder"),
        ("FRED (St. Louis Fed)", "National Days on Market baseline",
         datetime.now().strftime("%Y-%m-%d"), "fred.stlouisfed.org/series/MEDDAYONMARUS"),
        ("U.S. Census Bureau", "Demographics, population, income",
         "[If used]", "census.gov/quickfacts"),
        ("Bureau of Labor Statistics", "Employment, unemployment",
         "[If used]", "bls.gov/eag"),
        ("Redfin", "Housing market trends, prices",
         "[If used]", "redfin.com"),
    ]
    for i, (src, dtype, date, url) in enumerate(sources, 5):
        ws.cell(row=i, column=1, value=src)
        ws.cell(row=i, column=2, value=dtype)
        ws.cell(row=i, column=3, value=date)
        ws.cell(row=i, column=4, value=url)
        _apply_border(ws, i, 4)

    # Section B: Methodology
    row = 11
    ws.cell(row=row, column=1, value="Section B: Methodology").font = SUBSECTION_FONT
    row += 1
    _apply_header_row(ws, row, ["Analysis Component", "Description"])
    row += 1

    methods = [
        ("Wholesaling Score",
         "Composite: investor transactions (primary), DOM vs national (secondary), price accessibility (tertiary)"),
        ("DOM vs National",
         f"Comparison to FRED national median ({fred_dom} days as of {datetime.now():%b %Y})"),
        ("Price Spread",
         "(Median Sale Price - Median Home Value) / Median Home Value"),
        ("Supply Months",
         "Homes on Market / Homes Sold Last Month — under 3 = seller's, 3-6 = balanced, over 6 = buyer's"),
        ("Tier Classification",
         "Based on investor activity volume and market fundamentals"),
    ]
    for component, desc in methods:
        ws.cell(row=row, column=1, value=component)
        ws.cell(row=row, column=2, value=desc)
        _apply_border(ws, row, 2)
        row += 1

    # Section C: Disclaimers
    row += 1
    ws.cell(row=row, column=1, value="Section C: Disclaimers").font = SUBSECTION_FONT
    row += 1

    disclaimers = [
        "Data is current as of retrieval date and subject to change",
        "REI Sift data represents proprietary investor transaction tracking",
        "Crime statistics may be preliminary and pending audit",
        "Investment recommendations are for informational purposes only",
        "Always conduct independent due diligence before investing",
    ]
    for disclaimer in disclaimers:
        ws.cell(row=row, column=1, value=f"• {disclaimer}")
        row += 1

    _auto_fit_columns(ws)


# ── Main Generator ────────────────────────────────────────────────────

def generate_market_report(
    data: dict,
    fred_dom: int = 71,
    output_path: str | None = None,
) -> str:
    """Generate the 7-sheet Market Research Excel workbook.

    Args:
        data: Extracted market data dict (from extract_market_finder.py)
        fred_dom: FRED national Days on Market baseline
        output_path: Output file path (auto-generated if not provided)

    Returns:
        Path to the generated Excel file
    """
    wb = openpyxl.Workbook()

    _build_executive_summary(wb, data, fred_dom)
    _build_zip_analysis(wb, data, fred_dom)
    _build_neighborhood_analysis(wb, data, fred_dom)
    _build_economic_indicators(wb, data)
    _build_crime_safety(wb, data)
    _build_recommendations(wb, data, fred_dom)
    _build_data_sources(wb, data, fred_dom)

    if not output_path:
        county = data.get("county", "Unknown")
        state = data.get("state", "Unknown")
        # State abbreviation mapping (common states)
        state_abbrev = {
            "Tennessee": "TN", "Texas": "TX", "Florida": "FL",
            "Georgia": "GA", "North Carolina": "NC", "Ohio": "OH",
            "California": "CA",
        }
        st = state_abbrev.get(state, state[:2].upper())
        output_path = f"{county}_County_{st}_Market_Research.xlsx"

    wb.save(output_path)
    print(f"Generated: {output_path}")
    return output_path


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate Market Research Excel workbook from extracted JSON"
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to JSON file from extract_market_finder.py",
    )
    parser.add_argument(
        "--fred-dom", type=int, default=71,
        help="FRED national Days on Market baseline (default: 71)",
    )
    parser.add_argument(
        "--output",
        help="Output Excel file path (auto-generated if not provided)",
    )

    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    generate_market_report(data, fred_dom=args.fred_dom, output_path=args.output)


if __name__ == "__main__":
    main()
