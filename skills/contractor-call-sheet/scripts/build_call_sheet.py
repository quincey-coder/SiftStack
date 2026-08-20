#!/usr/bin/env python3
"""
build_call_sheet.py: turn a vetted contractor/vendor directory into a printable call sheet.

Usage:
    python build_call_sheet.py <directory.xlsx> <call_sheet.html> [--all]

Reads the Directory tab produced by the `vendor-directory-builder` skill (or any
spreadsheet with company + phone columns), and writes a clean, printable HTML call sheet:
top picks grouped by trade, cross-validated "call-first" providers banner'd at the top, and
blank Status / Notes / Next-step columns to work down.

  --all   include every provider, not just the starred top picks.

Column detection is fuzzy, so it also works on a generic spreadsheet whose headers merely
*contain* words like "company", "phone", "trade/category", "why/notes". Requires openpyxl.
"""
import sys
from datetime import date
from html import escape
import openpyxl


def find_col(headers, *needles, exclude=()):
    """Return the index of the first header containing any needle (and no exclude term)."""
    for i, h in enumerate(headers):
        hl = str(h or "").lower()
        if any(n in hl for n in needles) and not any(x in hl for x in exclude):
            return i
    return None


def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Directory"] if "Directory" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}, ""
    headers = [str(c or "") for c in rows[0]]
    idx = {
        "category":   find_col(headers, "category", "trade"),
        "company":    find_col(headers, "company", "provider", "name", exclude=("contact",)),
        "phone":      find_col(headers, "phone", exclude=("status",)),
        "serves":     find_col(headers, "serves"),
        "why":        find_col(headers, "why", "strength", "fit"),
        "cautions":   find_col(headers, "caution", "verify"),
        "source":     find_col(headers, "signal", "source", exclude=("rating",)),
        "confidence": find_col(headers, "confidence"),
        "top_pick":   find_col(headers, "top pick"),
        "phone_status": find_col(headers, "phone status"),
    }
    geo = ""
    if idx["serves"] is not None:
        geo = headers[idx["serves"]].replace("Serves", "").strip()
    data = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
    return data, idx, geo


def get(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    return str(row[i] or "").strip()


def is_top_pick(row, idx):
    return get(row, idx, "top_pick") == "★"


def is_call_first(row, idx):
    src = get(row, idx, "source").lower()
    return "2×" in src or "2x" in src or "recommended 2" in src


def confirm_note(row, idx):
    serves = get(row, idx, "serves")
    if serves.lower().startswith("confirm"):
        return serves  # e.g. "Confirm Blount"
    return ""


def build_html(rows, idx, geo, include_all):
    picks = rows if include_all else [r for r in rows if is_top_pick(r, idx)]
    if not picks and not include_all:
        # no stars found: fall back to everything so the sheet isn't empty
        picks = rows

    # group by category, preserving first-seen order
    groups, order = {}, []
    for r in picks:
        cat = get(r, idx, "category") or "Uncategorized"
        if cat not in groups:
            groups[cat] = []; order.append(cat)
        groups[cat].append(r)

    call_first = [r for r in picks if is_call_first(r, idx)]

    market = f" - {escape(geo)}" if geo else ""
    scope = "All providers" if include_all else "Top picks"
    out = []
    out.append(f"""<!doctype html><html><head><meta charset="utf-8">
<title>Contractor Call Sheet{market}</title>
<style>
 * {{ box-sizing: border-box; }}
 body {{ font-family: Arial, Helvetica, sans-serif; color:#1a2233; margin:24px; }}
 h1 {{ font-size:20px; margin:0 0 2px; color:#1F3A5F; }}
 .sub {{ color:#667085; font-size:12px; margin-bottom:16px; }}
 .banner {{ background:#FBF3D5; border:1px solid #C9A227; border-radius:6px;
            padding:10px 12px; margin:0 0 16px; font-size:12.5px; }}
 .banner b {{ color:#8a6d1a; }}
 h2 {{ font-size:13px; color:#2E6E6A; margin:18px 0 6px; text-transform:uppercase;
       letter-spacing:.04em; border-bottom:2px solid #2E6E6A; padding-bottom:3px; }}
 table {{ width:100%; border-collapse:collapse; margin-bottom:6px; }}
 th, td {{ border:1px solid #d0d7de; padding:6px 8px; font-size:12px; vertical-align:top; text-align:left; }}
 th {{ background:#1F3A5F; color:#fff; font-size:11px; }}
 td.company {{ font-weight:bold; white-space:nowrap; }}
 td.phone {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
 .confirm {{ color:#B7791F; font-weight:bold; font-size:11px; }}
 .cf {{ color:#8a6d1a; font-weight:bold; }}
 .status {{ width:70px; }} .notes {{ width:150px; }} .next {{ width:110px; }}
 tr:nth-child(even) td {{ background:#f7f9fb; }}
 .foot {{ margin-top:18px; color:#667085; font-size:11px; }}
 @media print {{ body {{ margin:10px; }} h2 {{ page-break-after:avoid; }} tr {{ page-break-inside:avoid; }} }}
</style></head><body>""")
    out.append(f"<h1>Contractor Call Sheet{market}</h1>")
    out.append(f'<div class="sub">{scope} &middot; generated {date.today().isoformat()} &middot; work top-down; call the flagged names first</div>')

    if call_first:
        names = ", ".join(escape(get(r, idx, "company")) for r in call_first)
        out.append(f'<div class="banner">📞 <b>Call first (recommended by 2+ people):</b> {names}</div>')

    for cat in order:
        out.append(f"<h2>{escape(cat)}</h2>")
        out.append("<table><tr>"
                   "<th>Company</th><th>Phone</th><th>Why / who vouched</th>"
                   "<th>Confirm on call</th><th class='status'>Status</th>"
                   "<th class='notes'>Notes</th><th class='next'>Next step</th></tr>")
        for r in groups[cat]:
            company = escape(get(r, idx, "company"))
            if is_call_first(r, idx):
                company = f'<span class="cf">★ {company}</span>'
            phone = escape(get(r, idx, "phone")) or '<span class="confirm">no #, get it first</span>'
            why = escape(get(r, idx, "why"))
            conf = get(r, idx, "confidence")
            if conf:
                why += f' <span style="color:#667085">({escape(conf)})</span>'
            confirm = escape(confirm_note(r, idx))
            confirm_html = f'<span class="confirm">{confirm}</span>' if confirm else ""
            out.append("<tr>"
                       f"<td class='company'>{company}</td>"
                       f"<td class='phone'>{phone}</td>"
                       f"<td>{why}</td>"
                       f"<td>{confirm_html}</td>"
                       "<td></td><td></td><td></td></tr>")
        out.append("</table>")

    out.append('<div class="foot">Vetting-call questions: serves my area? · work with investors / volume pricing? · '
               'realistic timeline, bid held in writing? · pulls permits (in the price)? · '
               'can you send a COI + license #? · payment terms (draws vs. up front)?</div>')
    out.append("</body></html>")
    return "\n".join(out)


def main():
    args = [a for a in sys.argv[1:] if a != "--all"]
    include_all = "--all" in sys.argv
    if len(args) != 2:
        print("Usage: python build_call_sheet.py <directory.xlsx> <call_sheet.html> [--all]")
        sys.exit(1)
    src, out_path = args
    rows, idx, geo = load_rows(src)
    if not rows:
        print("No rows found in the directory file."); sys.exit(1)
    html = build_html(rows, idx, geo, include_all)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    n_pick = len(rows) if include_all else sum(1 for r in rows if is_top_pick(r, idx))
    print(f"Saved {out_path}  ({n_pick} providers on the sheet"
          f"{' - top picks' if not include_all else ' - all'}"
          f"{', market: '+geo if geo else ''})")


if __name__ == "__main__":
    main()
