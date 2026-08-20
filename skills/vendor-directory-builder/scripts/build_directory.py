#!/usr/bin/env python3
"""
build_directory.py: turn a JSON config into a vetted vendor/contractor directory workbook.

Usage:
    python build_directory.py <config.json> <output.xlsx>

Produces up to four kinds of tab:
  1. "Directory"            : one row per provider (filterable, color-coded)
  2. "Top Picks by Category": best pick per category + a computed summary block
  3. "How to Vet + Method"  : methodology / vetting notes (from config "methodology")
  4. any number of reference tabs (from config "reference_tabs")

See assets/config_schema.md for the full config format, and assets/example_config.json
for a runnable example. Counts on the Top Picks tab are computed at build time and written
as values, so the file is correct the moment it's opened, no recalc step required.

Requires: openpyxl (preinstalled in this environment; `pip install openpyxl` otherwise).
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette -------------------------------------------------------------
NAVY="1F3A5F"; TEAL="2E6E6A"; GOLD="C9A227"; LTGOLD="FBF3D5"; LTGREY="F2F4F7"
WHITE="FFFFFF"; GREEN="1E7F3C"; AMBER="B7791F"; RED="9B2C2C"; GREY="6B7280"
BAND="4A5568"; BORDER="D0D7DE"
FONT="Arial"
NAMED={"navy":NAVY,"teal":TEAL,"gold":GOLD,"green":GREEN,"amber":AMBER,
       "red":RED,"redg":RED,"grey":GREY,"gray":GREY,"white":WHITE,"":"222222"}
_thin=Side(style="thin", color=BORDER)
BORDER_ALL=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)

# Directory columns. "Serves" header is filled in from config geo_label.
BASE_HEADERS=["Category","Company / Provider","Contact","Phone","Phone Status","Email",
              "Website","Service Area","Serves {geo}","Rating (source)",
              "License / Credential","BBB","Why (Strengths / Fit)",
              "Source / Signal","Cautions / Verify","Confidence","Top Pick"]
# JSON keys, in the same order as BASE_HEADERS:
FIELDS=["trade","company","contact","phone","phone_status","email","website",
        "service_area","serves_geo","rating","license","bbb","why","source",
        "cautions","confidence","top_pick"]
CENTER={5,9,17}                 # 1-indexed cols centered
WRAP={8,10,11,13,14,15,16}      # 1-indexed cols wrapped


def col(hexcode):
    """Resolve a color that may be a named color or a hex string."""
    if not hexcode:
        return "222222"
    return NAMED.get(str(hexcode).lower(), str(hexcode))


def conf_color(v):
    v=str(v)
    if v.startswith("High"): return GREEN
    if v.startswith("Med") or v.startswith("Medium"): return AMBER
    if v.startswith("Low"): return RED if "fit" not in v else GREY
    return "222222"


def serves_color(v):
    v=str(v).strip()
    if v.lower()=="yes": return GREEN
    if v.lower().startswith("confirm"): return AMBER
    if v.lower() in ("no","unverified","unconfirmed"): return RED
    return "222222"


def style_header(ws, ncols, row=1, fill=NAVY):
    for c in range(1, ncols+1):
        cell=ws.cell(row=row, column=c)
        cell.font=Font(name=FONT, bold=True, color=WHITE, size=10)
        cell.fill=PatternFill("solid", fgColor=fill)
        cell.alignment=Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border=BORDER_ALL


def build_directory_tab(wb, cfg):
    ws=wb.active; ws.title="Directory"
    ws.sheet_view.showGridLines=False
    geo=cfg.get("geo_label","target market")
    headers=[h.replace("{geo}", geo) for h in BASE_HEADERS]
    for ci,h in enumerate(headers,1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, len(headers))

    r=2
    for rec in cfg.get("records", []):
        vals=[rec.get(k,"") for k in FIELDS]
        top = str(vals[-1]).strip()=="★"
        for ci,val in enumerate(vals,1):
            cell=ws.cell(row=r, column=ci, value=val)
            color="222222"
            if ci==16: color=conf_color(val)
            elif ci==9: color=serves_color(val)
            cell.font=Font(name=FONT, size=9, color=color, bold=(ci in (2,9,17)))
            cell.alignment=Alignment(vertical="top", wrap_text=(ci in WRAP),
                                     horizontal=("center" if ci in CENTER else "left"))
            cell.border=BORDER_ALL
            if top:
                cell.fill=PatternFill("solid", fgColor=LTGOLD)
            elif r%2==0:
                cell.fill=PatternFill("solid", fgColor=LTGREY)
            if ci==17 and top:
                cell.font=Font(name=FONT, size=12, bold=True, color=GOLD)
        r+=1

    widths=[20,30,20,20,11,24,24,24,15,19,25,15,40,30,40,12,8]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="C2"
    if r>2:
        ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{r-1}"
    ws.row_dimensions[1].height=30
    for rr in range(2,r):
        ws.row_dimensions[rr].height=58
    return r-1  # last data row (or 1 if empty)


def build_top_picks_tab(wb, cfg, last_row):
    picks=cfg.get("top_picks", [])
    recs=cfg.get("records", [])
    geo=cfg.get("geo_label","target market")
    ws=wb.create_sheet("Top Picks by Category")
    ws.sheet_view.showGridLines=False
    t=ws.cell(row=1, column=1,
              value=f"Top Picks by Category: best verified, {geo}-serving option per category")
    t.font=Font(name=FONT, bold=True, size=13, color=NAVY); ws.merge_cells("A1:F1")
    s=ws.cell(row=2, column=1,
              value=cfg.get("top_picks_note",
                    "Starred (★) rows on the Directory tab. Confirm any credential/license "
                    "on the relevant board before a large-dollar job."))
    s.font=Font(name=FONT, italic=True, size=9, color="555555"); ws.merge_cells("A2:F2")

    hdr=4
    for ci,h in enumerate(["Category","Top Pick","Phone","Why it's the pick",
                           "Runner-up","Notes"],1):
        ws.cell(row=hdr, column=ci, value=h)
    style_header(ws, 6, row=hdr, fill=TEAL)
    rr=hdr+1
    for p in picks:
        row=[p.get("category",""),p.get("pick",""),p.get("phone",""),
             p.get("why",""),p.get("runner_up",""),p.get("note","")]
        for ci,val in enumerate(row,1):
            cell=ws.cell(row=rr, column=ci, value=val)
            cell.font=Font(name=FONT, size=9, bold=(ci==2), color="222222")
            cell.alignment=Alignment(vertical="top", wrap_text=(ci in (1,4,5,6)))
            cell.border=BORDER_ALL
            if rr%2==0: cell.fill=PatternFill("solid", fgColor=LTGREY)
        ws.row_dimensions[rr].height=46
        rr+=1
    for i,w in enumerate([24,30,20,44,40,30],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    # ---- summary block, computed as VALUES (no recalc needed) ----
    total=len(recs)
    stars=sum(1 for x in recs if str(x.get("top_pick","")).strip()=="★")
    serves=sum(1 for x in recs if str(x.get("serves_geo","")).strip().lower()=="yes")
    confirm=sum(1 for x in recs if str(x.get("serves_geo","")).strip().lower().startswith("confirm"))
    added=sum(1 for x in recs if str(x.get("source","")).strip().lower().startswith("added"))
    srow=rr+2
    ws.cell(row=srow, column=1, value="Directory at a glance").font=Font(
        name=FONT, bold=True, size=11, color=NAVY)
    stats=[("Total providers / leads", total),
           ("Top picks (★)", stars),
           (f"Confirmed serving {geo}", serves),
           (f"Metro-local, confirm coverage on call", confirm),
           ("Added from public sources", added)]
    for i,(label,val) in enumerate(stats):
        lr=srow+1+i
        ws.cell(row=lr, column=1, value=label).font=Font(name=FONT, size=10, color="222222")
        b=ws.cell(row=lr, column=2, value=val)
        b.font=Font(name=FONT, size=10, bold=True, color=TEAL)
        b.alignment=Alignment(horizontal="center")


def build_methodology_tab(wb, cfg):
    blocks=cfg.get("methodology")
    if not blocks:
        return
    ws=wb.create_sheet("How to Vet + Method")
    ws.column_dimensions["A"].width=4
    ws.column_dimensions["B"].width=112
    ws.sheet_view.showGridLines=False
    r=1
    for blk in blocks:
        typ=blk.get("type","p"); text=blk.get("text",""); color=col(blk.get("color"))
        if typ=="title":
            c=ws.cell(row=r, column=2, value=text)
            c.font=Font(name=FONT, bold=True, size=15, color=NAVY); r+=1
        elif typ=="subtitle":
            c=ws.cell(row=r, column=2, value=text)
            c.font=Font(name=FONT, italic=True, size=9, color="555555"); r+=2
        elif typ=="h1":
            c=ws.cell(row=r, column=2, value=text)
            c.font=Font(name=FONT, bold=True, size=13, color=WHITE)
            c.fill=PatternFill("solid", fgColor=NAVY)
            c.alignment=Alignment(vertical="center", indent=1)
            ws.row_dimensions[r].height=26; r+=2
        elif typ=="h2":
            c=ws.cell(row=r, column=2, value=text)
            c.font=Font(name=FONT, bold=True, size=11, color=TEAL); r+=1
        elif typ=="gap":
            r+=1
        else:  # paragraph
            c=ws.cell(row=r, column=2, value=text)
            c.font=Font(name=FONT, size=10, bold=blk.get("bold",False), color=color)
            c.alignment=Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height=15*max(1,(len(text)//110)+1); r+=1


def build_reference_tab(wb, spec):
    name=spec.get("name","Reference")[:31]
    ws=wb.create_sheet(name)
    ws.sheet_view.showGridLines=False
    headers=spec.get("headers", [])
    ncols=max(1,len(headers))
    r=1
    if spec.get("title"):
        c=ws.cell(row=r, column=1, value=spec["title"])
        c.font=Font(name=FONT, bold=True, size=13, color=NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols); r+=1
    if spec.get("subtitle"):
        c=ws.cell(row=r, column=1, value=spec["subtitle"])
        c.font=Font(name=FONT, italic=True, size=9, color="555555")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols); r+=1
    r+=1
    hdr=r
    for ci,h in enumerate(headers,1):
        ws.cell(row=hdr, column=ci, value=h)
    style_header(ws, ncols, row=hdr, fill=TEAL)
    r=hdr+1
    for row in spec.get("rows", []):
        if isinstance(row, dict) and "band" in row:
            c=ws.cell(row=r, column=1, value=row["band"])
            c.font=Font(name=FONT, bold=True, size=10, color=WHITE)
            c.fill=PatternFill("solid", fgColor=BAND)
            c.alignment=Alignment(indent=1, vertical="center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            ws.row_dimensions[r].height=18; r+=1; continue
        cells=row if isinstance(row, list) else [row.get(h,"") for h in headers]
        for ci,val in enumerate(cells,1):
            cell=ws.cell(row=r, column=ci, value=val)
            cell.font=Font(name=FONT, size=9, bold=(ci==1), color="222222")
            cell.alignment=Alignment(vertical="top", wrap_text=(ci!=1 or len(str(val))>30))
            cell.border=BORDER_ALL
            if r%2==0: cell.fill=PatternFill("solid", fgColor=LTGREY)
        ws.row_dimensions[r].height=spec.get("row_height",52); r+=1
    widths=spec.get("widths")
    if widths:
        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w
    if spec.get("footnote"):
        c=ws.cell(row=r+1, column=1, value=spec["footnote"])
        c.font=Font(name=FONT, italic=True, size=9, color="555555")
        c.alignment=Alignment(wrap_text=True)
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=ncols)


def main():
    if len(sys.argv)!=3:
        print("Usage: python build_directory.py <config.json> <output.xlsx>")
        sys.exit(1)
    cfg_path, out_path=sys.argv[1], sys.argv[2]
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg=json.load(f)

    wb=Workbook()
    last=build_directory_tab(wb, cfg)
    build_top_picks_tab(wb, cfg, last)
    # reference tabs sit before the methodology tab for logical flow
    for spec in cfg.get("reference_tabs", []):
        build_reference_tab(wb, spec)
    build_methodology_tab(wb, cfg)
    wb.save(out_path)
    print(f"Saved {out_path}  ({len(cfg.get('records', []))} providers, "
          f"{len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames)})")


if __name__=="__main__":
    main()
